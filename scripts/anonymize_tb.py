#!/usr/bin/env python3
"""Deterministic trial-balance anonymizer (GOLDEN CORPUS tooling).

Scrambles the IDENTIFYING text of a Romanian trial-balance export —
account labels ("Nume Cont" / "Denumire"), sheet titles, banner rows,
CUI / Reg-Com / CNP digit patterns — while preserving, byte-for-byte,
everything the deterministic engine consumes:

  · account CODES (the mapping key) — never touched;
  · every NUMERIC cell (all 8 columns of every pair) — never touched;
  · the file's own TOTALS ROW — never touched (the external
    conservation anchor must keep matching to the cent);
  · header rows / column labels — never touched (structure detection
    depends on their exact wording);
  · row order, sheet order, column layout — unchanged.

Determinism: seed = sha256(input bytes). Each distinct source string
gets its OWN substitution alphabet keyed by sha256(seed | string), so
the same label always scrambles identically within one file while
cross-string frequency analysis gets nothing to bite on. Diacritics are
folded to ASCII before substitution so the output survives any of the
parser's decode fallbacks (utf-8 / cp1250 / latin-1).

Residual-risk caveat (documented, not hidden): per-string substitution
preserves string LENGTH and repeated-letter patterns. For the public
OMFP account vocabulary that is irrelevant; for bespoke party names it
leaks shape only, never content. Files needing stronger guarantees
should not enter the corpus.

Formats: XLSX (all sheets) and delimited CSV/TXT. PDF is NOT written by
this tool — a PDF has no columns to key off, so its redaction is
lexicon-driven and lives in `scripts/pdf_scrambler.py`, which reuses
THIS module's seed rule, folding and substitution alphabet. `--verify`
on a PDF delegates there (see `verify_bytes`); `anonymize_bytes` still
refuses one, deliberately.

CLI:
  anonymize_tb.py INPUT -o OUTPUT [--self-check]
      Write the anonymized copy. --self-check re-parses BOTH files with
      the real engine parser and asserts every column-pair sum, every
      per-row numeric field, every account code and the totals-row
      anchor values are identical pre/post. Exit 1 on any mismatch.
  anonymize_tb.py --verify INPUT [INPUT ...]
      Corpus-replay mode: for each input, run the scramble transform
      IN MEMORY (nothing written) and assert the same invariants hold
      between the input and its scrambled image — proving the corpus
      input's numeric structure is anonymization-invariant. Exit 1 on
      any failure.

Stdlib + the engine's own parser deps (pandas / openpyxl). Python 3.9+.
"""
from __future__ import annotations

import argparse
import hashlib
import io
import random
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _find_repo_root() -> Path:
    here = Path(__file__).resolve().parent
    for candidate in [here, *here.parents][:6]:
        if (candidate / "pyproject.toml").is_file():
            return candidate
    return Path(__file__).resolve().parent.parent


REPO = _find_repo_root()
SRC = REPO / "src"
if SRC.is_dir() and str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from engine.country_packs.ro_romania import trial_balance_parser as tbp  # noqa: E402

_LOWER = "abcdefghijklmnopqrstuvwxyz"
_UPPER = _LOWER.upper()
_DIGITS = "0123456789"

# Number-shaped strings are NEVER scrambled (numeric structure is
# sacrosanct): optional sign, digits with ., , and space grouping.
_NUMBER_SHAPED_RE = re.compile(r"^\s*-?[\d.,\s ]*\d[\d.,\s ]*\s*$")

# Identifier patterns scrubbed inside SURVIVING text cells (cells that
# are not scrambled wholesale): Romanian CUI, Reg-Com number, CNP.
_ID_PATTERNS = (
    re.compile(r"(?i)\b(?:RO\s?|CUI\s*:?\s*|CIF\s*:?\s*)\d{2,10}\b"),
    re.compile(r"(?i)\bJ\s?\d{1,2}\s*/\s*\d{1,9}\s*/\s*\d{4}\b"),
    re.compile(r"\b\d{13}\b"),  # CNP
)


def content_seed(data: bytes) -> str:
    """The deterministic scramble seed — sha256 hex of the input bytes."""
    return hashlib.sha256(data or b"").hexdigest()


def _fold_ascii(text: str) -> str:
    """Fold diacritics to plain ASCII letters (ș→s, ă→a, ...) so the
    scrambled output is encodable under every parser decode fallback."""
    out = []
    for ch in text:
        decomposed = unicodedata.normalize("NFKD", ch)
        stripped = "".join(c for c in decomposed if not unicodedata.combining(c))
        out.append(stripped if stripped else ch)
    return "".join(out)


def scramble_text(seed_hex: str, text: str) -> str:
    """Deterministic per-string substitution. Letters map within case
    pools, digits within the digit pool, punctuation/whitespace pass
    through. Same (seed, string) → same output, always."""
    if not text:
        return text
    folded = _fold_ascii(text)
    key = hashlib.sha256(
        (seed_hex + "|" + folded).encode("utf-8", errors="replace")
    ).hexdigest()
    rng = random.Random(int(key[:16], 16))
    lower = list(_LOWER)
    upper = list(_UPPER)
    digits = list(_DIGITS)
    rng.shuffle(lower)
    rng.shuffle(upper)
    rng.shuffle(digits)
    lower_map = dict(zip(_LOWER, lower))
    upper_map = dict(zip(_UPPER, upper))
    digit_map = dict(zip(_DIGITS, digits))
    out = []
    for ch in folded:
        if ch in lower_map:
            out.append(lower_map[ch])
        elif ch in upper_map:
            out.append(upper_map[ch])
        elif ch in digit_map:
            out.append(digit_map[ch])
        else:
            out.append(ch)
    return "".join(out)


def _scrub_ids(seed_hex: str, text: str) -> str:
    """Replace CUI / Reg-Com / CNP digit patterns inside a surviving
    text cell with deterministically scrambled digits."""
    def _sub(match: "re.Match[str]") -> str:
        return scramble_text(seed_hex, match.group(0))

    out = text
    for pattern in _ID_PATTERNS:
        out = pattern.sub(_sub, out)
    return out


def _is_number_shaped(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        return bool(_NUMBER_SHAPED_RE.match(value))
    return False


# ── XLSX ───────────────────────────────────────────────────────────────


def _sheet_structures(data: bytes):
    """Per-sheet structure detection using the REAL engine detector.
    Returns [(sheet_index, structure_or_None, name_col_or_None)]."""
    import pandas as pd  # engine parser dependency

    fmt = tbp.detect_excel_format(data)
    engine = "openpyxl" if fmt == "xlsx" else "xlrd"
    frames = pd.read_excel(
        io.BytesIO(data), engine=engine, sheet_name=None, header=None, dtype=str
    )
    out = []
    for idx, (_sheet, df) in enumerate(frames.items()):
        try:
            structure = tbp.detect_trial_balance_structure(df)
            out.append((idx, structure, structure.columns.get("account_name")))
        except Exception:  # noqa: BLE001 — unrecognized sheet
            out.append((idx, None, None))
    return out


def anonymize_xlsx_bytes(data: bytes, seed_hex: Optional[str] = None) -> bytes:
    """Anonymize every sheet of an XLSX workbook. Recognized TB sheets:
    scramble ONLY data-row name-column cells + pre-header banner text
    cells, and ID-scrub other surviving text cells in data rows; the
    header row and the totals row are untouched. Unrecognized sheets:
    scramble every non-number-shaped text cell (no structure to trust).
    Sheet titles are always scrambled to a deterministic tag."""
    from openpyxl import load_workbook

    seed = seed_hex or content_seed(data)
    structures = _sheet_structures(data)
    wb = load_workbook(io.BytesIO(data), data_only=True)

    for sheet_idx, structure, name_col in structures:
        ws = wb.worksheets[sheet_idx]
        title_tag = hashlib.sha256(
            (seed + "|sheet|" + ws.title).encode("utf-8", errors="replace")
        ).hexdigest()[:10]
        ws.title = "Anon_%s" % title_tag

        if structure is None:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.strip() and not _is_number_shaped(v):
                        cell.value = scramble_text(seed, v)
            continue

        header_row = structure.header_row_index  # 0-based df row
        data_start = structure.data_start_index
        data_end = structure.data_end_index
        code_col = structure.columns.get("account_code")
        numeric_cols = {
            c for k, c in structure.columns.items()
            if k not in ("account_code", "account_name")
        }
        for df_row in range(0, ws.max_row):
            xl_row = df_row + 1
            if df_row == header_row:
                continue  # header wording is structural — never touched
            in_data = data_start <= df_row < data_end
            if not in_data and df_row >= data_start:
                continue  # totals block / footer rows — never touched
            for df_col in range(0, ws.max_column):
                cell = ws.cell(row=xl_row, column=df_col + 1)
                v = cell.value
                if not isinstance(v, str) or not v.strip() or _is_number_shaped(v):
                    continue
                if df_row < header_row:
                    cell.value = scramble_text(seed, v)  # banner rows
                elif df_col == name_col:
                    cell.value = scramble_text(seed, v)  # the label column
                elif df_col == code_col or df_col in numeric_cols:
                    continue  # codes + numeric columns: never touched
                else:
                    cell.value = _scrub_ids(seed, v)  # extra text columns

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CSV / delimited text ───────────────────────────────────────────────

_CSV_ENCODINGS = ("utf-8-sig", "cp1250", "latin-1")


def anonymize_csv_bytes(data: bytes, seed_hex: Optional[str] = None) -> bytes:
    """Anonymize a delimited CSV/TXT trial balance: scramble data-row
    name-column fields (+ banner rows, + ID scrub of other text fields);
    header row, code column, numeric fields and the totals row survive
    verbatim. Re-encoded with the SAME codec that decoded it (scrambled
    output is ASCII, so re-encoding never fails)."""
    seed = seed_hex or content_seed(data)
    text = None
    codec = None
    for encoding in _CSV_ENCODINGS:
        try:
            text = data.decode(encoding)
            codec = encoding
            break
        except UnicodeDecodeError:
            continue
    if text is None:  # unreachable given latin-1; guard anyway
        raise ValueError("could not decode the CSV under any known codec")

    df = tbp._grid_from_delimited_text(text, "CSV file")
    structure = tbp.detect_trial_balance_structure(df)
    delimiter = tbp._sniff_delimiter(text)
    name_col = structure.columns.get("account_name")
    code_col = structure.columns.get("account_code")
    numeric_cols = {
        c for k, c in structure.columns.items()
        if k not in ("account_code", "account_name")
    }

    newline = "\r\n" if "\r\n" in text else "\n"
    raw_lines = [ln for ln in text.splitlines()]
    # The grid skipped blank lines; map df row index → raw line index.
    nonblank_idx = [i for i, ln in enumerate(raw_lines) if ln.strip()]

    for df_row, raw_idx in enumerate(nonblank_idx):
        if df_row == structure.header_row_index:
            continue
        in_data = structure.data_start_index <= df_row < structure.data_end_index
        if not in_data and df_row >= structure.data_start_index:
            continue  # totals / footer rows untouched
        fields = raw_lines[raw_idx].split(delimiter)
        changed = False
        for col, field in enumerate(fields):
            if not field.strip() or _is_number_shaped(field):
                continue
            if df_row < structure.header_row_index:
                fields[col] = scramble_text(seed, field)
                changed = True
            elif col == name_col:
                fields[col] = scramble_text(seed, field)
                changed = True
            elif col == code_col or col in numeric_cols:
                continue
            else:
                scrubbed = _scrub_ids(seed, field)
                if scrubbed != field:
                    fields[col] = scrubbed
                    changed = True
        if changed:
            raw_lines[raw_idx] = delimiter.join(fields)

    out_text = newline.join(raw_lines)
    if text.endswith(("\n", "\r")):
        out_text += newline
    assert codec is not None
    return out_text.encode(codec)


# ── Dispatch + invariant checks ────────────────────────────────────────


def detect_format(data: bytes) -> str:
    if data[:4] == b"PK\x03\x04" or data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xlsx"
    if data[:5] == b"%PDF-":
        return "pdf"
    return "csv"


def anonymize_bytes(data: bytes, seed_hex: Optional[str] = None) -> bytes:
    kind = detect_format(data)
    if kind == "xlsx":
        return anonymize_xlsx_bytes(data, seed_hex)
    if kind == "pdf":
        raise ValueError(
            "PDF anonymization does not belong to the tabular scrambler — use "
            "scripts/pdf_scrambler.py (lexicon-driven content-stream rewrite, "
            "same seed and substitution alphabet)."
        )
    return anonymize_csv_bytes(data, seed_hex)


def _parse_rows(data: bytes) -> List[Dict[str, Any]]:
    kind = detect_format(data)
    if kind == "xlsx":
        return list(tbp.parse_trial_balance_file(data, "anonymize_check.xlsx"))
    return list(tbp.parse_trial_balance_csv(data, "anonymize_check.csv"))


def _parse_anchor(data: bytes) -> Dict[str, Any]:
    kind = detect_format(data)
    if kind == "xlsx":
        return dict(tbp.parse_trial_balance_file(data, "anonymize_check.xlsx").source_anchor)
    return dict(tbp.parse_trial_balance_csv(data, "anonymize_check.csv").source_anchor)


_NUM_FIELDS = ("si_d", "si_c", "r_d", "r_c", "st_d", "st_c", "sf_d", "sf_c")


def numeric_invariant_failures(original: bytes, transformed: bytes) -> List[str]:
    """Compare original vs transformed through the REAL parser. Returns
    a list of human-readable failures (empty == invariants hold):
      · same row count, same account-code sequence;
      · every numeric field of every row identical;
      · every column-pair sum identical (the --self-check contract);
      · the totals-row anchor file values identical (anchor preserved).
    """
    failures: List[str] = []
    rows_a = _parse_rows(original)
    rows_b = _parse_rows(transformed)
    if len(rows_a) != len(rows_b):
        failures.append("row count %d != %d" % (len(rows_a), len(rows_b)))
        return failures
    sums_a = {f: 0.0 for f in _NUM_FIELDS}
    sums_b = {f: 0.0 for f in _NUM_FIELDS}
    for i, (ra, rb) in enumerate(zip(rows_a, rows_b)):
        if str(ra.get("cont")) != str(rb.get("cont")):
            failures.append(
                "row %d account code %r != %r" % (i, ra.get("cont"), rb.get("cont"))
            )
            continue
        for f in _NUM_FIELDS:
            va = float(ra.get(f) or 0)
            vb = float(rb.get(f) or 0)
            sums_a[f] += va
            sums_b[f] += vb
            if va != vb:
                failures.append(
                    "row %d (%s) field %s: %r != %r"
                    % (i, ra.get("cont"), f, va, vb)
                )
    for f in _NUM_FIELDS:
        if round(sums_a[f], 2) != round(sums_b[f], 2):
            failures.append(
                "column sum %s: %.2f != %.2f" % (f, sums_a[f], sums_b[f])
            )
    anchor_a = _parse_anchor(original)
    anchor_b = _parse_anchor(transformed)
    pairs_a = anchor_a.get("pairs") or {}
    pairs_b = anchor_b.get("pairs") or {}
    for key in ("si", "rl", "rc", "sf"):
        pa, pb = pairs_a.get(key), pairs_b.get(key)
        if (pa is None) != (pb is None):
            failures.append("anchor pair %s presence changed" % key)
            continue
        if pa is None:
            continue
        for side in ("file_debit", "file_credit"):
            if pa.get(side) != pb.get(side):
                failures.append(
                    "anchor %s %s: %r != %r" % (key, side, pa.get(side), pb.get(side))
                )
    return failures


def verify_bytes(data: bytes) -> List[str]:
    """--verify contract (used by the corpus replay runner): scramble
    the input IN MEMORY and prove the numeric structure is invariant
    under anonymization. Returns failures (empty == pass).

    PDF inputs delegate to `pdf_scrambler.verify_bytes`: a PDF has no
    columns to scramble, so its anonymization is lexicon-driven and its
    replay-time contract is the residual-lexicon + convergence pair
    documented there. The WRITE path (`anonymize_bytes`) still refuses
    PDFs — that division is deliberate, one tool per file family."""
    if detect_format(data) == "pdf":
        import pdf_scrambler  # same scripts/ directory

        return pdf_scrambler.verify_bytes(data)
    try:
        transformed = anonymize_bytes(data)
    except Exception as exc:  # noqa: BLE001
        return ["anonymization transform failed: %s: %s" % (type(exc).__name__, exc)]
    return numeric_invariant_failures(data, transformed)


# ── CLI ────────────────────────────────────────────────────────────────


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("inputs", nargs="+", help="input file(s)")
    parser.add_argument("-o", "--output", help="output path (single-input mode)")
    parser.add_argument(
        "--self-check", action="store_true",
        help="re-parse input+output and assert numeric invariants",
    )
    parser.add_argument(
        "--verify", action="store_true",
        help="verify mode: in-memory scramble + invariant assertion, no writes",
    )
    args = parser.parse_args(argv)

    if args.verify:
        rc = 0
        for name in args.inputs:
            data = Path(name).read_bytes()
            failures = verify_bytes(data)
            if failures:
                rc = 1
                print("VERIFY FAIL %s" % name)
                for f in failures:
                    print("  ✗ %s" % f)
            else:
                print("VERIFY OK   %s" % name)
        return rc

    if len(args.inputs) != 1:
        print("anonymize mode takes exactly one input (use --verify for batches)")
        return 2
    in_path = Path(args.inputs[0])
    data = in_path.read_bytes()
    out_path = Path(args.output) if args.output else in_path.with_name(
        in_path.stem + ".anon" + in_path.suffix
    )
    result = anonymize_bytes(data)
    out_path.write_bytes(result)
    print("wrote %s (%d bytes, seed=%s)" % (out_path, len(result), content_seed(data)[:12]))

    if args.self_check:
        failures = numeric_invariant_failures(data, result)
        if failures:
            print("SELF-CHECK FAIL (%d):" % len(failures))
            for f in failures:
                print("  ✗ %s" % f)
            return 1
        print("SELF-CHECK OK — codes, per-row numerics, column-pair sums and "
              "totals-row anchor identical pre/post")
    return 0


if __name__ == "__main__":
    sys.exit(main())
