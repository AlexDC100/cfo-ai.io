"""Generate the canonical CFO AI upload template (.xlsx).

Why this script exists
----------------------
Before today, users uploaded random Excel files and we hoped the parser
figured them out. That's brittle and produced the kind of "DIO didn't
extract because the sheet name is YTD Oct'25 instead of YTD Mar'26" bug
that just burned five sessions to find.

The fix is a canonical template — one file with known sheet names, known
column names, known row offsets — that the parser knows it can lean on.
This script generates that template. The output (.xlsx) is committed at:

    scandi-desk-main/public/templates/cfo_ai_upload_template.xlsx

and served by the FE at:

    https://cfo-ai.io/templates/cfo_ai_upload_template.xlsx

It's regenerated on every `npm run build` (see package.json
`generate-template` script) so the artifact never drifts from the
parser's expectations.

What the parser actually reads (src/engine/api/_sales_extract.py)
-----------------------------------------------------------------
- "Trading" sheet — SKU-level rows. Columns identified by name:
    Canal, Client, Tip client, CLIENT_PARINTE, BU, Categ_Pr, Brand,
    Denumire_Produs, PackSize, Sold in KG, GR Gross Revenue finished
    goods, Net Invoice Value, Cost of Sales, Gross Margin.
- "DIO" sheet — two regions in ONE sheet:
    Region 1 (rows 2..N): per-category inventory snapshot.
        col A = category name (Grupa_Pr)
        col D = inventory value (Stoc Valoric Standard (RON))
        col H = inventory kg (Stoc KG)
    Region 2 (rows 28..52): per-category DIO days.
        col B = category name
        col D = DIO days
- "Trial Balance" sheet (optional) — Cont/Denumire/Sold debitor/Sold creditor.

The Region 2 row offset is what makes the layout look weird but it is
THE canonical source for DIO days the Products view reads. The empty
rows 22-27 between Region 1 and Region 2 are structural — the parser
locates Region 2 by row offset, so deleting them silently breaks DIO.

Verification
------------
After running this script, upload the generated file at /upload on prod
and confirm:
  - sku_aggregates ingests 3 example SKUs
  - Products view shows LEGUME CONSERVATE @ 73.5d, JELEURI @ 1001.7d
    (these example values match the Region 2 entries below — change here
    means change the example doc + screenshot if you're matching them).
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Output goes inside the FE workspace's `public/` so Vite serves it at
# /templates/cfo_ai_upload_template.xlsx without any extra config.
OUTPUT = Path(__file__).resolve().parents[1] / (
    "scandi-desk-main/public/templates/cfo_ai_upload_template.xlsx"
)


# ──────────────────────────────────────────────────────────────────────
# Styling helpers — kept tight so every sheet renders with the same
# typography. Hex colors mirror the v5 site palette (navy header / amber
# warnings) so the file feels like an extension of the product, not a
# random spreadsheet.
# ──────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F2937")   # slate-800
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ACCENT_FILL = PatternFill("solid", start_color="003366")   # navy
ACCENT_FONT = Font(bold=True, color="FFFFFF", size=18)
SUBTLE_FONT = Font(italic=True, color="6B7280", size=10)   # slate-500
NOTE_FONT = Font(italic=True, color="9CA3AF", size=9)      # slate-400
BOLD = Font(bold=True, size=11)


def _style_header_row(ws, row: int, headers: list[str], comments: dict[str, str] | None = None) -> None:
    """Apply the standard header style + optional cell comments."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if comments and header in comments:
            cell.comment = Comment(comments[header], "CFO AI")
    ws.row_dimensions[row].height = 32


def _set_widths(ws, widths: list[int]) -> None:
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ──────────────────────────────────────────────────────────────────────
# Sheet builders. Each function adds one sheet to the workbook. The order
# of `wb.create_sheet()` calls determines tab order — Instructions is
# created first so it's the sheet that opens by default.
# ──────────────────────────────────────────────────────────────────────


def add_instructions_sheet(wb: Workbook) -> None:
    """Plain-English/Romanian overview. First tab the user sees on open."""
    ws = wb.create_sheet("Instructions", 0)

    # Big navy header banner across A1:D1
    ws["A1"] = "CFO AI UPLOAD TEMPLATE"
    ws["A1"].font = ACCENT_FONT
    ws["A1"].fill = ACCENT_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 44

    ws["A3"] = (
        "This is the canonical format CFO AI expects. "
        "Acesta este formatul canonic pe care CFO AI îl așteaptă."
    )
    ws["A3"].font = SUBTLE_FONT
    ws.merge_cells("A3:D3")

    rows: list[tuple[str, str]] = [
        ("", ""),
        ("HOW TO USE", ""),
        ("1.", 'Fill in the "Trading" sheet with your SKU-level sales data.'),
        ("2.", 'Fill in the "DIO" sheet with your per-category inventory days.'),
        ("3.", '(Optional) Fill in "Trial Balance" if you want Dashboard financial reports.'),
        ("4.", "Save the file and upload it at https://cfo-ai.io/products"),
        ("", ""),
        ("SHEETS IN THIS WORKBOOK", ""),
        (
            "Trading",
            "Your SKU-level sales rows. REQUIRED. One row per SKU × channel × period.",
        ),
        (
            "DIO",
            "Days inventory outstanding per category. REQUIRED for DIO in Products view.",
        ),
        (
            "Trial Balance",
            "OPTIONAL. Standard Romanian chart of accounts. Only for Dashboard reports.",
        ),
        (
            "Field Reference",
            "Column dictionary — types, units, examples. Useful for ERP exports.",
        ),
        ("", ""),
        ("COMMON MISTAKES", ""),
        (
            "✗",
            'Renaming the sheets — keep them as "Trading", "DIO", "Trial Balance".',
        ),
        (
            "✗",
            'Capital mismatch between Trading and DIO (e.g., "Macrou" vs "MACROU"). '
            "Parser handles common variants but exact match is safest.",
        ),
        (
            "✗",
            "Removing the empty rows 22-27 in the DIO sheet — they are STRUCTURAL. "
            "The parser locates Region 2 by row offset; deleting them silently breaks DIO.",
        ),
        (
            "✗",
            'Adding columns before "Categ_Pr" in the Trading sheet — column order '
            "matters for the column-position fallback parser.",
        ),
        ("", ""),
        ("QUESTIONS?", "contact@cfo-ai.io · https://cfo-ai.io/docs/upload-format"),
    ]
    for i, (label, body) in enumerate(rows, start=4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = body
        if label in ("HOW TO USE", "SHEETS IN THIS WORKBOOK", "COMMON MISTAKES", "QUESTIONS?"):
            ws[f"A{i}"].font = BOLD
            ws[f"A{i}"].fill = PatternFill("solid", start_color="F3F4F6")
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        else:
            ws[f"A{i}"].alignment = Alignment(horizontal="right", vertical="top")
            ws[f"B{i}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        ws.row_dimensions[i].height = 22

    _set_widths(ws, [18, 32, 32, 24])

    # Hide gridlines on the cover — looks more like a printed page.
    ws.sheet_view.showGridLines = False


def add_trading_sheet(wb: Workbook) -> None:
    """SKU-level transactions. The parser's primary source of revenue,
    volume, gross margin, and category attribution.

    Column order matters: the parser has a positional fallback that maps
    column INDEX → semantic when header names don't match exactly. So
    even if a user renames "Categ_Pr" to "Category" we still pick it up
    — as long as it stays in position 6. Don't reorder these headers.
    """
    ws = wb.create_sheet("Trading")

    headers = [
        "Canal",                           # 1
        "Client",                          # 2
        "Tip client",                      # 3
        "CLIENT_PARINTE",                  # 4
        "BU",                              # 5
        "Categ_Pr",                        # 6  ← MUST match DIO sheet
        "Brand",                           # 7
        "Denumire_Produs",                 # 8
        "PackSize",                        # 9
        "Sold in KG",                      # 10
        "GR Gross Revenue finished goods", # 11
        "Net Invoice Value",               # 12
        "Cost of Sales",                   # 13
        "Gross Margin",                    # 14
    ]

    comments = {
        "Canal": "Sales channel. KA = key accounts, DIST = distributors, EXPORT, HORECA, etc.",
        "Categ_Pr": (
            "Category name. MUST match a row in the DIO sheet (case-insensitive). "
            "If a SKU's category doesn't appear in DIO, it inherits null DIO."
        ),
        "Sold in KG": "Total volume sold in this row, in kilograms.",
        "Net Invoice Value": "NIV in RON. Revenue gross of discounts, minus VAT.",
        "Cost of Sales": "COGS in RON for this row.",
        "Gross Margin": "GM in RON. Computed as NIV − Cost of Sales.",
    }

    _style_header_row(ws, 1, headers, comments)

    examples = [
        ["KA",   "AUCHAN",    "KA",   "AUCHAN",    "Diverse", "LEGUME CONSERVATE", "ROUA",    "Fasole alba in sapte ape 400g", "400g", 8420,  195000, 184000, 145600, 38400],
        ["KA",   "CARREFOUR", "KA",   "CARREFOUR", "Diverse", "LEGUME CONSERVATE", "BUCEGI",  "Porumb dulce 285g",             "285g", 6890,  172000, 162000, 130700, 31300],
        ["DIST", "METRO",     "DIST", "METRO",     "Diverse", "TON",               "NAVODUL", "Ton in ulei 80g",               "80g",  14200, 405000, 382000, 297600, 84400],
        ["KA",   "KAUFLAND",  "KA",   "KAUFLAND",  "Diverse", "JELEURI",           "SCANDIA", "Jeleu fructe 200g",             "200g", 320,    28000,  26400,  27800, -1400],
        ["KA",   "PROFI",     "KA",   "PROFI",     "Diverse", "MACROU",            "ROUA",    "Macrou in sos tomat 240g",      "240g", 4220,   89000,  83500,  68200, 15300],
    ]
    for row_data in examples:
        ws.append(row_data)

    # Cell-level number formatting on the numeric columns. Tabular-num
    # so values right-align cleanly when users add their own rows.
    money_fmt = '#,##0.00;-#,##0.00;"—"'
    int_fmt = '#,##0;-#,##0;"—"'
    for r in range(2, 2 + len(examples)):
        ws.cell(row=r, column=10).number_format = int_fmt   # Sold in KG
        for c in (11, 12, 13, 14):
            ws.cell(row=r, column=c).number_format = money_fmt

    _set_widths(ws, [8, 14, 10, 14, 10, 22, 12, 36, 10, 12, 24, 18, 14, 14])
    ws.freeze_panes = "A2"

    # Mark example rows with subtle styling so the user knows to replace
    # them rather than appending below.
    example_note_font = Font(italic=True, color="9CA3AF", size=10)
    for r in range(2, 2 + len(examples)):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).font = example_note_font


def add_dio_sheet(wb: Workbook) -> None:
    """Per-category inventory + DIO days. Two regions in one sheet.

    Region 1 (rows 1..end-of-data) — inventory snapshot.
    Region 2 (rows 28..52) — DIO days. Read by `extract_category_dio`
    from _sales_extract.py: col B = category name, col D = days.

    The empty rows 22-27 are STRUCTURAL. Without them, the parser's row
    offset search for Region 2 misses. Don't delete them.
    """
    ws = wb.create_sheet("DIO")
    today = date.today()

    # ──────────────────────────────────────────────────────────────
    # Row 1 — as-of dates across the value columns. Helps the
    # operator know when the snapshot was taken.
    # ──────────────────────────────────────────────────────────────
    ws["A1"] = ""
    for col in (2, 4, 5, 6, 7, 8):
        cell = ws.cell(row=1, column=col, value=today)
        cell.number_format = "dd.mm.yyyy"
        cell.font = BOLD
    ws.row_dimensions[1].height = 22

    # ──────────────────────────────────────────────────────────────
    # Row 2 — Region 1 headers. Wide spacing (empty cols) preserved
    # exactly because the parser reads named columns by header
    # match, not by column index — and the spacing makes Excel
    # render the snapshot readably.
    # ──────────────────────────────────────────────────────────────
    region1_headers = [
        "Grupa_Pr",                       # A
        "Stoc Cantitativ (UM)",           # B
        " ",                              # C
        "Stoc Valoric Standard (RON)",    # D
        " ",                              # E
        "Stoc RON / UM",                  # F
        " ",                              # G
        "Stoc KG",                        # H
    ]
    for col, h in enumerate(region1_headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = BOLD
        cell.fill = PatternFill("solid", start_color="F3F4F6")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32
    ws["A2"].comment = Comment(
        "Region 1: per-category inventory snapshot. "
        "Used for inventory value × financing cost in the Decision Rules engine.",
        "CFO AI",
    )

    # ──────────────────────────────────────────────────────────────
    # Region 1 data (rows 3..7) — same example categories as Trading
    # sheet so end-to-end smoke test shows matching values.
    # ──────────────────────────────────────────────────────────────
    region1_data = [
        ("LEGUME CONSERVATE", 71655, 0, 2986170.15, 0, 41.67,  0, 394797.70),
        ("JELEURI",            8693, 0,  457738.17, 0, 52.66,  0,  13908.80),
        ("TON",               12450, 0,  823412.55, 0, 66.14,  0,  19887.20),
        ("MACROU",             5210, 0,  274982.10, 0, 52.78,  0,   9402.50),
        ("ULEI",               1067, 0,  177280.82, 0, 166.15, 0,   6887.50),
    ]
    money_fmt = '#,##0.00;-#,##0.00;"—"'
    int_fmt = '#,##0;-#,##0;"—"'
    for i, row in enumerate(region1_data, start=3):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if col == 1:
                cell.font = BOLD
            elif col in (2, 8):
                cell.number_format = int_fmt
            elif col in (4, 6):
                cell.number_format = money_fmt

    # ──────────────────────────────────────────────────────────────
    # Rows 8-27 — structural padding. The parser locates Region 2
    # by scanning for the literal phrase pattern starting at row 28.
    # We add a few visible operator-context notes (Selection Status,
    # CompName, etc.) the way the source file does — these don't
    # affect parsing but make the file look familiar to operators
    # used to the legacy Trading_analysis_*.xlsx layout.
    # ──────────────────────────────────────────────────────────────
    operator_notes = [
        (23, "Selection Status:"),
        (24, "CompName: 40-SF"),
        (25, "Cod_Articol: 7*"),
        (26, "Subcategorie: 3.3. DIVERS"),
    ]
    for row_idx, label in operator_notes:
        ws.cell(row=row_idx, column=1, value=label).font = NOTE_FONT

    # ──────────────────────────────────────────────────────────────
    # Row 27 — annotation above Region 2. Plain English so the
    # operator knows what they're looking at.
    # ──────────────────────────────────────────────────────────────
    annotate = ws.cell(
        row=27,
        column=2,
        value="DIO days per category — canonical source (rows 28-52):",
    )
    annotate.font = Font(bold=True, italic=True, color="4B5563", size=10)
    annotate.comment = Comment(
        "Region 2: per-category DIO days. THIS is the canonical source "
        "the Products view reads from. Column B = category name, "
        "Column D = DIO days. Rows 28-52.",
        "CFO AI",
    )

    # ──────────────────────────────────────────────────────────────
    # Region 2 data (rows 28..52) — the famous JELEURI = 1001.7d
    # entry is preserved as a calibration anchor so the verification
    # screenshot in the docs page matches what the parser returns.
    # ──────────────────────────────────────────────────────────────
    region2_data = [
        ("LEGUME CONSERVATE",     73.5),
        ("MURATURI",              58.5),
        ("COMPOT",               167.9),
        ("DULCEATA",              29.2),
        ("JELEURI",            1001.7),  # the famous one
        ("PASTA TOMATE",          98.3),
        ("SUC",                  146.2),
        ("SUC DE ROSII",          34.7),
        ("SIROP",                 19.9),
        ("ZACUSCA",               27.8),
        ("OTET",                  17.6),
        ("MUSTAR",                24.1),
        ("SOSURI",                39.2),
        ("SUPE INSTANT NOODLES", 150.6),
        ("PET FOOD",              30.5),
        ("ULEI",                 286.6),
        ("TON",                  194.7),
        ("MACROU",                64.7),
        ("SARDINE",               84.9),
        ("HERING",               118.1),
        ("SOMON",                234.0),
        ("PASTRAV",              125.7),
        ("SPROT",                105.4),
        ("FRUCTE NOBILE",        222.1),
        ("ALTE MARFURI",         102.5),
    ]
    assert len(region2_data) == 25, "Region 2 must fill rows 28..52 exactly"

    for i, (cat, days) in enumerate(region2_data, start=28):
        name_cell = ws.cell(row=i, column=2, value=cat)
        name_cell.font = BOLD
        days_cell = ws.cell(row=i, column=4, value=days)
        days_cell.number_format = "0.0"
        # Highlight the "days" cell faintly so it stands out as the
        # actual value the operator should care about.
        days_cell.fill = PatternFill("solid", start_color="FEF3C7")  # amber-100

    _set_widths(ws, [22, 24, 4, 28, 4, 16, 4, 16])
    ws.freeze_panes = "A3"


def add_trial_balance_sheet(wb: Workbook) -> None:
    """Optional. Only needed for Dashboard financial reports — Products
    view runs entirely off the Trading + DIO sheets.
    """
    ws = wb.create_sheet("Trial Balance")
    headers = ["Cont", "Denumire", "Sold debitor", "Sold creditor"]
    comments = {
        "Cont": "Romanian chart-of-accounts code. e.g. 1012, 2131, 401.",
        "Sold debitor": "Debit closing balance (RON). Use 0 if not applicable.",
        "Sold creditor": "Credit closing balance (RON). Use 0 if not applicable.",
    }
    _style_header_row(ws, 1, headers, comments)

    examples = [
        ["1012", "Capital subscris vărsat",    0,       100000],
        ["2131", "Echipamente tehnologice",    250000,  0],
        ["371",  "Mărfuri",                    1200000, 0],
        ["401",  "Furnizori",                  0,       450000],
        ["411",  "Clienți",                    340000,  0],
        ["5121", "Conturi la bănci în lei",    320000,  0],
        ["707",  "Venituri din vânzarea măr.", 0,       2400000],
        ["607",  "Cheltuieli privind măr.",    1800000, 0],
    ]
    money_fmt = '#,##0.00;-#,##0.00;"—"'
    for row in examples:
        ws.append(row)
    for r in range(2, 2 + len(examples)):
        ws.cell(row=r, column=3).number_format = money_fmt
        ws.cell(row=r, column=4).number_format = money_fmt

    _set_widths(ws, [10, 40, 18, 18])
    ws.freeze_panes = "A2"

    # Subtle note on the Trial Balance tab — it's optional, and we
    # don't want users to think they need to fill it in to use Products.
    ws.cell(row=12, column=1, value="OPTIONAL").font = BOLD
    ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=4)
    ws.cell(row=13, column=1, value=(
        "This sheet is only needed for the Dashboard's P&L / Balance Sheet / "
        "Cash Flow reports. The Products view runs entirely off the Trading "
        "and DIO sheets, so you can leave this blank if you only need SKU "
        "intelligence."
    )).font = SUBTLE_FONT
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=4)
    ws.row_dimensions[13].height = 44
    ws.cell(row=13, column=1).alignment = Alignment(wrap_text=True, vertical="top")


def add_field_reference_sheet(wb: Workbook) -> None:
    """Machine-readable column dictionary. Useful for users wiring up
    automated ERP → CFO AI exports — they can map their ERP columns to
    our column names by reading this sheet programmatically.
    """
    ws = wb.create_sheet("Field Reference")
    headers = ["Sheet", "Column", "Type", "Required", "Unit", "Example", "Notes"]
    _style_header_row(ws, 1, headers)

    refs = [
        # Trading sheet
        ("Trading",        "Canal",                         "text",   "yes", "—",     "KA",                       "Sales channel: KA / DIST / EXPORT / HORECA"),
        ("Trading",        "Client",                        "text",   "yes", "—",     "AUCHAN",                   "Customer name"),
        ("Trading",        "Tip client",                    "text",   "no",  "—",     "KA",                       "Customer type — often duplicates Canal"),
        ("Trading",        "CLIENT_PARINTE",                "text",   "no",  "—",     "AUCHAN",                   "Parent customer (for chains)"),
        ("Trading",        "BU",                            "text",   "no",  "—",     "Diverse",                  "Business unit"),
        ("Trading",        "Categ_Pr",                      "text",   "yes", "—",     "LEGUME CONSERVATE",        "MUST match a category in DIO sheet"),
        ("Trading",        "Brand",                         "text",   "yes", "—",     "ROUA",                     "Product brand"),
        ("Trading",        "Denumire_Produs",               "text",   "yes", "—",     "Fasole alba 400g",         "Product name (SKU display label)"),
        ("Trading",        "PackSize",                      "text",   "no",  "—",     "400g",                     "Package size"),
        ("Trading",        "Sold in KG",                    "number", "yes", "kg",    "8420",                     "Total volume sold (kilograms)"),
        ("Trading",        "GR Gross Revenue finished goods", "number", "yes", "RON", "195000",                   "Gross revenue before discounts"),
        ("Trading",        "Net Invoice Value",             "number", "yes", "RON",   "184000",                   "NIV — revenue net of discounts, gross of VAT"),
        ("Trading",        "Cost of Sales",                 "number", "yes", "RON",   "145600",                   "COGS for this row"),
        ("Trading",        "Gross Margin",                  "number", "yes", "RON",   "38400",                    "GM = NIV − COGS"),
        # DIO Region 1
        ("DIO Region 1",   "Grupa_Pr",                      "text",   "yes", "—",     "LEGUME CONSERVATE",        "Same as Categ_Pr in Trading sheet"),
        ("DIO Region 1",   "Stoc Cantitativ (UM)",          "number", "yes", "units", "71655",                    "Stock quantity in units"),
        ("DIO Region 1",   "Stoc Valoric Standard (RON)",   "number", "yes", "RON",   "2986170",                  "Inventory value (used for financing-cost calc)"),
        ("DIO Region 1",   "Stoc RON / UM",                 "number", "no",  "RON",   "41.67",                    "Average per-unit inventory value"),
        ("DIO Region 1",   "Stoc KG",                       "number", "yes", "kg",    "394798",                   "Stock in kilograms"),
        # DIO Region 2 — the canonical source
        ("DIO Region 2",   "[col B] category name",         "text",   "yes", "—",     "JELEURI",                  "Category name. Rows 28-52."),
        ("DIO Region 2",   "[col D] DIO days",              "number", "yes", "days",  "1001.7",                   "DIO days. Rows 28-52. THE canonical source."),
        # Trial Balance
        ("Trial Balance",  "Cont",                          "text",   "no",  "—",     "1012",                     "Romanian CoA code"),
        ("Trial Balance",  "Denumire",                      "text",   "no",  "—",     "Capital subscris vărsat",  "Account name"),
        ("Trial Balance",  "Sold debitor",                  "number", "no",  "RON",   "0",                        "Debit closing balance"),
        ("Trial Balance",  "Sold creditor",                 "number", "no",  "RON",   "100000",                   "Credit closing balance"),
    ]
    for row in refs:
        ws.append(row)

    # Highlight required rows with a faint amber tint — quick visual scan
    # for "what do I HAVE to fill in".
    required_fill = PatternFill("solid", start_color="FEF3C7")  # amber-100
    for r in range(2, 2 + len(refs)):
        required = ws.cell(row=r, column=4).value
        if required == "yes":
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = required_fill

    _set_widths(ws, [14, 32, 10, 10, 8, 26, 56])
    ws.freeze_panes = "A2"


# ──────────────────────────────────────────────────────────────────────
# Driver
# ──────────────────────────────────────────────────────────────────────


def make_template() -> Path:
    """Build the workbook + write to disk. Returns the output path."""
    wb = Workbook()
    # `Workbook()` ships a default "Sheet" — drop it before we add ours.
    wb.remove(wb.active)

    add_instructions_sheet(wb)
    add_trading_sheet(wb)
    add_dio_sheet(wb)
    add_trial_balance_sheet(wb)
    add_field_reference_sheet(wb)

    # Make sure the Instructions tab is the one that opens by default.
    wb.active = 0

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = make_template()
    size_kb = path.stat().st_size / 1024
    print(f"Wrote {path}")
    print(f"  → {size_kb:.1f} KB · 5 sheets · ready to ship")
