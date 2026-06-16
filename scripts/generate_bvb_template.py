"""Generate a canonical BVB-financials admin upload template (.xlsx).

Why this exists
---------------
BVB Phase 1 seeded 20 BET-20 tickers with FY2024 numbers for 7 of them
(TLV, SNP, SNG, H2O, M, CFH, SFG). The remaining 13 have placeholder
``None`` cells. Rather than make the operator wrangle Supabase SQL by
hand, this template lays out one row per BVB ticker with the exact
columns the loader expects, plus inline cell comments explaining each
field.

Workflow
--------
1. Generate the template:
   ``python scripts/generate_bvb_template.py``
2. The output lands at:
   ``scandi-desk-main/public/templates/bvb_financials_template.xlsx``
3. Operator fills in the missing 13 rows (or refreshes the 7 already
   seeded with newer numbers).
4. Upload via the admin endpoint OR run the loader directly:
   ``python scripts/seed_bvb_companies.py --xlsx bvb_financials_template.xlsx``

What the loader reads
---------------------
- Sheet "BET-20" — one row per ticker. Columns identified by name:
    ticker, name, sector, industry, market_cap_b, revenue_b,
    revenue_growth_pct, ebitda_b, ebitda_margin_pct, net_income_b,
    net_margin_pct, equity_b, cash_b, gross_debt_b, net_debt_b,
    pe, ev_ebitda, dividend_yield_pct, roe_pct, latest_period, source_url.

All money values are in BILLIONS of RON. Margins / yields in percentage
POINTS (12.5 means 12.5%). Multiples dimensionless (12.5 means 12.5×).

Empty cells are interpreted as "no change" — the loader preserves the
seed default for that field. To explicitly null out a field, use the
sentinel string "NULL" (case-insensitive).
"""
from __future__ import annotations

import sys
from pathlib import Path

# Reuse the seed module so the template is always in lock-step with the
# canonical BET-20 composition. If the seed adds a 21st ticker later,
# the template picks it up automatically.
_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "src"))

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from engine.public.bvb_seed import bvb_universe  # noqa: E402


# ── Column spec ─────────────────────────────────────────────────────────
# (column_name, comment, optional default value extractor)
#
# The order here drives the column order in the template. The loader
# (scripts/seed_bvb_companies.py) reads by column name so a future
# re-order is safe.

_COLUMNS = [
    ("ticker",            "BET-20 ticker. EL.BVB is namespaced because EL collides with NASDAQ's Estée Lauder."),
    ("name",              "Issuer legal name (S.A. / N.V. / plc) — used in tables and drawer headers."),
    ("sector",            "One of: Financials, Energy, Utilities, Communication, Healthcare, Consumer Defensive, Consumer Discretionary, Industrials, Real Estate, Materials."),
    ("industry",          "Finer-grained sub-tag. Used by the metric quick-filter chips. Free-form text."),
    ("market_cap_b",      "Market cap in BILLIONS of RON (12.5 = RON 12.5B)."),
    ("revenue_b",         "Revenue in BILLIONS of RON. Banks: use net banking income (NII + fees)."),
    ("revenue_growth_pct","Y-o-Y revenue growth in PERCENTAGE POINTS (8.5 = +8.5%). Negative for declines."),
    ("ebitda_b",          "EBITDA in BILLIONS of RON. ABSOLUTE — preferred for banks where margin is not meaningful."),
    ("ebitda_margin_pct", "EBITDA margin in PERCENTAGE POINTS. Used only when ebitda_b is blank."),
    ("net_income_b",      "Net income (parent / group as documented) in BILLIONS of RON. ABSOLUTE."),
    ("net_margin_pct",    "Net margin in PERCENTAGE POINTS. Used only when net_income_b is blank."),
    ("equity_b",          "Book equity at period end in BILLIONS of RON."),
    ("cash_b",            "Cash + equivalents in BILLIONS of RON."),
    ("gross_debt_b",      "Gross debt (LT + ST + leasing) in BILLIONS of RON."),
    ("net_debt_b",        "Net debt = gross debt - cash in BILLIONS of RON."),
    ("pe",                "P/E ratio. 12.5 means 12.5×."),
    ("ev_ebitda",         "EV / EBITDA. 8.0 means 8.0×."),
    ("dividend_yield_pct","Trailing dividend yield in PERCENTAGE POINTS."),
    ("roe_pct",           "ROE in PERCENTAGE POINTS."),
    ("latest_period",     "Period label. Default 'FY2024'. Use 'H1 2025' / 'Q3 2025' for interim updates."),
    ("source_url",        "URL to the source filing (annual report / financial statements). Audit trail."),
]


_HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RICH_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # rows pre-filled from seed
_NUMERIC_COLS = {
    "market_cap_b", "revenue_b", "revenue_growth_pct",
    "ebitda_b", "ebitda_margin_pct", "net_income_b", "net_margin_pct",
    "equity_b", "cash_b", "gross_debt_b", "net_debt_b",
    "pe", "ev_ebitda", "dividend_yield_pct", "roe_pct",
}


def _seed_value(row: dict, key: str):
    """Extract the seed default for a given column. Maps the seed row's
    raw RON values back to BILLIONS for the template."""
    if key == "ticker":
        return row["ticker"]
    if key == "name":
        return row["companyName"]
    if key == "sector":
        return row.get("sector")
    if key == "industry":
        return row.get("industry")
    if key == "latest_period":
        return row.get("latestPeriod") or "FY2024"
    if key == "revenue_growth_pct":
        return row.get("revenueGrowth")
    if key == "ebitda_margin_pct":
        return row.get("ebitdaMargin")
    if key == "net_margin_pct":
        return row.get("netMargin")
    if key == "pe":
        return row.get("peRatio")
    if key == "ev_ebitda":
        return row.get("evToEbitda")
    if key == "dividend_yield_pct":
        return row.get("dividendYield")
    if key == "roe_pct":
        return row.get("roe")
    # Money fields: stored as raw RON in seed → present as RON billions.
    money_map = {
        "market_cap_b":  "marketCap",
        "revenue_b":     "revenue",
        "ebitda_b":      "ebitda",
        "net_income_b":  "netIncome",
        "equity_b":      "equity",
        "cash_b":        "cash",
        "gross_debt_b":  "grossDebt",
        "net_debt_b":    "netDebt",
    }
    if key in money_map:
        v = row.get(money_map[key])
        return None if v is None else round(v / 1_000_000_000, 4)
    if key == "source_url":
        return None  # Operator fills in
    return None


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "BET-20"

    # ── Header row ──
    for col_idx, (col_name, comment) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.comment = Comment(comment, "BVB seed")

    # ── Data rows ──
    bvb = bvb_universe()
    for row_idx, (ticker, row) in enumerate(bvb.items(), start=2):
        # Detect rich (revenue set) for visual tint.
        is_rich = row.get("revenue") is not None
        for col_idx, (col_name, _comment) in enumerate(_COLUMNS, start=1):
            v = _seed_value(row, col_name)
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            if is_rich:
                c.fill = _RICH_FILL
            if col_name in _NUMERIC_COLS and isinstance(v, (int, float)):
                # 4 decimal places — keeps precision for share-count
                # derived market caps without overflowing the display.
                c.number_format = "0.0000"

    # ── Column widths ──
    for col_idx, (col_name, _comment) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            14, len(col_name) + 4
        )
    ws.column_dimensions[get_column_letter(2)].width = 40  # name
    ws.column_dimensions[get_column_letter(4)].width = 28  # industry

    # Freeze header row so the operator can scroll without losing context.
    ws.freeze_panes = "B2"

    # ── README sheet ──
    readme = wb.create_sheet("README")
    readme["A1"] = "BVB Financials Upload Template"
    readme["A1"].font = Font(bold=True, size=14, color="003366")
    notes = [
        "",
        "How to use this template",
        "------------------------",
        "1. Fill missing values in the BET-20 sheet for rows that don't",
        "   already have a green tint. Green rows are pre-seeded from",
        "   FY2024 disclosures and can be left as-is.",
        "",
        "2. Unit conventions:",
        "   - Money fields (suffix _b): BILLIONS of RON",
        "   - Margins / yields (suffix _pct): PERCENTAGE POINTS",
        "   - Multiples (pe, ev_ebitda): dimensionless ratio",
        "",
        "3. Empty cell  = preserve seed default (no change)",
        "   'NULL'      = explicitly null out the field",
        "",
        "4. Run the loader to apply:",
        "   python scripts/seed_bvb_companies.py --xlsx <this-file>",
        "",
        "Source attribution",
        "------------------",
        "When you fill a row, please paste the URL of the source filing",
        "(annual report, press release, BVB filing) into source_url. The",
        "loader writes it to the public_companies row so any auditor can",
        "trace the number back to the issuer's disclosure.",
        "",
        "Why CFH is highlighted",
        "----------------------",
        "Cris-Tim Family Holding is Scandia Food's closest BVB-listed",
        "peer (Romanian meat processor, comparable scale). The Markets",
        "page leads with the CFH ↔ Scandia callout. Keeping CFH's row",
        "accurate is higher leverage than any other BVB ticker.",
    ]
    for i, line in enumerate(notes, start=2):
        readme.cell(row=i, column=1, value=line)
    readme.column_dimensions["A"].width = 70

    return wb


def main(argv: list[str]) -> int:
    if "--help" in argv or "-h" in argv:
        print(__doc__)
        return 0
    out_path = (
        _REPO_ROOT
        / "scandi-desk-main"
        / "public"
        / "templates"
        / "bvb_financials_template.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out_path)
    print(f"BVB template written: {out_path}")
    print(f"  Rows: {len(bvb_universe())} BET-20 tickers")
    rich = sum(1 for r in bvb_universe().values() if r.get("revenue") is not None)
    print(f"  Pre-seeded (FY2024 numbers): {rich}/{len(bvb_universe())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
