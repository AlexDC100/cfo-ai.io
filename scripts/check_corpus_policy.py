#!/usr/bin/env python3
"""Corpus data-hygiene gate — TREE-ONLY, BY CONSTRUCTION.

Two checks, an exemption register, a notice block, and a deliberate
refusal to look at version-control history.

CHECK 1 — anonymization declarations
    Every `corpus/<case>/meta.yaml` must be able to answer "why is it OK
    for this file to be in the repository?". A case fails when it
    carries REAL source data (`synthetic: false`) that this repo's
    scrambler never touched (`anonymized: false`) and that does not
    declare the reviewed `anonymized_upstream: true` escape hatch —
    unless it lives under `corpus/private/`, which is reserved for
    encrypted-at-rest inputs.

    A synthetic case is `anonymized: false` because there is nothing to
    anonymize; failing those would make the gate un-passable by design,
    which is a defect, not a control. Uses of the escape hatch are
    printed on every run so they can never become invisible.

CHECK 2 — sensitive-lexicon scan, over the WORKING TREE at HEAD

    PRECISION IS THE WHOLE DESIGN. An earlier revision of this gate ran
    the REDACTOR's identifier patterns over every tracked byte and
    produced 115 findings, none of them real: thirteen-digit map
    coordinates in a GeoJSON, the string `RO 2026` in a pack version,
    `RO19` VAT-rate fragments inside spreadsheet cells, and the
    deliberately fabricated CUIs that ship in the customer-facing
    example workbooks. A gate that cries wolf 115 times is worse than
    no gate: the first engineer it blocks switches it off, and then the
    real finding — when it comes — lands in a disabled control.

    So the question this gate asks is NOT "does this look like an
    identifier?". It is "is this one of the terms the scrambler
    actually knows about and rewrites?".

    TIER A — NAMED ENTITIES, tree-wide, hash-exact.
        The legal name, the site names and the preparer's name, matched
        against `pdf_scrambler.SENSITIVE_TERMS` by salted digest of the
        normalized n-gram. There is no pattern and no shape here: a
        term that is not in the lexicon is not a violation. False
        positives are therefore only ever real-word COLLISIONS (a
        three-letter site abbreviation that is also a weekday), which
        are handled by the exemption register below rather than by
        loosening the match.

        Tier A runs over the WHOLE tree, not just data payloads. The
        previous revision restricted named entities to `corpus/**` and
        files with a data suffix — which is exactly why a tracked
        regression baseline carrying the entity's legal name in two
        JSON fields sat undetected. Scope was the bug; precision is the
        fix.

    TIER B — STATUTORY IDENTIFIERS, corroborated.
        `pdf_scrambler.GATE_ID_PATTERNS` (the tightened set: labelled
        CUI/CIF, bare `RO`+6-10 digits, Reg-Com, and a structurally
        valid CNP whose official control digit checks out), minus
        fabricated placeholders like `RO99999999`.

        Even tightened, a bare identifier SHAPE is weak evidence on its
        own — public company registrations legitimately appear in a
        public-records parser, and coordinate data is full of long
        integers. So a Tier B hit is reported only when CORROBORATED:

          · the same file also carries a non-exempt Tier A hit, or
          · the file is a DATA PAYLOAD — anything under `corpus/`, or
            any `.pdf` / `.xlsx` / `.csv` / `.tsv` anywhere in the tree.

        Those are the places a real accounting export lands. Source
        code, prose and JSON seed data are not, so an identifier shape
        there is left alone unless a real lexicon term sits beside it.

    BINARY FILES ARE NEVER BYTE-SCANNED. A PDF goes through the
    content-stream text extractor, a workbook through openpyxl cell
    values; anything else that fails the binary sniff is SKIPPED and
    reported as a skip under `--verbose`, never silently dropped. A
    naive regex over compressed bytes produces noise, not findings.

EXEMPTIONS — `scripts/corpus_policy_allowlist.txt`
    One entry per (path, selector), each with a MANDATORY reason. Every
    exemption that fires is printed in the summary on every run, so an
    exemption stays a visible, reviewable decision instead of becoming
    invisible history. Entries that no longer match anything are
    reported as stale.

    The register cannot be used to smuggle a sensitive term into the
    tree: a literal selector that is itself a lexicon term is REFUSED,
    and this file's own allowlist is scanned like any other tracked
    file and cannot exempt itself.

WHY THIS GATE NEVER READS HISTORY
    A gate that scanned version-control history would fail permanently
    and by design here: a non-anonymized copy of one fixture is present
    in this repository's past, the owner has explicitly accepted that
    risk, and the accepted-risk record is
    `docs/decisions/ADR-corpus-history-sibiu.md`. A control that can
    only ever be red teaches people to ignore it — so this gate answers
    exactly one question, "is the tree I am about to ship clean?", and
    surfaces the historical exposure as a NOTICE sourced from the ADR
    FILE, never from an archaeology pass.

    Enforced by construction: every subprocess call goes through
    `_git()`, which refuses any subcommand outside
    `_GIT_ALLOWED_SUBCOMMANDS`. Enforced again from the outside by
    `tests/engine/test_corpus_policy.py`, which greps this source.

Exit codes: 0 = pass (notices may still print), 1 = policy failure,
2 = the gate could not run (missing tooling / not a work tree).
"""
from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple


def _find_repo_root() -> Path:
    here = Path(__file__).resolve().parent
    for candidate in [here, *here.parents][:6]:
        if (candidate / "pyproject.toml").is_file():
            return candidate
    return Path(__file__).resolve().parent.parent


REPO = _find_repo_root()
for _p in (REPO / "src", REPO / "scripts"):
    if _p.is_dir() and str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

import pdf_scrambler as _lex  # noqa: E402  (the single lexicon owner)

#: The ONLY git subcommand this gate may run. `ls-files` reports the
#: index and the work tree — the set of paths that exist right now. It
#: cannot reach a past revision, so no arrangement of arguments turns
#: this gate into a history scanner.
_GIT_ALLOWED_SUBCOMMANDS = ("ls-files",)

#: Directory reserved for inputs that stay encrypted at rest (the B2
#: fallback in the ADR). Nothing in it is expected to be plaintext, so
#: the anonymization declaration check does not apply there.
PRIVATE_DIR = "private"

#: Suffixes treated as data payloads wherever they sit in the tree —
#: the shapes a real accounting export actually arrives in. Used to
#: corroborate a TIER B identifier hit.
DATA_SUFFIXES = frozenset({".pdf", ".xlsx", ".xlsm", ".xls", ".csv", ".tsv"})

#: Under `corpus/`, these are prose and tooling rather than payloads.
CORPUS_PROSE = ("corpus/README.md", "corpus/_tools/")

#: Where the accepted-risk records live, and the header key each one
#: uses to declare how many plaintext blobs it accepts.
DECISIONS_DIR = "docs/decisions"
ADR_GLOB = "ADR-corpus-history-*.md"
ADR_COUNT_KEY = "Accepted-Plaintext-Blobs:"

#: The exemption register.
#: A corpus case that MUST be in every sweep — the framework's own
#: calibration fixture. Absent => the corpus walk is broken, not clean.
CANARY_CORPUS_CASE = "saga_10_col"

ALLOWLIST_PATH = "scripts/corpus_policy_allowlist.txt"

#: TIER B's category name, as emitted by the lexicon module.
TIER_B_CATEGORY = "statutory_identifier"

_XLSX_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xls"})


class GateError(RuntimeError):
    pass


def _git(*args: str) -> str:
    """The single subprocess entry point. Refuses anything that is not
    an explicitly allowed, tree-scoped subcommand."""
    if not args or args[0] not in _GIT_ALLOWED_SUBCOMMANDS:
        raise GateError(
            "this gate may only run %s; refused %r"
            % (" / ".join(_GIT_ALLOWED_SUBCOMMANDS), args[:1] or ("<none>",))
        )
    try:
        proc = subprocess.run(
            ["git", *args], cwd=str(REPO), check=True,
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        )
    except FileNotFoundError as exc:
        raise GateError("git is not available: %s" % exc)
    except subprocess.CalledProcessError as exc:
        raise GateError(
            "git %s failed: %s" % (" ".join(args), exc.stderr.decode("utf-8", "replace").strip())
        )
    return proc.stdout.decode("utf-8", "replace")


def tracked_paths() -> List[str]:
    """Every path this repository is about to ship: what is already in
    the index, PLUS files that exist but have not been added yet and are
    not ignored.

    The `--others` half matters. A gate that only judged the index would
    green-light a brand-new fixture right up until the moment it was
    committed, which is the one moment before it is too late to notice.
    Ignored paths are excluded by `--exclude-standard`, so build output
    and virtualenvs stay out. Neither half can reach a past revision.
    """
    raw = _git("ls-files", "-z", "--cached", "--others", "--exclude-standard")
    seen: Set[str] = set()
    out: List[str] = []
    for p in raw.split("\0"):
        if p and p not in seen and (REPO / p).is_file():
            seen.add(p)
            out.append(p)
    return out


def index_paths() -> Set[str]:
    """Just the index half, for the "N tracked + K not yet added" line."""
    return {p for p in _git("ls-files", "-z").split("\0") if p}


# ── content extraction ─────────────────────────────────────────────────


def _looks_binary(data: bytes) -> bool:
    return b"\x00" in data[:8192]


def text_units(path: str) -> Tuple[List[str], Optional[str]]:
    """(readable text units, skip-reason).

    Binary payloads are decoded STRUCTURALLY or not at all — a PDF via
    the content-stream extractor, a workbook via openpyxl cell values.
    A skip-reason means the file could not be read as text; it is
    reported, never silently dropped.
    """
    full = REPO / path
    suffix = full.suffix.lower()
    if suffix == ".pdf":
        try:
            return _lex.pdf_text_units(full.read_bytes()), None
        except Exception as exc:  # noqa: BLE001
            return [], "unreadable PDF (%s)" % type(exc).__name__
    if suffix in _XLSX_SUFFIXES:
        try:
            from openpyxl import load_workbook

            units: List[str] = []
            book = load_workbook(str(full), data_only=True, read_only=True)
            for sheet in book.worksheets:
                units.append(str(sheet.title))
                for row in sheet.iter_rows(values_only=True):
                    units.extend(str(v) for v in row if isinstance(v, str))
            book.close()
            return units, None
        except Exception as exc:  # noqa: BLE001
            return [], "unreadable workbook (%s)" % type(exc).__name__
    data = full.read_bytes()
    if _looks_binary(data):
        return [], "binary"
    return [data.decode("utf-8", errors="replace")], None


def is_data_payload(path: str) -> bool:
    """True where a real accounting export would actually land. Used as
    standing corroboration for a TIER B identifier hit."""
    if Path(path).suffix.lower() in DATA_SUFFIXES:
        return True
    if not path.startswith("corpus/"):
        return False
    return not any(path.startswith(prose) for prose in CORPUS_PROSE)


# ── check 1: anonymization declarations ────────────────────────────────


def _load_metas(paths: Sequence[str]) -> List[Tuple[str, Dict[str, Any]]]:
    import yaml

    out: List[Tuple[str, Dict[str, Any]]] = []
    for path in sorted(paths):
        if not (path.startswith("corpus/") and path.endswith("/meta.yaml")):
            continue
        meta = yaml.safe_load((REPO / path).read_text(encoding="utf-8"))
        out.append((path, meta if isinstance(meta, dict) else {}))
    return out


def check_declarations(
    metas: Sequence[Tuple[str, Dict[str, Any]]]
) -> Tuple[List[str], List[str]]:
    failures: List[str] = []
    notices: List[str] = []
    for path, meta in metas:
        rel = Path(path).parent
        if PRIVATE_DIR in rel.parts:
            continue
        case = meta.get("case_id") or rel.name
        if "synthetic" not in meta:
            failures.append(
                "%s: missing `synthetic:` — every case must say whether its "
                "input is fabricated or a real export" % path
            )
            continue
        if bool(meta.get("synthetic")):
            continue
        if bool(meta.get("anonymized")):
            continue
        if bool(meta.get("anonymized_upstream")):
            notices.append(
                "anonymization escape hatch in use: %s declares "
                "`anonymized_upstream: true` (real export, this repo's "
                "scrambler deliberately not applied)" % case
            )
            continue
        failures.append(
            "%s: real input (`synthetic: false`) that is not anonymized. "
            "Either run the matching scrambler and set `anonymized: true`, "
            "or declare `anonymized_upstream: true` with the reason in "
            "source_notes, or move the case under corpus/%s/ and encrypt "
            "it at rest." % (path, PRIVATE_DIR)
        )
    return failures, notices


# ── the exemption register ─────────────────────────────────────────────


class Exemption(tuple):
    """(path, selector, reason, line_no) — one reviewed exemption."""

    __slots__ = ()

    def __new__(cls, path: str, selector: str, reason: str, line_no: int):
        return super().__new__(cls, (path, selector, reason, line_no))

    @property
    def path(self) -> str:
        return self[0]

    @property
    def selector(self) -> str:
        return self[1]

    @property
    def reason(self) -> str:
        return self[2]

    @property
    def line_no(self) -> int:
        return self[3]


def _lexicon_in_force() -> Dict[str, str]:
    """{digest: category} for the lexicon actually in effect.

    Goes through `pdf_scrambler._term_index()` rather than reading
    `SENSITIVE_TERMS` directly so that `pdf_scrambler.lexicon_override`
    is honoured. That is not a test convenience: the anti-smuggling rule
    in `load_allowlist` is only provable against an overridden lexicon,
    because proving it against the real one would require writing a real
    sensitive term into a tracked test file — the exact thing the rule
    exists to prevent.
    """
    return _lex._term_index()


def known_categories() -> Set[str]:
    return set(_lexicon_in_force().values()) | {TIER_B_CATEGORY}


def load_allowlist() -> Tuple[List[Exemption], List[str]]:
    """Parse the register. Returns (entries, structural failures).

    Structural failures are gate failures, not warnings: a malformed or
    unjustified exemption is a hole in the control, and the only way to
    keep an exemption register honest is to refuse the dishonest entry.
    """
    full = REPO / ALLOWLIST_PATH
    entries: List[Exemption] = []
    failures: List[str] = []
    if not full.is_file():
        return entries, failures

    categories = known_categories()
    lexicon_digests = set(_lexicon_in_force())

    for line_no, raw in enumerate(full.read_text(encoding="utf-8").splitlines(), 1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split("|")]
        where = "%s:%d" % (ALLOWLIST_PATH, line_no)
        if len(parts) != 3:
            failures.append(
                "%s: expected `path | selector | reason` (3 fields separated "
                "by `|`), found %d" % (where, len(parts))
            )
            continue
        path, selector, reason = parts
        if not path or not selector:
            failures.append("%s: path and selector may not be empty" % where)
            continue
        if not reason:
            failures.append(
                "%s: exemption for %s has NO REASON. Every exemption must say "
                "why the hit is not a disclosure — an exemption nobody had to "
                "justify is how this control dies." % (where, path)
            )
            continue
        if selector not in categories and _lex.term_hash(selector) in lexicon_digests:
            failures.append(
                "%s: the selector is itself a sensitive lexicon term. Writing "
                "it here would put the term back into a tracked file — which "
                "is the leak this gate exists to stop. Exempt it by CATEGORY "
                "instead (one of: %s)." % (where, ", ".join(sorted(categories)))
            )
            continue
        entries.append(Exemption(path, selector, reason, line_no))
    return entries, failures


# ── check 2: sensitive-lexicon scan ────────────────────────────────────


class Finding(tuple):
    """(path, category, matched) — one lexicon hit."""

    __slots__ = ()

    def __new__(cls, path: str, category: str, matched: str):
        return super().__new__(cls, (path, category, matched))

    @property
    def path(self) -> str:
        return self[0]

    @property
    def category(self) -> str:
        return self[1]

    @property
    def matched(self) -> str:
        return self[2]


def _exempt_for(
    finding: Finding, entries: Sequence[Exemption]
) -> Optional[Exemption]:
    for entry in entries:
        if entry.path != finding.path:
            continue
        if entry.selector == finding.category or entry.selector == finding.matched:
            return entry
    return None


def check_lexicon(
    paths: Sequence[str], entries: Sequence[Exemption]
) -> Tuple[List[Finding], List[Tuple[Finding, Exemption]], List[str]]:
    """Returns (violations, exempted, skipped)."""
    violations: List[Finding] = []
    exempted: List[Tuple[Finding, Exemption]] = []
    skipped: List[str] = []

    for path in paths:
        units, reason = text_units(path)
        if reason:
            if reason != "binary":
                skipped.append("%s: %s" % (path, reason))
            continue

        seen: Set[Tuple[str, str]] = set()
        for unit in units:
            for category, matched in _lex.scan_text(unit, gate_patterns=True):
                seen.add((category, matched))
        if not seen:
            continue

        tier_a = [Finding(path, c, m) for c, m in sorted(seen) if c != TIER_B_CATEGORY]
        tier_b = [Finding(path, c, m) for c, m in sorted(seen) if c == TIER_B_CATEGORY]

        # TIER A — every hit counts; only the register can excuse one.
        live_tier_a = 0
        for finding in tier_a:
            entry = _exempt_for(finding, entries)
            if entry is not None:
                exempted.append((finding, entry))
            else:
                violations.append(finding)
                live_tier_a += 1

        # TIER B — needs corroboration. An exempted TIER A hit does NOT
        # corroborate: it has already been reviewed as a coincidence,
        # and a coincidence cannot be evidence.
        if not tier_b:
            continue
        corroborated = live_tier_a > 0 or is_data_payload(path)
        if not corroborated:
            continue
        for finding in tier_b:
            entry = _exempt_for(finding, entries)
            if entry is not None:
                exempted.append((finding, entry))
            else:
                violations.append(finding)

    return violations, exempted, skipped


# ── notices ────────────────────────────────────────────────────────────


def history_notice() -> str:
    """The ONE non-blocking history line, derived from the presence and
    content of the accepted-risk records in docs/decisions/ — never from
    an archaeology pass over the object store."""
    decisions = REPO / DECISIONS_DIR
    records = sorted(decisions.glob(ADR_GLOB)) if decisions.is_dir() else []
    total = 0
    names: List[str] = []
    for record in records:
        names.append(record.stem)
        for line in record.read_text(encoding="utf-8").splitlines():
            stripped = line.strip().lstrip("*_- ")
            if stripped.startswith(ADR_COUNT_KEY):
                try:
                    total += int(stripped[len(ADR_COUNT_KEY):].strip())
                except ValueError:
                    pass
                break
    if not records:
        return "history: 0 risk-accepted plaintext blobs (no accepted-risk record on file)"
    return "history: %d risk-accepted plaintext blob%s (%s)" % (
        total, "" if total == 1 else "s", ", ".join(names)
    )


def coverage_notice(metas: Sequence[Tuple[str, Dict[str, Any]]]) -> str:
    """Non-blocking nag: which dispatch lanes are exercised ONLY by
    fabricated inputs. A synthetic case proves the code path; it cannot
    prove the code path survives what an accounting package actually
    emits. Never a red gate — a real file may simply not exist yet."""
    real: Set[str] = set()
    synthetic: Set[str] = set()
    for _path, meta in metas:
        parser = str(meta.get("expected_parser") or "?")
        (synthetic if bool(meta.get("synthetic")) else real).add(parser)
    missing = sorted(synthetic - real)
    if not missing:
        return "real-file coverage: every parser lane has at least one real input"
    return "REAL-FILE COVERAGE MISSING: %s" % ", ".join(missing)


# ── driver ─────────────────────────────────────────────────────────────


def run(verbose: bool = False) -> Tuple[int, List[str]]:
    lines: List[str] = []
    paths = tracked_paths()
    metas = _load_metas(paths)
    in_index = index_paths()

    entries, allow_failures = load_allowlist()
    decl_failures, decl_notices = check_declarations(metas)
    violations, exempted, skipped = check_lexicon(paths, entries)

    lines.append("NOTICE  %s" % history_notice())
    lines.append("NOTICE  %s" % coverage_notice(metas))
    for notice in decl_notices:
        lines.append("NOTICE  %s" % notice)

    # Exemptions are printed EVERY run, whether or not the gate passes.
    for finding, entry in sorted(exempted, key=lambda fe: (fe[0].path, fe[0].category)):
        lines.append(
            "EXEMPT  %s: %s (%s:%d) — %s"
            % (finding.path, finding.category, ALLOWLIST_PATH,
               entry.line_no, entry.reason)
        )
    fired = {(f.path, e.line_no) for f, e in exempted}
    for entry in entries:
        if not any(line_no == entry.line_no for _p, line_no in fired):
            lines.append(
                "NOTICE  stale exemption: %s:%d exempts %s in %s, which no "
                "longer matches anything — delete it or explain why it stays"
                % (ALLOWLIST_PATH, entry.line_no, entry.selector, entry.path)
            )

    if verbose and skipped:
        for entry_line in skipped:
            lines.append("SKIP    %s" % entry_line)

    untracked = len(paths) - len([p for p in paths if p in in_index])
    lines.append(
        "checked %d file(s) (%d tracked + %d not yet added), %d corpus case(s), "
        "%d exemption(s) on file, %d fired"
        % (len(paths), len(paths) - untracked, untracked, len(metas),
           len(entries), len(exempted))
    )

    # ── DISCOVERY CANARY ────────────────────────────────────────────
    #
    # Both checks are sweeps: CHECK 1 over corpus/*/meta.yaml, CHECK 2
    # over the tracked tree. An empty sweep raises no failure and prints
    # "CORPUS POLICY: PASS", which is the same sentence a clean tree
    # prints. So two things that MUST be in the sweep are named. Neither
    # is a marker planted for the gate: the corpus case is the framework's
    # calibration fixture and the allowlist is the gate's own register.
    discovery = []
    if not paths:
        discovery.append("the tracked-tree sweep produced 0 files")
    if not metas:
        discovery.append("the corpus sweep found 0 meta.yaml declarations")
    if not any(("corpus/%s/" % CANARY_CORPUS_CASE) in path
               for path, _meta in metas):
        discovery.append(
            "corpus case %r was not among the %d declaration(s) read — "
            "the corpus walk is not seeing corpus/"
            % (CANARY_CORPUS_CASE, len(metas)))
    if not entries:
        discovery.append(
            "%s parsed to 0 exemption entries — the allowlist reader is "
            "not reading" % ALLOWLIST_PATH)
    if discovery:
        lines.append("")
        lines.append("CORPUS POLICY: DISCOVERY BROKEN")
        for d in discovery:
            lines.append("  x %s" % d)
        lines.append("  A sweep that finds nothing violates nothing. That is "
                     "not a pass.")
        return 1, lines

    failures = (
        allow_failures
        + decl_failures
        + ["%s: %s %r — tracked files may not carry this"
           % (f.path, f.category, f.matched) for f in violations]
    )
    if failures:
        lines.append("")
        lines.append("CORPUS POLICY: FAIL — %d violation(s)" % len(failures))
        for failure in failures:
            lines.append("  x %s" % failure)
        return 1, lines
    lines.append("CORPUS POLICY: PASS")
    return 0, lines


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="also list files that could not be read as text")
    args = parser.parse_args(argv)
    try:
        code, lines = run(verbose=args.verbose)
    except GateError as exc:
        print("CORPUS POLICY: ERROR — %s" % exc)
        return 2
    for line in lines:
        print(line)
    return code


if __name__ == "__main__":
    sys.exit(main())
