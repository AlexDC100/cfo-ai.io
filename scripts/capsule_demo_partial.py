#!/usr/bin/env python3
"""THE GROUNDED NUMERIC DEMO — ENGINE PATH ONLY. PARTIAL.

WHAT THIS IS. A real number, from a real company's real trial balance,
traced to the source cell it came from, with the arithmetic checked
rather than asserted.

WHAT THIS IS NOT, and the label is not decoration. The full demo shows a
user typing a question into the Capsule and clicking a provenance dot in
the browser. That needs a signed-in workspace, which needs Supabase —
and the local test path was just pinned AWAY from production because it
had created 8,880 junk organisations there. So the browser half cannot
run here until a non-production Supabase project exists.

This proves the CHAIN: served fact -> balance-sheet row -> account code
-> leaf account -> the cell in the uploaded file. It does not prove the
SURFACE. Both matter; only one is demonstrated here.

Source: corpus/saga_10_col_carniprod — a REAL export (`synthetic: false`),
anonymized by label only, with "codes and all numerics preserved to the
cent" per its own meta.yaml. TC-1: this is real engine output, not a
hand-built fixture.

Run: .venv/bin/python scripts/capsule_demo_partial.py
"""
import json
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CASE = os.path.join(ROOT, "corpus", "saga_10_col_carniprod")


def money(cents):
    return "{:,.2f}".format(cents / 100.0).replace(",", " ")


def main():
    meta = open(os.path.join(CASE, "meta.yaml"), encoding="utf-8").read()
    period = [l.split(":", 1)[1].strip().strip('"')
              for l in meta.splitlines() if l.startswith("period:")][0]
    period_end = [l.split(":", 1)[1].strip().strip('"')
                  for l in meta.splitlines() if l.startswith("period_end:")][0]
    synthetic = [l.split(":", 1)[1].strip()
                 for l in meta.splitlines() if l.startswith("synthetic:")][0]

    facts = json.load(open(os.path.join(CASE, "expected", "gateway_facts.json"),
                           encoding="utf-8"))
    env = json.load(open(os.path.join(CASE, "expected", "served_envelope.json"),
                         encoding="utf-8"))

    print("=" * 68)
    print("GROUNDED NUMERIC DEMO — PARTIAL (engine path, no browser)")
    print("=" * 68)
    print("  company    : Carniprod (labels anonymized, numerics preserved)")
    print("  period     : %s   ending %s" % (period, period_end))
    print("  synthetic  : %s   <- a REAL trial balance" % synthetic)
    print("  tier       : %s" % facts["tier"])
    print("  currency   : %s" % facts["currency"])
    print()

    # ── 1. THE NUMBER, as the facts gateway serves it ────────────────
    total_cents = facts["total_assets_cents"]
    print("1. THE FACT, as the gateway serves it")
    print("   total_assets = %s %s" % (money(total_cents), facts["currency"]))
    print("   (stored in CENTS as %d — integer, so no float drift)" % total_cents)
    print()

    # ── 2. THE ROWS IT IS COMPOSED OF ────────────────────────────────
    # DERIVE the asset sections from the envelope; do not hardcode them.
    #
    # My first version listed ("non_current_assets", "current_assets") by
    # hand and the arithmetic missed by 283,298.95 — exactly the
    # `prepaid_expenses` section I had not thought of. The demo REFUSED
    # to close rather than presenting a number whose chain did not add
    # up, which is the behaviour, and then the envelope's own `sections`
    # array said what I had left out.
    #
    # In this schema the asset sections are the ones that precede
    # `equity`; everything from `equity` onward is the funding side.
    order = [sec["id"] for sec in env.get("sections") or []]
    if "equity" not in order:
        print("   CANNOT DERIVE asset sections — the envelope declares no")
        print("   `equity` section, so the assets/funding split is unknown.")
        return 1
    asset_sections = tuple(order[:order.index("equity")])
    rows = [r for r in env["rows"] if r.get("section") in asset_sections]
    print("2. THE ROWS THAT COMPOSE IT  (%d rows across %d asset sections)"
          % (len(rows), len(asset_sections)))
    print("   sections (derived from the envelope): %s" % ", ".join(asset_sections))
    for r in sorted(rows, key=lambda x: -abs(x.get("amount") or 0))[:6]:
        print("   %-28s %18s   accounts %s"
              % (r["label"][:28], money(round((r["amount"] or 0) * 100)),
                 ",".join(r.get("account_codes") or []) or "-"))
    print("   ... %d more" % max(0, len(rows) - 6))
    print()

    # ── 3. THE ARITHMETIC, CHECKED ───────────────────────────────────
    summed = round(sum((r.get("amount") or 0) for r in rows) * 100)
    print("3. THE ARITHMETIC — checked, not asserted")
    print("   sum of asset rows : %s" % money(summed))
    print("   gateway total     : %s" % money(total_cents))
    delta = summed - total_cents
    print("   difference        : %s cent(s)" % delta)
    if delta != 0:
        print()
        print("   MISMATCH. The chain does not close, and this demo will not")
        print("   claim it does. Investigate before presenting this number.")
        return 1
    print("   -> the served fact IS the sum of the rows it cites. Closed.")
    print()

    # ── 4. THE PROVENANCE JUMP ───────────────────────────────────────
    biggest = max(rows, key=lambda r: abs(r.get("amount") or 0))
    print("4. THE PROVENANCE JUMP — one figure, all the way down")
    print("   figure        : %s  (%s)"
          % (money(round(biggest["amount"] * 100)), biggest["label"]))
    print("   row id        : %s" % biggest["id"])
    print("   section       : %s" % biggest["section"])
    print("   account codes : %s" % ", ".join(biggest.get("account_codes") or []))
    print("   leaf accounts : %s" % ", ".join(biggest.get("leaf_ids") or []))
    print("   source file   : corpus/saga_10_col_carniprod/input.xlsx")
    print()
    # ── 4b. THE LAST LINK, OPENED RATHER THAN ASSERTED ───────────────
    # "It resolves to a real cell" is a claim. Open the workbook and add
    # the cells up.
    cells = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(os.path.join(CASE, "input.xlsx"), data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        closing_col = next(i for i, h in enumerate(header)
                           if h and str(h).lower().startswith("sold final debit"))
        wanted = set(biggest.get("leaf_ids") or [])
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            code = str(row[0].value).strip() if row[0].value is not None else ""
            if code in wanted:
                v = row[closing_col].value or 0
                cells.append((row[closing_col].coordinate, code, float(v)))
    except Exception as exc:                      # noqa: BLE001
        print("   COULD NOT OPEN THE SOURCE FILE: %s" % exc)
        print("   The chain above still closes, but the last link is")
        print("   unverified — say so rather than implying otherwise.")
        cells = None

    if cells is not None:
        print("   THE CELLS THEMSELVES, from %s:" % os.path.basename(CASE))
        for coord, code, v in cells:
            print("     %-5s  account %-6s  Sold Final Debit  %s"
                  % (coord, code, money(round(v * 100))))
        cell_sum = round(sum(v for _, _, v in cells) * 100)
        want = round(biggest["amount"] * 100)
        print("     %-5s  %-16s %s" % ("", "sum of cells", money(cell_sum)))
        print("     %-5s  %-16s %s" % ("", "served row", money(want)))
        print("     %-5s  %-16s %s cent(s)" % ("", "difference", cell_sum - want))
        if cell_sum != want:
            print()
            print("   THE LAST LINK DOES NOT CLOSE. Not presenting this as a")
            print("   proven jump.")
            return 1
        print()
        print("   -> the served figure IS the sum of those cells. The jump a")
        print("      reader makes by clicking the provenance dot lands here.")
    print()

    # ── 5. WHAT IS NOT DEMONSTRATED ──────────────────────────────────
    print("5. NOT DEMONSTRATED HERE, stated plainly")
    print("   - the browser surface: typing the question, the fact tile,")
    print("     clicking the dot and landing on the cell")
    print("   - anything requiring a signed-in workspace")
    print("   Both need a non-production Supabase project. The local test")
    print("   path was pinned away from production after it created 8,880")
    print("   junk organisations there.")
    print()
    print("VERDICT: the numeric chain is PROVEN. The surface is NOT.")

    out = {
        "status": "PARTIAL — engine path only",
        "company": "Carniprod",
        "period": period,
        "period_end": period_end,
        "synthetic": synthetic,
        "currency": facts["currency"],
        "total_assets_cents": total_cents,
        "total_assets_display": money(total_cents),
        "asset_rows": len(rows),
        "sum_of_rows_cents": summed,
        "difference_cents": delta,
        "traced_figure": {
            "label": biggest["label"],
            "amount_cents": round(biggest["amount"] * 100),
            "row_id": biggest["id"],
            "account_codes": biggest.get("account_codes"),
            "leaf_ids": biggest.get("leaf_ids"),
            "source_file": "corpus/saga_10_col_carniprod/input.xlsx",
        },
        "source_cells": [
            {"cell": c, "account": a, "closing_debit": v} for c, a, v in (cells or [])
        ],
        "not_demonstrated": [
            "browser surface: question -> fact tile -> provenance click -> source cell",
            "any signed-in workspace path",
        ],
        "blocked_on": "a non-production Supabase project",
    }
    dest = os.path.join(ROOT, "design_review", "demo", "grounded_numeric_partial.json")
    with open(dest, "w", encoding="utf-8") as fh:
        json.dump(out, fh, indent=2)
        fh.write("\n")
    print("recorded: design_review/demo/grounded_numeric_partial.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
