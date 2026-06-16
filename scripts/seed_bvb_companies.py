"""Idempotent BVB seed loader.

Writes the BET-20 seed (src/engine/public/bvb_seed.py) into the live
``public_companies`` + ``public_company_periods`` tables. Designed to
run anywhere — locally against a dev Supabase, in CI as part of the
deploy pipeline, or one-off on the VPS during sprint cutover.

What it does
------------
For each of the 20 BET-index tickers:

1. Upsert into ``public_companies`` with:
     ticker, name, sector, industry,
     exchange='BVB', country='RO', currency='RON',
     is_active=True, last_synced_at=NOW(),
     last_sync_notes={'source': 'seed_bvb', 'confidence': 0.40..0.92}

2. If the seed row has revenue set (7 of 20 today), also upsert into
   ``public_company_periods`` for dimension='ARY',
   fiscal_period_end='2024-12-31', with an ``assembled_canonical_v1``
   JSONB envelope mirroring the canonical_schema_v1 shape used for
   private periods. FE renderers don't branch on source — the same
   bucket-level totals show up in the same surfaces.

Sparse rows (no revenue) get the public_companies entry but no
periods row. The FE renders them as "data pending" in the
RomanianListedCard.

Optional xlsx override
----------------------
Pass ``--xlsx PATH`` to overlay values from a filled-in template
(see scripts/generate_bvb_template.py). Empty cells in the xlsx are
left as-is (seed default preserved). "NULL" (case-insensitive)
explicitly nulls the field.

Environment
-----------
Requires SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY env vars (same
contract as every other script in this repo). Loader uses the
admin context so RLS doesn't apply.

Usage
-----
   # First time: seed everything
   python scripts/seed_bvb_companies.py

   # Apply operator-filled xlsx on top of seed
   python scripts/seed_bvb_companies.py --xlsx scandi-desk-main/public/templates/bvb_financials_template.xlsx

   # Dry-run (print what would change, don't write)
   python scripts/seed_bvb_companies.py --dry-run

Exit codes
----------
  0 — all 20 rows upserted (or dry-run completed)
  1 — env vars missing / Supabase unreachable
  2 — xlsx supplied but couldn't parse
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional


_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT / "src"))

from engine.public.bvb_seed import bvb_universe  # noqa: E402

logger = logging.getLogger("seed_bvb")


# ── XLSX overlay ─────────────────────────────────────────────────────────

def _load_xlsx_overrides(xlsx_path: Path) -> Dict[str, Dict[str, Any]]:
    """Read the BET-20 sheet of a filled-in template, return a map of
    ticker → {column_name: cell_value} for cells that the operator has
    populated. Empty cells are skipped (preserve seed default).
    "NULL" (case-insensitive) returns the sentinel value ``None``."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl",
              file=sys.stderr)
        sys.exit(2)
    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found: {xlsx_path}", file=sys.stderr)
        sys.exit(2)
    wb = load_workbook(xlsx_path, data_only=True)
    if "BET-20" not in wb.sheetnames:
        print(f"ERROR: xlsx has no 'BET-20' sheet: {xlsx_path}", file=sys.stderr)
        sys.exit(2)
    ws = wb["BET-20"]
    header = [c.value for c in ws[1]]
    overrides: Dict[str, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        ticker = str(row[0]).strip()
        row_overrides: Dict[str, Any] = {}
        for col_name, val in zip(header, row):
            if col_name == "ticker":
                continue
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            if isinstance(val, str) and val.strip().upper() == "NULL":
                row_overrides[col_name] = None
            else:
                row_overrides[col_name] = val
        if row_overrides:
            overrides[ticker] = row_overrides
    return overrides


# ── Envelope builder ─────────────────────────────────────────────────────
# We persist a minimal assembled_canonical_v1 envelope so FE consumers
# that read ``methodology.totals.*`` get coherent numbers. This is NOT
# the full F4.1 emission — it's the subset of bucket-level totals that
# Markets cards and Benchmark/Peer rendering rely on. Operator-curated
# detail can be backfilled later via the admin upload path.

def _build_envelope(seed_row: Dict[str, Any], overrides: Dict[str, Any]) -> Dict[str, Any]:
    """Build a tiny assembled_canonical_v1 envelope from the seed row +
    any xlsx overrides. Returns None if the row has no revenue (sparse).
    """
    def _value(b_key: str, raw_key: Optional[str] = None) -> Optional[float]:
        """Read a value, preferring the xlsx override over the seed.
        ``b_key`` is the BILLIONS-scale column name in the xlsx
        (e.g. 'revenue_b'). ``raw_key`` is the seed dict key in raw RON
        (e.g. 'revenue')."""
        if b_key in overrides:
            v = overrides[b_key]
            return None if v is None else float(v) * 1_000_000_000
        return seed_row.get(raw_key or b_key.replace("_b", ""))

    revenue = _value("revenue_b", "revenue")
    if revenue is None:
        return None
    ebitda = _value("ebitda_b", "ebitda")
    net_income = _value("net_income_b", "netIncome")
    equity = _value("equity_b", "equity")
    cash = _value("cash_b", "cash")
    gross_debt = _value("gross_debt_b", "grossDebt")
    net_debt = _value("net_debt_b", "netDebt")

    envelope: Dict[str, Any] = {
        "canonical_version": "v1.0",
        "source": "seed_bvb",
        "confidence": float(overrides.get("__confidence__", seed_row.get("confidence", 0.85))),
        "methodology": {
            "totals": {
                "revenue": revenue,
                "ebitda": ebitda,
                "net_income": net_income,
                "total_equity": equity,
                "cash": cash,
                "gross_debt": gross_debt,
                "net_debt": net_debt,
            },
            "currency": "RON",
            "period_label": "FY2024",
        },
        "lineage": {
            "loader": "scripts/seed_bvb_companies.py",
            "loaded_at": datetime.now(timezone.utc).isoformat(),
        },
    }
    return envelope


# ── Upsert pipeline ──────────────────────────────────────────────────────

def _build_company_row(
    ticker: str, seed_row: Dict[str, Any], overrides: Dict[str, Any]
) -> Dict[str, Any]:
    """Build the public_companies row. Maps seed → schema columns."""
    confidence = float(overrides.get("__confidence__",
                                     seed_row.get("confidence", 0.85)))
    return {
        "ticker": ticker,
        "name": overrides.get("name") or seed_row["companyName"],
        "sector": overrides.get("sector") or seed_row.get("sector"),
        "industry": overrides.get("industry") or seed_row.get("industry"),
        "exchange": "BVB",
        "country": "RO",
        "currency": "RON",
        "is_active": True,
        "last_synced_at": datetime.now(timezone.utc).isoformat(),
        "last_sync_notes": {
            "source": "seed_bvb",
            "confidence": confidence,
            "loader_version": "phase1-2026-05-31",
        },
    }


def upsert_all(
    overrides_by_ticker: Dict[str, Dict[str, Any]],
    *,
    dry_run: bool = False,
) -> int:
    """Run the full upsert pipeline. Returns number of rows that would
    be (or were) written."""
    bvb = bvb_universe()
    company_writes = 0
    period_writes = 0

    if dry_run:
        # In dry-run mode, just print what would happen and skip the
        # Supabase connection. Useful in CI to validate the seed/xlsx
        # without touching the DB.
        for ticker, seed_row in bvb.items():
            ov = overrides_by_ticker.get(ticker, {})
            company_row = _build_company_row(ticker, seed_row, ov)
            print(f"[dry-run] public_companies UPSERT {ticker:<7} "
                  f"({company_row['name']!r})")
            company_writes += 1
            envelope = _build_envelope(seed_row, ov)
            if envelope is not None:
                print(f"[dry-run]   + period ARY 2024-12-31 "
                      f"(rev={envelope['methodology']['totals']['revenue']:,.0f})")
                period_writes += 1
            else:
                print(f"[dry-run]   - period skipped (sparse — no revenue)")
        print(f"\n[dry-run] would write {company_writes} company rows, "
              f"{period_writes} period rows")
        return company_writes

    # Real-write path: deferred import so dry-run + xlsx-only validation
    # don't require the engine.api package (and its httpx / supabase deps)
    # to be installed locally.
    try:
        from engine.api import _supabase
    except ImportError:
        print("ERROR: engine.api._supabase module not found. Ensure src/ is on PYTHONPATH "
              "and engine deps are installed (httpx, etc).", file=sys.stderr)
        return -1

    with _supabase.admin() as ac:
        for ticker, seed_row in bvb.items():
            ov = overrides_by_ticker.get(ticker, {})
            company_row = _build_company_row(ticker, seed_row, ov)
            result = ac.upsert(
                "public_companies",
                company_row,
                on_conflict="ticker",
                returning=True,
            )
            if not result:
                logger.warning("%s: empty upsert response — skipping period", ticker)
                continue
            company_id = result[0]["id"]
            company_writes += 1
            logger.info("public_companies %s → %s", ticker, company_id)

            envelope = _build_envelope(seed_row, ov)
            if envelope is None:
                logger.info("  + %s: sparse row, period skipped", ticker)
                continue

            period_row = {
                "public_company_id": company_id,
                "dimension": "ARY",
                "fiscal_period_end": "2024-12-31",
                "assembled_canonical_v1": envelope,
                "source_payload": None,  # No raw payload for seed rows
                "synced_at": datetime.now(timezone.utc).isoformat(),
            }
            ac.upsert(
                "public_company_periods",
                period_row,
                on_conflict="public_company_id,dimension,fiscal_period_end",
            )
            period_writes += 1

    print(f"Wrote {company_writes} public_companies rows, "
          f"{period_writes} public_company_periods rows.")
    return company_writes


# ── Entry point ──────────────────────────────────────────────────────────

def main(argv: list[str]) -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    )
    parser = argparse.ArgumentParser(description="Seed BVB companies into Supabase")
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="Path to filled-in template xlsx to overlay onto seed")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print what would change without writing")
    args = parser.parse_args(argv)

    overrides: Dict[str, Dict[str, Any]] = {}
    if args.xlsx is not None:
        overrides = _load_xlsx_overrides(args.xlsx)
        print(f"Loaded {len(overrides)} ticker overrides from {args.xlsx}")

    if not args.dry_run:
        if not os.environ.get("SUPABASE_URL") or not os.environ.get("SUPABASE_SERVICE_ROLE_KEY"):
            print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set",
                  file=sys.stderr)
            return 1

    n = upsert_all(overrides, dry_run=args.dry_run)
    return 0 if n >= 0 else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
