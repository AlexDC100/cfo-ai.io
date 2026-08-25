"""map_guided front-end — mechanical execution of an AI-interpreted
StructuralMap (AI-first-reader mission, Lane 2 Part B).

Covers, red-test-first for every logic edit to an existing file:

  * `engine.ir.ledgerdoc` — the "mechanical_mapped" provenance method:
    admitted to PROVENANCE_METHODS, subject to the SAME confidence==1.0
    rule as "mechanical" (the mechanical FAMILY, not the literal), and
    constructible via `Provenance.mechanical_mapped(source_ref)`.
  * `engine.ir.schema` — `structural_interpretation_meta` is a VOLATILE
    source_meta key: run-varying interpretation metadata (model ids,
    cache keys, run stamps) must NOT shift `content_hash`, while the
    map-derived structure itself stays content-bearing.
  * `engine.frontends.map_guided` — MapGuidedFrontEnd: real per-atom
    cell SourceRefs, six Money slots from map semantics, the two
    DISTINCT cumulative side-channels (total_with_opening vs
    movement_cumulative), ABSENT != ZERO, subtotal/repeated-header
    skips, classic-parity dropped-row diagnostics, and the legacy
    bridge (10-key rows byte-comparable against the classic parser —
    the C1 consensus seed proof on agras).

NO AI calls anywhere: the StructuralMaps used here are hand-written
dicts (the interp lane's fixture maps in
tests/engine/fixtures/structmaps/ had not landed when this suite was
written — the dicts use the same vocabulary and can be swapped for the
fixture files when they exist).
"""
from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any, Dict

import pytest

import engine.country_packs.ro_romania  # noqa: F401 — registers RomaniaPack
from engine.core.country_pack_registry import get_pack
from engine.ir import (
    DocHeader,
    LedgerDoc,
    Money,
    Provenance,
    SourceRef,
    content_hash,
    deserialize,
    serialize,
)
from engine.ir.ledgerdoc import PROVENANCE_METHODS, LedgerDocError


REPO = Path(__file__).resolve().parents[2]
SIBIU_XLSX = REPO / "files" / "scandia_sibiu_tb_2019.xlsx"
AGRAS_XLSX = REPO / "files" / "agras_tb_2025.xlsx"
STRUCTMAPS_DIR = Path(__file__).resolve().parent / "fixtures" / "structmaps"


def _canon(obj: Any) -> str:
    return json.dumps(obj, sort_keys=True, ensure_ascii=False, allow_nan=False)


# ═══════════════════════════════════════════════════════════════════════
# 1. ir.ledgerdoc — the mechanical_mapped provenance method
#    (RED first: the method is rejected today; confidence<1.0 escapes
#    the mechanical rule today because :262 tests the literal.)
# ═══════════════════════════════════════════════════════════════════════


def test_mechanical_mapped_is_a_valid_provenance_method() -> None:
    assert "mechanical_mapped" in PROVENANCE_METHODS
    prov = Provenance(
        source_ref=SourceRef.cell("Sheet1", 3, 0), method="mechanical_mapped"
    )
    assert prov.confidence == 1.0


def test_mechanical_mapped_enforces_confidence_one() -> None:
    """The confidence==1.0 rule applies to the mechanical FAMILY, not
    the literal "mechanical" — a mechanical_mapped read is a
    deterministic cell read and exact by definition."""
    with pytest.raises(LedgerDocError, match="exact by definition"):
        Provenance(
            source_ref=SourceRef.cell("Sheet1", 3, 0),
            method="mechanical_mapped",
            confidence=0.9,
        )


def test_mechanical_literal_confidence_rule_unchanged() -> None:
    with pytest.raises(LedgerDocError, match="exact by definition"):
        Provenance(
            source_ref=SourceRef.pipeline_stage("x"),
            method="mechanical",
            confidence=0.5,
        )


def test_provenance_mechanical_mapped_classmethod() -> None:
    ref = SourceRef.cell("Sheet1", 7, 0)
    prov = Provenance.mechanical_mapped(ref)
    assert prov.method == "mechanical_mapped"
    assert prov.confidence == 1.0
    assert prov.source_ref is ref


def test_mechanical_mapped_roundtrips_through_schema() -> None:
    from engine.ir import AccountAtom

    doc = LedgerDoc(
        header=DocHeader(jurisdiction="RO", currency="RON"),
        atoms=(
            AccountAtom(
                atom_id="r00000:5121",
                account_code="5121",
                label="Banca",
                provenance=Provenance.mechanical_mapped(
                    SourceRef.cell("Sheet1", 1, 0)
                ),
                closing_debit=Money.from_minor("RON", 12345),
            ),
        ),
    )
    payload = serialize(doc)
    assert payload["atoms"][0]["provenance"]["method"] == "mechanical_mapped"
    again = serialize(deserialize(payload))
    assert _canon(payload) == _canon(again)


# ═══════════════════════════════════════════════════════════════════════
# 2. ir.schema — structural_interpretation_meta is hash-volatile
#    (RED first: today the key shifts content_hash.)
# ═══════════════════════════════════════════════════════════════════════


def _doc_with_meta(meta: Dict[str, Any]) -> LedgerDoc:
    return LedgerDoc(
        header=DocHeader(jurisdiction="RO", currency="RON", source_meta=meta)
    )


def test_structural_interpretation_meta_does_not_shift_content_hash() -> None:
    """Run-varying interpretation metadata (model ids, cache keys, run
    stamps of the two independent AI map runs) parks under the ONE
    top-level `structural_interpretation_meta` key, which is excluded
    from content_hash wholesale — two mechanical reads of the same
    bytes under the same map must hash identically across runs."""
    base = {"front_end": "map_guided@map_guided_v1"}
    a = _doc_with_meta(
        dict(base, structural_interpretation_meta={"model": "m1", "run": "r1"})
    )
    b = _doc_with_meta(
        dict(base, structural_interpretation_meta={"model": "m2", "run": "r2"})
    )
    c = _doc_with_meta(dict(base))
    assert content_hash(a) == content_hash(b) == content_hash(c)


def test_map_structure_stays_content_bearing() -> None:
    """The MAP-DERIVED structure itself (columns, header row, ...) is
    content-bearing: a different map over the same bytes is a different
    reading and MUST hash differently."""
    a = _doc_with_meta({"structural_map": {"header_row_index": 0}})
    b = _doc_with_meta({"structural_map": {"header_row_index": 3}})
    assert content_hash(a) != content_hash(b)


# ═══════════════════════════════════════════════════════════════════════
# 3. MapGuidedFrontEnd — mechanical execution over the real fixtures
# ═══════════════════════════════════════════════════════════════════════

from engine.frontends import derive_legacy, resolve_front_end  # noqa: E402
from engine.frontends.map_guided import (  # noqa: E402
    META_MOVEMENT_CUMULATIVE_PAIR,
    META_TOTAL_WITH_OPENING_PAIR,
    MapGuidedFrontEnd,
    _parse_cell,
)
from engine.frontends.saga10 import FrontEndError  # noqa: E402


def _load_map(name: str) -> Dict[str, Any]:
    return json.loads((STRUCTMAPS_DIR / name).read_text(encoding="utf-8"))


def _parse_fixture(path: Path, map_name: str, **extra_hints: Any):
    fe = resolve_front_end("map_guided")
    smap = _load_map(map_name)
    hints = {
        "structural_map": smap,
        "jurisdiction": "RO",
        "filename": path.name,
        "interpreter_roles": ["structural_interpreter_a",
                              "structural_interpreter_b"],
        "map_prompt_versions": {"structural_interpreter_a": "fixture",
                                "structural_interpreter_b": "fixture"},
    }
    hints.update(extra_hints)
    return fe.parse(path.read_bytes(), hints)


# ── (a) sibiu: a file the classic parser REJECTS, read via its map ─────


def test_sibiu_is_rejected_by_the_classic_parser() -> None:
    """The C1/C2 boundary motivation, kept honest: sibiu's abbreviated
    D/C headers defeat the classic header detector — the map lane reads
    a file the deterministic lane cannot."""
    from engine.country_packs.ro_romania.trial_balance_parser import ParseError

    with pytest.raises(ParseError):
        get_pack("RO").parse_trial_balance(
            SIBIU_XLSX.read_bytes(), SIBIU_XLSX.name
        )


def test_sibiu_map_guided_ledgerdoc() -> None:
    doc, diagnostics = _parse_fixture(SIBIU_XLSX, "scandia_sibiu_tb_2019.json")

    # 249 data rows, every code passes the shape discipline.
    assert len(doc.atoms) == 249
    codes = {d["code"] for d in diagnostics}
    assert "cell_provenance_unavailable" not in codes
    assert "dropped_rows" not in codes  # nothing to drop in this file

    # Real per-atom cell provenance — fresh Provenance per atom.
    first = doc.atoms[0]
    assert first.provenance.method == "mechanical_mapped"
    assert first.provenance.confidence == 1.0
    assert first.provenance.source_ref.sheet == "Sheet1"
    assert first.provenance.source_ref.row == 1
    assert first.provenance.source_ref.col == 0
    assert doc.atoms[5].provenance.source_ref.row == 6
    assert doc.atoms[0].provenance is not doc.atoms[1].provenance

    # Six Money slots populated from map semantics; the total_with_opening
    # pair rides ITS OWN side-channel (never the six slots), and the
    # movement_cumulative channel is absent — this file has no such pair.
    meta = doc.header.source_meta
    assert first.opening_debit is not None and first.period_debit is not None
    assert first.closing_debit is not None
    two_pair = meta[META_TOTAL_WITH_OPENING_PAIR]
    assert len(two_pair) > 0
    assert META_MOVEMENT_CUMULATIVE_PAIR not in meta

    # Row 3 fact from the hand-verified map: opening 1,549,139.17
    # reappears verbatim in the total_with_opening pair.
    atom_117 = doc.atoms[2]
    assert atom_117.account_code == "117101"
    assert atom_117.opening_debit == Money.from_minor("RON", 154913917)
    assert two_pair[atom_117.atom_id][0] == "1549139.17"

    # Extraction stamp — method never empty, map identity pinned.
    extraction = dict(meta["extraction"])
    assert extraction["method"] == "mechanical_mapped"
    assert extraction["parser_version"] == "map_guided_v1"
    assert extraction["source_format"] == "map_guided"
    assert extraction["sheet"] == "Sheet1"
    assert extraction["header_row_index"] == 0
    assert extraction["map_version"] == "smap1"
    assert extraction["map_hash"]
    assert list(extraction["interpreter_roles"]) == [
        "structural_interpreter_a", "structural_interpreter_b"
    ]

    # No totals row anywhere in the sheet — no document_totals, and the
    # anchor must say so honestly.
    assert doc.header.document_totals is None


def test_sibiu_121_closing_and_anchor_via_legacy_bridge() -> None:
    doc, _ = _parse_fixture(SIBIU_XLSX, "scandia_sibiu_tb_2019.json")
    view = derive_legacy(doc)
    tb = view.tb_rows
    assert len(tb) == 249
    anchor = tb.source_anchor
    # Honest anchor: the file has NO totals row, so there is no third leg.
    assert anchor["anchor_status"] == "NO_ANCHOR"
    assert anchor["totals_row_found"] is False
    # The 121 closing net, in exact cents, from the REAL pack seam.
    # (File fact: the single 121 row carries sf_c 650,887.06 — verified
    # against the workbook; 650887.00 in the lane brief was imprecise.)
    closing = anchor["closing_result"]
    assert closing["p121_cents"] == 65088706
    # And the same value straight from the IR atoms, integer cents.
    net_cents = 0
    for atom in doc.atoms:
        if atom.account_code.startswith("121"):
            for money, sign in ((atom.closing_credit, 1), (atom.closing_debit, -1)):
                assert money is not None and money.scale == 2
                net_cents += sign * money.amount_minor
    assert net_cents == 65088706
    # Extraction block flows through the carrier for C2.
    assert tb.extraction["method"] == "mechanical_mapped"


# ── (b) agras: the C1 seed proof — mapped leg vs classic leg ───────────


def test_agras_c1_rows_byte_equal_to_classic() -> None:
    """Both legs read the same cells of the same workbook — the mapped
    lane through the hand-verified StructuralMap, the classic lane
    through header detection. The 10-key rows must agree byte-for-byte
    (a single ulp of drift fails), and so must the recomputed anchor
    and the pack's closing_result. This is the seed of every future C1
    consensus check."""
    content = AGRAS_XLSX.read_bytes()
    classic = get_pack("RO").parse_trial_balance(content, AGRAS_XLSX.name)

    doc, diagnostics = _parse_fixture(AGRAS_XLSX, "agras_tb_2025.json")
    view = derive_legacy(doc)
    tb = view.tb_rows

    # Row-count parity (the classic parser drops non-code rows silently;
    # the mapped lane drops the same rows and SAYS so).
    assert len(tb) == len(classic) == len(doc.atoms)
    assert any(d["code"] == "dropped_rows" for d in diagnostics)

    assert _canon(list(tb)) == _canon(list(classic))
    assert _canon(tb.source_anchor) == _canon(classic.source_anchor)

    # agras has NO totals row: both legs must degrade explicitly.
    assert tb.source_anchor["anchor_status"] == "NO_ANCHOR"

    # The mapped lane keeps the two cumulative semantics DISTINCT: this
    # file's pair is movement-cumulative (movements only), so it rides
    # that channel — while the legacy bridge still lands it in st_*
    # (where the classic parser files it today) for byte-parity.
    meta = doc.header.source_meta
    assert META_MOVEMENT_CUMULATIVE_PAIR in meta
    assert META_TOTAL_WITH_OPENING_PAIR not in meta


# ── (c) determinism ────────────────────────────────────────────────────


def test_same_bytes_and_map_twice_is_byte_identical() -> None:
    doc_a, _ = _parse_fixture(SIBIU_XLSX, "scandia_sibiu_tb_2019.json")
    doc_b, _ = _parse_fixture(SIBIU_XLSX, "scandia_sibiu_tb_2019.json")
    assert _canon(serialize(doc_a)) == _canon(serialize(doc_b))
    assert content_hash(doc_a) == content_hash(doc_b)


def test_interpretation_run_metadata_is_hash_volatile_end_to_end() -> None:
    doc_a, _ = _parse_fixture(
        SIBIU_XLSX, "scandia_sibiu_tb_2019.json",
        interpretation_meta={"model": "m1", "run_id": "r1"},
    )
    doc_b, _ = _parse_fixture(
        SIBIU_XLSX, "scandia_sibiu_tb_2019.json",
        interpretation_meta={"model": "m2", "run_id": "r2"},
    )
    assert _canon(serialize(doc_a)) != _canon(serialize(doc_b))
    assert content_hash(doc_a) == content_hash(doc_b)


# ── (d) absent pairs, subtotal skips, dropped rows, totals row ─────────


def _mini_workbook() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "TB"
    for row in (
        ["Cont", "Denumire", "SiD", "SiC", "RD", "RC", "SfD", "SfC"],
        ["TOTAL", "", 100, 100, 50, 50, 100, 100],          # totals (SAGA-first)
        ["1012", "Capital", 0, 100, 0, 0, 0, 100],
        ["SUBTOTAL cls 1", "", 0, 100, 0, 0, 0, 100],       # map-marked subtotal
        ["5121", "Banca", 100, 0, 50, 50, 100, 0],
        [None, "junk", 1, 2, 0, 0, 0, 0],                   # blank code
        ["Banner row", "x", 0, 0, 0, 0, 0, 0],              # fails code shape
    ):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mini_map(with_closing: bool, with_totals: bool) -> Dict[str, Any]:
    columns = [
        {"index": 0, "semantic": "account_code"},
        {"index": 1, "semantic": "account_name"},
        {"index": 2, "semantic": "opening_debit"},
        {"index": 3, "semantic": "opening_credit"},
        {"index": 4, "semantic": "movement_period_debit"},
        {"index": 5, "semantic": "movement_period_credit"},
    ]
    if with_closing:
        columns += [
            {"index": 6, "semantic": "closing_debit"},
            {"index": 7, "semantic": "closing_credit"},
        ]
    return {
        "map_version": "smap1",
        "map_hash": "test-mini",
        "sheet": "TB",
        "header_row_index": 0,
        "account_code_col": 0,
        "columns": columns,
        "totals_row_indexes": [1] if with_totals else [],
        "subtotal_row_indexes": [3],
        "repeated_header_rows": [],
        "number_locale": {"decimal_sep": ".", "thousands_sep": None},
        "currency": "RON",
    }


def test_mini_skips_drops_and_totals() -> None:
    data = _mini_workbook()
    doc, diagnostics = MapGuidedFrontEnd().parse(data, {
        "structural_map": _mini_map(with_closing=True, with_totals=True),
        "jurisdiction": "RO",
        "filename": "mini.xlsx",
    })
    assert [a.account_code for a in doc.atoms] == ["1012", "5121"]

    by_code = {d["code"]: d["detail"] for d in diagnostics}
    assert "dropped_rows" in by_code
    assert "2 data-region row(s) dropped: 1 blank/absent account code, 1 " \
        in by_code["dropped_rows"]
    assert "skipped_marked_rows" in by_code
    assert "1 row(s) skipped as map-marked (subtotal)" in by_code["skipped_marked_rows"]

    totals = doc.header.document_totals
    assert totals is not None
    assert totals.opening_debit == Money.from_minor("RON", 10000)
    assert totals.period_credit == Money.from_minor("RON", 5000)
    assert totals.closing_credit == Money.from_minor("RON", 10000)

    view = derive_legacy(doc)
    anchor = view.tb_rows.source_anchor
    assert anchor["anchor_status"] == "MATCHED"
    assert anchor["totals_row_index"] == 1
    # No cumulative pair mapped -> the rc pair is honestly null.
    assert anchor["pairs"]["rc"] is None
    assert anchor["source_balanced"] is True


def test_mini_absent_closing_pair_stays_absent_then_synthesizes() -> None:
    data = _mini_workbook()
    doc, _ = MapGuidedFrontEnd().parse(data, {
        "structural_map": _mini_map(with_closing=False, with_totals=False),
        "jurisdiction": "RO",
        "filename": "mini.xlsx",
    })
    # ABSENT pair => None on every atom, and the serialized form OMITS
    # the keys (never Money.zero, never null).
    for atom in doc.atoms:
        assert atom.closing_debit is None and atom.closing_credit is None
    payload = serialize(doc)
    for raw_atom in payload["atoms"]:
        assert "closing_debit" not in raw_atom
        assert "closing_credit" not in raw_atom

    # The legacy bridge re-runs the classic Layout-B identity synthesis.
    view = derive_legacy(doc)
    rows = list(view.tb_rows)
    assert rows[0]["cont"] == "1012"
    assert (rows[0]["sf_d"], rows[0]["sf_c"]) == (0.0, 100.0)
    assert (rows[1]["sf_d"], rows[1]["sf_c"]) == (100.0, 0.0)
    sf_pair = view.tb_rows.source_anchor["pairs"]["sf"]
    assert sf_pair["synthesized_from_identity"] is True


def test_map_validation_errors() -> None:
    data = _mini_workbook()
    fe = MapGuidedFrontEnd()
    with pytest.raises(FrontEndError, match="structural_map"):
        fe.parse(data, {"jurisdiction": "RO"})
    with pytest.raises(FrontEndError, match="jurisdiction"):
        smap = _mini_map(True, False)
        smap.pop("currency")
        fe.parse(data, {"structural_map": dict(smap, currency="RON")})
    half = _mini_map(True, False)
    half["columns"] = [c for c in half["columns"]
                       if c["semantic"] != "closing_credit"]
    with pytest.raises(FrontEndError, match="one side of the closing pair"):
        fe.parse(data, {"structural_map": half, "jurisdiction": "RO"})
    scaled = dict(_mini_map(True, False), scale=1000)
    with pytest.raises(FrontEndError, match="scale"):
        fe.parse(data, {"structural_map": scaled, "jurisdiction": "RO"})


def test_local_cell_parser_is_map_driven() -> None:
    """The separators come from the MAP — no per-cell grammar voting."""
    assert _parse_cell("1.234,56", ",", ".") == (1234.56, "ok")
    assert _parse_cell("1,234.56", ".", ",") == (1234.56, "ok")
    assert _parse_cell("(1 234,50)", ",", ".") == (-1234.5, "ok")
    assert _parse_cell("-17.5", ".", None) == (-17.5, "ok")
    assert _parse_cell("", ".", None) == (0.0, "blank")
    assert _parse_cell("-", ".", None) == (0.0, "blank")
    assert _parse_cell(None, ".", None) == (0.0, "blank")
    assert _parse_cell("abc", ".", None) == (0.0, "unparseable")
