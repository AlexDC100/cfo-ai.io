"""ARTIFACT EXPORT — gates A1..A5.

The builder's contract, asserted. Its five guarantees are stated at the
top of ``src/engine/api/_artifact_export.py``; this file is where each
one is made falsifiable.

Two conventions shape the file:

  TC-6  a recorded expectation PER COMPONENT. The components here are
        the THREE FORMATS. A single "an export was produced" assertion
        would let .pptx collapse to an empty package while .xlsx kept
        the suite green, so every format carries its own part census and
        its own floor, asserted after the loop.

  TC-9  would a clean result be distinguishable from "no subject"? A
        package that opens as a valid zip and contains nothing passes
        every structural check ever written. So each format asserts a
        MINIMUM PART COUNT and a minimum count of the payload it was
        given — the figures, the slides, the paragraphs.

The plant log is in design_review/artifacts/GATES.md.
"""
from __future__ import annotations

import io
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, List, Optional

import pytest

_SRC = Path(__file__).resolve().parents[2] / "src"
if str(_SRC) not in sys.path:  # pragma: no cover — import wiring
    sys.path.insert(0, str(_SRC))

from engine.api import _artifact_export as ax  # noqa: E402


# ──────────────────────────────────────────────────────────────────────
# Fixtures — the payload shape `frontend/lib/artifactExport.ts` sends
# ──────────────────────────────────────────────────────────────────────


def cell(
    value: Optional[float],
    minor: Optional[int],
    fact: str,
    unit: str = "money",
    currency: str = "RON",
) -> Dict[str, Any]:
    if value is None:
        return {
            "value": None,
            "minor": None,
            "unit": None,
            "currency": None,
            "fact": fact,
            "periodLabel": None,
            "provenance": None,
        }
    return {
        "value": value,
        "minor": minor,
        "unit": unit,
        "currency": currency,
        "fact": fact,
        "periodLabel": "Dec 2025",
        "provenance": {
            "source": "assembled_canonical_v1",
            "method": "served",
            "snapshot": "snap-a1b2c3d4",
        },
    }


CITATION = {
    "periods": ["Dec 2025"],
    "snapshots": ["snap-a1b2c3d4"],
    "sources": ["assembled_canonical_v1"],
    "currency": "RON",
    "trust": "reconciled",
    "incomplete": False,
}


def faithful_sheet() -> Dict[str, Any]:
    """Rows that DO add up to the served total, with one nested child so
    the live SUM has to skip it."""
    return {
        "name": "Balance sheet",
        "liveTotals": True,
        "columns": [
            {"label": "Line", "role": "label"},
            {"label": "Amount", "role": "value"},
        ],
        "rows": [
            {"label": "Current assets", "depth": 0, "accounts": ["411"],
             "cells": [cell(174229000.33, 17422900033, "current_assets")]},
            {"label": "of which trade", "depth": 1, "accounts": [],
             "cells": [cell(90000000.00, 9000000000, "trade_rec")]},
            {"label": "Non-current assets", "depth": 0, "accounts": [],
             "cells": [cell(118821084.78, 11882108478, "non_current_assets")]},
        ],
        "totalRow": {"label": "Total assets", "depth": 0, "accounts": [],
                     "cells": [cell(293050085.11, 29305008511, "total_assets")]},
    }


def request_for(fmt: str, **extra: Any) -> Dict[str, Any]:
    req = {
        "version": ax.ARTIFACT_EXPORT_VERSION,
        "format": fmt,
        "title": "Balance sheet — Dec 2025",
        "citation": CITATION,
    }
    req.update(extra)
    return req


XLSX_REQ = request_for("xlsx", sheets=[faithful_sheet()])
DOCX_REQ = request_for(
    "docx",
    sections=[
        {"heading": "Summary", "paragraphs": [
            "Total assets closed at 293.050.085,11 RON.",
            "The balance sheet reconciled.",
        ]},
    ],
)
PPTX_REQ = request_for(
    "pptx",
    slides=[
        {"heading": "Where the period landed", "blocks": [
            {"block": "headline", "lines": ["The balance sheet closed reconciled."]},
            {"block": "metrics", "metrics": [
                {"label": "Total assets", "cell": cell(293050085.11, 29305008511, "total_assets")},
                {"label": "Not retrieved", "cell": cell(None, None, "missing_x")},
            ]},
            {"block": "bullets", "lines": ["Liquidity held.", "Leverage unchanged."]},
        ]},
    ],
)

#: Minimum parts a valid package carries. Measured, then rounded down.
PART_FLOOR = {"xlsx": 8, "docx": 3, "pptx": 12}


# ──────────────────────────────────────────────────────────────────────
# A4 — DETERMINISTIC BYTES
# ──────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize("req", [XLSX_REQ, DOCX_REQ, PPTX_REQ], ids=["xlsx", "docx", "pptx"])
def test_a4_the_same_request_produces_the_same_bytes(req, monkeypatch):
    """Two clocks had to be removed for this: openpyxl stamps
    ``docProps/core.xml`` with ``datetime.now()`` and ``zipfile`` stamps
    every entry with ``time.localtime()``. A byte-identical export is
    what makes "this is the file I sent in March" checkable.

    THE CLOCK IS MOVED BETWEEN THE TWO BUILDS, and that is the whole
    test. The first draft simply built twice and compared — and a
    refuter proved it green with the wall clock restored, because two
    builds a few milliseconds apart read the SAME second and produced
    the same bytes. A determinism test that only passes because it ran
    fast is TC-9's shape exactly: its clean result is indistinguishable
    from "the clock never had a chance to move"."""
    first, _, _ = ax.build_export(req)

    import time as _time

    real = _time.localtime
    monkeypatch.setattr(_time, "localtime", lambda *a: real(1234567890))
    second, _, _ = ax.build_export(req)
    monkeypatch.setattr(_time, "localtime", lambda *a: real(1600000000))
    third, _, _ = ax.build_export(req)

    assert first == second == third
    # TC-9 — identical EMPTY outputs would satisfy the line above.
    assert len(first) > 800


@pytest.mark.parametrize("req", [XLSX_REQ, DOCX_REQ, PPTX_REQ], ids=["xlsx", "docx", "pptx"])
def test_a4_no_entry_carries_a_wall_clock_timestamp(req):
    data, _, _ = ax.build_export(req)
    archive = zipfile.ZipFile(io.BytesIO(data))
    assert archive.infolist(), "the package is empty"
    for info in archive.infolist():
        assert info.date_time == ax.FIXED_ZIP_DATE, info.filename
        assert info.create_system == 0, info.filename


# ──────────────────────────────────────────────────────────────────────
# PACKAGE INTEGRITY — per format, per component (TC-6)
# ──────────────────────────────────────────────────────────────────────


def _relationship_targets(archive: zipfile.ZipFile, rels_path: str) -> List[str]:
    root = ET.fromstring(archive.read(rels_path))
    return [r.get("Target", "") for r in root]


@pytest.mark.parametrize(
    "fmt,req", [("xlsx", XLSX_REQ), ("docx", DOCX_REQ), ("pptx", PPTX_REQ)]
)
def test_package_is_structurally_sound(fmt, req):
    """Every part parses, every part is declared in [Content_Types].xml,
    and every relationship points at a part that exists.

    A hand-built OOXML package fails in exactly these three ways, and
    all three fail SILENTLY — the file downloads, and the reader's
    application refuses it. Nothing here proves PowerPoint opens the
    deck; it proves the package is not malformed, which is the class of
    breakage a test can see."""
    data, media, ext = ax.build_export(req)
    assert ext == fmt
    assert media == ax.MEDIA_TYPES[fmt]

    archive = zipfile.ZipFile(io.BytesIO(data))
    names = archive.namelist()
    assert archive.testzip() is None
    assert len(names) >= PART_FLOOR[fmt], "%s: %d parts, floor %d" % (fmt, len(names), PART_FLOOR[fmt])

    for name in names:
        if name.endswith(".xml") or name.endswith(".rels"):
            ET.fromstring(archive.read(name))  # raises on malformed XML

    # Content types cover every XML part.
    types = ET.fromstring(archive.read("[Content_Types].xml"))
    defaults = {e.get("Extension") for e in types if e.tag.endswith("Default")}
    overrides = {e.get("PartName") for e in types if e.tag.endswith("Override")}
    for name in names:
        if name == "[Content_Types].xml":
            continue
        ext_of = name.rsplit(".", 1)[-1]
        assert ext_of in defaults or ("/" + name) in overrides, name

    # Every relationship target resolves to a real part.
    for name in names:
        if not name.endswith(".rels"):
            continue
        base = name.rsplit("_rels/", 1)[0]
        for target in _relationship_targets(archive, name):
            if target.startswith("http"):
                continue
            # A target may be PACKAGE-ABSOLUTE ("/xl/comments/comment1.xml",
            # which is how openpyxl writes its comment relationships) or
            # part-relative with any number of "../" hops. The first
            # draft handled only the second form and reported a real
            # part as missing — a false RED, which teaches the next
            # person to delete the check.
            if target.startswith("/"):
                candidate = target[1:]
            else:
                resolved = target
                prefix = base
                while resolved.startswith("../"):
                    resolved = resolved[3:]
                    trimmed = prefix.rstrip("/")
                    prefix = trimmed.rsplit("/", 1)[0] + "/" if "/" in trimmed else ""
                candidate = (prefix + resolved).lstrip("/")
            assert candidate in names, "%s -> %s (resolved %s)" % (name, target, candidate)


def test_pptx_declares_every_slide_it_writes():
    """A deck whose `sldIdLst` and its slide parts disagree opens with
    the slides missing — the package is valid and the content is gone."""
    data, _, _ = ax.build_export(PPTX_REQ)
    archive = zipfile.ZipFile(io.BytesIO(data))
    slides = [n for n in archive.namelist() if n.startswith("ppt/slides/slide")]
    slide_parts = [n for n in slides if n.endswith(".xml")]
    presentation = ET.fromstring(archive.read("ppt/presentation.xml"))
    listed = [e for e in presentation.iter() if e.tag.endswith("}sldId")]
    # One cover slide plus every slide in the request.
    assert len(slide_parts) == len(PPTX_REQ["slides"]) + 1
    assert len(listed) == len(slide_parts)


# ──────────────────────────────────────────────────────────────────────
# A3 — PROVENANCE SURVIVES THE EXPORT
# ──────────────────────────────────────────────────────────────────────


def _xlsx_cells(data: bytes):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data))
    ws = wb[wb.sheetnames[0]]
    return ws, [c for row in ws.iter_rows() for c in row]


def test_a3_every_fact_cell_carries_its_source_and_snapshot():
    data, _, _ = ax.build_export(XLSX_REQ)
    _, cells = _xlsx_cells(data)
    commented = [c for c in cells if c.comment is not None and c.value is not None]
    # TC-9 — a census that found nothing would pass the loop below.
    assert len(commented) >= 4, "only %d commented cell(s)" % len(commented)
    facts = 0
    for c in commented:
        text = c.comment.text
        if "Absent in the source" in text:
            continue
        assert "Source:" in text, c.coordinate
        assert "Snapshot:" in text, c.coordinate
        assert "Period:" in text, c.coordinate
        facts += 1
    assert facts >= 4, "only %d fact cell(s) carried provenance" % facts


def test_a3_the_citation_block_leads_the_sheet():
    data, _, _ = ax.build_export(XLSX_REQ)
    ws, _ = _xlsx_cells(data)
    head = [ws.cell(row=r, column=1).value for r in range(1, 8)]
    joined = " ".join([str(v) for v in head if v])
    for expected in ("Dec 2025", "snap-a1b2c3d4", "assembled_canonical_v1", "RON", "reconciled"):
        assert expected in joined, expected


# ──────────────────────────────────────────────────────────────────────
# A2 — ABSENT IS NOT ZERO
# ──────────────────────────────────────────────────────────────────────


def test_a2_an_absent_cell_is_a_glyph_not_a_zero_and_not_an_empty_cell():
    sheet = faithful_sheet()
    sheet["rows"].append(
        {"label": "Not retrieved", "depth": 0, "accounts": [],
         "cells": [cell(None, None, "missing_x")]}
    )
    data, _, _ = ax.build_export(request_for("xlsx", sheets=[sheet]))
    _, cells = _xlsx_cells(data)
    glyphs = [c for c in cells if c.value == ax.MISSING_GLYPH]
    assert len(glyphs) == 1
    # An empty numeric cell reads as ZERO inside every SUM that crosses
    # it — that is the same lie with a spreadsheet's authority behind it.
    assert glyphs[0].comment is not None
    assert "Not zero" in glyphs[0].comment.text
    assert not any(c.value == 0 for c in cells)


def test_a2_a_slide_prints_the_glyph_for_a_metric_that_was_not_retrieved():
    data, _, _ = ax.build_export(PPTX_REQ)
    archive = zipfile.ZipFile(io.BytesIO(data))
    body = "".join(
        archive.read(n).decode("utf-8")
        for n in archive.namelist()
        if n.startswith("ppt/slides/slide") and n.endswith(".xml")
    )
    assert ax.MISSING_GLYPH in body
    assert "Not retrieved" in body


# ──────────────────────────────────────────────────────────────────────
# A1 — NO DERIVATION: the served total always wins
# ──────────────────────────────────────────────────────────────────────


def test_a1_a_live_sum_is_written_only_when_it_reproduces_the_served_total():
    data, _, _ = ax.build_export(XLSX_REQ)
    ws, cells = _xlsx_cells(data)
    formulas = [c for c in cells if isinstance(c.value, str) and c.value.startswith("=SUM(")]
    assert len(formulas) == 1, [c.value for c in cells if isinstance(c.value, str)]
    formula = formulas[0].value
    # The nested child must NOT be in the sum — adding a parent and its
    # own detail together double-counts the parent.
    assert formula.count(",") == 1, formula
    assert "Live SUM" in formulas[0].comment.text


def test_a1_a_disagreeing_total_is_written_static_and_says_why():
    """THE PLANT: a served total the rows do not add up to. The builder
    must keep the SERVED figure and explain the withheld formula — never
    silently replace the engine's number with the spreadsheet's."""
    sheet = faithful_sheet()
    sheet["totalRow"]["cells"][0]["value"] = 999999999.99
    sheet["totalRow"]["cells"][0]["minor"] = 99999999999
    data, _, _ = ax.build_export(request_for("xlsx", sheets=[sheet]))
    _, cells = _xlsx_cells(data)
    assert not any(isinstance(c.value, str) and c.value.startswith("=SUM(") for c in cells)
    total = [c for c in cells if c.value == 999999999.99]
    assert len(total) == 1
    assert "does not reproduce it" in total[0].comment.text


def test_a1_a_gap_in_the_rows_withholds_the_formula():
    """A sum across an absence is not a sum."""
    sheet = faithful_sheet()
    sheet["rows"][0]["cells"] = [cell(None, None, "current_assets")]
    data, _, _ = ax.build_export(request_for("xlsx", sheets=[sheet]))
    _, cells = _xlsx_cells(data)
    assert not any(isinstance(c.value, str) and c.value.startswith("=SUM(") for c in cells)


def test_a1_the_faithfulness_test_is_exact_on_integer_minor_units():
    # Equal in minor units → faithful.
    assert ax._live_total_is_faithful([1.0, 2.0], [100, 200], 3.0, 300)
    # One cent out → NOT faithful, even though the floats round the same.
    assert not ax._live_total_is_faithful([1.0, 2.0], [100, 200], 3.0, 301)
    # An absent operand refuses outright.
    assert not ax._live_total_is_faithful([1.0, None], [100, None], 3.0, 300)


# ──────────────────────────────────────────────────────────────────────
# NUMBER FORMATS — declared units only, never a guess
# ──────────────────────────────────────────────────────────────────────


def test_a_unit_this_build_does_not_know_never_acquires_a_currency():
    assert ax._number_format("money", "RON") == '#,##0.00 "RON"'
    assert ax._number_format("percent", None) == "0.0%"
    assert ax._number_format("ratio", None) == '0.00"×"'
    # The refusal: no unit → General, NOT a currency mask.
    assert ax._number_format(None, "RON") == "General"
    assert ax._number_format("wat", "RON") == "General"


def test_the_contract_version_is_enforced_rather_than_assumed():
    with pytest.raises(ax.ExportRefused):
        ax.build_export(dict(XLSX_REQ, version="ax0"))
    with pytest.raises(ax.ExportRefused):
        ax.build_export(dict(XLSX_REQ, format="pdf"))
    # CSV is built on the client and must never be silently accepted here.
    with pytest.raises(ax.ExportRefused):
        ax.build_export(dict(XLSX_REQ, format="csv"))


def test_an_empty_payload_is_refused_rather_than_producing_an_empty_file():
    for fmt, key in (("xlsx", "sheets"), ("pptx", "slides"), ("docx", "sections")):
        with pytest.raises(ax.ExportRefused):
            ax.build_export(request_for(fmt, **{key: []}))


def test_the_filename_is_deterministic_and_carries_no_clock():
    a = ax.safe_filename("Balance sheet — Dec 2025", CITATION, "xlsx")
    b = ax.safe_filename("Balance sheet — Dec 2025", CITATION, "xlsx")
    assert a == b
    assert a == "Balance_sheet_Dec_2025_Dec_2025.xlsx"


def test_sheet_names_are_excel_legal_and_never_collide():
    used = set()
    first = ax._sheet_name("Profit/Loss [2025]", used)
    second = ax._sheet_name("Profit/Loss [2025]", used)
    assert first != second
    for name in (first, second):
        assert len(name) <= 31
        assert not set(name) & set("[]:*?/\\")


# ──────────────────────────────────────────────────────────────────────
# THE ROUTE — the one defect a builder test cannot see
# ──────────────────────────────────────────────────────────────────────


def _client():
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    app = FastAPI()
    app.include_router(ax.build_router())
    return TestClient(app)


def test_the_route_returns_bytes_not_a_422():
    """THE DEFECT THIS EXISTS FOR, found by exercising the route and by
    nothing else.

    ``_artifact_export`` carries ``from __future__ import annotations``,
    so annotations are STRINGS that FastAPI resolves against the
    MODULE's globals — and fastapi is imported INSIDE ``build_router``
    so the pure builders stay importable without it. A handler annotated
    ``request: Request`` therefore resolved to nothing, FastAPI read it
    as an unknown QUERY parameter, and every POST returned

        422 {"detail":[{"type":"missing","loc":["query","request"], …}]}

    The module imported cleanly, the builders were all green, and the
    typecheck was clean. Only a request could see it.

    TC-5 — ``follow_redirects=False``, because the URL itself is under
    test and the default would read the redirect TARGET's status.
    """
    res = _client().post("/api/artifacts/export", json=XLSX_REQ, follow_redirects=False)
    assert res.status_code == 200, res.text
    assert res.headers["content-type"] == ax.MEDIA_TYPES["xlsx"]
    assert "attachment; filename=" in res.headers["content-disposition"]
    assert res.content[:2] == b"PK", "the response body is not a zip package"
    assert zipfile.ZipFile(io.BytesIO(res.content)).testzip() is None


def test_the_route_refuses_rather_than_guessing():
    client = _client()
    for payload, why in (
        (dict(XLSX_REQ, version="ax0"), "contract"),
        (dict(XLSX_REQ, format="pdf"), "format"),
        (dict(XLSX_REQ, format="csv"), "csv is built on the client"),
        (dict(XLSX_REQ, sheets=[]), "empty payload"),
    ):
        res = client.post("/api/artifacts/export", json=payload, follow_redirects=False)
        assert res.status_code == 400, "%s: %d" % (why, res.status_code)
        assert isinstance(res.json().get("detail"), str)


# ──────────────────────────────────────────────────────────────────────
# THE CENSUS — what these gates examined (TC-3 / TC-9)
# ──────────────────────────────────────────────────────────────────────


def test_gate_work_census():
    """Emitted so a reader of a green run can see WHAT was examined, and
    asserted per FORMAT so one format collapsing cannot hide behind the
    other two."""
    counts = {}
    for fmt, req in (("xlsx", XLSX_REQ), ("docx", DOCX_REQ), ("pptx", PPTX_REQ)):
        data, _, _ = ax.build_export(req)
        archive = zipfile.ZipFile(io.BytesIO(data))
        counts[fmt] = len(archive.namelist())
        print("GATE-WORK artifact-export-%s parts=%d bytes=%d" % (fmt, counts[fmt], len(data)))
    for fmt, floor in PART_FLOOR.items():
        assert counts[fmt] >= floor, "%s: %d parts, floor %d" % (fmt, counts[fmt], floor)
    assert sorted(counts) == ["docx", "pptx", "xlsx"]
