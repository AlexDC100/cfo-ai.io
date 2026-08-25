"""run_structural_interpretation — one model call, one StructuralMap.

TWO DISTINCT PROMPT FRAMINGS, two registry roles:

  framing "a" → role "structural_interpreter_a" — COLUMN-SEMANTICS-FIRST:
    walk the header and assign every column a semantic, then describe
    the row topology.
  framing "b" → role "structural_interpreter_b" — ROW-TOPOLOGY-FIRST:
    identify totals / subtotal / repeated-header rows first, then assign
    column semantics.

Independent framings exist so a downstream consumer can cross-check two
maps of the same file produced by differently-biased walkthroughs.

Registry guard: :func:`engine.ai.registry.params_for` is called FIRST
and its RegistryError re-raised as a typed :class:`InterpError` —
because ``engine.ai_lane._client.call_strict_json`` silently falls back
to the lane's default model on an unknown stage/role, which would turn a
typo'd role into a silent misconfiguration.

Spend armor: when no client/factory is injected, the default factory is
``engine.ai.breaker.guarded_client_factory(role)`` — the breaker is
consulted BEFORE an Anthropic client is constructed, and BreakerOpen
degrades to the typed :class:`InterpUnavailable` (never a serving
block). Each real call is best-effort recorded against the role.

Document rendering: XLSX workbooks are rendered by a local renderer that
PREFIXES EVERY LINE WITH ITS 0-BASED WORKBOOK ROW INDEX and keeps blank
rows — deliberately NOT ``engine.ai_lane.format_detect
.spreadsheet_to_text``, which drops fully-blank rows and therefore
shifts row indices relative to the workbook (a StructuralMap's row
indexes must match openpyxl coordinates). Non-XLSX kinds reuse the
lane's jurisdiction-blind ``document_to_text`` verbatim.

E1 at the parse seam: a model response whose map carries any cell value
is rejected by :meth:`StructuralMap.from_json_dict`, retried ONCE with
the validation error fed back, then raised as :class:`InterpError`.

AI calls never live in tests — inject ``client`` (scripted) or
``client_factory``; the production default is lazy and breaker-guarded.
"""
from __future__ import annotations

import io
from typing import Any, Callable, Dict, List, Optional, Tuple

from engine.ai import breaker as _breaker
from engine.ai import registry as _registry
from engine.ai_lane import config as _lane_config
from engine.ai_lane._client import call_strict_json
from engine.ai_lane.format_detect import document_to_text
from engine.ai_lane.schemas import AiLaneError

from .structmap import COLUMN_SEMANTICS, MAP_VERSION, StructMapError, StructuralMap

#: framing letter -> registry role.
ROLE_BY_FRAMING: Dict[str, str] = {
    "a": "structural_interpreter_a",
    "b": "structural_interpreter_b",
}


class InterpError(RuntimeError):
    """Structural interpretation failed (registry gap, unreadable
    document, or a model that twice returned an invalid / value-bearing
    map). No partial map is ever returned."""


class InterpUnavailable(InterpError):
    """The interpreter cannot run right now (breaker open, no client).

    Callers degrade honestly — "no structural map" — never to an error
    surface or a serving block."""

    def __init__(self, role: str, reason: str) -> None:
        super().__init__(
            "structural interpreter unavailable for role '%s': %s" % (role, reason)
        )
        self.role = role
        self.reason = reason


def role_for_framing(framing: str) -> str:
    role = ROLE_BY_FRAMING.get(framing)
    if role is None:
        raise InterpError(
            "unknown structural-interpretation framing %r (known: %s)"
            % (framing, ", ".join(sorted(ROLE_BY_FRAMING)))
        )
    return role


# ── document rendering ─────────────────────────────────────────────────


def _render_xlsx_with_row_indexes(
    file_bytes: bytes, max_chars: int = _lane_config.MAX_DOC_CHARS
) -> str:
    """XLSX -> text with 0-based WORKBOOK row indexes on every line.

    Blank rows are KEPT (rendered as a bare index) so the indexes the
    model reports match openpyxl coordinates exactly."""
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    out: List[str] = []
    total = 0
    for sheet_name in wb.sheetnames:
        if total > max_chars:
            break
        ws = wb[sheet_name]
        section = ["=== Sheet: %s ===" % sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = ["" if v is None else str(v).replace("\t", " ") for v in row]
            line = "r%d\t%s" % (i, "\t".join(cells))
            if total + len(line) > max_chars:
                section.append("[truncated]")
                break
            section.append(line)
            total += len(line) + 1
        out.append("\n".join(section))
    return "\n\n".join(out)


def _kind_of(filename: str) -> str:
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return "xlsx"
    if lower.endswith(".pdf"):
        return "pdf"
    return "csv"


def render_document_text(file_bytes: bytes, filename: str = "") -> str:
    """The interpreter's prompt payload for a document."""
    kind = _kind_of(filename)
    if kind == "xlsx":
        try:
            text = _render_xlsx_with_row_indexes(file_bytes)
        except Exception as e:  # noqa: BLE001 — wrap honestly, never empty
            raise InterpError(
                "structural interpreter could not read the spreadsheet: %s" % e
            )
        text = text.strip()
        if not text:
            raise InterpError(
                "the spreadsheet produced no readable text — the structural "
                "interpreter cannot process it."
            )
        return text
    try:
        return document_to_text(kind, file_bytes, filename)
    except AiLaneError as e:
        raise InterpError(str(e))


# ── prompts ────────────────────────────────────────────────────────────

_SEMANTICS_GLOSSARY = (
    "Column semantics vocabulary (assign each column EXACTLY one):\n"
    "  account_code — the account identifier column\n"
    "  account_name — the account label/description column\n"
    "  opening_debit / opening_credit — balances at period START\n"
    "  movement_period_debit / movement_period_credit — movements of the "
    "single current period only (e.g. the month)\n"
    "  movement_cumulative_debit / movement_cumulative_credit — CUMULATIVE "
    "movements since the start of the year, EXCLUDING opening balances "
    "(movements-only running totals)\n"
    "  total_with_opening_debit / total_with_opening_credit — opening "
    "balances PLUS cumulative movements (an opening-inclusive 'total "
    "sums' pair)\n"
    "  closing_debit / closing_credit — balances at period END\n"
    "  marker — a per-row categorical marker column (e.g. a balance-sheet "
    "vs profit-and-loss flag)\n"
    "  hint_classification — a source-system classification hint column "
    "(statutory account-type label, reporting-item name, ...)\n"
    "  ignore — a column with no extraction relevance\n"
    "DISTINGUISH the two cumulative kinds carefully: verify against a few "
    "rows whether the pair INCLUDES the opening balances "
    "(opening + cumulative movements) or EXCLUDES them (movements only). "
    "Do this by comparing MAGNITUDE RELATIONSHIPS between columns, and "
    "report ONLY the semantic — never the figures themselves.\n"
)

_SCHEMA_SPEC = (
    "Respond with ONLY a strict JSON object — no prose, no markdown "
    "fences — with exactly these keys:\n"
    "{\n"
    '  "map_version": "%s",\n'
    '  "sheet": <sheet name string, or null>,\n'
    '  "header_row_index": <0-based row index of the column-header row>,\n'
    '  "columns": [{"index": <0-based column index>, "semantic": <one of '
    "the vocabulary>}, ... one entry per column],\n"
    '  "totals_row_indexes": [<0-based row indexes of grand-totals rows; '
    "EMPTY list when the file has no totals row — never invent one>],\n"
    '  "subtotal_row_indexes": [<0-based row indexes of subtotal rows>],\n'
    '  "repeated_header_rows": [<0-based row indexes where the header row '
    "repeats, e.g. page-break repeats>],\n"
    '  "account_code_col": <the index of the account_code column>,\n'
    '  "analytic_structure": {"separator": <the character separating the '
    'synthetic account from its analytic suffix, e.g. ".", or null when '
    "analytic codes are positional (digits appended without a "
    'separator)>, "synthetic_digits": <for positional analytics: the '
    "digit count of the synthetic prefix, or null>},\n"
    '  "number_locale": {"thousands_sep": <string or null>, '
    '"decimal_sep": <string or null>},\n'
    '  "currency": <ISO 4217 code or null>,\n'
    '  "scale": <integer multiplier the figures are expressed in: 1 for '
    "units, 1000 for thousands>,\n"
    '  "anomaly_notes": [<short advisory strings about layout oddities: '
    "trailing junk rows, merged cells, formula remnants, ...>]\n"
    "}\n"
    "CRITICAL RULE: the map describes STRUCTURE ONLY. It must carry "
    "coordinates, indices, enum names and short strings — NEVER any "
    "monetary amount, balance, or other cell value. Any numeric field "
    "other than row/column indexes, scale, and synthetic_digits will be "
    "REJECTED and your response discarded.\n"
) % MAP_VERSION

_SYSTEM_A = (
    "You are a financial-document layout analyst. You are given the raw "
    "text rendering of a tabular accounting export (trial balance, ledger "
    "summary, or statement); spreadsheet lines are prefixed with their "
    "0-based workbook row index as 'r<N>'. Report row indexes exactly as "
    "those prefixes give them.\n\n"
    "Work COLUMN-SEMANTICS-FIRST: (1) find the header row; (2) walk the "
    "columns left to right and assign each one exactly one semantic from "
    "the vocabulary, checking your assignment against several data rows; "
    "(3) only then describe the row topology (totals rows, subtotal rows, "
    "repeated headers); (4) finish with analytic-code structure, number "
    "locale, currency and scale.\n\n"
    + _SEMANTICS_GLOSSARY + "\n" + _SCHEMA_SPEC
)

_SYSTEM_B = (
    "You are a financial-document layout analyst. You are given the raw "
    "text rendering of a tabular accounting export (trial balance, ledger "
    "summary, or statement); spreadsheet lines are prefixed with their "
    "0-based workbook row index as 'r<N>'. Report row indexes exactly as "
    "those prefixes give them.\n\n"
    "Work ROW-TOPOLOGY-FIRST: (1) scan the rows top to bottom and classify "
    "them — the header row, any REPEATED header rows (page-break repeats), "
    "grand-totals rows (do not invent one if none exists), subtotal rows, "
    "trailing junk/blank rows; (2) only after the row topology is settled, "
    "assign each column exactly one semantic from the vocabulary, checking "
    "against several data rows; (3) finish with analytic-code structure, "
    "number locale, currency and scale.\n\n"
    + _SEMANTICS_GLOSSARY + "\n" + _SCHEMA_SPEC
)

_SYSTEM_BY_FRAMING = {"a": _SYSTEM_A, "b": _SYSTEM_B}

_REJECTION_FEEDBACK = (
    "\n\nYour previous structural map was REJECTED by the schema "
    "validator: %s\nReturn ONLY structure — coordinates, indices, enum "
    "names and short strings. Never include any monetary amount or cell "
    "value anywhere in the JSON."
)


# ── the entry point ────────────────────────────────────────────────────


def run_structural_interpretation(
    file_bytes: bytes,
    filename: str,
    *,
    jurisdiction: str,
    framing: str,
    client: Any = None,
    client_factory: Optional[Callable[[], Any]] = None,
    audit: Optional[Dict[str, Any]] = None,
) -> Tuple[StructuralMap, Dict[str, Any]]:
    """Interpret one document's structure. Returns (map, audit_dict).

    Raises :class:`InterpUnavailable` when no model can be consulted
    (breaker open / client construction failure) and :class:`InterpError`
    on registry gaps, unreadable documents, or a model that twice failed
    E1 validation.
    """
    role = role_for_framing(framing)

    # LOUD registry guard — call_strict_json falls back silently on an
    # unknown role, which would mask a misconfigured registry.
    try:
        params = _registry.params_for(role)
    except _registry.RegistryError as e:
        raise InterpError(
            "structural interpreter role '%s' is not configured in the "
            "model registry: %s" % (role, e)
        )
    prompt_version = params["prompt_version"]
    max_tokens = params["max_tokens"]

    doc_text = render_document_text(file_bytes, filename)
    system = _SYSTEM_BY_FRAMING[framing]
    user_text = (
        "Jurisdiction hint: %s. Filename: %s.\n\nDocument text follows.\n\n%s"
        % (jurisdiction, filename or "unknown", doc_text)
    )

    constructed_here = client is None
    if client is None:
        factory = client_factory or _breaker.guarded_client_factory(role)
        try:
            client = factory()
        except _breaker.BreakerOpen as e:
            raise InterpUnavailable(role, "breaker_open: %s" % e.reason)
        except Exception as e:  # noqa: BLE001 — missing key/SDK → degrade
            raise InterpUnavailable(role, "client_unavailable: %s" % e)

    audit_stages: List[Dict[str, Any]] = []
    last_error: Optional[str] = None
    for attempt in (1, 2):
        payload = user_text if attempt == 1 else (
            user_text + _REJECTION_FEEDBACK % last_error
        )
        try:
            data = call_strict_json(
                client,
                stage=role,
                prompt_version=prompt_version,
                system=system,
                user_text=payload,
                max_tokens=max_tokens,
                audit_stages=audit_stages,
            )
        except AiLaneError as e:
            raise InterpError(
                "structural interpretation failed for role '%s': %s" % (role, e)
            )
        finally:
            if constructed_here:
                # Best-effort spend counting (cost armor, never a gate).
                try:
                    _breaker.record(role)
                except Exception:  # noqa: BLE001
                    pass
        try:
            smap = StructuralMap.from_json_dict(data)
        except StructMapError as e:
            last_error = str(e)
            continue
        audit_dict: Dict[str, Any] = {
            "role": role,
            "framing": framing,
            "model_id": params["model_id"],
            "prompt_version": prompt_version,
            "map_version": MAP_VERSION,
            "map_hash": smap.map_hash,
            "attempts": attempt,
            "column_semantics_vocabulary": list(COLUMN_SEMANTICS),
            "stages": audit_stages,
        }
        if isinstance(audit, dict):
            audit.update(audit_dict)
        return smap, audit_dict

    raise InterpError(
        "structural interpretation failed for role '%s': the model twice "
        "returned a map that failed strict validation (last error: %s). "
        "No structural map was produced." % (role, last_error)
    )
