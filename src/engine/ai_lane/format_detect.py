"""AI lane stage 1 — document format detection (strict JSON).

Detects the tabular layout (columns, header row, totals row), currency,
scale and number locale of a non-RO trial balance / statement export.
The result drives the extract prompt AND the self-check (the totals row
index tells the checker a file anchor exists).
"""
from __future__ import annotations

import io
from typing import Any, Callable, Dict, List, Optional

from . import config
from ._client import call_strict_json
from .schemas import AiLaneError, FormatDetectResult

_SYSTEM = (
    "You are a financial-document layout analyst. You are given the raw "
    "text rendering of an accounting export (trial balance, ledger "
    "summary, or statement). Respond with ONLY a strict JSON object — no "
    "prose, no markdown fences — with exactly these keys:\n"
    "{\n"
    '  "layout": {"columns": [<ordered column labels as strings>],\n'
    '             "header_row": <0-based row index of the header, or null>,\n'
    '             "totals_row_index": <0-based row index of the document\'s own '
    "grand-totals row, or null when the file carries none>},\n"
    '  "currency": <ISO 4217 code such as "HUF", "EUR", "RON", or null>,\n'
    '  "scale": <numeric multiplier the figures are expressed in: 1 for units, '
    "1000 for thousands (e.g. Hungarian 'ezer Ft'), or 1 when unknown>,\n"
    '  "thousands_sep": <the thousands separator character, e.g. " ", ".", ",", or null>,\n'
    '  "decimal_sep": <the decimal separator character, e.g. "," or ".", or null>,\n'
    '  "language": <ISO 639-1 language code of the document, e.g. "hu", or null>\n'
    "}\n"
    "Hungarian exports typically use SPACE as the thousands separator and "
    "COMMA as the decimal separator (e.g. '1 234 567,89'). Read the "
    "document, do not guess a locale that contradicts the visible numbers."
)


def spreadsheet_to_text(file_bytes: bytes, max_chars: int = config.MAX_DOC_CHARS) -> str:
    """Render an XLSX workbook as TSV text (sheets separated by markers).
    Local copy of the pipeline's renderer so the lane never imports
    engine.api (which would drag the whole FastAPI app into tests)."""
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    out: List[str] = []
    total = 0
    for sheet_name in wb.sheetnames:
        if total > max_chars:
            break
        ws = wb[sheet_name]
        section = ["=== Sheet: %s ===" % sheet_name]
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).replace("\t", " ") for v in row]
            if not any(c.strip() for c in cells):
                continue
            line = "\t".join(cells)
            if total + len(line) > max_chars:
                section.append("[truncated]")
                break
            section.append(line)
            total += len(line) + 1
        out.append("\n".join(section))
    return "\n\n".join(out)


def document_to_text(kind: str, file_bytes: bytes, filename: str = "") -> str:
    """Text payload for the lane's prompts, by file kind (the pipeline's
    `_classify_file` vocabulary). Raises AiLaneError for inputs the lane
    cannot read — never silently empty."""
    text = ""
    if kind == "xlsx":
        try:
            text = spreadsheet_to_text(file_bytes)
        except Exception as e:  # noqa: BLE001
            raise AiLaneError("AI lane could not read the spreadsheet: %s" % e)
    elif kind == "pdf":
        try:
            try:
                from pypdf import PdfReader  # type: ignore
            except ImportError:
                from PyPDF2 import PdfReader  # type: ignore
            reader = PdfReader(io.BytesIO(file_bytes))
            text = "\n".join((p.extract_text() or "") for p in reader.pages)
        except Exception as e:  # noqa: BLE001
            raise AiLaneError("AI lane could not read the PDF: %s" % e)
    elif kind in ("image_jpeg", "image_png"):
        raise AiLaneError(
            "Image uploads are not supported by the AI extraction lane yet — "
            "upload the spreadsheet, CSV or PDF export instead."
        )
    else:  # csv / text / unknown
        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1", errors="replace")
    text = (text or "").strip()
    if not text:
        raise AiLaneError(
            "The document produced no readable text — the AI extraction "
            "lane cannot process it."
        )
    if len(text) > config.MAX_DOC_CHARS:
        text = text[:config.MAX_DOC_CHARS] + "\n[truncated]"
    return text


def run_format_detect(
    doc_text: str,
    *,
    jurisdiction: str,
    filename: str = "",
    client: Any = None,
    client_factory: Optional[Callable[[], Any]] = None,
    audit_stages: Optional[List[Dict[str, Any]]] = None,
) -> FormatDetectResult:
    if client is None:
        client = (client_factory or config.default_client_factory)()
    user_text = (
        "Jurisdiction: %s. Filename: %s.\n\nDocument text follows.\n\n%s"
        % (jurisdiction, filename or "unknown", doc_text)
    )
    data = call_strict_json(
        client,
        stage="format_detect",
        prompt_version=config.FORMAT_DETECT_PROMPT_VERSION,
        system=_SYSTEM,
        user_text=user_text,
        max_tokens=config.FORMAT_DETECT_MAX_TOKENS,
        audit_stages=audit_stages,
    )
    layout = data.get("layout") if isinstance(data.get("layout"), dict) else {}

    def _opt_int(v: Any) -> Optional[int]:
        try:
            return None if v is None else int(v)
        except (TypeError, ValueError):
            return None

    try:
        scale = float(data.get("scale") or 1.0)
    except (TypeError, ValueError):
        scale = 1.0
    return FormatDetectResult(
        columns=[str(c) for c in (layout.get("columns") or [])],
        header_row=_opt_int(layout.get("header_row")),
        totals_row_index=_opt_int(layout.get("totals_row_index")),
        currency=(str(data["currency"]) if data.get("currency") else None),
        scale=scale if scale > 0 else 1.0,
        thousands_sep=(str(data["thousands_sep"]) if data.get("thousands_sep") is not None else None),
        decimal_sep=(str(data["decimal_sep"]) if data.get("decimal_sep") is not None else None),
        language=(str(data["language"]) if data.get("language") else None),
        raw=data,
    )
