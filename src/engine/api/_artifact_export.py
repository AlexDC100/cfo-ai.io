"""ARTIFACT EXPORT — .xlsx / .pptx / .docx built from RESOLVED figures.

``POST /api/artifacts/export``. The request carries figures that the
facts gateway already resolved on the client (value, integer minor
units, declared unit, currency, provenance). This module writes them
into a file. It does not compute, it does not convert, and it does not
interpret prose.

WHAT THIS MODULE GUARANTEES
===========================

A1 — NO DERIVATION.
    There is no arithmetic here that produces a figure a cell then
    shows. The single sum this module performs
    (:func:`_live_total_is_faithful`) exists ONLY to decide whether a
    live ``SUM()`` formula may be written, and its answer is used as a
    boolean: when the sum of the rows does not reproduce the total the
    engine served, the total is written STATIC and the disagreement is
    named in the cell comment. The served figure always wins; the
    formula is a convenience that is withheld the moment it would
    replace an authority.

A2 — ABSENT IS NOT ZERO.
    A cell whose ``value`` is ``None`` is written as the missing glyph
    (a STRING), never as ``0`` and never as an empty cell. An empty
    numeric cell reads as zero inside every ``SUM`` that crosses it,
    which is the same lie with a spreadsheet's authority behind it.

A3 — PROVENANCE SURVIVES THE EXPORT.
    Every fact-bearing cell carries a comment naming its source, its
    period and its snapshot. A workbook that leaves this product with
    the numbers but without their provenance is precisely the artifact
    this product exists to replace, and it is the reason the workbook
    is built here rather than with the frontend's bundled SheetJS.

A4 — DETERMINISTIC BYTES.
    The same request produces the same bytes, twice, on any machine.
    Two clocks had to be removed for that: ``openpyxl`` stamps
    ``docProps/core.xml`` with ``datetime.now()`` and ``zipfile``
    stamps every entry with ``time.localtime()``. Both are pinned —
    the document properties to a fixed instant, the archive through
    :func:`_rewrite_zip_deterministic`. A byte-identical export is what
    makes "this is the file I sent the bank in March" a checkable
    claim rather than a hope.

A5 — NO NETWORK, NO DATABASE, NO MODEL.
    The builders are pure functions of the request. They can be called
    from a test with a dict and no server, which is how the gate runs.

PYTHON 3.9 — no ``match``, no ``X | Y`` unions anywhere in this file.
"""
from __future__ import annotations

import datetime as _dt
import io
import re
import zipfile
from typing import Any, Dict, List, Optional, Sequence, Tuple

#: Contract version the client stamps. A request without it is refused
#: rather than assumed current.
ARTIFACT_EXPORT_VERSION = "ax1"

#: What a resolved-but-absent cell prints. Matches the frontend's
#: ``artifact.missing`` glyph so the file and the screen agree.
MISSING_GLYPH = "—"

#: Every timestamp in every produced file. 1980-01-01 is the earliest
#: instant the ZIP format can represent, so it is the one value that
#: cannot be mistaken for a real authoring time.
FIXED_ZIP_DATE = (1980, 1, 1, 0, 0, 0)
FIXED_DOC_INSTANT = _dt.datetime(1980, 1, 1, 0, 0, 0)

#: Excel forbids these in a sheet name, and caps it at 31 characters.
_SHEET_FORBIDDEN = re.compile(r"[\[\]:*?/\\]")

#: Excel's own cap on arguments to one function call.
_MAX_SUM_ARGS = 250


class ExportRefused(ValueError):
    """The request cannot be built into a file, and saying so is the
    correct outcome. Never downgraded into a partial file."""


# ──────────────────────────────────────────────────────────────────────
# REQUEST READING — transcription, not interpretation
# ──────────────────────────────────────────────────────────────────────


def _s(value: Any, default: str = "") -> str:
    return value if isinstance(value, str) else default


def _cells_of(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    cells = row.get("cells")
    return [c for c in cells if isinstance(c, dict)] if isinstance(cells, list) else []


def _cell_value(cell: Dict[str, Any]) -> Optional[float]:
    v = cell.get("value")
    if isinstance(v, bool):  # bool is an int subclass; never a figure
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _cell_minor(cell: Dict[str, Any]) -> Optional[int]:
    v = cell.get("minor")
    if isinstance(v, bool):
        return None
    return int(v) if isinstance(v, int) else None


def _provenance_note(cell: Dict[str, Any]) -> str:
    """The comment body. Assembled from what the gateway actually sent —
    a field the payload did not carry is OMITTED, never filled with a
    placeholder that would read as a recorded fact."""
    parts = []  # type: List[str]
    fact = _s(cell.get("fact"))
    if fact:
        parts.append("Fact: %s" % fact)
    prov = cell.get("provenance")
    if isinstance(prov, dict):
        source = _s(prov.get("source"))
        method = _s(prov.get("method"))
        snapshot = _s(prov.get("snapshot"))
        if source:
            parts.append("Source: %s" % source)
        if method:
            parts.append("Method: %s" % method)
        if snapshot:
            parts.append("Snapshot: %s" % snapshot)
    period = _s(cell.get("periodLabel"))
    if period:
        parts.append("Period: %s" % period)
    currency = _s(cell.get("currency"))
    if currency:
        parts.append("Currency: %s" % currency)
    return "\n".join(parts)


def _citation_lines(citation: Dict[str, Any]) -> List[str]:
    out = []  # type: List[str]
    if not isinstance(citation, dict):
        return out

    def _join(key: str) -> str:
        v = citation.get(key)
        if isinstance(v, list):
            return " · ".join([x for x in v if isinstance(x, str) and x])
        return ""

    periods = _join("periods")
    snapshots = _join("snapshots")
    sources = _join("sources")
    if periods:
        out.append("Period: %s" % periods)
    if snapshots:
        out.append("Snapshot: %s" % snapshots)
    if sources:
        out.append("Source: %s" % sources)
    currency = _s(citation.get("currency"))
    if currency:
        out.append("Currency: %s" % currency)
    trust = _s(citation.get("trust"))
    if trust:
        out.append("Trust: %s" % trust)
    if citation.get("incomplete") is True:
        out.append("Partial — the retrieval reported a gap.")
    return out


# ──────────────────────────────────────────────────────────────────────
# DETERMINISM
# ──────────────────────────────────────────────────────────────────────


def _rewrite_zip_deterministic(raw: bytes) -> bytes:
    """Re-emit an archive with every entry stamped at :data:`FIXED_ZIP_DATE`.

    ``zipfile.writestr`` with a *string* name takes ``time.localtime()``,
    so a library that writes its package that way (openpyxl does)
    produces different bytes every second AND different bytes in
    different timezones. Entry ORDER is preserved rather than sorted:
    OOXML readers tolerate any order, but ``[Content_Types].xml`` first
    is the convention every producer follows and there is no reason to
    be the exception.
    """
    src = zipfile.ZipFile(io.BytesIO(raw), "r")
    buf = io.BytesIO()
    out = zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED)
    try:
        for info in src.infolist():
            data = src.read(info.filename)
            clone = zipfile.ZipInfo(info.filename, date_time=FIXED_ZIP_DATE)
            clone.compress_type = zipfile.ZIP_DEFLATED
            clone.external_attr = info.external_attr
            # create_system 0 (MS-DOS) rather than the host's own value —
            # the same file built on macOS and in the Linux container
            # must not differ by a platform byte.
            clone.create_system = 0
            out.writestr(clone, data)
    finally:
        out.close()
        src.close()
    return buf.getvalue()


def _write_parts(parts: Sequence[Tuple[str, str]]) -> bytes:
    """Build an OOXML package from (path, xml) pairs, deterministically."""
    buf = io.BytesIO()
    zf = zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED)
    try:
        for path, xml in parts:
            info = zipfile.ZipInfo(path, date_time=FIXED_ZIP_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            info.create_system = 0
            zf.writestr(info, xml.encode("utf-8"))
    finally:
        zf.close()
    return buf.getvalue()


def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ──────────────────────────────────────────────────────────────────────
# XLSX
# ──────────────────────────────────────────────────────────────────────


def _number_format(unit: Optional[str], currency: Optional[str]) -> str:
    """An Excel format string per DECLARED unit.

    The unit comes from the engine's own declaration; it is never
    guessed from magnitude. An undeclared unit gets ``General`` — the
    format that adds nothing — rather than a currency mask, because a
    mask is a claim about what the number is.
    """
    if unit == "money":
        symbol = currency or ""
        if symbol:
            return '#,##0.00 "%s"' % symbol
        return "#,##0.00"
    if unit == "percent":
        return "0.0%"
    if unit == "ratio":
        return '0.00"×"'
    if unit == "days":
        return '0" d"'
    if unit == "count":
        return "#,##0"
    return "General"


def _column_letter(index_zero_based: int) -> str:
    n = index_zero_based + 1
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _live_total_is_faithful(
    row_cells: Sequence[Optional[float]],
    row_minors: Sequence[Optional[int]],
    total_value: Optional[float],
    total_minor: Optional[int],
) -> bool:
    """Would ``SUM(rows)`` reproduce the total the engine served?

    Integer minor units when every operand carries them — that is an
    EXACT comparison and the only one worth making about money. Falls
    back to a relative epsilon on floats, and refuses outright when any
    operand is absent: a sum across a gap is not a sum.
    """
    if total_value is None:
        return False
    if any(v is None for v in row_cells):
        return False
    if total_minor is not None and all(m is not None for m in row_minors) and row_minors:
        return sum(int(m) for m in row_minors if m is not None) == total_minor
    total_rows = sum(float(v) for v in row_cells if v is not None)
    scale = max(abs(total_rows), abs(total_value), 1e-12)
    return abs(total_rows - float(total_value)) / scale <= 1e-9


def build_xlsx(request: Dict[str, Any]) -> bytes:
    """One sheet per ``sheets[]`` entry. Values native, formats by
    declared unit, provenance in a comment on every fact cell."""
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, Side
    except ImportError as exc:  # pragma: no cover — declared in pyproject
        raise ExportRefused("openpyxl is not installed: %s" % exc)

    sheets = request.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ExportRefused("An .xlsx export needs at least one sheet.")

    wb = Workbook()
    wb.remove(wb.active)
    # A4 — pin both clocks openpyxl would otherwise read.
    wb.properties.created = FIXED_DOC_INSTANT
    wb.properties.modified = FIXED_DOC_INSTANT
    wb.properties.creator = "CFO AI"
    wb.properties.lastModifiedBy = "CFO AI"

    title = _s(request.get("title"), "Artifact")
    citation = request.get("citation") if isinstance(request.get("citation"), dict) else {}
    header_lines = _citation_lines(citation)

    head_font = Font(bold=True, size=11)
    label_font = Font(size=10)
    note_font = Font(size=9, italic=True)
    total_font = Font(bold=True, size=10)
    right = Alignment(horizontal="right")
    double_top = Border(top=Side(style="double"))

    used_names = set()  # type: set
    for sheet in sheets:
        if not isinstance(sheet, dict):
            continue
        name = _sheet_name(_s(sheet.get("name"), "Sheet"), used_names)
        ws = wb.create_sheet(title=name)

        r = 1
        ws.cell(row=r, column=1, value=title).font = head_font
        r += 1
        for line in header_lines:
            ws.cell(row=r, column=1, value=line).font = note_font
            r += 1
        r += 1  # one blank line between the citation block and the table

        columns = sheet.get("columns")
        columns = [c for c in columns if isinstance(c, dict)] if isinstance(columns, list) else []
        if not columns:
            raise ExportRefused("Sheet '%s' has no columns." % name)

        header_row = r
        for ci, col in enumerate(columns):
            cell = ws.cell(row=header_row, column=ci + 1, value=_s(col.get("label")))
            cell.font = head_font
            if _s(col.get("role")) != "label":
                cell.alignment = right
        r += 1

        rows = sheet.get("rows")
        rows = [x for x in rows if isinstance(x, dict)] if isinstance(rows, list) else []

        # Excel row numbers of the DEPTH-0 rows, per value column. A live
        # SUM must not add a parent and its children together, so nested
        # detail is written but never summed.
        top_level_rows = []  # type: List[int]
        for row in rows:
            depth = row.get("depth")
            depth = int(depth) if isinstance(depth, int) and not isinstance(depth, bool) else 0
            label = ("    " * depth) + _s(row.get("label"))
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = label_font
            for ci, cell in enumerate(_cells_of(row)):
                _write_value_cell(ws, r, ci + 2, cell, Comment, right)
            accounts = row.get("accounts")
            if isinstance(accounts, list) and accounts:
                codes = " · ".join([a for a in accounts if isinstance(a, str)])
                if codes:
                    ac = ws.cell(row=r, column=len(columns) + 2, value=codes)
                    ac.font = note_font
            if depth == 0:
                top_level_rows.append(r)
            r += 1

        total_row = sheet.get("totalRow")
        if isinstance(total_row, dict):
            live = sheet.get("liveTotals") is True
            tc = ws.cell(row=r, column=1, value=_s(total_row.get("label"), "Total"))
            tc.font = total_font
            tc.border = double_top
            for ci, cell in enumerate(_cells_of(total_row)):
                col_index = ci + 2
                column_values = [
                    _cell_value(_cells_of(row)[ci]) if ci < len(_cells_of(row)) else None
                    for row in rows
                    if (row.get("depth") in (0, None))
                ]
                column_minors = [
                    _cell_minor(_cells_of(row)[ci]) if ci < len(_cells_of(row)) else None
                    for row in rows
                    if (row.get("depth") in (0, None))
                ]
                faithful = _live_total_is_faithful(
                    column_values, column_minors, _cell_value(cell), _cell_minor(cell)
                )
                _write_total_cell(
                    ws,
                    r,
                    col_index,
                    cell,
                    Comment,
                    right,
                    total_font,
                    double_top,
                    top_level_rows if (live and faithful and len(top_level_rows) <= _MAX_SUM_ARGS) else None,
                    faithful,
                )
            r += 1

        ws.column_dimensions["A"].width = 40
        for ci in range(1, len(columns) + 1):
            ws.column_dimensions[_column_letter(ci)].width = 18

    raw = io.BytesIO()
    wb.save(raw)
    return _rewrite_zip_deterministic(raw.getvalue())


def _sheet_name(raw: str, used: set) -> str:
    name = _SHEET_FORBIDDEN.sub("-", raw).strip() or "Sheet"
    name = name[:31]
    candidate = name
    n = 2
    while candidate.lower() in used:
        suffix = " (%d)" % n
        candidate = name[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _write_value_cell(ws, row: int, col: int, cell: Dict[str, Any], Comment, right) -> None:
    value = _cell_value(cell)
    if value is None:
        # A2 — the glyph, as a STRING. An empty cell would be summed as
        # zero by the next person who drags a SUM across it.
        target = ws.cell(row=row, column=col, value=MISSING_GLYPH)
        target.alignment = right
        target.comment = Comment("Absent in the source. Not zero.", "CFO AI")
        return
    target = ws.cell(row=row, column=col, value=value)
    target.alignment = right
    target.number_format = _number_format(_s(cell.get("unit")) or None, _s(cell.get("currency")) or None)
    note = _provenance_note(cell)
    if note:
        target.comment = Comment(note, "CFO AI")


def _write_total_cell(
    ws,
    row: int,
    col: int,
    cell: Dict[str, Any],
    Comment,
    right,
    total_font,
    double_top,
    sum_rows: Optional[List[int]],
    faithful: bool,
) -> None:
    value = _cell_value(cell)
    letter = _column_letter(col - 1)
    if sum_rows:
        formula = "=SUM(%s)" % ",".join(["%s%d" % (letter, n) for n in sum_rows])
        target = ws.cell(row=row, column=col, value=formula)
        note = _provenance_note(cell)
        note = (note + "\n" if note else "") + "Live SUM over the rows above; it reproduces the served total exactly."
    else:
        target = ws.cell(row=row, column=col, value=value if value is not None else MISSING_GLYPH)
        note = _provenance_note(cell)
        if value is not None and not faithful:
            # A1 — the served figure wins, and the reason the formula was
            # withheld is written down rather than left as a mystery.
            note = (note + "\n" if note else "") + (
                "Written as the SERVED total, not a formula: summing the rows above "
                "does not reproduce it. Neither figure was adjusted."
            )
    target.alignment = right
    target.font = total_font
    target.border = double_top
    if value is not None:
        target.number_format = _number_format(
            _s(cell.get("unit")) or None, _s(cell.get("currency")) or None
        )
    if note:
        target.comment = Comment(note, "CFO AI")


# ──────────────────────────────────────────────────────────────────────
# DOCX — minimal, valid WordprocessingML
# ──────────────────────────────────────────────────────────────────────

_DOCX_CONTENT_TYPES = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
    '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
    '<Default Extension="xml" ContentType="application/xml"/>'
    '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
    "</Types>"
)

_DOCX_ROOT_RELS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
    "</Relationships>"
)

_W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _docx_paragraph(text: str, size_half_points: int, bold: bool, italic: bool) -> str:
    """One paragraph, DIRECT-formatted.

    Direct formatting rather than named styles keeps the package to three
    parts: a ``styles.xml`` would have to define every style the document
    references, and a document referencing a style that is not defined is
    the single most common way a hand-built .docx opens blank.
    """
    props = []  # type: List[str]
    if bold:
        props.append("<w:b/>")
    if italic:
        props.append("<w:i/>")
    props.append('<w:sz w:val="%d"/>' % size_half_points)
    props.append('<w:szCs w:val="%d"/>' % size_half_points)
    return (
        "<w:p><w:r><w:rPr>%s</w:rPr><w:t xml:space=\"preserve\">%s</w:t></w:r></w:p>"
        % ("".join(props), _xml_escape(text))
    )


def build_docx(request: Dict[str, Any]) -> bytes:
    sections = request.get("sections")
    if not isinstance(sections, list) or not sections:
        raise ExportRefused("A .docx export needs at least one section.")

    title = _s(request.get("title"), "Artifact")
    citation = request.get("citation") if isinstance(request.get("citation"), dict) else {}

    body = [_docx_paragraph(title, 40, True, False)]
    for line in _citation_lines(citation):
        body.append(_docx_paragraph(line, 16, False, True))
    body.append(_docx_paragraph("", 20, False, False))

    for section in sections:
        if not isinstance(section, dict):
            continue
        heading = _s(section.get("heading"))
        if heading:
            body.append(_docx_paragraph(heading, 28, True, False))
        paragraphs = section.get("paragraphs")
        if isinstance(paragraphs, list):
            for para in paragraphs:
                if isinstance(para, str) and para.strip():
                    body.append(_docx_paragraph(para, 22, False, False))

    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="%s"><w:body>%s'
        '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1134" w:right="1134" w:bottom="1134" w:left="1134" '
        'w:header="708" w:footer="708" w:gutter="0"/></w:sectPr>'
        "</w:body></w:document>" % (_W_NS, "".join(body))
    )

    return _write_parts(
        [
            ("[Content_Types].xml", _DOCX_CONTENT_TYPES),
            ("_rels/.rels", _DOCX_ROOT_RELS),
            ("word/document.xml", document),
        ]
    )


# ──────────────────────────────────────────────────────────────────────
# PPTX — minimal, valid PresentationML
# ──────────────────────────────────────────────────────────────────────

_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_P_NS = "http://schemas.openxmlformats.org/presentationml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

#: 16:9, in EMU. 914400 EMU = 1 inch.
_SLIDE_CX = 12192000
_SLIDE_CY = 6858000
_MARGIN = 685800


def _pptx_theme() -> str:
    """A complete-enough theme.

    ``fmtScheme`` is not optional and is not lenient: PowerPoint requires
    exactly three fill styles, three line styles, three effect styles and
    three background fill styles, and a master whose theme is short one
    of them fails to open rather than falling back. That is why this
    function is long — every element below is load-bearing.
    """
    def _srgb(name, value):
        return '<a:%s><a:srgbClr val="%s"/></a:%s>' % (name, value, name)

    clr = (
        "<a:clrScheme name=\"Instrument\">"
        + _srgb("dk1", "0B0E0D")
        + _srgb("lt1", "FAFAF7")
        + _srgb("dk2", "16211E")
        + _srgb("lt2", "EFEFE9")
        + _srgb("accent1", "0E7C6B")
        + _srgb("accent2", "0E7C6B")
        + _srgb("accent3", "5B6B66")
        + _srgb("accent4", "8A9793")
        + _srgb("accent5", "B7C0BD")
        + _srgb("accent6", "C62828")
        + _srgb("hlink", "0E7C6B")
        + _srgb("folHlink", "0B5B4F")
        + "</a:clrScheme>"
    )
    font = (
        '<a:fontScheme name="Instrument">'
        '<a:majorFont><a:latin typeface="Arial"/><a:ea typeface=""/><a:cs typeface=""/></a:majorFont>'
        '<a:minorFont><a:latin typeface="Arial"/><a:ea typeface=""/><a:cs typeface=""/></a:minorFont>'
        "</a:fontScheme>"
    )
    solid = '<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
    line = (
        '<a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">%s'
        '<a:prstDash val="solid"/></a:ln>' % solid
    )
    fmt = (
        '<a:fmtScheme name="Instrument">'
        "<a:fillStyleLst>%s%s%s</a:fillStyleLst>"
        "<a:lnStyleLst>%s%s%s</a:lnStyleLst>"
        "<a:effectStyleLst>"
        "<a:effectStyle><a:effectLst/></a:effectStyle>"
        "<a:effectStyle><a:effectLst/></a:effectStyle>"
        "<a:effectStyle><a:effectLst/></a:effectStyle>"
        "</a:effectStyleLst>"
        "<a:bgFillStyleLst>%s%s%s</a:bgFillStyleLst>"
        "</a:fmtScheme>" % (solid, solid, solid, line, line, line, solid, solid, solid)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<a:theme xmlns:a="%s" name="Instrument">'
        "<a:themeElements>%s%s%s</a:themeElements>"
        "<a:objectDefaults/><a:extraClrSchemeLst/>"
        "</a:theme>" % (_A_NS, clr, font, fmt)
    )


def _pptx_sp_tree_open() -> str:
    return (
        "<p:spTree>"
        '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
        "<p:grpSpPr><a:xfrm>"
        '<a:off x="0" y="0"/><a:ext cx="0" cy="0"/>'
        '<a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/>'
        "</a:xfrm></p:grpSpPr>"
    )


def _pptx_text_shape(
    shape_id: int,
    name: str,
    x: int,
    y: int,
    cx: int,
    cy: int,
    lines: Sequence[Tuple[str, int, bool]],
) -> str:
    """A text box. ``lines`` are (text, size in hundredths of a point,
    bold)."""
    paragraphs = []  # type: List[str]
    for text, size, bold in lines:
        paragraphs.append(
            '<a:p><a:r><a:rPr lang="en-US" sz="%d" b="%d" dirty="0"/>'
            "<a:t>%s</a:t></a:r></a:p>" % (size, 1 if bold else 0, _xml_escape(text))
        )
    if not paragraphs:
        paragraphs.append("<a:p/>")
    return (
        "<p:sp>"
        '<p:nvSpPr><p:cNvPr id="%d" name="%s"/>'
        '<p:cNvSpPr><a:spLocks noGrp="1"/></p:cNvSpPr><p:nvPr/></p:nvSpPr>'
        "<p:spPr><a:xfrm>"
        '<a:off x="%d" y="%d"/><a:ext cx="%d" cy="%d"/>'
        '</a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></p:spPr>'
        '<p:txBody><a:bodyPr wrap="square"><a:normAutofit/></a:bodyPr><a:lstStyle/>%s</p:txBody>'
        "</p:sp>" % (shape_id, _xml_escape(name), x, y, cx, cy, "".join(paragraphs))
    )


def _pptx_slide(heading: str, lines: Sequence[Tuple[str, int, bool]]) -> str:
    shapes = _pptx_text_shape(
        2, "Heading", _MARGIN, _MARGIN, _SLIDE_CX - _MARGIN * 2, 800000,
        [(heading, 2800, True)],
    ) + _pptx_text_shape(
        3, "Body", _MARGIN, _MARGIN + 900000, _SLIDE_CX - _MARGIN * 2,
        _SLIDE_CY - _MARGIN * 2 - 900000, lines,
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:sld xmlns:a="%s" xmlns:r="%s" xmlns:p="%s">'
        "<p:cSld>%s%s</p:spTree></p:cSld>"
        "<p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr>"
        "</p:sld>" % (_A_NS, _R_NS, _P_NS, _pptx_sp_tree_open(), shapes)
    )


def _pptx_slide_master() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:sldMaster xmlns:a="%s" xmlns:r="%s" xmlns:p="%s">'
        "<p:cSld>%s</p:spTree></p:cSld>"
        '<p:clrMap bg1="lt1" tx1="dk1" bg2="lt2" tx2="dk2" accent1="accent1" '
        'accent2="accent2" accent3="accent3" accent4="accent4" accent5="accent5" '
        'accent6="accent6" hlink="hlink" folHlink="folHlink"/>'
        '<p:sldLayoutIdLst><p:sldLayoutId id="2147483649" r:id="rId1"/></p:sldLayoutIdLst>'
        "</p:sldMaster>" % (_A_NS, _R_NS, _P_NS, _pptx_sp_tree_open())
    )


def _pptx_slide_layout() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:sldLayout xmlns:a="%s" xmlns:r="%s" xmlns:p="%s" type="blank" preserve="1">'
        "<p:cSld name=\"Blank\">%s</p:spTree></p:cSld>"
        "<p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr>"
        "</p:sldLayout>" % (_A_NS, _R_NS, _P_NS, _pptx_sp_tree_open())
    )


def _pptx_lines_for_block(block: Dict[str, Any]) -> List[Tuple[str, int, bool]]:
    """A block becomes text lines. Figures were resolved on the client;
    this only formats what it was handed, and prints the glyph where a
    value is absent."""
    kind = _s(block.get("block"))
    out = []  # type: List[Tuple[str, int, bool]]
    if kind in ("headline", "bullets"):
        lines = block.get("lines")
        if isinstance(lines, list):
            for line in lines:
                if isinstance(line, str) and line.strip():
                    out.append((("• " + line) if kind == "bullets" else line, 1800, kind == "headline"))
    elif kind == "metrics":
        metrics = block.get("metrics")
        if isinstance(metrics, list):
            for metric in metrics:
                if not isinstance(metric, dict):
                    continue
                cell = metric.get("cell") if isinstance(metric.get("cell"), dict) else {}
                out.append((
                    "%s   %s" % (_s(metric.get("label")), _format_cell_text(cell)),
                    2000,
                    True,
                ))
    elif kind == "table":
        rows = block.get("rows")
        if isinstance(rows, list):
            for row in rows:
                if not isinstance(row, dict):
                    continue
                values = "   ".join([_format_cell_text(c) for c in _cells_of(row)])
                out.append(("%s   %s" % (_s(row.get("label")), values), 1600, False))
    return out


def _format_cell_text(cell: Dict[str, Any]) -> str:
    """The one place a figure becomes a string in this module.

    Deliberately plain: grouping with a thin space, a dot decimal, and
    the currency spelled out after the number. A slide leaves this
    product and is read where no locale dial exists, so the NATIVE
    rendering with its currency named is the only form that stays true.
    """
    value = _cell_value(cell)
    if value is None:
        return MISSING_GLYPH
    unit = _s(cell.get("unit"))
    if unit == "percent":
        return "%.1f%%" % (value * 100.0)
    if unit == "ratio":
        return "%.2f×" % value
    if unit == "days":
        return "%.0f d" % value
    text = "{:,.2f}".format(value).replace(",", " ")
    currency = _s(cell.get("currency"))
    return ("%s %s" % (text, currency)).strip()


def build_pptx(request: Dict[str, Any]) -> bytes:
    slides = request.get("slides")
    if not isinstance(slides, list) or not slides:
        raise ExportRefused("A .pptx export needs at least one slide.")

    title = _s(request.get("title"), "Artifact")
    citation = request.get("citation") if isinstance(request.get("citation"), dict) else {}

    # Slide 1 is the cover, and it carries the citation. A deck that
    # leaves the product without naming its period and snapshot is the
    # board pack nobody can check.
    cover_lines = [(line, 1400, False) for line in _citation_lines(citation)]
    slide_xml = [_pptx_slide(title, cover_lines)]

    for slide in slides:
        if not isinstance(slide, dict):
            continue
        blocks = slide.get("blocks")
        lines = []  # type: List[Tuple[str, int, bool]]
        if isinstance(blocks, list):
            for block in blocks:
                if isinstance(block, dict):
                    lines.extend(_pptx_lines_for_block(block))
        slide_xml.append(_pptx_slide(_s(slide.get("heading")), lines))

    n = len(slide_xml)

    content_types = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
        '<Default Extension="xml" ContentType="application/xml"/>',
        '<Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>',
        '<Override PartName="/ppt/slideMasters/slideMaster1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideMaster+xml"/>',
        '<Override PartName="/ppt/slideLayouts/slideLayout1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideLayout+xml"/>',
        '<Override PartName="/ppt/theme/theme1.xml" ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/>',
    ]
    for i in range(1, n + 1):
        content_types.append(
            '<Override PartName="/ppt/slides/slide%d.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>' % i
        )
    content_types.append("</Types>")

    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="%s/officeDocument" Target="ppt/presentation.xml"/>'
        "</Relationships>" % _R_NS
    )

    pres_rels = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                 '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">',
                 '<Relationship Id="rId1" Type="%s/slideMaster" Target="slideMasters/slideMaster1.xml"/>' % _R_NS]
    slide_ids = []  # type: List[str]
    for i in range(1, n + 1):
        rid = "rId%d" % (i + 1)
        pres_rels.append(
            '<Relationship Id="%s" Type="%s/slide" Target="slides/slide%d.xml"/>' % (rid, _R_NS, i)
        )
        slide_ids.append('<p:sldId id="%d" r:id="%s"/>' % (255 + i, rid))
    pres_rels.append(
        '<Relationship Id="rId%d" Type="%s/theme" Target="theme/theme1.xml"/>' % (n + 2, _R_NS)
    )
    pres_rels.append("</Relationships>")

    presentation = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<p:presentation xmlns:a="%s" xmlns:r="%s" xmlns:p="%s" saveSubsetFonts="1">'
        '<p:sldMasterIdLst><p:sldMasterId id="2147483648" r:id="rId1"/></p:sldMasterIdLst>'
        "<p:sldIdLst>%s</p:sldIdLst>"
        '<p:sldSz cx="%d" cy="%d"/><p:notesSz cx="6858000" cy="9144000"/>'
        "</p:presentation>" % (_A_NS, _R_NS, _P_NS, "".join(slide_ids), _SLIDE_CX, _SLIDE_CY)
    )

    master_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="%s/slideLayout" Target="../slideLayouts/slideLayout1.xml"/>'
        '<Relationship Id="rId2" Type="%s/theme" Target="../theme/theme1.xml"/>'
        "</Relationships>" % (_R_NS, _R_NS)
    )
    layout_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="%s/slideMaster" Target="../slideMasters/slideMaster1.xml"/>'
        "</Relationships>" % _R_NS
    )
    slide_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="%s/slideLayout" Target="../slideLayouts/slideLayout1.xml"/>'
        "</Relationships>" % _R_NS
    )

    parts = [
        ("[Content_Types].xml", "".join(content_types)),
        ("_rels/.rels", root_rels),
        ("ppt/presentation.xml", presentation),
        ("ppt/_rels/presentation.xml.rels", "".join(pres_rels)),
        ("ppt/slideMasters/slideMaster1.xml", _pptx_slide_master()),
        ("ppt/slideMasters/_rels/slideMaster1.xml.rels", master_rels),
        ("ppt/slideLayouts/slideLayout1.xml", _pptx_slide_layout()),
        ("ppt/slideLayouts/_rels/slideLayout1.xml.rels", layout_rels),
        ("ppt/theme/theme1.xml", _pptx_theme()),
    ]
    for i, xml in enumerate(slide_xml, start=1):
        parts.append(("ppt/slides/slide%d.xml" % i, xml))
        parts.append(("ppt/slides/_rels/slide%d.xml.rels" % i, slide_rels))
    return _write_parts(parts)


# ──────────────────────────────────────────────────────────────────────
# DISPATCH
# ──────────────────────────────────────────────────────────────────────

MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}

BUILDERS = {"xlsx": build_xlsx, "pptx": build_pptx, "docx": build_docx}


def build_export(request: Dict[str, Any]) -> Tuple[bytes, str, str]:
    """(bytes, media type, extension). Refuses on an unknown version or
    format rather than guessing which one was meant."""
    if not isinstance(request, dict):
        raise ExportRefused("The export request must be an object.")
    version = _s(request.get("version"))
    if version != ARTIFACT_EXPORT_VERSION:
        raise ExportRefused(
            "Unsupported export contract '%s' (this build speaks '%s')."
            % (version, ARTIFACT_EXPORT_VERSION)
        )
    fmt = _s(request.get("format"))
    builder = BUILDERS.get(fmt)
    if builder is None:
        raise ExportRefused(
            "Unsupported format '%s'. CSV is built on the client and never reaches here." % fmt
        )
    return builder(request), MEDIA_TYPES[fmt], fmt


def safe_filename(title: str, citation: Dict[str, Any], extension: str) -> str:
    """Deterministic, filesystem-safe. No clock: two exports of the same
    artifact must produce the same name, or a reader comparing them sees
    two artifacts."""
    def _slug(text: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9]+", "_", text or "").strip("_")
        return cleaned[:60]

    base = _slug(title) or "artifact"
    periods = citation.get("periods") if isinstance(citation, dict) else None
    tail = ""
    if isinstance(periods, list) and periods:
        tail = "_" + _slug("_".join([p for p in periods if isinstance(p, str)]))
    return "%s%s.%s" % (base, tail, extension)


def build_router():
    """``POST /api/artifacts/export`` — bytes out, nothing written.

    ⚠ THE ANNOTATION HAZARD, and why the body is typed ``Dict[str, Any]``.

    This module carries ``from __future__ import annotations``, so every
    annotation is a STRING that FastAPI resolves with ``get_type_hints``
    against the MODULE's globals. FastAPI's own imports live inside this
    factory (so the pure builders above stay importable without
    fastapi), which means a handler annotated ``request: Request``
    resolves to nothing — FastAPI then treats it as an unknown QUERY
    PARAMETER and every POST returns:

        422 {"detail":[{"type":"missing","loc":["query","request"], …}]}

    That is the same forward-ref-in-a-route-factory-closure trap this
    repo already hit on ``CreateCheckoutRequest`` (root CLAUDE.md,
    "Backend cleanup"), and it is invisible to an import check, to a
    unit test of the builders, and to a typecheck. Only a request finds
    it, which is why ``test_the_route_returns_bytes_not_a_422`` exists.

    ``Dict`` and ``Any`` ARE module-level imports, so this annotation
    resolves. ``Body`` is a default VALUE, never an annotation.
    """
    from fastapi import APIRouter, Body, HTTPException
    from fastapi.responses import Response

    router = APIRouter(prefix="/api/artifacts", tags=["artifacts"])

    @router.post("/export")
    async def export_artifact(payload: Dict[str, Any] = Body(...)):  # noqa: B008
        try:
            data, media_type, extension = build_export(payload)
        except ExportRefused as exc:
            raise HTTPException(400, str(exc))
        citation = payload.get("citation") if isinstance(payload, dict) else {}
        filename = safe_filename(
            _s(payload.get("title"), "artifact") if isinstance(payload, dict) else "artifact",
            citation if isinstance(citation, dict) else {},
            extension,
        )
        return Response(
            content=data,
            media_type=media_type,
            headers={"Content-Disposition": 'attachment; filename="%s"' % filename},
        )

    return router
