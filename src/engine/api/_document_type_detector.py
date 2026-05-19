"""Document type detection — routes uploads to the correct parser.

The platform now supports two distinct Romanian financial-statement
shapes, each with a completely different parser:

  - ``trial_balance``      raw balanță de verificare (800+ accounts).
                           The existing path used by EEI + Scandia.
  - ``statutory_f30_f10``  ANAF statutory filing: Formular F30 (P&L)
                           + Formular F10 (Bilanț). Fixed-row layout.
  - ``unknown``            neither — reject with a helpful error.

Detection is deterministic (regex + structural inspection of the first
~30 rows of each sheet). NO LLM calls — adding one here would add
latency, cost, and unreliability for a problem with stable shapes.

Default on ambiguity: ``trial_balance`` — keeps the existing EEI /
Scandia path working when a new edge case shows up. Statutory
recognition demands strong evidence (≥3 F30 anchors, or F30+F10
combined, or F30 + statutory sheet name).

Public API:

    detect_document_type(file_bytes, filename) -> tuple[DocType, dict]

`metadata` carries the anchor-match counts + reason string so the
pipeline can log + the UI can show why the routing decision was made.
"""

from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Literal, Tuple

import openpyxl

try:
    import xlrd  # type: ignore
except ImportError:  # xlrd is optional — legacy .xls files only
    xlrd = None  # type: ignore


DocType = Literal["trial_balance", "statutory_f30_f10", "unknown"]


# Detection anchors for Formular F30 (Contul de Profit și Pierdere).
# These strings appear in the Romanian statutory P&L and effectively
# never in a raw trial-balance export. Patterns are lenient on
# diacritics (Ş/S, ș/s, Â/A) since exports vary.
F30_PL_ANCHORS: List[str] = [
    r"CONTUL\s+DE\s+PROFIT.{0,5}I\s+PIERDERE",       # form title (handles Ș/S)
    r"\bCifra\s+de\s+afaceri\s+net",                 # row 1
    r"\bProduc.ia\s+v.ndut",                         # row 2
    r"VENITURI\s+DIN\s+EXPLOATARE\s*[-–]\s*TOTAL",   # row 16 / row 13 (template-dependent)
    r"CHELTUIELI\s+DE\s+EXPLOATARE\s*[-–]\s*TOTAL",  # row 42
    r"PROFITUL\s+SAU\s+PIERDEREA\s+DIN\s+EXPLOATARE",  # rows 43-44 header
    r"PROFITUL\s+SAU\s+PIERDEREA\s+NET",             # net result header
]

# Detection anchors for Formular F10 (Bilanț).
F10_BS_ANCHORS: List[str] = [
    r"BILAN.{0,2}\s*(?:CONSO|CONSOLIDAT|PRESCURTAT)?",
    r"ACTIVE\s+IMOBILIZATE",
    r"IMOBILIZ.RI\s+NECORPORALE",
    r"IMOBILIZ.RI\s+CORPORALE",
    r"DATORII.{0,40}SUMELE\s+CARE\s+TREBUIE",        # DATORII: SUMELE CARE TREBUIE PLATITE…
    r"CAPITALURI\s+PROPRII",
    r"ACTIVE\s+CIRCULANTE",
]

# Trial-balance anchors — what a raw balanță looks like. Multiple of
# these together is strong evidence we're looking at a TB and NOT a
# statutory return (which doesn't have movement columns).
TB_ANCHORS: List[str] = [
    r"\bbalan.{1,3}\s*de\s*verificare\b",
    r"\bsold\s+(initial|final|init|fin)\b",
    r"\brulaj\s+(debit|credit|cumulat|perioad)\b",
    r"\bsimbol\s*cont\b",
    r"\bcont\s*sintetic\b",
    r"\b(sold|rulaj)\s*(debitor|creditor)\b",
]

# Sheet-name tokens that hint at statutory format. Supplementary
# signal — must combine with at least 2 F30 anchors to count as
# statutory (sheet name alone is too weak).
STATUTORY_SHEET_TOKENS = (
    "P&L", "P_L", "PL ", " PL", "PROFIT_PIERDERE", "PROFIT-PIERDERE",
    "BS", "BILANT", "BILANŢ", "BILANȚ", "F30", "F10", "F-30", "F-10",
)


# ─── Workbook readers (bytes-based, defensive) ──────────────────────────────


def _detect_format(file_bytes: bytes) -> str:
    """Magic-byte format detection. Returns 'xlsx', 'xls', or 'other'.
    Filename extensions lie; the first 8 bytes don't (same convention
    `_trial_balance_parser.detect_excel_format` uses)."""
    if len(file_bytes) < 8:
        return "other"
    if file_bytes[:4] == b"PK\x03\x04":
        return "xlsx"
    if file_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    return "other"


def _xlsx_sheet_names(file_bytes: bytes) -> List[str]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def _xlsx_first_rows(file_bytes: bytes, sheet_idx: int, max_rows: int = 30) -> List[List[str]]:
    """Read first `max_rows` of the given sheet as plain strings. Resilient."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        try:
            if sheet_idx >= len(wb.sheetnames):
                return []
            ws = wb[wb.sheetnames[sheet_idx]]
            out: List[List[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                out.append(["" if c is None else str(c) for c in row])
            return out
        finally:
            wb.close()
    except Exception:
        return []


def _xls_first_rows(file_bytes: bytes, sheet_idx: int, max_rows: int = 30) -> List[List[str]]:
    """Legacy .xls reader (Scandia's path)."""
    if xlrd is None:
        return []
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
        if sheet_idx >= wb.nsheets:
            return []
        ws = wb.sheet_by_index(sheet_idx)
        out: List[List[str]] = []
        for r in range(min(max_rows, ws.nrows)):
            row = []
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                row.append("" if v is None else str(v))
            out.append(row)
        return out
    except Exception:
        return []


# ─── Anchor counting ────────────────────────────────────────────────────────


def _count_matches(rows: List[List[str]], patterns: List[str]) -> int:
    """Count how many distinct patterns appear anywhere in the rows."""
    if not rows:
        return 0
    text = " | ".join(" ".join(r) for r in rows)
    hits = 0
    for pat in patterns:
        if re.search(pat, text, re.IGNORECASE):
            hits += 1
    return hits


def _sheet_name_indicates_statutory(sheet_names: List[str]) -> bool:
    if not sheet_names:
        return False
    joined = " ".join(sheet_names).upper()
    return any(tok in joined for tok in STATUTORY_SHEET_TOKENS)


# ─── Public detector ────────────────────────────────────────────────────────


def detect_document_type(
    file_bytes: bytes,
    filename: str = "",
) -> Tuple[DocType, Dict[str, Any]]:
    """Classify an uploaded file as ``trial_balance`` / ``statutory_f30_f10`` /
    ``unknown``. Returns ``(doc_type, metadata)``.

    metadata keys:
      - reason: human-readable rationale (logged + surfaced to UI)
      - f30_anchors_matched: int
      - f10_anchors_matched: int
      - tb_anchors_matched: int
      - sheet_names: list[str]
      - is_xls_legacy: bool
      - sheet_name_indicates_statutory: bool
    """
    fmt = _detect_format(file_bytes)
    lower = (filename or "").lower()

    # CSVs are exclusively trial balances in Romanian accounting practice.
    # No statutory return is ever distributed as CSV.
    if fmt == "other":
        if lower.endswith(".csv"):
            return "trial_balance", {
                "reason": "CSV format — assumed trial balance.",
                "f30_anchors_matched": 0,
                "f10_anchors_matched": 0,
                "tb_anchors_matched": 0,
                "sheet_names": [],
                "is_xls_legacy": False,
                "sheet_name_indicates_statutory": False,
            }
        return "unknown", {
            "reason": "Unsupported file format (not xlsx/xls/csv).",
            "f30_anchors_matched": 0,
            "f10_anchors_matched": 0,
            "tb_anchors_matched": 0,
            "sheet_names": [],
            "is_xls_legacy": False,
            "sheet_name_indicates_statutory": False,
        }

    is_legacy = fmt == "xls"
    sheet_names = [] if is_legacy else _xlsx_sheet_names(file_bytes)
    statutory_sheet = _sheet_name_indicates_statutory(sheet_names)

    # Scan the first ~3 sheets — statutory files often put F30 + F10 on
    # separate sheets, so we have to look past sheet 0 to detect both.
    sheets_to_check = min(3, max(1, len(sheet_names))) if sheet_names else 1

    f30_hits = 0
    f10_hits = 0
    tb_hits = 0

    for sidx in range(sheets_to_check):
        rows = (
            _xls_first_rows(file_bytes, sidx, max_rows=30)
            if is_legacy
            else _xlsx_first_rows(file_bytes, sidx, max_rows=30)
        )
        f30_hits = max(f30_hits, _count_matches(rows, F30_PL_ANCHORS))
        f10_hits = max(f10_hits, _count_matches(rows, F10_BS_ANCHORS))
        tb_hits = max(tb_hits, _count_matches(rows, TB_ANCHORS))

    metadata: Dict[str, Any] = {
        "f30_anchors_matched": f30_hits,
        "f10_anchors_matched": f10_hits,
        "tb_anchors_matched": tb_hits,
        "sheet_names": sheet_names,
        "is_xls_legacy": is_legacy,
        "sheet_name_indicates_statutory": statutory_sheet,
    }

    # ── Decision logic — conservative thresholds.
    # Statutory needs strong evidence (this is the new path; false
    # positives break the EEI/Scandia regression).
    is_statutory = (
        f30_hits >= 3
        or (f30_hits >= 2 and f10_hits >= 2)
        or (f30_hits >= 2 and statutory_sheet)
    )
    is_trial_balance = tb_hits >= 2

    # When BOTH signals fire (rare — a statutory export that happens to
    # mention "balanță" or vice versa), trust the trial-balance hit. The
    # TB parser is the safe path; if statutory data actually was present
    # the user will re-upload after they see headline-only numbers.
    if is_trial_balance:
        metadata["reason"] = (
            f"Trial balance detected ({tb_hits} TB anchors)."
            + (f" (also saw {f30_hits} F30 anchors — TB wins.)" if is_statutory else "")
        )
        return "trial_balance", metadata

    if is_statutory:
        metadata["reason"] = (
            f"Statutory F30+F10 detected: {f30_hits} F30 anchors, "
            f"{f10_hits} F10 anchors, sheets={sheet_names}."
        )
        return "statutory_f30_f10", metadata

    # Default on ambiguity: trial_balance. The existing pipeline has
    # its own Claude-based fallback, so an ambiguous file still gets
    # a chance. Don't break what works.
    metadata["reason"] = (
        f"Ambiguous — defaulting to trial balance. "
        f"F30={f30_hits}, F10={f10_hits}, TB={tb_hits}."
    )
    return "trial_balance", metadata
