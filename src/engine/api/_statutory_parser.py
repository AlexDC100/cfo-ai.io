"""Parser for the Romanian statutory financial-statement filing.

Scope: Formular F30 (Contul de Profit și Pierdere) + Formular F10
(Bilanț). These follow OMFP 1802/2014 with FIXED ROW NUMBERS — the
"Nr rd." column is standardized so every Romanian company's filing
uses the same row layout for the same canonical line.

Parsing strategy:

  1. Open the workbook with openpyxl. Walk sheets, identify a P&L
     sheet and a BS sheet by name tokens (P&L / Profit_Pierdere /
     F30 for the P&L; BS / Bilant / F10 for the balance sheet).
     Falls back to "first sheet = P&L, second = BS" if names don't
     match.

  2. Locate the "Nr rd." column (defaults to column B / index 2)
     and the value columns (current period + prior period) by
     scanning the header for year tokens.

  3. Walk rows, look up "Nr rd." in F30_ROW_MAP / F10_ROW_MAP,
     extract the value into a dict.

  4. Synthesize a canonical-RO-code accounts list ({code, name,
     amount}) from the F30+F10 line totals. This list flows into
     the existing `_ro_coa.assemble_statements` mapper unchanged,
     so the rest of the pipeline (line items, ratios, valuation,
     briefing) works without modification.

Module is bytes-based and side-effect-free, matching the
`_trial_balance_parser` convention. No DB writes here — the pipeline
orchestrator persists results.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


logger = logging.getLogger(__name__)


# ─── Row maps (OMFP 1802/2014) ──────────────────────────────────────────────


# Formular F30 — Contul de Profit și Pierdere.
# Row numbers ("Nr rd.") are fixed by the regulator. Multiple template
# variants exist (consolidated vs separate; pre-2018 vs current) so a
# handful of rows shift by 1-2 between templates; the synthesizer below
# always prefers the explicit total rows when present, falling back to
# the sum of components otherwise.
F30_ROW_MAP: Dict[int, str] = {
    # Revenue side
    1:  "cifra_afaceri_neta",                  # Cifra de afaceri netă
    2:  "productia_vanduta",                   # ct 701-704 + 707-708
    3:  "venituri_marfuri",                    # ct 707 — wholesale
    4:  "reduceri_comerciale_acordate",        # ct 709 (contra-revenue)
    7:  "variatie_stocuri_credit",             # Sold C (production-side)
    8:  "variatie_stocuri_debit",              # Sold D (consumption-side)
    9:  "venituri_imobilizari",                # ct 722 — capitalized own work
    10: "venituri_reevaluare",
    11: "venituri_investitii_imobiliare",
    12: "venituri_subventii_exploatare",
    13: "alte_venituri_exploatare",            # ct 751 + 758
    16: "venituri_exploatare_total",           # VENITURI DIN EXPLOATARE — TOTAL

    # Cost side
    17: "cheltuieli_materii_prime",            # ct 601 + 602
    18: "alte_cheltuieli_materiale",           # ct 603 + 604 + 606 + 608
    19: "cheltuieli_energie_apa",              # ct 605
    20: "cheltuieli_marfuri",                  # ct 607
    21: "reduceri_comerciale_primite",         # ct 609 (contra-expense)
    22: "cheltuieli_personal_total",           # ct 64
    23: "salarii",                             # ct 641-644
    24: "asigurari_sociale",                   # ct 645
    25: "ajustari_imobilizari",
    26: "cheltuieli_amortizare",               # ct 6811 + 6813 + 6817
    27: "venituri_amortizare",                 # ct 7813
    28: "ajustari_active_circulante_net",
    31: "alte_cheltuieli_exploatare_total",
    32: "cheltuieli_prestatii_externe",        # ct 611-628
    33: "cheltuieli_chirii_redevente",         # ct 612
    42: "cheltuieli_exploatare_total",         # CHELTUIELI DIN EXPLOATARE — TOTAL

    # Operating result
    43: "profit_exploatare",                   # if positive
    44: "pierdere_exploatare",                 # if loss — POSITIVE NUMBER

    # Financial items
    47: "venituri_dobanzi",                    # ct 766
    50: "alte_venituri_financiare",
    52: "venituri_financiare_total",           # VENITURI FINANCIARE — TOTAL
    56: "cheltuieli_dobanzi",                  # ct 666
    57: "cheltuieli_financiare_total",         # one template variant
    62: "cheltuieli_financiare_total_alt",     # other template variant

    # Net result
    65: "profit_brut",                         # gross result before tax
    66: "pierdere_bruta",
    69: "impozit_profit",                      # ct 691
    70: "profit_net",                          # PROFIT NET / pierdere
    71: "pierdere_neta",
}


# Formular F10 — Bilanț.
F10_ROW_MAP: Dict[int, str] = {
    # ASSETS
    25: "active_imobilizate_total",            # ACTIVE IMOBILIZATE — TOTAL
    30: "stocuri_total",                       # STOCURI — TOTAL
    36: "creante_total",                       # CREANȚE — TOTAL
    41: "investitii_financiare_termen_scurt",  # ct 501 + 505 + 506
    42: "casa_si_conturi_la_banci",            # ct 5xx — CASH
    43: "active_circulante_total",             # ACTIVE CIRCULANTE — TOTAL
    45: "cheltuieli_in_avans",

    # LIABILITIES — current
    56: "datorii_pana_la_un_an_total",         # DATORII < 1 AN — TOTAL
    57: "active_circulante_nete_minus_datorii",

    # LIABILITIES — long-term
    58: "total_active_minus_datorii_curente",
    66: "datorii_peste_un_an_total",           # DATORII > 1 AN — TOTAL

    # Provisions and accruals
    73: "provizioane_total",
    77: "venituri_in_avans",

    # EQUITY
    81: "capital_subscris_varsat",             # ct 1012
    91: "rezerve_total",                       # ct 106
    97: "profit_pierdere_reportat",            # ct 117 — retained earnings
    103: "profit_pierdere_exercitiu",          # ct 121 — current year
    106: "capitaluri_proprii_total",           # CAPITALURI PROPRII — TOTAL
    109: "capitaluri_total",                   # CAPITALURI — TOTAL
}


# ─── Label maps (template-agnostic regulatory strings) ─────────────────────


def _pat(rx: str) -> re.Pattern:
    """Compile a regex with IGNORECASE + DOTALL (DOTALL because some
    template cells embed newlines inside the label, e.g.
    `"11.4. Cheltuieli de management (ct. 617),\\ndin care:"`)."""
    return re.compile(rx, re.IGNORECASE | re.DOTALL)


# Regulatory labels in column A. Patterns are matched top-to-bottom per
# row; FIRST match wins. Subset / drilldown rows are EXPLICITLY excluded
# with negative-lookahead guards (`(?!din care)`, `(?!Sume de reluat)`)
# so we never clobber a parent total with one of its components.
#
# Order matters: more specific patterns must come BEFORE more general
# ones (e.g. "Pierdere (rd. 65 + 66" must match before generic "Pierdere
# (rd."). The label_map is consulted before the row-number fallback.

F30_LABEL_MAP: List[Tuple[re.Pattern, str]] = [
    # Revenue side. The "din care" guard excludes subset rows like
    # "din care, cifra de afaceri netă corespunzătoare …" which carry
    # the same monetary digits but are a subset of row 1.
    (_pat(r"^\s*(?!din\b)1\.\s*Cifra\s+de\s+afaceri"),  "cifra_afaceri_neta"),
    (_pat(r"Produc.ia\s+v.ndut"),                       "productia_vanduta"),
    (_pat(r"Venituri\s+din\s+v.nzarea\s+m.rfurilor"),   "venituri_marfuri"),
    (_pat(r"Reduceri\s+comerciale\s+acordate"),         "reduceri_comerciale_acordate"),
    (_pat(r"Venituri\s+din\s+produc.ia\s+de\s+imobiliz"), "venituri_imobilizari"),
    (_pat(r"Venituri\s+din\s+subven.ii\s+de\s+exploatare\s*\(ct"), "venituri_subventii_exploatare"),
    (_pat(r"^\s*7\.\s*Alte\s+venituri\s+din\s+exploatare"),"alte_venituri_exploatare"),
    (_pat(r"VENITURI\s+DIN\s+EXPLOATARE.{0,5}TOTAL"),   "venituri_exploatare_total"),

    # Cost side.
    (_pat(r"Cheltuieli\s+cu\s+materiile?\s+prime"),      "cheltuieli_materii_prime"),
    (_pat(r"Alte\s+cheltuieli\s+materiale"),             "alte_cheltuieli_materiale"),
    (_pat(r"Alte\s+cheltuieli\s+externe.{0,40}energie"), "cheltuieli_energie_apa"),
    (_pat(r"^[\s,]*cheltuieli\s+privind\s+consumul"),    "_skip_energy_subset"),
    (_pat(r"Cheltuieli\s+privind\s+m.rfurile"),          "cheltuieli_marfuri"),
    (_pat(r"Reduceri\s+comerciale\s+primite"),           "reduceri_comerciale_primite"),
    (_pat(r"Cheltuieli\s+cu\s+personalul\s*\(rd"),       "cheltuieli_personal_total"),
    (_pat(r"Salarii\s+.i\s+indemniza.ii"),               "salarii"),
    (_pat(r"Cheltuieli\s+cu\s+asigur.rile"),             "asigurari_sociale"),
    (_pat(r"Ajust.ri\s+de\s+valoare\s+privind\s+imobiliz.rile\s+corporale"),
                                                          "ajustari_imobilizari"),
    (_pat(r"a\.1\)\s*Cheltuieli\s*\(ct\.\s*6811"),       "cheltuieli_amortizare"),
    (_pat(r"a\.2\)\s*Venituri\s*\(ct\.\s*7813"),         "venituri_amortizare"),
    (_pat(r"Ajust.ri\s+de\s+valoare\s+privind\s+activele\s+circulante"),
                                                          "ajustari_active_circulante_net"),
    (_pat(r"Alte\s+cheltuieli\s+de\s+exploatare\s*\(rd"),"alte_cheltuieli_exploatare_total"),
    (_pat(r"Cheltuieli\s+privind\s+presta.iile\s+externe"),"cheltuieli_prestatii_externe"),
    (_pat(r"Cheltuieli\s+cu\s+redeven.ele|cheltuieli\s+cu\s+chiriile\s*\(ct\.\s*6123"),
                                                          "cheltuieli_chirii_redevente"),
    (_pat(r"CHELTUIELI\s+DE\s+EXPLOATARE.{0,5}TOTAL"),   "cheltuieli_exploatare_total"),

    # Operating result. Match by formula reference — distinguishes
    # operating (rd. 16 − 42) from gross / net.
    (_pat(r"Profit\s*\(\s*rd\.\s*16\s*-\s*42"),          "profit_exploatare"),
    (_pat(r"Pierdere\s*\(\s*rd\.\s*42\s*-\s*16"),        "pierdere_exploatare"),

    # Financials.
    (_pat(r"Venituri\s+din\s+dob.nzi\s*\(ct\.\s*766"),   "venituri_dobanzi"),
    (_pat(r"Alte\s+venituri\s+financiare"),              "alte_venituri_financiare"),
    (_pat(r"VENITURI\s+FINANCIARE.{0,5}TOTAL"),          "venituri_financiare_total"),
    (_pat(r"Cheltuieli\s+privind\s+dob.nzile"),          "cheltuieli_dobanzi"),
    (_pat(r"CHELTUIELI\s+FINANCIARE.{0,5}TOTAL"),        "cheltuieli_financiare_total"),

    # Gross + net result. Formula references uniquely identify which
    # template variant we're on — "rd. 62 − 63" is the consolidated
    # gross, "rd. 64 − 65 − 66 − 67" is the consolidated net.
    (_pat(r"Profit\s*\(\s*rd\.\s*62\s*-\s*63"),          "profit_brut"),
    (_pat(r"Pierdere\s*\(\s*rd\.\s*63\s*-\s*62"),        "pierdere_bruta"),
    (_pat(r"Impozitul?\s+pe\s+profit\s*\(ct\.\s*691"),   "impozit_profit"),
    (_pat(r"Profit\s*\(\s*rd\.\s*64\s*-\s*65"),          "profit_net"),
    (_pat(r"Pierdere\s*\(\s*rd\.\s*65\s*\+\s*66"),       "pierdere_neta"),
]

F10_LABEL_MAP: List[Tuple[re.Pattern, str]] = [
    (_pat(r"ACTIVE\s+IMOBILIZATE\s*[-–]\s*TOTAL"),                "active_imobilizate_total"),
    # STOCURI total — the parent of "1. Materii prime…2. Producția…3. Produse finite".
    # The label is literally "TOTAL (rd. 26 la 29)" in the consolidated template
    # or "STOCURI - TOTAL" in others; we match on either.
    (_pat(r"^\s*STOCURI\s*[-–]\s*TOTAL|TOTAL\s*\(rd\.\s*26\s*la\s*29"),
                                                                  "stocuri_total"),
    (_pat(r"^\s*CREAN.E\s*[-–]\s*TOTAL|TOTAL\s*\(rd\.\s*31\s*la\s*35"),
                                                                  "creante_total"),
    (_pat(r"INVESTI.II\s+PE\s+TERMEN\s+SCURT.{0,20}TOTAL|TOTAL\s*\(rd\.\s*37\s*\+\s*38"),
                                                                  "investitii_financiare_termen_scurt"),
    (_pat(r"CASA\s+.I\s+CONTURI\s+LA\s+B.NCI"),                   "casa_si_conturi_la_banci"),
    (_pat(r"ACTIVE\s+CIRCULANTE\s*[-–]\s*TOTAL"),                 "active_circulante_total"),
    (_pat(r"CHELTUIELI\s+.N\s+AVANS\s*\(ct\.\s*471"),             "cheltuieli_in_avans"),

    # The DATORII totals are the trickiest — the same label "TOTAL"
    # appears twice (short-term at "rd. 45 la 52", long-term at
    # "rd. 56 la 63"). Use the formula reference to disambiguate.
    (_pat(r"TOTAL\s*\(rd\.\s*45\s+la\s+52"),                      "datorii_pana_la_un_an_total"),
    (_pat(r"ACTIVE\s+CIRCULANTE\s+NETE.{0,20}DATORII\s+CURENTE"), "active_circulante_nete_minus_datorii"),
    (_pat(r"TOTAL\s+ACTIVE\s+MINUS\s+DATORII\s+CURENTE"),         "total_active_minus_datorii_curente"),
    (_pat(r"TOTAL\s*\(rd\.\s*56\s+la\s+63|TOTAL\s*\(rd\.56\s+la\s+63"),
                                                                  "datorii_peste_un_an_total"),

    (_pat(r"TOTAL\s*\(rd\.\s*65\s+la\s+67"),                      "provizioane_total"),
    (_pat(r"^\s*2\.\s*Venituri\s+.nregistrate\s+.n\s+avans"),     "venituri_in_avans"),

    # Equity. Capital subscris vărsat carries `(ct. 1012)`.
    (_pat(r"Capital\s+subscris\s+v.rsat\s*\(ct\.\s*1012"),        "capital_subscris_varsat"),
    (_pat(r"REZERVE\s+DIN\s+REEVALUARE\s*\(ct\.\s*105"),          "rezerve_din_reevaluare"),
    (_pat(r"TOTAL\s*\(rd\.\s*88\s+la\s+90"),                      "rezerve_total"),
    (_pat(r"PROFITUL\s+SAU\s+PIERDEREA\s+REPORTAT|SOLD\s+C\s*\(ct\.\s*117"),
                                                                  "profit_pierdere_reportat"),
    (_pat(r"SOLD\s+C\s*\(ct\.\s*121"),                            "profit_pierdere_exercitiu_credit"),
    (_pat(r"SOLD\s+D\s*\(ct\.\s*121"),                            "profit_pierdere_exercitiu_debit"),
    (_pat(r"CAPITALURI\s+PROPRII\s*[-–]\s*TOTAL"),                "capitaluri_proprii_total"),
    (_pat(r"CAPITALURI\s*[-–]\s*TOTAL\s*\(rd\.100\+101\+102|CAPITALURI\s*[-–]\s*TOTAL\s*\(rd\.\s*103"),
                                                                  "capitaluri_total"),
]


# ─── Result container ───────────────────────────────────────────────────────


@dataclass
class StatutoryExtractionResult:
    """Output of `parse_statutory_file`. Mirrors `_trial_balance_parser`'s
    convention — pure data, no side effects."""
    pl_data: Dict[str, float] = field(default_factory=dict)
    bs_data: Dict[str, float] = field(default_factory=dict)
    period_end: Optional[str] = None
    period_prior: Optional[str] = None
    pl_data_prior: Dict[str, float] = field(default_factory=dict)
    bs_data_prior: Dict[str, float] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    detected_industry: Optional[str] = None
    company_name: Optional[str] = None
    pl_sheet_name: Optional[str] = None
    bs_sheet_name: Optional[str] = None


# ─── Sheet identification ───────────────────────────────────────────────────


_PL_SHEET_TOKENS = ("P&L", "PROFIT", "F30", "F-30", "PL", "PROFIT_PIERDERE", "PROFIT-PIERDERE")
_BS_SHEET_TOKENS = ("BILANT", "BILANŢ", "BILANȚ", "F10", "F-10", "BS", "BILANCE")


def _identify_pl_bs_sheets(wb: openpyxl.Workbook) -> Tuple[Optional[str], Optional[str], List[str]]:
    """Return (pl_sheet_name, bs_sheet_name, warnings).
    Falls back to first two sheets when names don't match the tokens."""
    warnings: List[str] = []
    pl_name: Optional[str] = None
    bs_name: Optional[str] = None

    for sn in wb.sheetnames:
        upper = sn.upper().strip()
        # Match BS first — "PROFIT" appears in some BS sheet names but the
        # F30 marker is unambiguous, so explicit-F30 wins on the second
        # pass.
        if bs_name is None and any(tok in upper for tok in _BS_SHEET_TOKENS):
            bs_name = sn

    for sn in wb.sheetnames:
        upper = sn.upper().strip()
        if pl_name is None and any(tok in upper for tok in _PL_SHEET_TOKENS):
            # Don't reuse the same sheet as both P&L and BS.
            if sn != bs_name:
                pl_name = sn

    # Fallback: first sheet = P&L, second = BS. We warn so the UI can
    # surface that the sheet-name detection didn't lock cleanly.
    if pl_name is None or bs_name is None:
        warnings.append(
            f"Could not identify P&L + BS sheets by name. Found: {wb.sheetnames}. "
            "Falling back to first sheet = P&L, second = BS."
        )
        if len(wb.sheetnames) >= 1 and pl_name is None:
            pl_name = wb.sheetnames[0]
        if len(wb.sheetnames) >= 2 and bs_name is None:
            bs_name = wb.sheetnames[1] if wb.sheetnames[1] != pl_name else (
                wb.sheetnames[0] if pl_name != wb.sheetnames[0] else None
            )

    return pl_name, bs_name, warnings


# ─── Header / column discovery ──────────────────────────────────────────────


_ISO_DATE = re.compile(r"(20\d{2})-(\d{2})-(\d{2})")
_EU_DATE = re.compile(r"(\d{2})[./-](\d{2})[./-](20\d{2})")
_YEAR = re.compile(r"\b(20\d{2})\b")


def _looks_like_year_header(cell_value: Any) -> Optional[int]:
    """Conservative year extraction for column-detection.

    A cell counts as a "year header" only if it's short (≤ 30 chars
    once stringified) AND contains a 20XX in the valid period
    (2010-2099). This excludes account-label cells that happen to
    contain account numbers like `ct.2071-2807` — those previously
    produced false-positive year=2071 hits and routed the parser to
    read labels as values."""
    if cell_value is None:
        return None
    # openpyxl returns native datetime objects for typed date cells —
    # always trust those.
    if hasattr(cell_value, "year") and isinstance(getattr(cell_value, "year", None), int):
        y = cell_value.year
        return y if 2010 <= y <= 2099 else None
    s = str(cell_value).strip()
    if not s or len(s) > 30:
        return None
    m = _YEAR.search(s)
    if not m:
        return None
    y = int(m.group(1))
    return y if 2010 <= y <= 2099 else None


def _extract_period_dates(ws) -> Tuple[Optional[str], Optional[str]]:
    """Scan the first ~8 rows for a current-period and prior-period date.
    Returns ISO strings (YYYY-MM-DD) when found. The most recent date is
    the current period; the older one is the prior comparative.

    Datetime cells are trusted directly; string cells are scanned with
    ISO / DD.MM.YYYY patterns. The year fallback only fires when nothing
    else matched, and is gated by `_looks_like_year_header` so we don't
    pick "2071" out of an account-code label."""
    found: List[str] = []
    for r in range(1, min(9, ws.max_row + 1)):
        for c in range(1, min(20, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            # Native datetime cells — most reliable signal.
            if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
                try:
                    y, m, d = int(v.year), int(v.month), int(v.day)
                    if 2010 <= y <= 2099 and 1 <= m <= 12 and 1 <= d <= 31:
                        found.append(f"{y:04d}-{m:02d}-{d:02d}")
                        continue
                except Exception:
                    pass
            s = str(v)
            m = _ISO_DATE.search(s)
            if m:
                found.append(f"{m.group(1)}-{m.group(2)}-{m.group(3)}")
                continue
            m = _EU_DATE.search(s)
            if m:
                found.append(f"{m.group(3)}-{m.group(2)}-{m.group(1)}")

    if not found:
        years: List[int] = []
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                y = _looks_like_year_header(ws.cell(r, c).value)
                if y is not None:
                    years.append(y)
        if years:
            uniq = sorted(set(years), reverse=True)
            current = f"{uniq[0]}-12-31"
            prior = f"{uniq[1]}-12-31" if len(uniq) > 1 else None
            return current, prior
        return None, None

    uniq = sorted(set(found), reverse=True)
    current = uniq[0]
    prior = uniq[1] if len(uniq) > 1 else None
    return current, prior


def _find_value_columns(ws) -> Tuple[int, Optional[int]]:
    """Decide which columns hold the current-period and prior-period numbers.

    Strategy: tag each header column with its year (via
    `_looks_like_year_header`, which rejects long label cells with
    embedded account codes). The column with the latest year is
    current; the next-latest is prior. When the header has no year
    markers, fall back to the OMFP 1802 layout: col D = current,
    col C = prior."""
    col_year: Dict[int, int] = {}
    for r in range(1, min(5, ws.max_row + 1)):
        for c in range(1, min(20, ws.max_column + 1)):
            y = _looks_like_year_header(ws.cell(r, c).value)
            if y is None:
                continue
            if y > col_year.get(c, 0):
                col_year[c] = y

    if len(col_year) >= 2:
        ordered = sorted(col_year.items(), key=lambda kv: -kv[1])
        return ordered[0][0], ordered[1][0]
    if len(col_year) == 1:
        only_col = next(iter(col_year.keys()))
        return only_col, max(1, only_col - 1)
    return 4, 3


def _find_row_number_col(ws) -> int:
    """Find the "Nr rd." column. OMFP 1802 puts it in column B (index 2),
    but we scan the first few rows just in case the export shifted layout."""
    for c in range(1, min(6, ws.max_column + 1)):
        for r in range(1, min(8, ws.max_row + 1)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if ("nr" in s and "rd" in s) or s == "rd" or s == "nr.rd.":
                return c
    return 2


def _find_company_name(ws, max_rows: int = 8) -> Optional[str]:
    """Look for a "Denumirea entității"/"Entitatea" cell in the header.
    Returns the value of the cell immediately to its right (or the
    cell value itself if it already contains the company name)."""
    pat = re.compile(r"(denumirea\s+entit|entitatea|denumire\s+raportor|cui|cod\s+unic)", re.IGNORECASE)
    for r in range(1, min(max_rows, ws.max_row + 1)):
        for c in range(1, min(12, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if pat.search(str(v)):
                # Check the cell to the right + one down for the value.
                for cand in ((r, c + 1), (r, c + 2), (r + 1, c)):
                    rr, cc = cand
                    if rr <= ws.max_row and cc <= ws.max_column:
                        cv = ws.cell(rr, cc).value
                        if cv is not None and str(cv).strip():
                            name = str(cv).strip()
                            # Reject if it's another label
                            if not pat.search(name):
                                return name
    return None


# ─── Value parsing ──────────────────────────────────────────────────────────


def _to_float(v: Any) -> Optional[float]:
    """Robust float conversion. Handles None, blanks, numerics, Romanian-
    locale strings (1.234,56) and Anglo-locale strings (1,234.56)."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Strip currency / non-breaking spaces / unit suffixes.
    s = s.replace("RON", "").replace("EUR", "").replace("\xa0", "").replace(" ", "")
    # Parentheses = negative (common in old templates).
    negate = False
    if s.startswith("(") and s.endswith(")"):
        negate = True
        s = s[1:-1]
    if "," in s and "." in s:
        # Two separators — last one wins as decimal.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and 1 <= len(parts[1]) <= 2:
            s = parts[0] + "." + parts[1]
        else:
            s = s.replace(",", "")
    try:
        x = float(s)
        return -x if negate else x
    except ValueError:
        return None


_BANK_DEBT_LABEL = _pat(r"Sume\s+datorate\s+institu.iilor\s+de\s+credit")
# Section headers — accept "UN AN" (spelled) or "1 AN" (digit). The
# food-manufacturer file uses the digit form.
_ST_SECTION_HEADER = _pat(r"DATORII.{0,5}SUMELE\s+CARE\s+TREBUIE\s+PL.TITE\s+.NTR.O\s+PERIOAD.{0,5}DE\s+P.N.\s+LA\s+(?:UN|1)\s+AN")
_LT_SECTION_HEADER = _pat(r"DATORII.{0,5}SUMELE\s+CARE\s+TREBUIE\s+PL.TITE\s+.NTR.O\s+PERIOAD.{0,5}MAI\s+MARE\s+DE\s+(?:UN|1)\s+AN")
_TRADE_PAYABLES_LABEL = _pat(r"Datorii\s+comerciale.{0,5}furnizori")
_EFFECTS_PAYABLE_LABEL = _pat(r"Efecte\s+de\s+come?r.\s+de\s+pl.tit")


def _extract_bank_debt_split(
    ws,
    current_col: int,
    prior_col: Optional[int],
    out_current: Dict[str, float],
    out_prior: Dict[str, float],
) -> None:
    """Walk the BS sheet, find the ST liabilities section header and
    LT liabilities section header, then capture the "Sume datorate
    instituţiilor de credit" rows inside each section as `bank_debt_st`
    and `bank_debt_lt`. Also captures `trade_payables_st` (Datorii
    comerciale - furnizori) and `effects_payable_st` so we can route
    those to AP (code 401) instead of letting them inflate `total_debt`."""
    section: Optional[str] = None  # None / "st" / "lt"

    def _write(key: str, r: int) -> None:
        if key in out_current:
            return
        cv = _to_float(ws.cell(r, current_col).value)
        if cv is not None:
            out_current[key] = cv
        if prior_col is not None:
            pv = _to_float(ws.cell(r, prior_col).value)
            if pv is not None:
                out_prior[key] = pv

    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        s = str(label)
        if _ST_SECTION_HEADER.search(s):
            section = "st"
            continue
        if _LT_SECTION_HEADER.search(s):
            section = "lt"
            continue
        if section is None:
            continue
        # Bank debt — the row we most care about.
        if _BANK_DEBT_LABEL.search(s):
            _write(f"bank_debt_{section}", r)
            continue
        # Trade payables — ST only (LT trade payables aren't a thing
        # in OMFP RAS — long-term suppliers don't get their own line).
        if section == "st" and _TRADE_PAYABLES_LABEL.search(s):
            _write("trade_payables_st", r)
            continue
        if section == "st" and _EFFECTS_PAYABLE_LABEL.search(s):
            _write("effects_payable_st", r)


def _combine_profit_loss(profit: float, loss: float) -> float:
    """Combine a profit cell + a loss cell into a single signed result.

    Two conventions exist in OMFP statutory templates:

      (a) Both profit and loss carry positive magnitudes; only one is
          non-zero. Net result = profit − loss (e.g. profit=100, loss=0
          → +100; profit=0, loss=50 → −50).
      (b) Loss is stored as a NEGATIVE number (signed), profit stays 0.
          The food-manufacturer file uses this convention:
          profit_exploatare=0, pierdere_exploatare=−40,734,209.
          Net result must be the signed loss (−40,734,209), so we must
          treat the magnitude of loss, not its raw sign.

    Combining `profit - abs(loss)` covers both cases without needing
    to know which template the file came from."""
    p = profit or 0.0
    l = loss or 0.0
    return p - abs(l)


def _parse_sheet(
    ws,
    row_map: Dict[int, str],
    label_map: List[Tuple[re.Pattern, str]],
    current_col: int,
    prior_col: Optional[int],
) -> Tuple[Dict[str, float], Dict[str, float]]:
    """Walk every row, identify it by LABEL first, fall back to "Nr rd."
    only when label matching fails.

    Why label-first: OMFP statutory templates ship in multiple variants
    (individual-entity Formular F30/F10 vs consolidated BILANŢ_CONSO).
    Row numbers shift between variants (e.g. "CAPITALURI PROPRII TOTAL"
    is rd=106 in individual filings, rd=100 in consolidated). The
    Romanian regulatory LABEL strings are stable across variants —
    "VENITURI DIN EXPLOATARE - TOTAL" reads identically everywhere.

    Subset rows (rd=1a / 33b / 35a) and label rows that look like
    subsets ("din care…", "Sume de reluat…") are skipped so they
    don't clobber the parent total. Label patterns can include
    negative-lookahead guards to enforce this where needed."""
    rd_col = _find_row_number_col(ws)
    current: Dict[str, float] = {}
    prior: Dict[str, float] = {}

    def _commit(key: str, r: int) -> None:
        # Only write the FIRST match for a key — once committed, later
        # subset / re-stated rows can't overwrite.
        if key in current and current[key] is not None:
            return
        cur = _to_float(ws.cell(r, current_col).value)
        if cur is not None:
            current[key] = cur
        if prior_col:
            prv = _to_float(ws.cell(r, prior_col).value)
            if prv is not None:
                prior[key] = prv

    for r in range(1, ws.max_row + 1):
        # Label match (preferred). Column A holds the label.
        label_raw = ws.cell(r, 1).value
        label = "" if label_raw is None else str(label_raw)
        if label:
            for pat, key in label_map:
                if pat.search(label):
                    _commit(key, r)
                    break  # first match wins per row
        # Row-number fallback — only when no label match committed AND
        # the rd value is a pure integer (no "1a" suffix).
        rd_value = ws.cell(r, rd_col).value
        if rd_value is None:
            continue
        rd_str = str(rd_value).strip()
        m = re.match(r"^\s*(\d+)\s*$", rd_str)
        if not m and re.match(r"^\s*\d+\.0\s*$", rd_str):
            m = re.match(r"^\s*(\d+)", rd_str)
        if not m:
            continue
        try:
            rd_int = int(m.group(1))
        except ValueError:
            continue
        if rd_int not in row_map:
            continue
        _commit(row_map[rd_int], r)

    return current, prior


# ─── Industry hint (heuristic — must be user-confirmed) ─────────────────────


def _detect_industry(pl: Dict[str, float], bs: Dict[str, float]) -> str:
    """Cheap heuristic from the P&L shape. NOT authoritative — the user
    should confirm in onboarding (this is a starting hint only)."""
    revenue = pl.get("cifra_afaceri_neta", 0) or 0
    materii_prime = pl.get("cheltuieli_materii_prime", 0) or 0
    if revenue <= 0:
        return "unknown"
    ratio = materii_prime / revenue
    if ratio > 0.25:
        # Food / agri tend to run ratios above 40% with meaningful energy
        # costs. We don't have category-level signal here, so flag as
        # consumer manufacturing only when raw-material intensity is high.
        if ratio > 0.40:
            return "manufacturing_consumer"
        return "manufacturing_general"
    if ratio < 0.05:
        return "real_estate_commercial"
    return "manufacturing_general"


# ─── Public API ─────────────────────────────────────────────────────────────


def parse_statutory_file(file_bytes: bytes, filename: str = "") -> StatutoryExtractionResult:
    """Parse a Formular F30 + F10 .xlsx workbook from raw bytes.
    Bytes-in, dataclass-out — matches `_trial_balance_parser` convention.

    Raises ValueError on a workbook that can't be opened. Returns a result
    object with empty pl_data on a workbook that opens but has no rows we
    recognize; callers should treat that as a failure (the spec mandates
    raising an error in the pipeline router)."""
    result = StatutoryExtractionResult()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Could not open workbook: {e}") from e

    try:
        pl_name, bs_name, warnings = _identify_pl_bs_sheets(wb)
        result.warnings.extend(warnings)
        result.pl_sheet_name = pl_name
        result.bs_sheet_name = bs_name

        if pl_name is None:
            result.warnings.append("No P&L sheet identified — extraction skipped.")
            return result

        pl_ws = wb[pl_name]
        # Company name + periods from the P&L header.
        result.company_name = _find_company_name(pl_ws)
        result.period_end, result.period_prior = _extract_period_dates(pl_ws)

        pl_cur_col, pl_prv_col = _find_value_columns(pl_ws)
        pl_current, pl_prior = _parse_sheet(pl_ws, F30_ROW_MAP, F30_LABEL_MAP, pl_cur_col, pl_prv_col)
        result.pl_data = pl_current
        result.pl_data_prior = pl_prior

        if bs_name is not None:
            bs_ws = wb[bs_name]
            bs_cur_col, bs_prv_col = _find_value_columns(bs_ws)
            bs_current, bs_prior = _parse_sheet(bs_ws, F10_ROW_MAP, F10_LABEL_MAP, bs_cur_col, bs_prv_col)
            # Post-pass: split bank debt (ct. 1621 + …) into ST and LT
            # buckets. OMFP F10 has TWO rows with identical label
            # "Sume datorate instituţiilor de credit (ct. 1621 + …)" —
            # the first sits inside the < 1 an block, the second inside
            # the > 1 an block. Without this split, `total_debt` in
            # calculated_metrics conflates trade payables + social +
            # fiscal with bank debt and wildly overstates leverage
            # (~1.0B vs the ~565M of actual bank loans on the test file).
            _extract_bank_debt_split(bs_ws, bs_cur_col, bs_prv_col, bs_current, bs_prior)
            result.bs_data = bs_current
            result.bs_data_prior = bs_prior
        else:
            result.warnings.append("No BS sheet identified — balance-sheet metrics missing.")

        result.detected_industry = _detect_industry(result.pl_data, result.bs_data)

    finally:
        wb.close()

    return result


# ─── Canonical-RO-code synthesizer ──────────────────────────────────────────


# Lookup tables for the synthesized accounts. Codes match the existing
# `_ro_coa.py` MAPPING_RULES so `assemble_statements` consumes the output
# without any change to the COA layer.
#
# Naming convention: when a statutory line aggregates several RO accounts
# (e.g. "cheltuieli_personal_total" = 64*), pick the most representative
# leaf code (641 — salaries). The mapper buckets ALL 64x to
# `operatingExpenses` anyway, so picking 641 is correct as long as we
# don't double-count by also emitting 645 + 646 + ….


def _synth_accounts_from_extraction(
    pl: Dict[str, float],
    bs: Dict[str, float],
) -> List[Dict[str, Any]]:
    """Build {code, name, amount} accounts matching the trial-balance
    parser output shape. Skips zero/null values so downstream doesn't
    spam empty rows."""
    rows: List[Dict[str, Any]] = []

    def add(code: str, name: str, amount: float) -> None:
        if amount is None:
            return
        # Cap noise: skip vanishingly small lines (< 1 RON) — almost
        # always template-rounding artifacts.
        if abs(amount) < 1:
            return
        rows.append({"code": code, "name": name, "amount": round(float(amount), 2)})

    # ── P&L ───────────────────────────────────────────────────────────────
    # Revenue: prefer the explicit Cifra de afaceri (row 1) plus
    # complementary income lines so total_operating_revenue derived
    # downstream matches what the F30 reports.
    cifra = float(pl.get("cifra_afaceri_neta", 0) or 0)
    if cifra:
        # Bucket: revenue. 706 sums both Romanian sale-of-goods (701-704)
        # and services. The mapper routes 706 → revenue.
        add("706", "Cifra de afaceri netă", cifra)

    venituri_marfuri = float(pl.get("venituri_marfuri", 0) or 0)
    # Row 3 is the wholesale subset of row 1. It's already in cifra above,
    # so DON'T re-emit it as 707 — that would double-count revenue.

    # Capitalized own work (account 722) — IMPORTANT semantic choice.
    # In the trial-balance pipeline, 722 maps to `capitalizedOwnWork`
    # (a memo bucket excluded from cash EBITDA — the CFO-AI convention
    # that a non-cash internal-construction credit shouldn't inflate
    # operating profit).
    #
    # For the STATUTORY pipeline, however, the user explicitly expects
    # the F30's reported EBITDA (which DOES include 722 because OMFP
    # treats it as part of venituri din exploatare). Per the spec's
    # "ONE HONEST CAUTION": the two pipelines produce slightly
    # different metrics; the statutory pipeline mirrors what F30 row
    # 43 + amortization says. So we route 722's value through code 758
    # (otherIncome) here — that makes statutory EBITDA = F30 EBITDA.
    cap_own = float(pl.get("venituri_imobilizari", 0) or 0)
    if cap_own:
        add("758", "Venituri din producția de imobilizări (722, included in EBITDA)", cap_own)

    # Inventory variation memo (711) — the F30 nets it on rows 7-8. The
    # downstream cash-EBITDA computation excludes 711 by design, so we
    # surface the net value under code 711 for visibility but the mapper
    # routes it to inventoryVariationMemo (excluded from cash EBITDA).
    var_credit = float(pl.get("variatie_stocuri_credit", 0) or 0)
    var_debit = float(pl.get("variatie_stocuri_debit", 0) or 0)
    var_net = var_credit - var_debit
    if abs(var_net) > 1:
        add("711", "Variația stocurilor (rd 7-8 net)", var_net)

    alte_ven = float(pl.get("alte_venituri_exploatare", 0) or 0)
    if alte_ven:
        add("758", "Alte venituri din exploatare", alte_ven)

    subv = float(pl.get("venituri_subventii_exploatare", 0) or 0)
    if subv:
        # Code 740 is the COA's canonical "subvenții" rule → otherIncome.
        # 741 isn't mapped (no Rule("741", …) in _ro_coa); use 740.
        add("740", "Venituri din subvenții de exploatare", subv)

    # Costs — materials.
    materii = float(pl.get("cheltuieli_materii_prime", 0) or 0)
    if materii:
        add("602", "Cheltuieli cu materialele consumabile", materii)
    alte_mat = float(pl.get("alte_cheltuieli_materiale", 0) or 0)
    if alte_mat:
        add("603", "Cheltuieli cu materialele de natura obiectelor de inventar", alte_mat)
    energie = float(pl.get("cheltuieli_energie_apa", 0) or 0)
    if energie:
        add("605", "Cheltuieli cu energia și apa", energie)
    marfuri = float(pl.get("cheltuieli_marfuri", 0) or 0)
    if marfuri:
        add("607", "Cheltuieli cu mărfurile", marfuri)

    # Personnel — we emit the aggregate to 641 (mapped to operatingExpenses)
    # rather than splitting salarii / asigurari / etc, because the
    # individual components would either (a) overlap if both row 22 total
    # and rows 23-24 components are present, or (b) miss subcategories
    # entirely when only the total row is filled. Picking the total row
    # avoids both pitfalls.
    personal = float(pl.get("cheltuieli_personal_total", 0) or 0)
    if personal:
        add("641", "Cheltuieli cu salariile + asigurări (rd 22 total)", personal)

    # "Alte cheltuieli de exploatare" — F30 rd 31 is the TOTAL of rd 32
    # (prestații externe) + rd 33 (chirii + redevențe) + rd 34 (protecția
    # mediului) + rd 36 (calamități) + rd 37 (alte). Emitting the
    # sub-lines individually misses rd 34/36/37 and undercounts opex by
    # ~20M on the test file. Emitting the total under code 658 (which
    # routes to operatingExpenses via the Class-65 catchall rule)
    # captures everything once. Trade-off: line-item drilldown loses the
    # prestații/chirii split — but the statutory pipeline is explicitly
    # aggregate-only, so this is consistent with the spec.
    alte_chelt_total = float(pl.get("alte_cheltuieli_exploatare_total", 0) or 0)
    if alte_chelt_total:
        add("658", "Alte cheltuieli de exploatare — TOTAL (F30 rd 31)", alte_chelt_total)
    else:
        # Older template that doesn't expose the rd-31 total — fall back
        # to the components.
        servicii = float(pl.get("cheltuieli_prestatii_externe", 0) or 0)
        if servicii:
            add("628", "Alte cheltuieli cu serviciile prestate de terți", servicii)
        chirii = float(pl.get("cheltuieli_chirii_redevente", 0) or 0)
        if chirii:
            add("612", "Cheltuieli cu chiriile și redevențele", chirii)

    # Depreciation / amortization (net: row 26 expense − row 27 reversal).
    amort_exp = float(pl.get("cheltuieli_amortizare", 0) or 0)
    amort_rev = float(pl.get("venituri_amortizare", 0) or 0)
    amort_net = amort_exp - amort_rev
    if amort_net:
        add("6811", "Cheltuieli de exploatare privind amortizarea imobilizărilor", amort_net)

    # Provisions for current assets (net) — route to depreciation bucket so
    # cash EBITDA strips it out (matches trial-balance treatment of 6814).
    ajust_ca = float(pl.get("ajustari_active_circulante_net", 0) or 0)
    if ajust_ca:
        add("6814", "Ajustări pentru deprecierea activelor circulante", ajust_ca)

    # Interest.
    chelt_dob = float(pl.get("cheltuieli_dobanzi", 0) or 0)
    if chelt_dob:
        add("666", "Cheltuieli privind dobânzile", chelt_dob)
    ven_dob = float(pl.get("venituri_dobanzi", 0) or 0)
    if ven_dob:
        add("766", "Venituri din dobânzi", ven_dob)

    # Tax (current-year).
    tax = float(pl.get("impozit_profit", 0) or 0)
    if tax:
        add("691", "Cheltuieli cu impozitul pe profit curent", tax)

    # NOTE: we intentionally do NOT emit the current-year net result
    # under any equity code here. The COA assembler (`_ro_coa.py:777`)
    # automatically adds `net_income_statutory` (computed from the
    # synthesized P&L lines below) back into retainedEarnings.
    # Emitting it under 1171 would double-count. The cash-EBITDA value
    # the assembler computes may differ slightly from F30's reported
    # net result (cash vs statutory definitional differences) — that's
    # the documented "ONE HONEST CAUTION" divergence between pipelines.

    # ── Balance sheet ─────────────────────────────────────────────────────
    cash = float(bs.get("casa_si_conturi_la_banci", 0) or 0)
    if cash:
        add("5121", "Casa și conturi la bănci (F10 rd 42)", cash)

    inventory = float(bs.get("stocuri_total", 0) or 0)
    if inventory:
        add("371", "Stocuri — total (F10 rd 30)", inventory)

    receivables = float(bs.get("creante_total", 0) or 0)
    if receivables:
        add("4111", "Creanțe — total (F10 rd 36)", receivables)

    fixed_assets = float(bs.get("active_imobilizate_total", 0) or 0)
    if fixed_assets:
        add("212", "Active imobilizate — total (F10 rd 25)", fixed_assets)

    cheltuieli_avans = float(bs.get("cheltuieli_in_avans", 0) or 0)
    if cheltuieli_avans:
        add("471", "Cheltuieli înregistrate în avans (F10 rd 45)", cheltuieli_avans)

    # Short-term investments — route to cash bucket via 505 (the mapper's
    # closest match for "cash equivalents"). Conservative when the value
    # is small; if it's large, the user can re-classify post-upload.
    st_inv = float(bs.get("investitii_financiare_termen_scurt", 0) or 0)
    if st_inv:
        add("508", "Alte investiții financiare pe termen scurt (F10 rd 41)", st_inv)

    # Split the ST/LT liability totals into bank debt vs trade payables
    # vs "other current/non-current". `bank_debt_*` and
    # `trade_payables_st` were captured in the post-pass; everything
    # else in the total goes to the "other" residual buckets so the
    # books balance.
    st_total = float(bs.get("datorii_pana_la_un_an_total", 0) or 0)
    lt_total = float(bs.get("datorii_peste_un_an_total", 0) or 0)
    bank_st = float(bs.get("bank_debt_st", 0) or 0)
    bank_lt = float(bs.get("bank_debt_lt", 0) or 0)
    trade_ap = float(bs.get("trade_payables_st", 0) or 0)
    effects_ap = float(bs.get("effects_payable_st", 0) or 0)

    # Bank debt — these are the only lines that should drive
    # `total_debt` in calculated_metrics. 5191 maps to stDebt, 1621
    # maps to ltDebt per `_ro_coa.py`.
    if bank_st:
        add("5191", "Sume datorate instituțiilor de credit — ST (F10)", bank_st)
    if bank_lt:
        add("1621", "Sume datorate instituțiilor de credit — LT (F10)", bank_lt)

    # Trade payables — route to 401 (AP bucket).
    if trade_ap:
        add("401", "Datorii comerciale - furnizori (F10)", trade_ap)
    if effects_ap:
        add("403", "Efecte de comerț de plătit (F10)", effects_ap)

    # Residual ST liabilities (social, fiscal, group, other) → other
    # current liabilities. Falls back to the full total if the
    # post-pass couldn't split (older templates that don't expose
    # the bank-debt line distinctly).
    if st_total:
        residual_st = st_total - bank_st - trade_ap - effects_ap
        if abs(residual_st) >= 1:
            # 462 maps to otherCurrentLiab per _ro_coa.
            add("462", "Alte datorii curente — residual ST (F10)", residual_st)

    # Residual LT liabilities → other non-current.
    if lt_total:
        residual_lt = lt_total - bank_lt
        if abs(residual_lt) >= 1:
            # 167 is "Alte împrumuturi și datorii asimilate" — other LT.
            add("167", "Alte datorii > 1 an — residual (F10)", residual_lt)

    venituri_avans = float(bs.get("venituri_in_avans", 0) or 0)
    if venituri_avans:
        add("472", "Venituri înregistrate în avans (F10 rd 77)", venituri_avans)

    provizioane = float(bs.get("provizioane_total", 0) or 0)
    if provizioane:
        add("151", "Provizioane — total (F10 rd 73)", provizioane)

    # Equity: emit components, NOT the total, so the mapper's
    # shareCapital/retainedEarnings buckets each get their own value.
    capital = float(bs.get("capital_subscris_varsat", 0) or 0)
    if capital:
        add("1012", "Capital subscris vărsat (F10 rd 81)", capital)

    rezerve = float(bs.get("rezerve_total", 0) or 0)
    if rezerve:
        add("106", "Rezerve (F10 rd 91)", rezerve)

    # Rezerve din reevaluare (ct.105) — non-cash revaluation reserve.
    # Routes via the mapper's `equity_revaluation` → otherEquity bucket.
    # Material on this test file (~359M), so omitting it under-states
    # equity by half.
    rezerve_reev = float(bs.get("rezerve_din_reevaluare", 0) or 0)
    if rezerve_reev:
        add("105", "Rezerve din reevaluare (F10)", rezerve_reev)

    # Retained earnings carry-forward (ct 117) — emit ONLY when
    # POSITIVE (entity has accumulated profits). When the carry-forward
    # is NEGATIVE (accumulated deficit), emitting it triggers the COA
    # assembler's defensive sign-flip on retainedEarnings (_ro_coa.py
    # line 599), which inverts an accurate negative deficit into a
    # spurious positive credit. The assembler then adds the computed
    # negative net_income_statutory on top, producing nonsense.
    #
    # When the carry-forward is negative we skip it here; the
    # assembler still gets the current-year net result via the P&L
    # path, so retainedEarnings = abs(net_income_loss) reflects at
    # least the current period's contribution. Total equity will not
    # match F10's CAPITALURI PROPRII TOTAL exactly for entities with
    # accumulated deficits — this is the same kind of pipeline
    # divergence the spec calls out under "ONE HONEST CAUTION".
    # Resolving it cleanly would require modifying _ro_coa.py to
    # special-case statutory-sourced negative equity (forbidden by
    # the task constraints).
    reportat = float(bs.get("profit_pierdere_reportat", 0) or 0)
    if reportat > 1:
        add("1171", "Rezultatul reportat (F10, accumulated profit)", reportat)

    return rows


def accounts_to_assemble_shape(extraction: StatutoryExtractionResult) -> List[Dict[str, Any]]:
    """Public synthesizer: turn a parsed statutory extraction into the same
    `{code, name, amount}` accounts list `_trial_balance_parser.accounts_
    to_assemble_shape` produces. Fed into `_ro_coa.assemble_statements`
    unchanged."""
    return _synth_accounts_from_extraction(extraction.pl_data, extraction.bs_data)
