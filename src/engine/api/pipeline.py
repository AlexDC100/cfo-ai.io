"""Phase 3 pipeline orchestrator — turn an uploaded document into a populated
financial period the UI can render.

Endpoints:
  POST /api/pipeline/run     { document_id }   → 202 (async kicked off)
  POST /api/pipeline/retry   { document_id }   → 202 (resets + reruns)
  GET  /api/period/:id                          → consolidated payload

Pipeline stages (each updates documents.status as it starts):
  queued → extracting → mapping → computing → narrating → analyzed | failed

Stages:
  detect   — filename + LLM classifier
  ocr      — Claude Opus 4.7 reads PDF directly (no separate OCR call)
  extract  — Same call returns structured RO accounts
  map      — _ro_coa.assemble_statements() rolls accounts into BS/PL buckets
  assemble — Statements blob + statement_line_items rows
  validate — BS-balance + sanity checks → alerts (data_quality)
  compute  — calculated_metrics rows (revenue, EBITDA, margin, leverage, ratios…)
  narrate  — Opus 4.7 → briefing + recommendations + alerts
  finalize — documents.status = 'analyzed'; documents.period_id set

CRITICAL: every stage is idempotent. retry() wipes prior derivatives before
re-running so the user can re-attempt without ghost data.
"""

from __future__ import annotations

import json
import logging
import os
import threading
import time
import traceback
import uuid
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional

import httpx
from fastapi import APIRouter, Depends, Header, HTTPException
from pydantic import BaseModel, Field

from . import _detect
from . import _org
from . import _supabase
from . import _usage_limits
from . import _valuation

# F3.1c — Country-pack dispatch. The pack is reached via the registry
# rather than direct module imports so country-specific behaviour
# (account-code mapping, parsing, industry classification) can be
# swapped at runtime in F3.3+. Importing the ro_romania subpackage as
# a side-effect registers the Romania pack with the registry.
import engine.country_packs.ro_romania  # noqa: F401  — side-effect: registers RomaniaPack
# F4.5 — Hungary pack (SKELETON / UNCALIBRATED). Registered so F4.4 fan-out
# routing can exercise multi-pack logic. detect_from_content() returns 0.0
# confidence so the RO pack always wins on RO uploads.
try:
    import engine.country_packs.hu_hungary  # noqa: F401
except Exception:  # noqa: BLE001
    # Skeleton may not be deployed in every environment; non-fatal.
    pass


# F4.6 — deprecated fields list (static, cheap to re-emit per request).
# Imported once at module load; returned in /api/period responses to give
# consumers migration targets before the 2Q sunset.
def _deprecated_fields_for_response() -> List[Dict[str, Any]]:
    try:
        from .deprecated_fields import deprecated_fields_list
        return deprecated_fields_list()
    except Exception:  # noqa: BLE001
        return []
from engine.core.country_pack_registry import get_pack as _get_country_pack


def _ro_pack():
    """Lazy resolver for the Romania country pack. Cached on first
    call; raises a clear error if the pack failed to register (which
    indicates an import-time failure inside the pack module, not a
    routine missing-country case)."""
    pack = _get_country_pack("RO")
    if pack is None:
        raise RuntimeError(
            "Romania country pack not registered. Check that "
            "engine.country_packs.ro_romania imported cleanly."
        )
    return pack


logger = logging.getLogger(__name__)


# ── Loud-failure state for the DIO-columns-missing fallback ──────────
# When sku_aggregates lacks days_inventory_on_hand / inventory_value_krn
# / cogs_krn, the insert retry path silently drops those values to keep
# the upload alive. Without this tracker the warning log scrolled past
# unnoticed for weeks. We capture: when it last fired, how many rows were
# affected, and the dataset_id, so /api/health can surface the degraded
# state and the FE can render a "schema migration pending" banner.
_DIO_DROP_STATE: Dict[str, Any] = {
    "degraded": False,
    "last_drop_at": None,
    "last_dataset_id": None,
    "total_rows_dropped": 0,
    "drop_event_count": 0,
    "first_error_sample": None,
}


def _record_dio_drop(*, dataset_id: str, rows_dropped: int, first_error: str) -> None:
    """Mark DIO persistence as degraded. Reset by the next clean upload."""
    _DIO_DROP_STATE["degraded"] = True
    _DIO_DROP_STATE["last_drop_at"] = _now_iso() if "_now_iso" in globals() else None
    _DIO_DROP_STATE["last_dataset_id"] = dataset_id
    _DIO_DROP_STATE["total_rows_dropped"] += rows_dropped
    _DIO_DROP_STATE["drop_event_count"] += 1
    if not _DIO_DROP_STATE.get("first_error_sample"):
        _DIO_DROP_STATE["first_error_sample"] = first_error


def get_dio_persistence_state() -> Dict[str, Any]:
    """Read-only snapshot for /api/health. Public — imported by _health.py."""
    return dict(_DIO_DROP_STATE)


def reset_dio_persistence_state() -> None:
    """Clear the degraded flag — called after a clean upload that
    succeeds without hitting the column-missing fallback. Operator can
    also clear via a one-shot admin endpoint if needed."""
    _DIO_DROP_STATE["degraded"] = False


class PublicRecordsUnparseableError(Exception):
    """Raised when stage_extract detects a listafirme.ro / termene.ro / firme.info
    public-records PDF (via `_public_records_parser.looks_like_public_records`)
    but the year-row extractor recovers zero rows from the document.

    Without this guard, the document falls through to the Claude TB extractor,
    which hallucinates trial-balance numbers from a 6-aggregate annual table
    (the original failure mode: revenue = EBITDA = net income, or 1,921 RON
    total assets on ELIT). Surfacing as an exception lets the outer pipeline
    handler at `_run_pipeline_sync` mark the document `status='failed'` with
    a user-readable `error` string, and STOPS the Claude TB path from ever
    seeing this PDF.
    """


# ─── Helpers ────────────────────────────────────────────────────────────────


def _require_jwt(authorization: Optional[str]) -> str:
    # PUBLIC_TEST_MODE bypass — see `_test_mode.py`. When the env flag is
    # on, every request is treated as authenticated as the shared test
    # user (id from TEST_USER_ID). Real customer orgs remain isolated
    # because the test user has membership only in TEST_ORG_ID.
    from . import _test_mode
    if _test_mode.is_test_mode():
        # Mint (and cache) a real Supabase access_token for the synthetic
        # test user — keeps RLS active (scoped to test org via the test
        # user's membership). See _test_mode.get_test_user_jwt().
        try:
            return _test_mode.get_test_user_jwt()
        except Exception:  # noqa: BLE001
            logger.exception("[test_mode] JWT mint failed; falling back to placeholder.")
            return _test_mode.JWT_BYPASS_PLACEHOLDER
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "Missing Bearer token.")
    return authorization.split(" ", 1)[1].strip()


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _detect_period_end_from_filename(filename: Optional[str]) -> str:
    """Extract the trial-balance date from a Romanian ERP export filename.

    Patterns we've actually seen in customer uploads:
      - 'Balanta Scandia Food_31.12.2025 LV.xls'  → '2025-12-31'
      - 'Balanta_EEI_dec_2025.pdf'                → '2025-12-31' (month-name)
      - 'scandia trial balance 2025.xlsx'         → '2025-12-31' (year only)
      - 'balanta_verificare_dec_2025.xlsx'        → '2025-12-31'
      - 'TB-2024-09-30.xlsx'                      → '2024-09-30'

    Falls back to today's date when no pattern matches — logs a warning so
    the misclassification is visible. Never raises.
    """
    import re
    if not filename:
        logger.warning("[period_end] no filename — defaulting to today")
        return date.today().isoformat()

    name = filename.strip()
    # Pattern: DD.MM.YYYY or DD_MM_YYYY or DD-MM-YYYY (Romanian convention)
    m = re.search(r"(\d{1,2})[._\-](\d{1,2})[._\-](\d{4})", name)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            pass
    # Pattern: YYYY-MM-DD or YYYY_MM_DD
    m = re.search(r"(\d{4})[._\-](\d{1,2})[._\-](\d{1,2})", name)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            pass
    # Pattern: month-name + year ("dec 2025", "decembrie 2025")
    months_ro = {
        "ian": 1, "feb": 2, "mar": 3, "apr": 4, "mai": 5, "iun": 6,
        "iul": 7, "aug": 8, "sep": 9, "oct": 10, "noi": 11, "dec": 12,
    }
    # Replace separators with spaces so the regex sees real word boundaries
    # ('balanta_EEI_dec_2025' → 'balanta eei dec 2025').
    name_lower = re.sub(r"[._\-]+", " ", name.lower())
    for tag, mo in months_ro.items():
        m = re.search(rf"\b{tag}\w*\b.*?(\d{{4}})", name_lower)
        if m:
            y = int(m.group(1))
            try:
                # Default to end of month — that's the trial-balance convention.
                import calendar
                last_day = calendar.monthrange(y, mo)[1]
                return date(y, mo, last_day).isoformat()
            except ValueError:
                pass
    # Year-only pattern — assume end of year.
    m = re.search(r"\b(20\d{2})\b", name)
    if m:
        return date(int(m.group(1)), 12, 31).isoformat()

    logger.warning("[period_end] no date pattern in filename %r — defaulting to today", name)
    return date.today().isoformat()


def _verify_user_owns_document(jwt: str, document_id: str) -> Dict[str, Any]:
    """Returns the document row when the JWT-bearing user has read access.
    Raises 403 otherwise.
    """
    with _supabase.per_user(jwt) as client:
        rows = client.select(
            "documents",
            filters={"id": f"eq.{document_id}"},
            single=True,
        )
        if not rows:
            raise HTTPException(404, f"Document {document_id} not found or not visible to you.")
        return rows[0]


def _user_id_from_jwt(jwt: str) -> str:
    """Resolve the calling user's id from the JWT (same pattern as _billing).
    Raises 401 if the JWT is malformed."""
    # PUBLIC_TEST_MODE bypass — return the shared test user_id.
    from . import _test_mode
    if _test_mode.is_bypass_token(jwt):
        return _test_mode.test_user_id()
    with _supabase.per_user(jwt) as client:
        user = client.get_user(jwt)
    user_id = user.get("id") if user else None
    if not user_id:
        raise HTTPException(401, "Could not resolve user from JWT.")
    return user_id


def _admin_set_status(doc_id: str, status: str, *, error: Optional[str] = None,
                      duration_ms: Optional[int] = None,
                      period_id: Optional[str] = None,
                      pipeline_started_at: Optional[str] = None) -> None:
    patch: Dict[str, Any] = {"status": status}
    if error is not None:
        patch["error"] = error
    elif status != "failed":
        patch["error"] = None
    if duration_ms is not None:
        patch["duration_ms"] = duration_ms
    if period_id is not None:
        patch["period_id"] = period_id
    if pipeline_started_at is not None:
        patch["pipeline_started_at"] = pipeline_started_at
    with _supabase.admin() as client:
        client.update("documents", patch, filters={"id": f"eq.{doc_id}"})


# ─── Pipeline stages ────────────────────────────────────────────────────────


# ─── Multi-format extraction ────────────────────────────────────────────────
# stage_extract dispatches on file type. The original PDF-only path lives at
# /api/financial-statements/parse and is reused for actual PDFs. For everything
# else (XLSX, CSV, JPG, PNG, plain text) we call Claude directly with the
# appropriate content block shape, using a broader extraction prompt that
# accepts trial balances, balance sheets, P&Ls, invoice registers, sales
# analyses, product catalogs, and bank statements.

_BROAD_SYSTEM_PROMPT = """You are a forensic accountant analyzing a financial business document.

Your job: extract structured financial data from whatever the user uploaded.
The document might be a Romanian trial balance ("balanță de verificare"), a
bilanț (balance sheet), a P&L, an invoice register (e-Factura / SAF-T /
SmartBill / generic CSV), a sales-by-product analysis, a bank statement, or
any other accounting / business document.

ADDITIONAL EXTRACTION FOR SKU/SALES DOCUMENTS:
If the document contains SKU-level or product-level rollups (sales by
product, trading analysis, invoice register with line items, inventory
report), populate the `skus` array with one row per distinct SKU.
Each SKU row: { sku, brand, category, channel, volume, volume_unit,
units_sold, revenue, cogs, gross_margin, gross_margin_pct,
inventory_value, days_inventory_on_hand }. All numeric fields default to
null when not in the source. Volume in tonnes when the data is in tonnes,
otherwise "units". For trading analyses with NIV (Net Invoiced Value),
treat NIV as `revenue`. For documents with only category-level rollups
(no individual SKUs), emit one row PER category as the sku ("CATEGORY:
PESTE") with category=name and brand=null — these synthetic rows still
let the engine classify portfolio segments.

CRITICAL RULES — read these before extracting:

1. Output STRICT JSON matching <schema>. No prose, no preamble, no markdown
   fences. The first character of your reply must be '{'.

2. Identify the document type and set `detected_type` accordingly. Acceptable
   values: trial_balance | bilant | pl | annual_report | invoice_register |
   sales_analysis | bank_statement | unknown.

3. If the document IS a trial balance / bilanț / P&L (financial statements):
   extract account rows into the `accounts` array using Romanian OMFP-1802
   account codes. Each account: closing balance (sold final). For accounts
   with separate debit/credit columns, follow standard signing — Class 1, 4
   (passive), 5 (passive), 7 take credit, emit positive; Class 2, 3, 4
   (active), 5 (active cash), 6 take debit, emit positive. If the codes
   aren't shown, map line items to canonical RO codes (5121 cash, 4111 AR,
   371 inventory, 212 PPE, 401 AP, 1621 LT debt, 1012 share capital, 117
   retained earnings, 704/706 revenue, 602 materials, 628 services, 641
   salaries, 681 D&A, 666 interest, 691 tax).

4. If the document IS NOT a financial statement (e.g. invoice register,
   sales analysis, product list): leave `accounts` as an empty array. Do
   NOT invent accounts. Populate `summary` with what the document IS and
   what it contains so downstream tabs can render it. Suggested fields:
       summary.row_count            — number of detail rows
       summary.headline_total       — RON total if obvious
       summary.headline_label       — what the total represents
       summary.top_records          — array of up to 10 string descriptors of
                                      the most material rows (e.g. customer
                                      names, product SKUs, invoice numbers)
       summary.warnings             — anything you couldn't resolve

5. Romanian numbers use ',' or '.' as decimal separator with '.' or space
   thousand grouping. Always emit clean decimal numbers.

6. Confidence rubric (emit a number 0..1):
   0.95 — clear balanță de verificare with all rows mapped
   0.85 — bilanț + P&L extracted to canonical RO codes
   0.70 — invoice register / sales analysis with structured rows captured
   0.60 — scanned/OCR'd image, some rows unclear
   0.40 — heavily inferred from narrative
   <0.40 — only headline figures

<schema>
{
  "company_name": string | null,
  "period_label": string,
  "period_end": string | null,         // ISO yyyy-mm-dd if discoverable
  "currency": string,                  // default "RON"
  "confidence": number,                // 0..1
  "detected_type": string,             // see rule 2
  "accounts": [{ "code": "5121", "name": "...", "amount": 1494836.00 }],
  "summary": {
    "row_count": number | null,
    "headline_total": number | null,
    "headline_label": string | null,
    "top_records": [string],
    "warnings": [string]
  },
  "skus": [
    {
      "sku": "string (full descriptor incl. weight/format)",
      "brand": "string | null",
      "category": "string | null",
      "channel": "string | null",
      "volume": "number | null",
      "volume_unit": "tons | units | kg | l | null",
      "units_sold": "number | null",
      "revenue": "number | null (NIV in source currency)",
      "cogs": "number | null",
      "gross_margin": "number | null (revenue - cogs)",
      "gross_margin_pct": "number | null (gm / revenue, decimal)",
      "inventory_value": "number | null",
      "days_inventory_on_hand": "number | null"
    }
  ],
  "warnings": [string]
}
</schema>

Begin."""


def _classify_file(doc: Dict[str, Any]) -> str:
    """Returns 'pdf' | 'xlsx' | 'csv' | 'image_jpeg' | 'image_png' | 'text' | 'unknown'."""
    mime = (doc.get("mime_type") or "").lower()
    name = (doc.get("original_filename") or "").lower()
    if mime == "application/pdf" or name.endswith(".pdf"):
        return "pdf"
    if "spreadsheet" in mime or name.endswith(".xlsx") or name.endswith(".xls"):
        return "xlsx"
    if mime == "text/csv" or name.endswith(".csv"):
        return "csv"
    if mime == "image/jpeg" or name.endswith((".jpg", ".jpeg")):
        return "image_jpeg"
    if mime == "image/png" or name.endswith(".png"):
        return "image_png"
    if mime.startswith("text/") or name.endswith(".txt"):
        return "text"
    return "unknown"


def _detect_spreadsheet_format(spreadsheet_bytes: bytes) -> str:
    """Identify the actual spreadsheet container by magic bytes — file
    extensions can lie (an .xls file may really be xlsx, or vice versa).

    Returns one of: 'xlsx' (zip-based OOXML), 'xls' (legacy OLE2/CFB
    binary), 'unknown'.
    """
    if len(spreadsheet_bytes) < 8:
        return "unknown"
    head = spreadsheet_bytes[:8]
    # XLSX / XLSB / ODS / DOCX etc. are all ZIP archives. We only handle
    # XLSX here; XLSB would also start with PK but openpyxl can't read it.
    if head[:4] == b"PK\x03\x04":
        return "xlsx"
    # XLS (Excel 97-2003) and other CFB/OLE2 documents.
    if head == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return "xls"
    return "unknown"


def _xls_to_text(xls_bytes: bytes, *, max_chars: int = 200_000) -> str:
    """Render a legacy .xls workbook (OLE2/CFB) as TSV. Uses pandas+xlrd
    since openpyxl only reads zip-based .xlsx files."""
    import io
    import pandas as pd  # type: ignore

    # engine='xlrd' is required for .xls; pandas would otherwise try
    # openpyxl and crash with BadZipFile.
    sheets = pd.read_excel(
        io.BytesIO(xls_bytes), sheet_name=None, header=None, engine="xlrd"
    )
    out: List[str] = []
    total_chars = 0
    for sheet_name, df in sheets.items():
        if total_chars > max_chars:
            out.append(f"\n[truncated — additional sheets omitted at {max_chars} chars]")
            break
        section = [f"=== Sheet: {sheet_name} ==="]
        for _, row in df.iterrows():
            cells = [
                "" if (pd.isna(v) or v is None) else str(v).replace("\t", " ")
                for v in row.values
            ]
            if not any(c.strip() for c in cells):
                continue
            line = "\t".join(cells)
            if total_chars + len(line) > max_chars:
                section.append("[truncated]")
                break
            section.append(line)
            total_chars += len(line) + 1
        out.append("\n".join(section))
    return "\n\n".join(out)


def _xlsx_to_text(spreadsheet_bytes: bytes, *, max_chars: int = 200_000) -> str:
    """Render a spreadsheet workbook as TSV text Claude can read.
    Each sheet becomes a labeled section: '=== Sheet: <name> ===' followed by
    rows joined with tabs. Truncated at max_chars to stay under context limits.

    Format-aware: detects xlsx vs legacy .xls by magic bytes and dispatches
    to the right reader. Trial balances from Romanian ERP exports often come
    as .xls (Excel 97-2003); openpyxl can't read those.

    Trial-balance fast-path: tries the deterministic structure detector
    first. When that succeeds, returns a canonical TSV (10 fixed columns,
    explicit labels) so Claude doesn't have to guess at paired-Debit/Credit
    columns in Crystal Reports / SAP exports. Falls through to the raw-
    sheet rendering when the file isn't a trial balance.
    """
    import io
    from openpyxl import load_workbook  # type: ignore

    fmt = _detect_spreadsheet_format(spreadsheet_bytes)
    if fmt == "unknown":
        raise RuntimeError(
            "Unrecognized spreadsheet format. Supported: .xlsx (Excel 2007+) "
            "and .xls (Excel 97-2003). If your file was downloaded from "
            "an ERP, try Save As → Excel Workbook (.xlsx) before uploading."
        )

    # Trial-balance fast-path. Best-effort: any failure here just falls
    # through to the legacy raw-render. Never raise from this branch.
    try:
        pack = _ro_pack()
        accounts = pack.parse_trial_balance(spreadsheet_bytes, "")
        if accounts:
            canonical = pack.accounts_to_canonical_tsv(accounts)
            preamble = (
                "Trial balance pre-parsed into canonical layout "
                f"({len(accounts)} accounts). Columns are explicit; no "
                "column-pairing inference needed.\n\n"
            )
            payload = preamble + canonical
            if len(payload) > max_chars:
                payload = payload[:max_chars] + "\n[truncated at canonical-tsv cap]"
            logger.info(
                "[xlsx_to_text] trial-balance fast-path: %d accounts, %d chars",
                len(accounts), len(payload),
            )
            return payload
    except Exception as e:  # noqa: BLE001
        logger.info(
            "[xlsx_to_text] trial-balance fast-path skipped (%s); using raw render",
            type(e).__name__,
        )

    if fmt == "xls":
        return _xls_to_text(spreadsheet_bytes, max_chars=max_chars)

    wb = load_workbook(io.BytesIO(spreadsheet_bytes), data_only=True, read_only=True)
    out: List[str] = []
    total_chars = 0
    for sheet_name in wb.sheetnames:
        if total_chars > max_chars:
            out.append(f"\n[truncated — additional sheets omitted at {max_chars} chars]")
            break
        ws = wb[sheet_name]
        section = [f"=== Sheet: {sheet_name} ==="]
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).replace("\t", " ") for v in row]
            if not any(c.strip() for c in cells):
                continue
            line = "\t".join(cells)
            if total_chars + len(line) > max_chars:
                section.append("[truncated]")
                break
            section.append(line)
            total_chars += len(line) + 1
        out.append("\n".join(section))
    return "\n\n".join(out)


def stage_extract(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Format-aware extraction. Sends the document to Claude Opus 4.7 in the
    most appropriate content-block shape for its type, then validates the
    JSON response and returns it.

    Supported inputs:
      - PDF      → document content block (handled by the legacy parser)
      - XLSX     → openpyxl-rendered TSV text content block
      - CSV      → raw text content block
      - JPG/PNG  → image content block
      - Plain text → text content block
    Anything else falls back to text after best-effort UTF-8 decoding.
    """
    storage_path: str = doc["storage_path"]
    kind = _classify_file(doc)

    # PDF still uses the existing parser (it has the canonical RO trial-balance
    # rubric in its system prompt — re-using avoids drift between two prompts).
    if kind == "pdf":
        # ── Public-records short-circuit ─────────────────────────────
        # Before handing off to the TB-rubric parser, check whether the
        # PDF is a listafirme.ro / termene.ro / firme.info / risco.ro
        # public-records summary. These PDFs carry a 6-aggregate × N-year
        # table that CANNOT be reconstructed as a trial balance — pushing
        # them through the TB pipeline produces nonsense (the Claude
        # extractor stuffs cifra de afaceri into every empty slot:
        # revenue = EBITDA = net income = 1.14B for PRO TV's case).
        # Detection runs locally on extracted text — zero Claude tokens
        # spent on the wrong format.
        try:
            with _supabase.admin() as admin_client:
                signed_for_text = admin_client.signed_url("documents", storage_path, expires_in=300)
            with httpx.Client(timeout=30.0) as http:
                _r = http.get(signed_for_text)
                _r.raise_for_status()
                _pdf_bytes = _r.content
            try:
                from pypdf import PdfReader  # type: ignore
            except ImportError:
                from PyPDF2 import PdfReader  # type: ignore
            import io
            _txt = ""
            _reader = PdfReader(io.BytesIO(_pdf_bytes))
            for _p in _reader.pages:
                _txt += "\n" + (_p.extract_text() or "")
            from . import _public_records_parser as _prp
            # Sentinel captured INSIDE the try, acted on OUTSIDE the broad
            # except. When `looks_like_public_records` is True but the
            # year-row extractor returns no rows (or low confidence), we
            # MUST NOT fall through to the Claude TB extractor — it
            # hallucinates trial-balance numbers from this format. The
            # raise lives outside this try so it isn't swallowed by the
            # blanket "detection skipped (non-fatal)" handler below.
            _pr_unparseable: Optional[Dict[str, Any]] = None
            if _prp.looks_like_public_records(_txt):
                # Pass the raw PDF bytes so the parser can use the
                # geometry-aware extractor (pdfplumber word-coordinates +
                # 10 px gap threshold). That path handles both dense
                # (PRO TV, 20/20) and sparse (ELIT, 17/17) layouts
                # uniformly. Falls back to text-only parsing if pdfplumber
                # isn't available or geometry fails.
                _extract = _prp.parse_public_records_pdf(_txt, pdf_bytes=_pdf_bytes)
                if _extract.confidence >= 0.5 and _extract.years:
                    logger.info(
                        "[stage_extract] public_records_summary detected: %s "
                        "CUI=%s CAEN=%s years=%d confidence=%.2f",
                        _extract.company_name, _extract.cui, _extract.caen_code,
                        len(_extract.years), _extract.confidence,
                    )
                    # Return a marker payload — the orchestrator routes this
                    # to a dedicated persistence path that skips TB stages.
                    return {
                        "detected_type": "public_records_summary",
                        "company_name": _extract.company_name,
                        "cui": _extract.cui,
                        "reg_com": _extract.reg_com,
                        "caen_code": _extract.caen_code,
                        "caen_description": _extract.caen_description,
                        "source_site": _extract.source_site,
                        "confidence": _extract.confidence,
                        "years": [
                            {
                                "year": r.year,
                                "cifra_afaceri": r.cifra_afaceri,
                                "profit_net": r.profit_net,
                                "datorii_totale": r.datorii_totale,
                                "active_imobilizate": r.active_imobilizate,
                                "active_circulante": r.active_circulante,
                                "capitaluri_proprii": r.capitaluri_proprii,
                                "total_assets": r.total_assets,
                                "salariati": r.salariati,
                                "net_margin_pct": r.net_margin_pct,
                            }
                            for r in _extract.years
                        ],
                        "raw_text": _txt[:5000],  # for downstream detection caching
                        # Empty placeholders so the orchestrator's downstream
                        # detection/persist stages don't crash on missing keys.
                        "accounts": [],
                        "warnings": [],
                    }
                # Header matched but year-row extractor came up empty (or
                # confidence < 0.5). Capture context for the post-try raise.
                _pr_unparseable = {
                    "confidence": float(_extract.confidence),
                    "years_found": len(_extract.years),
                    "company_name": _extract.company_name,
                    "cui": _extract.cui,
                    "source_site": _extract.source_site,
                }
                logger.warning(
                    "[stage_extract] public-records detected but unparseable: "
                    "name=%r cui=%r confidence=%.2f years=%d — STOPPING "
                    "before Claude TB fall-through.",
                    _extract.company_name, _extract.cui,
                    _extract.confidence, len(_extract.years),
                )
        except Exception:  # noqa: BLE001
            logger.exception("[stage_extract] public-records detection skipped (non-fatal)")
            _pr_unparseable = None  # detection itself crashed — let the TB path try

        # OUTSIDE the broad-except: if header detection succeeded but no
        # rows came out, fail loudly. The outer pipeline handler in
        # `_run_pipeline_sync` will catch this and mark documents.status
        # = 'failed' with the message below as documents.error, which
        # surfaces to the FE upload panel. Document stays orphan-free —
        # no period created, no statement_line_items written, no garbage.
        if _pr_unparseable is not None:
            raise PublicRecordsUnparseableError(
                "This looks like a public-records financial summary "
                "(listafirme.ro / termene.ro / firme.info), but we couldn't "
                "read the financial-year rows from this specific PDF layout. "
                "Try: (1) re-download the PDF directly from the source site "
                "using the browser's 'Print → Save as PDF' option, or "
                "(2) upload a trial balance (balanță de verificare) Excel "
                f"instead. [cui={_pr_unparseable.get('cui')}, "
                f"confidence={_pr_unparseable['confidence']:.2f}, "
                f"years_found={_pr_unparseable['years_found']}]"
            )

        # ── F3.8c — Deterministic PDF trial-balance fast-path ───────
        # Romanian RAS PDF trial balances (WinMENTOR / SAGA / Ciel /
        # generic-RAS) parse cleanly via the PyMuPDF position-based
        # ingester landed in F3.8a. When the parser yields a non-
        # empty account list AND captures the account-121 statutory
        # anchor, skip Claude entirely and feed the structured output
        # straight into the mapper — same fast-path the XLSX kind
        # has used since the F3.1 country-pack refactor.
        #
        # If the PDF doesn't parse cleanly (rare layouts, scanned
        # pages, non-RAS PDFs), fall through to the existing
        # Claude-based extractor in financial_statements.parse_document.
        try:
            pack = _ro_pack()
            tb_rows = pack.parse_trial_balance(
                _pdf_bytes, doc.get("original_filename") or "",
            )
            if tb_rows:
                shaped = pack.accounts_to_assemble_shape(tb_rows)
                statutory_anchor = pack.compute_statutory_net_profit_anchor(tb_rows)
                # F3.9 / F3.11 — source-data quality telemetry. Computed
                # from raw tb_rows (sf_d/sf_c) before any engine routing
                # so the operator can see source-data fault vs engine
                # fault on the WARN badge. Threaded through parsed →
                # stage_map → assemble_statements → API response.
                source_quality = pack.compute_source_imbalance(tb_rows)
                if shaped and (statutory_anchor or len(shaped) >= 50):
                    logger.info(
                        "[stage_extract] deterministic PDF TB path: "
                        "%d raw rows → %d mapped accounts "
                        "(ct 121 anchor = %s RON, source imbalance %.4f%% %s)",
                        len(tb_rows), len(shaped),
                        f"{statutory_anchor:,.0f}" if statutory_anchor else "n/a",
                        float(source_quality.get("raw_imbalance_pct") or 0),
                        "WARN" if source_quality.get("warn") else "ok",
                    )
                    return {
                        "company_name": (doc.get("original_filename") or "Imported entity").rsplit(".", 1)[0],
                        "period_label": "Imported period",
                        "period_end": _detect_period_end_from_filename(doc.get("original_filename")),
                        "currency": "RON",
                        "confidence": 0.95,
                        "detected_type": "trial_balance",
                        "accounts": shaped,
                        "warnings": [],
                        "statutory_net_profit_anchor": statutory_anchor,
                        "source_data_quality": source_quality,
                    }
        except Exception as e:  # noqa: BLE001
            logger.info(
                "[stage_extract] PDF TB fast-path skipped (%s) — "
                "falling back to Claude/statements parser",
                type(e).__name__,
            )

        with _supabase.admin() as admin_client:
            signed = admin_client.signed_url("documents", storage_path, expires_in=300)
        from .financial_statements import (  # type: ignore
            ParseRequest,
            build_router as _build_fs_router,
        )
        fs_router = _build_fs_router()
        parse_handler = None
        for route in fs_router.routes:
            if getattr(route, "name", None) == "parse_document":
                parse_handler = route.endpoint  # type: ignore[attr-defined]
                break
        if parse_handler is None:
            raise RuntimeError("financial_statements router missing parse_document route")
        parsed = parse_handler(ParseRequest(
            pdf_url=signed,
            original_filename=doc.get("original_filename"),
        ))
        return parsed.model_dump() if hasattr(parsed, "model_dump") else dict(parsed)

    # Everything else: download bytes, build a Claude message, parse the JSON.
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY not configured.")

    with _supabase.admin() as admin_client:
        signed = admin_client.signed_url("documents", storage_path, expires_in=300)
    with httpx.Client(timeout=30.0) as http:
        r = http.get(signed)
        r.raise_for_status()
        file_bytes = r.content

    if len(file_bytes) > 25 * 1024 * 1024:
        raise RuntimeError(f"File too large ({len(file_bytes)/1_000_000:.1f} MB) — 25 MB ceiling.")

    # Fix 9 — dual-pipeline routing. BEFORE the trial-balance fast-path
    # runs, detect whether this xlsx/xls is a statutory ANAF filing
    # (Formular F30 P&L + F10 Bilanț) and, if so, parse it via the
    # statutory parser instead. The trial-balance parser CANNOT read a
    # statutory return — it only sees account 722 (capitalized own work)
    # and silently drops Cifra de afaceri (the actual revenue line),
    # which produced a 20× revenue understatement on the food-manufacturer
    # case. Detection is regex + structural; on ambiguity we fall through
    # to the TB path so EEI / Scandia regression stays at zero. See
    # `_document_type_detector.py` for the anchor list.
    # `_classify_file` collapses both .xlsx and legacy .xls under "xlsx";
    # the document detector itself does magic-byte format detection.
    if kind == "xlsx":
        try:
            from . import _document_type_detector as _dtd
            doc_type, dt_meta = _dtd.detect_document_type(
                file_bytes, doc.get("original_filename") or "",
            )
            logger.info(
                "[stage_extract] document type detected: %s (%s)",
                doc_type, dt_meta.get("reason"),
            )
            if doc_type == "statutory_f30_f10":
                from . import _statutory_parser as _sp
                extraction = _sp.parse_statutory_file(
                    file_bytes, doc.get("original_filename") or "",
                )
                if not extraction.pl_data:
                    # Detection said statutory, but the row map produced
                    # nothing — most likely a layout variant we don't
                    # support yet. Fall through to TB / Claude rather
                    # than crashing so the user still gets *something*.
                    logger.warning(
                        "[stage_extract] statutory detection fired but no rows "
                        "extracted — falling through to TB path. Sheets=%s",
                        extraction.warnings,
                    )
                else:
                    shaped = _sp.accounts_to_assemble_shape(extraction)
                    period_end = extraction.period_end or _detect_period_end_from_filename(
                        doc.get("original_filename")
                    )
                    logger.info(
                        "[stage_extract] statutory F30/F10 path: %d P&L rows + %d BS rows "
                        "→ %d synth accounts (period_end=%s industry=%s)",
                        len(extraction.pl_data),
                        len(extraction.bs_data),
                        len(shaped),
                        period_end,
                        extraction.detected_industry,
                    )
                    return {
                        "company_name": (
                            extraction.company_name
                            or (doc.get("original_filename") or "Imported entity").rsplit(".", 1)[0]
                        ),
                        "period_label": (period_end or "Imported period"),
                        "period_end": period_end,
                        "currency": "RON",
                        "confidence": 0.88,  # high but below TB (less granular)
                        "detected_type": "statutory_f30_f10",
                        "accounts": shaped,
                        "warnings": extraction.warnings,
                        "statutory": {
                            "pl_data": extraction.pl_data,
                            "bs_data": extraction.bs_data,
                            "period_prior": extraction.period_prior,
                            "industry_hint": extraction.detected_industry,
                            "pl_sheet": extraction.pl_sheet_name,
                            "bs_sheet": extraction.bs_sheet_name,
                            "detection": dt_meta,
                        },
                    }
        except Exception as e:  # noqa: BLE001
            # Detection / statutory parsing should NEVER block the TB
            # path. Log and fall through silently.
            logger.info(
                "[stage_extract] statutory branch skipped (%s: %s) — falling back to TB path",
                type(e).__name__, str(e)[:120],
            )

    # Fix 8 — deterministic trial-balance fast-path. When the file parses
    # cleanly as a Romanian trial balance, skip Claude entirely and feed
    # the structured 809-account output straight into the mapper. Closes
    # the class-65 / sub-class-64 extraction gap (Claude was silently
    # dropping minor accounts and inflating downstream EBITDA).
    if kind == "xlsx":
        try:
            pack = _ro_pack()
            tb_rows = pack.parse_trial_balance(
                file_bytes, doc.get("original_filename") or "",
            )
            if tb_rows:
                shaped = pack.accounts_to_assemble_shape(tb_rows)
                # Account 121 closing balance = the statutory net profit
                # anchor (the legally filed figure on Romanian books).
                # Threaded through parsed → stage_compute so the platform
                # cites the SAME number the user sees on their own account
                # 121 instead of a reconstructed approximation that's off
                # by 1-6% on real data.
                statutory_anchor = pack.compute_statutory_net_profit_anchor(tb_rows)
                # F3.9 / F3.11 — source-data quality telemetry. See PDF
                # fast-path above for the rationale.
                source_quality = pack.compute_source_imbalance(tb_rows)
                logger.info(
                    "[stage_extract] deterministic TB path: %d raw rows → %d mapped accounts "
                    "(ct 121 anchor = %s RON, source imbalance %.4f%% %s)",
                    len(tb_rows), len(shaped),
                    f"{statutory_anchor:,.0f}" if statutory_anchor else "n/a",
                    float(source_quality.get("raw_imbalance_pct") or 0),
                    "WARN" if source_quality.get("warn") else "ok",
                )
                return {
                    "company_name": (doc.get("original_filename") or "Imported entity").rsplit(".", 1)[0],
                    "period_label": "Imported period",
                    "period_end": _detect_period_end_from_filename(doc.get("original_filename")),
                    "currency": "RON",
                    "confidence": 0.95,
                    "detected_type": "trial_balance",
                    "accounts": shaped,
                    "warnings": [],
                    "source_data_quality": source_quality,
                    "statutory_net_profit_anchor": statutory_anchor,
                }
        except Exception as e:  # noqa: BLE001
            logger.info(
                "[stage_extract] TB fast-path skipped (%s) — falling back to Claude",
                type(e).__name__,
            )

    try:
        from anthropic import Anthropic  # type: ignore
    except ImportError:
        raise RuntimeError("anthropic SDK not installed.")
    # max_retries=5 (vs SDK default 2) + extended timeout — covers the
    # sustained-Opus-overload case (HTTP 529) without surfacing it to
    # the user. Exponential backoff ~1s → 2s → 4s → 8s → 16s.
    client = Anthropic(api_key=api_key, max_retries=5, timeout=180.0)

    import base64 as _b64
    user_content: List[Dict[str, Any]]
    user_text = (
        f"Extract this document into the JSON schema. "
        f"Filename: {doc.get('original_filename') or 'unknown'}. "
        "Pick the most appropriate detected_type. Return JSON only."
    )

    if kind in ("image_jpeg", "image_png"):
        media_type = "image/jpeg" if kind == "image_jpeg" else "image/png"
        b64 = _b64.standard_b64encode(file_bytes).decode("ascii")
        user_content = [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
            {"type": "text", "text": user_text},
        ]
    else:
        # XLSX, CSV, plain text, unknown — coerce to text.
        if kind == "xlsx":
            text_payload = _xlsx_to_text(file_bytes)
            preamble = "Workbook rendered as TSV (sheets separated by '=== Sheet:' markers)."
        elif kind == "csv":
            text_payload = file_bytes.decode("utf-8", errors="replace")
            preamble = "CSV content."
        else:
            text_payload = file_bytes.decode("utf-8", errors="replace")
            preamble = "File content as text (best-effort decode)."

        # Cap the text payload so we don't blow the context window.
        if len(text_payload) > 250_000:
            text_payload = text_payload[:250_000] + "\n[truncated at 250k chars]"

        user_content = [
            {"type": "text", "text": f"{preamble}\n\n{text_payload}"},
            {"type": "text", "text": user_text},
        ]

    try:
        resp = client.messages.create(
            model="claude-opus-4-7",
            max_tokens=8000,
            system=[
                {"type": "text", "text": _BROAD_SYSTEM_PROMPT, "cache_control": {"type": "ephemeral"}},
            ],
            messages=[{"role": "user", "content": user_content}],
            output_config={"effort": "high"},
        )
    except Exception as e:  # noqa: BLE001
        # After max_retries on the Anthropic client are exhausted, a 529
        # (overloaded) becomes user-visible. Re-raise with a clear message
        # so the upstream orchestrator can show actionable copy instead of
        # the raw SDK error.
        status = getattr(e, "status_code", None)
        err_str = str(e)
        if status == 529 or "overloaded_error" in err_str or "529" in err_str:
            raise RuntimeError(
                "Claude is temporarily overloaded after multiple retries. Try again in 1-2 minutes — your document is fine."
            )
        if status == 429 or "rate_limit" in err_str:
            raise RuntimeError("Rate limit reached on the Claude API. Try again in a minute.")
        raise RuntimeError(f"Claude extraction failed: {e}")

    text = "".join(getattr(b, "text", "") for b in resp.content if getattr(b, "type", None) == "text").strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.startswith("json"):
            text = text[4:]
        text = text.strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"Claude returned invalid JSON. First 200 chars: {text[:200]!r}. Error: {e}")

    # Repair / defaults
    data.setdefault("company_name", None)
    data.setdefault("period_label", "Imported period")
    data.setdefault("period_end", None)
    data.setdefault("currency", "RON")
    data.setdefault("confidence", 0.5)
    data.setdefault("detected_type", "unknown")
    data.setdefault("accounts", [])
    data.setdefault("warnings", [])
    data.setdefault("summary", {})
    data.setdefault("skus", [])

    # Coerce account amounts
    cleaned: List[Dict[str, Any]] = []
    for raw in data.get("accounts", []):
        try:
            amt = float(raw.get("amount", 0) or 0)
            cleaned.append({
                "code": str(raw.get("code", "")).strip(),
                "name": str(raw.get("name", "")).strip(),
                "amount": amt,
            })
        except (TypeError, ValueError):
            data["warnings"].append(f"Dropped malformed account row: {raw}")
    data["accounts"] = cleaned

    # F3.16-3b.2 Path X (Path B variant) — synthesize the statutory
    # 121 anchor from Claude's cleaned accounts so the kwarg in
    # `stage_map`'s call to `assemble_statements` can fire the
    # net-income override even when the TB fast-path was bypassed.
    # Claude usually emits account 121 as a regular row; when it
    # does, sum every 121-prefixed code's amount. When Claude omits
    # 121 entirely (some periods), this stays None and the in-loop
    # capture inside `assemble_statements` is the only safety net
    # (and since Claude doesn't set bucket metadata, the in-loop
    # capture also yields None — fallback to reconstruction). This is
    # best-effort: it makes Claude path PARITY with TB fast-path when
    # Claude surfaces 121; otherwise the period gets reconstruction-
    # only behavior, which is what it had before this fix anyway.
    _claude_121_amounts = [
        float(a.get("amount") or 0)
        for a in cleaned
        if str(a.get("code") or "").strip().startswith("121")
    ]
    if _claude_121_amounts:
        _synth_anchor = sum(_claude_121_amounts)
        logger.info(
            "[stage_extract] Claude path: synthesized statutory_net_profit_anchor"
            "=%s RON from %d 121-prefixed rows (Path X / Path B)",
            f"{_synth_anchor:,.0f}", len(_claude_121_amounts),
        )
        data["statutory_net_profit_anchor"] = _synth_anchor
    else:
        data.setdefault("statutory_net_profit_anchor", None)

    return data


def stage_map(doc: Dict[str, Any], parsed: Dict[str, Any], industry: Optional[str]) -> Dict[str, Any]:
    """Roll accounts into BS/PL buckets via the country pack's mapper.
    Pre-F3.1c this called `_ro_coa.assemble_statements` directly; now
    dispatched through the Romania pack so the same call site will
    transparently work for other country packs once they're added.

    F3.11 — `source_data_quality` (when computed upstream in
    stage_extract and stored in `parsed`) is forwarded through so the
    engine envelope surfaces source-imbalance telemetry to the FE.
    Default None preserves byte-identical for paths that don't set it
    (Claude-extracted fallback, statutory F30/F10).
    """
    # F3.16-3b.2 Path X — thread the 121-anchor through. `stage_extract`
    # captures `statutory_net_profit_anchor` from raw TB rows BEFORE any
    # preprocessing that might drop the 121 line item (Path A: TB fast-
    # path via `compute_statutory_net_profit_anchor`; Path B: Claude
    # path via the post-parse 121-prefix sum). When neither path
    # surfaced an anchor, the kwarg is None and `assemble_statements`
    # falls back to its in-loop capture, preserving byte-identical
    # behavior on the fixture path and on any caller that doesn't ship
    # the anchor in `parsed`. See chart_of_accounts.py:1796-1804 for
    # the F3.16 forbidden-pattern marker explaining why this kwarg
    # exists (and why a comment-without-test was the original bug).
    return _ro_pack().assemble_statements(
        parsed.get("accounts") or [],
        company_name=parsed.get("company_name") or "Imported entity",
        currency=parsed.get("currency") or "RON",
        period_label=parsed.get("period_label") or "Imported period",
        industry=industry,
        source_data_quality=parsed.get("source_data_quality"),
        account_121_anchor_override=parsed.get("statutory_net_profit_anchor"),
    )


def stage_persist(doc: Dict[str, Any], parsed: Dict[str, Any], assembled: Dict[str, Any]) -> str:
    """Lookup-or-create the financial_period for this document's
    (org, period_end, source_document_id) tuple, then refresh its
    statement_line_items from the extracted statements. Returns the
    resolved period_id.

    Period-container discipline (post Bug-A fix — May 2026):
      · ONE period per (org_id, period_end, source_document_id) — enforced
        by the `financial_periods_org_period_doc_unique` constraint.
      · Same-document re-runs UPDATE the same period (the SELECT on the
        3-col tuple finds the existing row).
      · Two different documents on the same date — even for the same
        company — each get their OWN period row. This prevents the
        historical collision where a same-date upload from a different
        company would hijack and wipe the first company's line items.
      · `financial_periods.source_document_id` is the canonical identity
        key, immutable, FK-bound to documents(id).

    Pre-Bug-A history: the lookup-or-create previously filtered only by
    (org_id, period_end), which meant ANY second upload sharing a date
    found and reused the first period's row, then wiped its line items
    in step 4 below. Adding source_document_id to the SELECT closes that
    collision. (See Bug A fix: src/engine/api/pipeline.py:910-922.)
    """
    period_end_str = parsed.get("period_end")
    if period_end_str:
        try:
            period_end = date.fromisoformat(period_end_str).isoformat()
        except (TypeError, ValueError):
            period_end = _detect_period_end_from_filename(doc.get("original_filename"))
    else:
        # Claude didn't surface a period_end; fall back to filename detection
        # (Romanian ERP exports almost always include the trial-balance date
        # in the filename: "Balanta_..._31.12.2025_..." or "..._2025.xlsx").
        # Last resort: today's date with a warning so the period doesn't get
        # silently misclassified.
        period_end = _detect_period_end_from_filename(doc.get("original_filename"))

    period_start = period_end  # we don't have start info — treat as point-in-time

    with _supabase.admin() as admin_client:
        # 1. Lookup existing period for this (org, period_end, source_document_id).
        # Post-Bug-A: the DB enforces UNIQUE (org_id, period_end, source_document_id);
        # the 3-col SELECT here finds same-document re-runs (so we UPDATE
        # the same period row) but NOT different-document uploads sharing
        # the same date (each gets its own period row via the INSERT branch).
        existing = admin_client.select(
            "financial_periods",
            filters={
                "org_id": f"eq.{doc['org_id']}",
                "period_end": f"eq.{period_end}",
                "source_document_id": f"eq.{doc['id']}",
            },
            single=True,
        )
        if existing:
            period_id = existing[0]["id"]
            # Keep source_document_id pointing at the original. Update mutable
            # fields only — extraction_confidence reflects the latest analysis.
            admin_client.update(
                "financial_periods",
                {
                    "currency": parsed.get("currency") or existing[0].get("currency") or "RON",
                    "extraction_confidence": parsed.get("confidence", 0.5),
                    "updated_at": _now_iso(),
                },
                filters={"id": f"eq.{period_id}"},
            )
        else:
            # 2. Insert a fresh period row. The unique constraint on
            #    (org_id, period_end) prevents races — if another concurrent
            #    upload races to insert, this raises 23505 and we re-select.
            try:
                inserted = admin_client.insert(
                    "financial_periods",
                    {
                        "org_id": doc["org_id"],
                        "source_document_id": doc["id"],
                        "period_start": period_start,
                        "period_end": period_end,
                        "currency": parsed.get("currency") or "RON",
                        "extraction_confidence": parsed.get("confidence", 0.5),
                    },
                    returning=True,
                )
                period_id = inserted[0]["id"]
            except Exception:
                # Race-loser: re-select on the 3-col tuple — matches the
                # post-Bug-A unique constraint. Reuse the winner. Only ever
                # collides on a re-run of the same document; different
                # documents on the same date are not in conflict.
                rows = admin_client.select(
                    "financial_periods",
                    filters={
                        "org_id": f"eq.{doc['org_id']}",
                        "period_end": f"eq.{period_end}",
                        "source_document_id": f"eq.{doc['id']}",
                    },
                    single=True,
                )
                if not rows:
                    raise
                period_id = rows[0]["id"]

        # 3. Pin the document to the resolved period. Documents drive period
        #    ownership now — multiple docs per period.
        admin_client.update(
            "documents",
            {"period_id": period_id},
            filters={"id": f"eq.{doc['id']}"},
        )

        # 4. Wipe + re-insert statement line items for this period. The new
        #    document's extraction becomes the canonical analysis until
        #    another doc on this period is re-run.
        admin_client.delete("statement_line_items", filters={"period_id": f"eq.{period_id}"})
        line_items = assembled.get("lineItems") or []
        if line_items:
            # Whitelist columns that exist in statement_line_items. The
            # assembler emits an extra `canonical_bucket` field for the
            # sub-aggregate audit trail; PostgREST 400s on unknown columns,
            # so strip it here. When the DB migration adds a sub_bucket
            # column, this whitelist gets extended (or removed).
            _ALLOWED_COLS = {
                "period_id", "statement", "bucket",
                "ro_account_code", "ro_account_name",
                "amount", "is_derived",
            }
            rows = [
                {k: v for k, v in {"period_id": period_id, **item}.items() if k in _ALLOWED_COLS}
                for item in line_items
            ]
            for i in range(0, len(rows), 500):
                admin_client.insert("statement_line_items", rows[i:i+500], returning=False)

        # 5. F4.1e — persist the canonical envelope to the JSONB column on
        #    financial_periods. Best-effort: any failure here doesn't break
        #    the pipeline (assembled_bs/pl/cf and line items already landed).
        #    The read path also recomputes canonical on-the-fly via
        #    assemble_statements, so a NULL persisted value doesn't block
        #    consumers — DB persistence is for archive + offline analytics +
        #    audit (matches the comment on the column).
        #
        #    Note: only writes when the engine actually emitted a canonical
        #    envelope. Old engine builds without F4.1c return assembled
        #    without the key, in which case we leave the column NULL rather
        #    than UPDATE to NULL (which would overwrite any prior canonical
        #    captured by a newer engine — preserving the most-recent envelope).
        canonical = assembled.get("assembled_canonical_v1")
        if isinstance(canonical, dict):
            try:
                admin_client.update(
                    "financial_periods",
                    {"assembled_canonical_v1": canonical},
                    filters={"id": f"eq.{period_id}"},
                )
            except Exception:  # noqa: BLE001
                logger.exception(
                    "[stage_persist] failed to persist assembled_canonical_v1 "
                    "for period %s (non-fatal; read path still recomputes)",
                    period_id,
                )

    return period_id


# ── F1.h — composite-score → letter grade mapping (locked per SPEC §10). ──
# Single source of truth for both stage_compute (logging path) AND the
# get_period endpoint's `assembled_metrics.credit.letter_grade` emission.
# The FE's `compositeToGrade()` in `CreditScoreCard.tsx` MUST mirror these
# exact bands; mirror by the band-table comment block below, not by reading
# from a remote source (FE has no network access at component-render time).
#
#   Score        Grade
#   ≥ 90         AAA
#   80 – 89      AA      (NEW — the AA notch added by F1.h re-banding)
#   70 – 79      A
#   60 – 69      BBB
#   50 – 59      BB
#   40 – 49      B
#   25 – 39      CCC
#   < 25         CC
def _composite_to_letter_grade(composite: float) -> str:
    if composite >= 90: return "AAA"
    if composite >= 80: return "AA"
    if composite >= 70: return "A"
    if composite >= 60: return "BBB"
    if composite >= 50: return "BB"
    if composite >= 40: return "B"
    if composite >= 25: return "CCC"
    return "CC"


def stage_compute(doc: Dict[str, Any], assembled: Dict[str, Any], period_id: str) -> List[Dict[str, Any]]:
    """Compute headline metrics + ratios from the assembled statements.
    Persists to calculated_metrics (idempotent: wipe + re-insert).
    """
    s = assembled["statements"]
    bs = s["balanceSheet"]
    pl = s["incomeStatement"]

    revenue = pl["revenue"]
    cogs = pl["costOfGoodsSold"]
    opex = pl["operatingExpenses"]
    depreciation = pl["depreciationAmortization"]
    interest = pl["interestExpense"]
    other_inc = pl["otherIncome"]
    fin_inc = pl["financialIncome"]
    fin_exp = pl["financialExpense"]
    tax = pl["taxExpense"]
    # Inventory variation memo (RAS 711) — non-cash; EXCLUDED from cash EBITDA.
    # `other_inc` no longer contains it after the _ro_coa.py mapping fix; this
    # field is surfaced separately so the statutory-view metric can re-add it.
    inv_var_memo = pl.get("inventoryVariationMemo", 0.0)

    gross_profit = revenue - cogs
    operating_profit = gross_profit - opex - depreciation + other_inc
    ebitda = operating_profit + depreciation  # CASH view — primary
    ebitda_statutory_with_711 = ebitda + inv_var_memo  # IFRS / total-production view
    pretax = operating_profit + fin_inc - fin_exp - interest
    net_income = pretax - tax  # OPERATIONAL view — excludes account 722

    # ── Statutory net profit (anchors to account 121 closing balance) ──
    # Two valid views of "net profit" coexist in Romanian books:
    #   · net_income_operational — excludes 722 capitalized own-work. The
    #     "cash earnings" view a buyer or lender uses for EV / coverage.
    #   · net_income_statutory   — includes 722. This is what account 121
    #     closes to at year-end, what gets FILED with ANAF, and what a
    #     Romanian CFO sees on their own legal accounts.
    # Both are correct. We persist BOTH metrics so neither view is hidden;
    # the briefing + P&L tab default to the statutory figure (it's the
    # number a Romanian CFO recognizes from their filed accounts), and
    # operational stays available for valuation work.
    # For Scandia: 722 = ~2.21M → statutory 36.79M, operational 34.57M.
    # The values come from the canonical pl assembly (`_ro_coa.py`),
    # which already separates the two; pulling from there keeps the
    # arithmetic in ONE place rather than re-deriving here.
    capitalized_own_work = float(pl.get("capitalizedOwnWork", 0) or 0)
    net_income_statutory = net_income + capitalized_own_work
    # Companion statutory views — symmetric with net_income_statutory.
    # ebitda_statutory  = cash EBITDA + 722 (includes capitalized own-work);
    # total_operating_revenue = revenue + 722 + 711 + other_income (operating
    # view that the FE KPI tiles and benchmark engine consume against).
    # Surfacing as named metrics so regression checks can query by exact name.
    ebitda_statutory = ebitda + capitalized_own_work
    total_operating_revenue = revenue + capitalized_own_work + inv_var_memo + other_inc

    current_assets = bs["cash"] + bs["accountsReceivable"] + bs["inventory"] + bs["otherCurrentAssets"]
    non_current_assets = bs["propertyPlantEquipment"] + bs["intangibles"] + bs["otherNonCurrentAssets"]
    total_assets = current_assets + non_current_assets

    current_liab = bs["accountsPayable"] + bs["shortTermDebt"] + bs["otherCurrentLiabilities"]
    non_current_liab = bs["longTermDebt"] + bs["otherNonCurrentLiabilities"]
    total_debt = bs["shortTermDebt"] + bs["longTermDebt"]
    total_equity = bs["shareCapital"] + bs["retainedEarnings"] + bs["otherEquity"]

    def safe(num: float, denom: float) -> Optional[float]:
        return None if denom == 0 else round(num / denom, 4)

    metrics: List[Dict[str, Any]] = [
        {"name": "revenue",            "value": round(revenue, 2),         "unit": "RON",   "direction": "higher"},
        {"name": "gross_profit",       "value": round(gross_profit, 2),    "unit": "RON",   "direction": "higher"},
        {"name": "ebitda",             "value": round(ebitda, 2),          "unit": "RON",   "direction": "higher"},
        # Explicit cash + statutory views so consumers don't need to recompute.
        {"name": "ebitda_cash",        "value": round(ebitda, 2),          "unit": "RON",   "direction": "higher"},
        {"name": "ebitda_statutory_with_711", "value": round(ebitda_statutory_with_711, 2), "unit": "RON", "direction": "higher"},
        {"name": "inventory_variation_memo",  "value": round(inv_var_memo, 2),               "unit": "RON", "direction": "neutral"},
        {"name": "operating_profit",   "value": round(operating_profit, 2),"unit": "RON",   "direction": "higher"},
        # `net_income` (existing) is the OPERATIONAL view — kept under the
        # existing key for back-compat. Old FE / briefing consumers see the
        # same number they always did. NEW callers should prefer the
        # explicit `net_income_statutory` (matches account 121 / oracle) or
        # `net_income_operational` (alias of net_income).
        {"name": "net_income",         "value": round(net_income, 2),      "unit": "RON",   "direction": "higher"},
        {"name": "net_income_operational", "value": round(net_income, 2),  "unit": "RON",   "direction": "higher"},
        {"name": "net_income_statutory",   "value": round(net_income_statutory, 2), "unit": "RON", "direction": "higher"},
        {"name": "ebitda_statutory",       "value": round(ebitda_statutory, 2),     "unit": "RON", "direction": "higher"},
        {"name": "total_operating_revenue","value": round(total_operating_revenue, 2),"unit": "RON","direction": "higher"},
        {"name": "capitalized_own_work_memo", "value": round(capitalized_own_work, 2), "unit": "RON", "direction": "neutral"},
        {"name": "gross_margin",       "value": safe(gross_profit, revenue),"unit": "ratio","direction": "higher"},
        {"name": "ebitda_margin",      "value": safe(ebitda, revenue),     "unit": "ratio", "direction": "higher"},
        {"name": "net_margin",         "value": safe(net_income, revenue), "unit": "ratio", "direction": "higher"},
        {"name": "total_assets",       "value": round(total_assets, 2),    "unit": "RON",   "direction": "neutral"},
        {"name": "total_debt",         "value": round(total_debt, 2),      "unit": "RON",   "direction": "lower"},
        {"name": "total_equity",       "value": round(total_equity, 2),    "unit": "RON",   "direction": "higher"},
        {"name": "current_ratio",      "value": safe(current_assets, current_liab), "unit": "ratio", "direction": "higher"},
        {"name": "debt_to_equity",     "value": safe(total_debt, total_equity),     "unit": "ratio", "direction": "lower"},
        {"name": "debt_to_ebitda",     "value": safe(total_debt, ebitda),           "unit": "ratio", "direction": "lower"},
        {"name": "interest_coverage",  "value": safe(ebitda, interest),             "unit": "ratio", "direction": "higher"},
        {"name": "roa",                "value": safe(net_income, total_assets),     "unit": "ratio", "direction": "higher"},
        {"name": "roe",                "value": safe(net_income, total_equity),     "unit": "ratio", "direction": "higher"},
        {"name": "roic",               "value": safe(operating_profit * (1 - 0.16), max(total_debt + total_equity, 1)), "unit": "ratio", "direction": "higher"},
        {"name": "cash",               "value": round(bs["cash"], 2),              "unit": "RON",   "direction": "higher"},
        {"name": "free_cash_flow",     "value": round(net_income + depreciation, 2),"unit": "RON",  "direction": "higher"},
    ]

    # F3.11 — F3.9 source-data quality telemetry. Persisted as numeric
    # metrics so the FE can pick them up via remotePeriod.metrics without
    # a DB-schema migration (financial_periods has no metadata JSONB).
    # Two metrics — pct + abs — plus the warn flag derived FE-side from
    # pct > 2.0. Only emitted when the upstream stage_extract attached
    # source_data_quality (TB fast-paths; Claude-extracted path falls back
    # to None and skips these metrics entirely — `assembled.get(...) or {}`
    # makes the absence safe). See chart_of_accounts.assemble_statements
    # and pack.compute_source_imbalance for the upstream wiring.
    # F3.14 (3b) — Adjusted EBITDA + reconciliation components, read from
    # canonical assembled_pl (where the engine already computed them per
    # ADR_F3_14_DEFERRED_ITEMS.md). Operating EBITDA stays the headline;
    # adjusted_ebitda sits next to it on the dashboard with the two
    # subtracted lines (758 + 781) shown as the reconciliation bridge.
    pl_canonical_for_adj = s.get("assembled_pl") or {}
    if isinstance(pl_canonical_for_adj, dict):
        adj_ebitda_val = pl_canonical_for_adj.get("adjusted_ebitda")
        oi_758_val = pl_canonical_for_adj.get("other_income_758", 0) or 0
        oi_781_val = pl_canonical_for_adj.get("other_income_781_reversals", 0) or 0
        if adj_ebitda_val is None:
            # Fallback: compute from `ebitda` (cash view) − 758 − 781 if
            # canonical hasn't surfaced it yet (older cached fixtures).
            adj_ebitda_val = ebitda - float(oi_758_val) - float(oi_781_val)
        metrics.extend([
            {"name": "adjusted_ebitda",               "value": round(float(adj_ebitda_val), 2),
             "unit": "RON",   "direction": "higher"},
            {"name": "other_income_758",              "value": round(float(oi_758_val), 2),
             "unit": "RON",   "direction": "neutral"},
            {"name": "other_income_781_reversals",    "value": round(float(oi_781_val), 2),
             "unit": "RON",   "direction": "neutral"},
        ])

    sq = assembled.get("source_data_quality") or {}
    if sq:
        metrics.extend([
            {"name": "source_imbalance_pct", "value": round(float(sq.get("raw_imbalance_pct") or 0.0), 4),
             "unit": "ratio", "direction": "lower"},
            {"name": "source_imbalance_abs", "value": round(float(sq.get("raw_imbalance_abs") or 0.0), 2),
             "unit": "RON", "direction": "lower"},
            {"name": "source_closing_debit_sum",  "value": round(float(sq.get("sum_closing_debit") or 0.0), 2),
             "unit": "RON", "direction": "neutral"},
            {"name": "source_closing_credit_sum", "value": round(float(sq.get("sum_closing_credit") or 0.0), 2),
             "unit": "RON", "direction": "neutral"},
        ])

    # ── F1.d additions (SPEC §3) — every ratio the FE displays, emitted
    # once here so the Tier-1 FE recompute sites in
    # AUDIT_FE_CANONICAL_CONFORMANCE.md can become pure readers (F3).
    # Each row consumes values that are now on assembled_pl / assembled_bs
    # (F1.a / F1.b), or that are already in scope as locals from the
    # stage_compute body. No new math; every field is either a ratio of
    # existing canonical values or a direct alias.
    ar = bs.get("accountsReceivable", 0.0)
    inventory = bs.get("inventory", 0.0)
    ap = bs.get("accountsPayable", 0.0)
    st_debt = bs.get("shortTermDebt", 0.0)
    lt_debt = bs.get("longTermDebt", 0.0)
    # `total_operating_expense` per the B1 closure: full opex (cogs + opex
    # + D&A), NOT narrow COGS. Mirrors the same value emitted on
    # assembled_pl in F1.a.
    total_operating_expense = cogs + pl.get("operatingExpenses", 0.0) + depreciation
    # F3.15 Chunk 1 — `core_ebitda` is now READ from the canonical
    # `assembled_pl.core_ebitda` field, NOT recomputed locally. The
    # canonical definition (chart_of_accounts.py:1418) is `ebitda_statutory
    # − 758 − 781` (per-account from line_items). The prior local
    # redefinition `ebitda_statutory − pl["otherIncome"]` quietly diverged
    # whenever otherIncome contained items beyond 758/781 (e.g., F3.10
    # semantic-fallback-routed accounts named "venituri din subventii"
    # → otherIncome, or any 74/75/77/78 catchall hits from F3.8). The
    # two definitions agreed when otherIncome ≡ {758, 781}, drifted
    # otherwise. F3.15 fix: canonical chart_of_accounts is the single
    # source of truth; pipeline becomes a reader. Fall back to the
    # local arithmetic only when the canonical field is missing (pre-
    # F3.14 cached re-assemblies that bypass the F1.a code path).
    pl_canonical_for_core = s.get("assembled_pl") or {}
    if isinstance(pl_canonical_for_core, dict) and "core_ebitda" in pl_canonical_for_core:
        core_ebitda = float(pl_canonical_for_core.get("core_ebitda") or 0.0)
    else:
        # Fallback: narrow definition matching chart_of_accounts.py:1418
        # exactly (the two named-account sums), so even the fallback path
        # agrees with the canonical answer.
        oi_758_fb = float(pl_canonical_for_core.get("other_income_758", 0.0) if isinstance(pl_canonical_for_core, dict) else 0.0)
        oi_781_fb = float(pl_canonical_for_core.get("other_income_781_reversals", 0.0) if isinstance(pl_canonical_for_core, dict) else 0.0)
        core_ebitda = ebitda_statutory - oi_758_fb - oi_781_fb
    net_debt = total_debt - bs.get("cash", 0.0)

    metrics.extend([
        # Liquidity
        {"name": "quick_ratio",          "value": safe(bs.get("cash", 0.0) + ar, current_liab),       "unit": "ratio", "direction": "higher"},
        {"name": "cash_ratio",           "value": safe(bs.get("cash", 0.0), current_liab),            "unit": "ratio", "direction": "higher"},
        {"name": "working_capital",      "value": round(current_assets - current_liab, 2),            "unit": "RON",   "direction": "higher"},
        {"name": "net_debt",             "value": round(net_debt, 2),                                 "unit": "RON",   "direction": "lower"},
        {"name": "net_debt_to_ebitda",   "value": safe(net_debt, ebitda_statutory),                   "unit": "ratio", "direction": "lower"},
        # Leverage
        {"name": "equity_ratio",         "value": safe(total_equity, total_assets),                   "unit": "ratio", "direction": "higher"},
        {"name": "debt_to_assets",       "value": safe(total_debt, total_assets),                     "unit": "ratio", "direction": "lower"},
        {"name": "lt_debt_to_equity",    "value": safe(lt_debt, total_equity),                        "unit": "ratio", "direction": "lower"},
        # Coverage
        {"name": "ebitda_to_interest",   "value": safe(ebitda_statutory, interest),                   "unit": "ratio", "direction": "higher"},
        {"name": "dscr",                 "value": safe(ebitda_statutory, interest + st_debt),         "unit": "ratio", "direction": "higher"},
        # 8-year amortization proxy for LT principal — methodology cheat
        # sheet calls this the "DSCR with LT principal" view.
        {"name": "dscr_with_lt_principal","value": safe(ebitda_statutory, interest + lt_debt / 8.0),   "unit": "ratio", "direction": "higher"},
        # Efficiency
        {"name": "asset_turnover",       "value": safe(revenue, total_assets),                        "unit": "ratio", "direction": "higher"},
        {"name": "dso",                  "value": safe(ar * 365, revenue),                            "unit": "days",  "direction": "lower"},
        {"name": "dio",                  "value": safe(inventory * 365, total_operating_expense),     "unit": "days",  "direction": "lower"},
        {"name": "dpo",                  "value": safe(ap * 365, total_operating_expense),            "unit": "days",  "direction": "higher"},
        # CCC composes the three above. None-safe in the FE; here we
        # surface only when all three components are computable.
        {"name": "ccc",                  "value": (
            None if (current_liab == 0 or revenue == 0 or total_operating_expense == 0)
            else round(
                (ar * 365 / revenue)
                + (inventory * 365 / total_operating_expense)
                - (ap * 365 / total_operating_expense),
                4,
            )
        ), "unit": "days", "direction": "lower"},
        {"name": "inventory_turnover",   "value": safe(total_operating_expense, inventory),           "unit": "ratio", "direction": "higher"},
        # Margins — operating_margin (operational EBIT) + core_ebitda_margin.
        # `operating_margin` here uses `operating_profit` (= OPERATIONAL
        # EBIT, 722-excluded). The statutory operating margin would use
        # `ebit_statutory` if we surfaced it; not in this batch.
        {"name": "operating_margin",     "value": safe(operating_profit, revenue),                    "unit": "ratio", "direction": "higher"},
        {"name": "core_ebitda_margin",   "value": safe(core_ebitda, revenue),                         "unit": "ratio", "direction": "higher"},
        # Core EBITDA itself — surfaced as a RON figure so the FE can
        # render the dual-basis card without recomputing.
        {"name": "core_ebitda",          "value": round(core_ebitda, 2),                              "unit": "RON",   "direction": "higher"},
    ])

    # ── Altman Z″ score (emerging-markets variant) + composite credit ──
    # Ported from archive/calibration_toolkit/financial_analysis.py build_credit_score().
    # The Z″ formula uses BOOK retained earnings (not net income), so this
    # is robust against a single-year loss; the composite score blends Z″
    # with profitability, leverage, coverage, DSCR, liquidity, and equity
    # ratio into a single 0-100 number, mapped to a letter grade A-D.
    # Surfaced on the dashboard's CreditScoreCard so the user sees one
    # headline trust signal before drilling into individual ratios.
    altman_z = None
    composite = None
    letter_grade = None
    if total_assets > 0:
        x1 = (current_assets - current_liab) / total_assets
        x2 = bs["retainedEarnings"] / total_assets
        x3 = operating_profit / total_assets
        # X4 = book equity / total liabilities. max(1, ...) so we never
        # divide by zero on an asset-only entity (no debt).
        total_liab_safe = max(current_liab + non_current_liab, 1)
        x4 = total_equity / total_liab_safe
        altman_z = 6.56 * x1 + 3.26 * x2 + 6.72 * x3 + 1.05 * x4

        # Map Z″ to a 0-100 sub-score with the same three-zone reading the
        # methodology uses (>2.60 safe, 1.10-2.60 grey, <1.10 distress).
        if altman_z >= 2.60:
            altman_subscore = min(100, 70 + (altman_z - 2.60) * 15)
        elif altman_z >= 1.10:
            altman_subscore = 40 + (altman_z - 1.10) * 20
        else:
            altman_subscore = max(0, altman_z * 36)

        # Profitability sub-score: blend ROE + net margin. ROE weighted 0.5×
        # to keep margin-led growth companies from looking weak.
        roe_val = net_income / total_equity if total_equity > 0 else 0
        net_margin_val = net_income / revenue if revenue > 0 else 0
        prof_subscore = min(100, max(0, (roe_val * 100 * 0.5 + net_margin_val * 100 * 5) / 1.5))

        # Leverage sub-score (lower Net Debt/EBITDA = higher score). The
        # net_debt_ebitda formula uses CASH ebitda not statutory.
        net_debt = total_debt - bs["cash"]
        nde = net_debt / ebitda if ebitda > 0 else 999
        if nde <= 0:
            lev_subscore = 100
        elif nde <= 1.5:
            lev_subscore = 90
        elif nde <= 3.0:
            lev_subscore = 70
        elif nde <= 5.0:
            lev_subscore = 50
        else:
            lev_subscore = max(0, 50 - (nde - 5) * 10)

        # Interest coverage sub-score: EBIT / interest. 999 sentinel when
        # there's no interest (zero-debt company).
        ic = (operating_profit / interest) if interest > 0 else 999
        if ic >= 8:
            ic_subscore = 95
        elif ic >= 4:
            ic_subscore = 80
        elif ic >= 2:
            ic_subscore = 60
        elif ic >= 1:
            ic_subscore = 40
        else:
            ic_subscore = max(0, ic * 30)

        # DSCR — EBITDA / (interest + LT debt principal / 8 years).
        dscr = (ebitda / (interest + bs["longTermDebt"] / 8)) if interest > 0 else 999
        if dscr >= 2:
            dscr_subscore = 90
        elif dscr >= 1.25:
            dscr_subscore = 70
        else:
            dscr_subscore = max(0, dscr * 50)

        # Liquidity sub-score: blend of current / quick / cash ratios.
        cur_ratio = current_assets / current_liab if current_liab > 0 else 0
        quick_ratio = (current_assets - bs["inventory"]) / current_liab if current_liab > 0 else 0
        cash_ratio = bs["cash"] / current_liab if current_liab > 0 else 0
        liq_subscore = (
            min(100, cur_ratio * 50) + min(100, quick_ratio * 80) + min(100, cash_ratio * 250)
        ) / 3

        # Equity ratio sub-score.
        eq_subscore = min(100, (total_equity / total_assets) * 200) if total_assets > 0 else 0

        # Composite — same weights as the methodology.
        composite = (
            0.30 * altman_subscore
            + 0.20 * prof_subscore
            + 0.15 * lev_subscore
            + 0.10 * ic_subscore
            + 0.10 * dscr_subscore
            + 0.10 * liq_subscore
            + 0.05 * eq_subscore
        )

        letter_grade = _composite_to_letter_grade(composite)

        # Surface Altman + composite + sub-scores as calculated_metrics so the
        # CreditScoreCard renders without a separate DB query.
        metrics.extend([
            {"name": "altman_z_score",       "value": round(altman_z, 2),         "unit": "ratio", "direction": "higher"},
            {"name": "altman_x1",            "value": round(x1, 4),               "unit": "ratio", "direction": "higher"},
            {"name": "altman_x2",            "value": round(x2, 4),               "unit": "ratio", "direction": "higher"},
            {"name": "altman_x3",            "value": round(x3, 4),               "unit": "ratio", "direction": "higher"},
            {"name": "altman_x4",            "value": round(x4, 4),               "unit": "ratio", "direction": "higher"},
            {"name": "credit_composite",     "value": round(composite, 1),        "unit": "score", "direction": "higher"},
            {"name": "credit_subscore_altman",       "value": round(altman_subscore, 1),  "unit": "score", "direction": "higher"},
            {"name": "credit_subscore_profitability","value": round(prof_subscore, 1),    "unit": "score", "direction": "higher"},
            {"name": "credit_subscore_leverage",     "value": round(lev_subscore, 1),     "unit": "score", "direction": "higher"},
            {"name": "credit_subscore_coverage",     "value": round(ic_subscore, 1),      "unit": "score", "direction": "higher"},
            {"name": "credit_subscore_dscr",         "value": round(dscr_subscore, 1),    "unit": "score", "direction": "higher"},
            {"name": "credit_subscore_liquidity",    "value": round(liq_subscore, 1),     "unit": "score", "direction": "higher"},
            {"name": "credit_subscore_equity",       "value": round(eq_subscore, 1),      "unit": "score", "direction": "higher"},
        ])
        logger.info(
            "[pipeline] credit: Altman Z″=%.2f composite=%.0f → grade %s",
            altman_z, composite, letter_grade,
        )

    # Persist
    org_id = doc["org_id"]
    with _supabase.admin() as admin_client:
        admin_client.delete("calculated_metrics", filters={"period_id": f"eq.{period_id}"})
        rows = [
            {
                "period_id": period_id,
                "org_id": org_id,
                "name": m["name"],
                "value": m["value"],
                "unit": m["unit"],
                "direction": m["direction"],
            }
            for m in metrics
        ]
        admin_client.insert("calculated_metrics", rows, returning=False)
    return metrics


def stage_validate(doc: Dict[str, Any], assembled: Dict[str, Any], period_id: str) -> List[Dict[str, Any]]:
    """Generate deterministic alerts from the canonical `period_facts`-shaped
    views (`assembled_pl`, `assembled_bs`, `assembled_cf`). Every rule has a
    unique `alert_key`; duplicates are removed structurally before insert.

    LLM-generated alerts are NO LONGER persisted — narrative is reserved for
    the briefing + decision rationale. This eliminates the "15 duplicate
    critical alerts" problem (LLM emitting multiple variations of the same
    concern, each with a fresh fallback key on every rerun).

    Sign-bug guard: rules that interpret negative numbers (equity below half
    of capital, negative FCF, etc.) read values directly from the canonical
    views, which store credit-natural amounts with the SAME sign convention
    as the source books (positive equity = positive equity). No inversion.
    """
    s = assembled["statements"]
    bs = s["balanceSheet"]
    pl = s["incomeStatement"]
    pl_canonical = s.get("assembled_pl", {}) or {}
    bs_canonical = s.get("assembled_bs", {}) or {}
    cf_canonical = s.get("assembled_cf", {}) or {}
    sub_agg = s.get("subAggregates", {}) or {}
    industry_key = doc.get("industry_key") or "generic"

    # ── Canonical inputs ─────────────────────────────────────────────────
    # Pull the numbers we need ONCE up front, with safe fallbacks back to
    # the legacy assembled shape for the rare case where canonical views
    # aren't populated. Every rule below reads from these locals — no rule
    # re-derives a metric from the raw `bs` / `pl` blobs.
    ebitda_statutory = float(pl_canonical.get("ebitda_statutory") or 0)
    ebitda_operational = float(pl_canonical.get("ebitda_operational") or 0)
    capitalized = float(
        pl_canonical.get("capitalized_own_work_memo", pl.get("capitalizedOwnWork", 0)) or 0
    )
    rental_revenue = float(pl_canonical.get("revenue", pl.get("revenue", 0)) or 0)
    interest_expense = float(pl_canonical.get("interest_expense", pl.get("interestExpense", 0)) or 0)
    total_assets = float(bs_canonical.get("total_assets") or 0)
    total_liabilities = float(bs_canonical.get("total_liabilities") or 0)
    total_equity = float(bs_canonical.get("total_equity") or 0)
    share_capital = float(bs_canonical.get("share_capital", bs.get("shareCapital", 0)) or 0)
    revaluation_reserves = float(bs_canonical.get("revaluation_reserves") or 0)
    bank_debt_total = float(bs_canonical.get("total_debt") or 0)
    cash_val = float(bs_canonical.get("cash", bs.get("cash", 0)) or 0)
    ap_dividends = float(bs_canonical.get("ap_dividends") or sub_agg.get("ap_dividends", 0) or 0)
    intercompany = float(bs_canonical.get("intercompany_loans") or sub_agg.get("ar_intercompany", 0) or 0)
    cf_cfo = float(cf_canonical.get("cash_from_operating") or 0)
    cf_capex = float(cf_canonical.get("capex_real") or 0)
    cf_fcf = float(cf_canonical.get("free_cash_flow") or 0)
    cip_capex = float(cf_canonical.get("capitalized_construction") or 0)

    # Fallback computation for total_assets when canonical isn't populated
    # — keeps the BS-imbalance rule meaningful for legacy callers.
    if total_assets == 0 and bs:
        total_assets = (
            bs.get("cash", 0) + bs.get("accountsReceivable", 0) + bs.get("inventory", 0)
            + bs.get("otherCurrentAssets", 0) + bs.get("propertyPlantEquipment", 0)
            + bs.get("intangibles", 0) + bs.get("otherNonCurrentAssets", 0)
        )
    bs_delta = float(bs_canonical.get("bs_balance_delta") or 0)

    # Industry-aware Debt/EBITDA threshold table.
    _DTE_THRESHOLDS = {
        "real_estate_commercial":  (8.0, 12.0),
        "real_estate_residential": (7.0, 10.0),
        "real_estate":             (8.0, 12.0),
        "manufacturing":           (4.0, 6.0),
        "wholesale_distribution":  (3.5, 5.5),
        "fmcg":                    (3.5, 5.5),
        "saas":                    (3.0, 5.0),
        "b2b_saas":                (3.0, 5.0),
    }
    dte_high, dte_critical = _DTE_THRESHOLDS.get(industry_key, (4.0, 6.0))

    candidates: List[Dict[str, Any]] = []

    def _add(rule_key: str, severity: str, category: str, title: str, body: str,
             facts: Dict[str, float]) -> None:
        """Append a candidate alert. Same rule_key can only land once per
        period (deduped before persist)."""
        candidates.append({
            "alert_key": f"{rule_key}:{period_id}",
            "rule_key": rule_key,
            "severity": severity,
            "category": category,
            "title": title,
            "body": body,
            "facts_cited": facts,
            "industry": industry_key,
        })

    # ── R1. Data quality — BS imbalance ──────────────────────────────────
    drift = abs(bs_delta) if bs_delta else abs(total_assets - total_liabilities - total_equity)
    if total_assets > 0 and drift / max(total_assets, 1) > 0.01:
        sev = "critical" if drift > total_assets * 0.05 else "high"
        _add(
            "data_quality_bs_imbalance", sev, "data_quality",
            f"Balance sheet does not balance — drift RON {drift:,.0f}",
            f"Total assets RON {total_assets:,.0f} vs liabilities + equity RON "
            f"{total_liabilities + total_equity:,.0f} differ by RON {drift:,.0f} "
            f"({drift / max(total_assets, 1) * 100:.1f}% of assets). Pipeline "
            f"integrity issue — every downstream metric is suspect until resolved.",
            {"total_assets": total_assets, "total_liabilities": total_liabilities,
             "total_equity": total_equity, "drift": drift},
        )

    # ── R2. Data quality — empty P&L ─────────────────────────────────────
    if rental_revenue == 0 and total_assets > 1_000_000 and capitalized == 0:
        _add(
            "data_quality_pnl_zero", "critical", "data_quality",
            f"Zero revenue with RON {total_assets:,.0f} asset base — extraction gap",
            f"No revenue recorded for the period despite material assets. Likely a "
            f"P&L extraction issue (assembler reading closing balances for income "
            f"accounts instead of YTD movements). Investigate before relying on any "
            f"P&L metric.",
            {"rental_revenue": rental_revenue, "total_assets": total_assets},
        )

    # ── R3. Leverage — Debt/EBITDA above threshold ───────────────────────
    if ebitda_statutory > 0 and bank_debt_total > 0:
        dte = bank_debt_total / ebitda_statutory
        if dte > dte_critical:
            _add(
                "leverage_debt_to_ebitda_high", "critical", "leverage",
                f"Debt/EBITDA at {dte:.2f}× exceeds {dte_critical:.1f}× critical threshold for {industry_key}",
                f"Bank debt RON {bank_debt_total:,.0f} divided by statutory EBITDA "
                f"RON {ebitda_statutory:,.0f} = {dte:.2f}×, above the {dte_critical:.1f}× "
                f"critical threshold typical for this industry. Covenant breach risk.",
                {"debt_to_ebitda": dte, "bank_debt_total": bank_debt_total,
                 "ebitda_statutory": ebitda_statutory, "threshold": dte_critical},
            )
        elif dte > dte_high:
            _add(
                "leverage_debt_to_ebitda_high", "high", "leverage",
                f"Debt/EBITDA at {dte:.2f}× above {dte_high:.1f}× comfort zone for {industry_key}",
                f"Bank debt RON {bank_debt_total:,.0f} on statutory EBITDA "
                f"RON {ebitda_statutory:,.0f} = {dte:.2f}×, above typical comfort but "
                f"below covenant alarm for {industry_key}.",
                {"debt_to_ebitda": dte, "bank_debt_total": bank_debt_total,
                 "ebitda_statutory": ebitda_statutory, "threshold": dte_high},
            )

    # ── R4. Equity below half of share capital (Art. 153^24) ─────────────
    # SIGN-BUG GUARD: only fires when total_equity (as stored, positive
    # convention) is actually below half of share_capital. Previously the
    # platform inverted the sign and tripped this rule on a positive-equity
    # company.
    if share_capital > 0 and total_equity < share_capital / 2:
        sev = "critical" if total_equity < 0 else "high"
        title = (
            f"Negative book equity RON {total_equity:,.0f} — Romanian Company Law requires review"
            if total_equity < 0
            else f"Equity (RON {total_equity:,.0f}) below half of share capital (RON {share_capital:,.0f})"
        )
        _add(
            "equity_below_half_capital", sev, "compliance",
            title,
            f"Under Romanian Company Law Art. 153^24, when equity falls below half of "
            f"registered share capital the administrator must convene the general "
            f"meeting to decide on recapitalisation or dissolution.",
            {"total_equity": total_equity, "share_capital": share_capital,
             "ratio": total_equity / max(share_capital, 1)},
        )

    # ── R5. Capitalized own-work earnings-quality observation ────────────
    # Info-level only — the operating-view P&L already accounts for 722.
    # This card explains the 722/628 wash to the analyst.
    if capitalized > 100_000 and rental_revenue > 0:
        pct = capitalized / rental_revenue
        if pct > 0.5:
            _add(
                "earnings_quality_capitalized_own_work", "info", "data_quality",
                f"Capitalized own-work RON {capitalized:,.0f} = {pct*100:.0f}% of rental revenue",
                f"Account 722 (Producția imobilizări corporale) carries RON {capitalized:,.0f} of "
                f"capitalized own-work, mirrored by a roughly equal cost on 628 — net P&L "
                f"effect is approximately zero. Statutory EBITDA RON {ebitda_statutory:,.0f} "
                f"(with 722) vs operational view RON {ebitda_operational:,.0f} (without). "
                f"Bank covenants typically use the statutory view.",
                {"capitalized_own_work_memo": capitalized,
                 "ebitda_statutory": ebitda_statutory,
                 "ebitda_operational": ebitda_operational,
                 "pct_of_rental_revenue": pct},
            )

    # ── R6. Revaluation reserves — equity quality ────────────────────────
    if total_equity > 0 and abs(revaluation_reserves) > total_equity * 0.25:
        share_pct = abs(revaluation_reserves) / total_equity * 100
        _add(
            "equity_quality_revaluation_reserves", "info", "data_quality",
            f"Revaluation reserves are {share_pct:.0f}% of equity",
            f"Account 105 (Rezerve din reevaluare) of RON {abs(revaluation_reserves):,.0f} "
            f"represents {share_pct:.0f}% of total equity RON {total_equity:,.0f}. "
            f"This is a non-cash accounting reserve from upward revaluation of property — "
            f"equity quality is materially lower than the balance sheet suggests for lender / "
            f"buyer analysis.",
            {"revaluation_reserves": revaluation_reserves, "total_equity": total_equity,
             "pct_of_equity": share_pct / 100},
        )

    # ── R7. Concentration — intercompany receivable ──────────────────────
    if total_assets > 0 and intercompany > 100_000:
        pct = intercompany / total_assets
        if pct > 0.10:
            sev = "high" if pct > 0.20 else "medium"
            _add(
                "concentration_intercompany_loan", sev, "data_quality",
                f"Intercompany receivable RON {intercompany:,.0f} = {pct*100:.1f}% of total assets",
                f"Account 461 (Debitori diverși) holds RON {intercompany:,.0f} due from "
                f"related parties — {pct*100:.1f}% of total assets RON {total_assets:,.0f}. "
                f"Recoverability and intent on settlement should be confirmed. Lenders "
                f"typically haircut related-party receivables during covenant measurement.",
                {"intercompany_loans": intercompany, "total_assets": total_assets,
                 "pct_of_assets": pct},
            )

    # ── R8. Dividends declared but unpaid ────────────────────────────────
    if ap_dividends > 1000:
        _add(
            "cash_dividends_declared_unpaid", "medium", "liquidity",
            f"RON {ap_dividends:,.0f} dividends declared but not paid in cash",
            f"Account 457 (Dividende de plătit) carries RON {ap_dividends:,.0f} liability. "
            f"Dividends were debited to retained earnings but no cash distribution occurred. "
            + ("Operating cash flow is positive — could service this if distribution is planned."
               if cf_cfo > 0
               else "Operating cash flow is negative; distribution would strain liquidity."),
            {"dividends_payable": ap_dividends, "cash": cash_val, "cash_from_operating": cf_cfo},
        )

    # ── R9. FCF negative — development phase vs ongoing burn ─────────────
    if cf_fcf < 0 and cf_capex < 0:
        cip_dominant = abs(cip_capex) > abs(cf_capex) * 0.7
        if cip_dominant:
            _add(
                "fcf_negative_development_phase", "medium", "liquidity",
                f"Free cash flow RON {cf_fcf:,.0f} — one-time CIP capex",
                f"Operating cash flow RON {cf_cfo:,.0f} minus capex RON {abs(cf_capex):,.0f} "
                f"(RON {abs(cip_capex):,.0f} into account 231 Construction in Progress) "
                f"produces negative FCF this period. Development-phase drag, not ongoing burn — "
                f"stabilized FCF should be positive once CIP delivers.",
                {"cash_from_operating": cf_cfo, "capex_real": cf_capex,
                 "capitalized_construction": cip_capex, "free_cash_flow": cf_fcf},
            )
        else:
            _add(
                "fcf_negative_development_phase", "high", "liquidity",
                f"Free cash flow RON {cf_fcf:,.0f} — ongoing burn",
                f"Operating cash flow does not cover capex; cash buffer is being eroded.",
                {"cash_from_operating": cf_cfo, "capex_real": cf_capex,
                 "free_cash_flow": cf_fcf},
            )

    # ── R10. Valuation — EBITDA non-positive ─────────────────────────────
    # Single alert covers it (no longer 6 variations from the LLM).
    if ebitda_statutory <= 0:
        _add(
            "valuation_ebitda_negative", "high", "data_quality",
            f"Statutory EBITDA RON {ebitda_statutory:,.0f} — earnings-based valuation not applicable",
            f"With EBITDA at or below zero, EV/EBITDA multiples produce meaningless values. "
            f"The platform uses asset-based and revenue-multiple methods for valuation; see "
            f"the Valuation tab.",
            {"ebitda_statutory": ebitda_statutory},
        )

    # ── RISK INVENTORY — 5-8 named risks per analysis ───────────────────
    # Ported from archive/calibration_toolkit/financial_analysis.py build_risk_inventory().
    # These are the structural risks a CFO scans for when reading a deal
    # memo: receivables-allowance quality, liquidity tightness, raw-material
    # exposure, affiliate-income dependency, asset maturity, and leverage.
    # Each one is a separate alert with category='risk_inventory' so the
    # FE can group them into a distinct section (Section 7 in the
    # comprehensive report) without interleaving with data-quality alerts.
    revenue_local = float(pl_canonical.get("revenue", pl.get("revenue", 0)) or 0)
    net_income_local = float(pl_canonical.get("net_income_statutory") or pl_canonical.get("net_income_operational") or 0)
    trade_rec_local = float(bs_canonical.get("ar_net") or 0)
    rec_provisions_local = float(bs_canonical.get("ar_provisions") or 0)
    inventory_local = float(bs.get("inventory", 0) or 0)
    exp_601 = float(sub_agg.get("cogs_601") or 0)
    exp_602 = float(sub_agg.get("cogs_602") or 0)
    materials_pct = ((exp_601 + exp_602) / revenue_local) if revenue_local > 0 else 0
    cur_liab_local = float(
        bs.get("accountsPayable", 0) + bs.get("shortTermDebt", 0)
        + bs.get("otherCurrentLiabilities", 0)
    )
    cash_ratio_local = (cash_val / cur_liab_local) if cur_liab_local > 0 else 0
    ppe_gross_proxy = abs(float(bs.get("propertyPlantEquipment", 0) or 0))
    ppe_amort_proxy = max(0.0, float(sub_agg.get("ppe_amort") or 0))
    asset_maturity = (ppe_amort_proxy / ppe_gross_proxy) if ppe_gross_proxy > 0 else 0
    affiliate_income = float(sub_agg.get("financial_income") or 0)
    affiliate_dep = (affiliate_income / net_income_local) if net_income_local > 0 else 0
    net_debt_local = bank_debt_total - cash_val
    nde_local = (net_debt_local / ebitda_statutory) if ebitda_statutory > 0 else 0

    # R-RI-1 — Receivables allowance elevated. Historical credit issues
    # or affiliated-party stale balances. The Scandia case had 9.78M of
    # 491+496 provisions on 51M of gross trade rec = 19%, which fires.
    # Category is `working_capital` (a DB-allowed category); the
    # `rule_key` prefix `risk_inventory_*` is what the FE filters by to
    # show Section-7 risks. The DB CHECK constraint doesn't accept
    # `risk_inventory` so we route via rule_key instead.
    if trade_rec_local > 0 and rec_provisions_local > 0:
        prov_pct = rec_provisions_local / trade_rec_local
        if prov_pct > 0.15:
            _add(
                "risk_inventory_receivables_quality", "high", "working_capital",
                f"Receivables allowance {prov_pct*100:.0f}% of gross — historical credit issues",
                f"Provisions at {prov_pct*100:.0f}% of gross trade receivables suggest stale "
                f"balances (often affiliated parties or one-off customer defaults). Pull the "
                f"491/496 aging by counterparty; write off uncollectibles to clean the BS.",
                {"prov_pct": prov_pct, "trade_rec": trade_rec_local, "rec_provisions": rec_provisions_local},
            )

    # R-RI-2 — Tight cash liquidity (cash ratio <0.10).
    if cur_liab_local > 0 and cash_ratio_local < 0.10:
        _add(
            "risk_inventory_cash_tight", "high", "liquidity",
            f"Tight cash liquidity — cash ratio {cash_ratio_local:.2f}×",
            f"Cash covers only {cash_ratio_local*100:.1f}% of current liabilities — heavy "
            f"dependence on revolvers. A 15-day disruption could push the company past "
            f"covenants or payment terms.",
            {"cash_ratio": cash_ratio_local, "cash": cash_val, "cur_liab": cur_liab_local},
        )

    # R-RI-3 — Raw-material price exposure (>30% of revenue).
    if materials_pct > 0.30:
        _add(
            "risk_inventory_raw_materials", "medium", "margin",
            f"Raw material exposure — {materials_pct*100:.0f}% of turnover",
            f"Materials cost {materials_pct*100:.0f}% of revenue. A 10% commodity-price spike "
            f"compresses EBITDA margin by ~{materials_pct*10:.1f} pp. Consider hedging the "
            f"next 6-12 months of volume via forward contracts.",
            {"materials_pct": materials_pct},
        )

    # R-RI-4 — Affiliate income dependency (>15% of net profit).
    if net_income_local > 0 and affiliate_dep > 0.15:
        _add(
            "risk_inventory_affiliate_dep", "medium", "opportunity",
            f"Affiliate income dependency — {affiliate_dep*100:.0f}% of net profit",
            f"Affiliate dividends + interest produce {affiliate_dep*100:.0f}% of net profit. "
            f"Concentration risk if any single affiliate stops distributing. Entity-by-entity "
            f"yield review recommended.",
            {"affiliate_dep": affiliate_dep, "affiliate_income": affiliate_income, "net_income": net_income_local},
        )

    # R-RI-5 — Mature asset base (accumulated dep >55% of gross PP&E).
    if ppe_gross_proxy > 0 and asset_maturity > 0.55:
        _add(
            "risk_inventory_asset_maturity", "medium", "leverage",
            f"Mature asset base — accumulated depreciation {asset_maturity*100:.0f}% of gross PP&E",
            f"Equipment approaching end of useful life. Capex pressure ahead — plan a "
            f"3-5 year replacement program; consider EU grants for eligible modernizations.",
            {"asset_maturity": asset_maturity},
        )

    # R-RI-6 — Elevated leverage (Net Debt/EBITDA >4).
    if nde_local > 4 and ebitda_statutory > 0:
        _add(
            "risk_inventory_leverage", "high", "leverage",
            f"Elevated leverage — Net Debt/EBITDA {nde_local:.1f}×",
            f"Leverage at {nde_local:.1f}× EBITDA is above the typical 3× safety threshold. "
            f"Covenant pressure likely; refinancing risk if rates rise. Build a covenant "
            f"dashboard with the lender.",
            {"net_debt_ebitda": nde_local, "net_debt": net_debt_local, "ebitda": ebitda_statutory},
        )

    # R-RI-7 — FX exposure (FX cash >10% of total cash). Proxy via cash_fx sub_agg.
    fx_cash_local = float(sub_agg.get("cash_fx") or 0)
    if cash_val > 0 and fx_cash_local / cash_val > 0.10:
        _add(
            "risk_inventory_fx_exposure", "medium", "liquidity",
            f"FX exposure — {(fx_cash_local/cash_val)*100:.0f}% of cash in foreign currency",
            f"Significant FX cash position. Movements in EUR/RON or USD/RON create P&L "
            f"volatility. Consider an FX hedging policy or natural-hedge alignment with "
            f"foreign-currency liabilities.",
            {"fx_cash_pct": fx_cash_local/cash_val, "fx_cash": fx_cash_local, "total_cash": cash_val},
        )

    # ── Deduplicate by rule_key (structural guarantee) ──────────────────
    seen: set[str] = set()
    deduped: List[Dict[str, Any]] = []
    for alert in candidates:
        if alert["rule_key"] in seen:
            continue
        seen.add(alert["rule_key"])
        deduped.append(alert)

    # Sort critical → info so the FE renders the right order on insert.
    _severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    deduped.sort(key=lambda a: _severity_order.get(a["severity"], 99))
    return deduped


def _convert_currency(
    value: Any,
    source_currency: str,
    display_currency: str,
    fx_rates: Optional[Dict[str, float]],
) -> Any:
    """Convert a monetary value from source to display currency.

    `fx_rates` follows the BNR shape from `fx_rates.get_fx_rates()` —
    EUR-base, where rates["RON"]=4.97 means 1 EUR ≈ 4.97 RON. Conversion
    formula: `amount_dst = amount_src * (rates[dst] / rates[src])`.
    Non-numeric inputs pass through unchanged; mismatched/missing rates
    return the source value so the briefing never silently shows garbage.
    """
    if value is None or not isinstance(value, (int, float)):
        return value
    src = (source_currency or "RON").upper()
    dst = (display_currency or src).upper()
    if dst == src or not fx_rates:
        return value
    src_rate = fx_rates.get(src)
    dst_rate = fx_rates.get(dst)
    if not src_rate or not dst_rate:
        return value
    return value * (dst_rate / src_rate)


def _convert_briefing_facts(
    facts: Dict[str, Any],
    source_currency: str,
    display_currency: str,
    fx_rates: Optional[Dict[str, float]],
) -> Dict[str, Any]:
    """Walk briefing_facts and convert every monetary field. Ratios + pct
    fields (anything ending in `_pct`, `_margin`, `_ratio`, `_pct_*`, or
    inside the `ratios` sub-dict that isn't itself a money amount) stay
    untouched — they're dimensionless and don't change with FX."""
    if display_currency == source_currency or not fx_rates:
        return facts
    out: Dict[str, Any] = {}
    for k, v in facts.items():
        if k == "ratios" and isinstance(v, dict):
            # Sub-dict: only `net_debt` is currency-denominated; ratios are dimensionless.
            sub: Dict[str, Any] = {}
            for rk, rv in v.items():
                if rk == "net_debt":
                    sub[rk] = _convert_currency(rv, source_currency, display_currency, fx_rates)
                else:
                    sub[rk] = rv
            out[k] = sub
        elif isinstance(v, (int, float)):
            out[k] = _convert_currency(v, source_currency, display_currency, fx_rates)
        else:
            out[k] = v
    return out


def stage_council(doc: Dict[str, Any], parsed: Dict[str, Any],
                  assembled: Dict[str, Any]) -> Dict[str, Any]:
    """Advisory AI-council review of EXTRACTION INTEGRITY (added 2026-07-20).

    A panel of independent Claude personas — reconciliation / completeness /
    classification auditors — scans the freshly-assembled statements; a
    deterministic chair returns a consensus verdict (pass / warn / fail).

    Advisory by design: this stage NEVER blocks the pipeline and NEVER
    raises. Any failure (no API key, provider down, malformed output)
    degrades to the deterministic baseline inside `_ai_council.run_council`,
    which is itself exception-safe. The caller surfaces the verdict + any
    findings by appending `council_findings_as_alerts(...)` to
    `validation_alerts`, so they flow through the existing 'data_quality'
    alerts channel with no schema migration.

    Returns the full council result dict, or {} if the stage itself errors.
    """
    from . import _ai_council
    try:
        return _ai_council.run_council(assembled, parsed, document_id=doc.get("id"))
    except Exception:  # noqa: BLE001 — advisory stage must never break analysis
        logger.exception("[stage_council] failed (non-fatal)")
        return {}


def stage_narrate(doc: Dict[str, Any], assembled: Dict[str, Any], metrics: List[Dict[str, Any]],
                  org: Dict[str, Any], period_id: str,
                  parsed: Optional[Dict[str, Any]] = None,
                  valuation: Optional[Dict[str, Any]] = None,
                  display_currency: Optional[str] = None,
                  fx_rates: Optional[Dict[str, float]] = None) -> Dict[str, Any]:
    """Call Opus 4.7 with the metrics + industry context. Returns:
       { briefing: str, recommendations: [...], alerts: [...] }

    For non-financial documents (invoice register, sales analysis, product
    catalog) the prompt switches modes: instead of ratio-based CFO commentary,
    Claude describes what's in the document and surfaces the headline
    findings from `parsed.summary`.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return {"briefing": "Set ANTHROPIC_API_KEY on the backend to enable AI narrative.", "recommendations": [], "alerts": []}

    try:
        from anthropic import Anthropic  # type: ignore
    except ImportError:
        return {"briefing": "anthropic SDK not installed on backend.", "recommendations": [], "alerts": []}

    # max_retries=5 covers transient Opus 529 overloads on the narrate stage.
    client = Anthropic(api_key=api_key, max_retries=5, timeout=180.0)

    industry_key = org.get("industry_key") or "generic"
    industry_display = org.get("industry_display_name") or industry_key

    accounts_count = len((parsed or {}).get("accounts") or [])
    detected_type = (parsed or {}).get("detected_type") or "unknown"
    is_financial = accounts_count > 0 or detected_type in (
        "trial_balance", "statutory_f30_f10", "bilant", "pl", "annual_report",
    )

    # Output language — explicit on document row, falls back to English. The
    # /api/pipeline/run endpoint stores the user's UI language onto
    # documents.detected_language when they trigger a run; later, the real
    # auto-detection stage (Phase 4 Step 2) will overwrite this from the
    # document itself, so an English UI user uploading a German Saldenliste
    # still gets German narrative.
    output_language = (doc.get("detected_language") or "en").lower()[:2]
    language_instructions = {
        "en": "Reply in English.",
        "ro": "Răspunde în limba română.",
        "de": "Antworten Sie auf Deutsch.",
        "fr": "Répondez en français.",
        "es": "Responde en español.",
        "it": "Rispondi in italiano.",
        "pt": "Responda em português.",
        "nl": "Antwoord in het Nederlands.",
        "pl": "Odpowiedz po polsku.",
    }
    lang_instruction = language_instructions.get(output_language, language_instructions["en"])

    # ── Display currency — drives both the FX-converted numbers in
    # briefing_facts AND the currency suffix the LLM uses in prose. The
    # source currency comes from the statements; display defaults to it
    # unless the regenerate endpoint passes an explicit override from
    # the user's TopHeader currency toggle (RON/EUR/USD).
    source_currency = (assembled["statements"].get("currency") or "RON").upper()
    effective_display = (display_currency or source_currency).upper()
    # Currency formatting hint — uses the display currency code, with
    # locale-appropriate thousand/decimal separators per the output language.
    _example_amount = "1,234,567" if output_language in ("en",) else (
        "1 234 567" if output_language == "fr" else "1.234.567"
    )
    currency_hint = (
        f"Numbers in {output_language.upper()} locale; cite currency as "
        f"'{effective_display}' (e.g. '{_example_amount} {effective_display}'). "
        f"Every monetary figure in `briefing_facts` is pre-converted to "
        f"{effective_display} — do NOT re-convert."
    )

    if is_financial:
        system = (
            "You are a senior CFO advising the management team of a European SME.\n"
            "You receive standardized financial statements and computed ratios.\n"
            "You DO NOT compute numbers — explain them in industry context.\n\n"
            "═══════════════════════════════════════════════════════════════\n"
            "CANONICAL FACTS — SINGLE SOURCE OF TRUTH\n"
            "═══════════════════════════════════════════════════════════════\n"
            "The user payload contains a `briefing_facts` block. Every\n"
            "headline number you cite in the briefing — revenue, EBITDA,\n"
            "net profit, total debt, equity, leverage, key ratios — MUST\n"
            "come from `briefing_facts` verbatim. Do NOT derive your own\n"
            "EBITDA from `income_statement.revenue − operatingExpenses` —\n"
            "that path drops capitalized own-work (account 722) and gives\n"
            "the wrong sign. The frontend, KPI tiles, P&L tab, and balance\n"
            "sheet all read the operating-view numbers; the briefing must\n"
            "match them or the dashboard contradicts itself.\n\n"
            "Specifically, when citing P&L numbers:\n"
            " - Revenue → `briefing_facts.total_operating_revenue`\n"
            " - EBITDA → `briefing_facts.operating_ebitda` (NOT `metrics.ebitda`,\n"
            "   which is the older operational view that excludes 722)\n"
            " - Net profit → `briefing_facts.net_income_statutory` (NOT\n"
            "   `net_income_operational`)\n"
            " - Total debt → `briefing_facts.total_debt`\n"
            " - Equity → `briefing_facts.total_equity`\n"
            " - Cash → `briefing_facts.cash`\n\n"
            # ── F3.16-3b.6 EBITDA RULE — binding constraint on prose ─────
            # Closes the Carniprod −6.87M briefing-prose problem (the LLM
            # combined PL line items to produce an EBITDA value that
            # appeared nowhere in the engine's canonical fields). Per
            # docs/F3.16-3b6-f42-hardening-plan.md §4 — the engine is the
            # source of truth; LLM prose may reference canonical fields
            # by name but MUST NOT compute new values.
            "EBITDA RULE — EBITDA values referenced in the briefing MUST come\n"
            "from the canonical fields in the input dict. Specifically: use\n"
            "`briefing_facts.operating_ebitda` (equivalent to\n"
            "`methodology.ebitda.reported`) for headline statements;\n"
            "reference `methodology.ebitda.strict` / `cash` / `adjusted` by\n"
            "name when comparing methodologies. DO NOT compute new EBITDA\n"
            "values in prose. Do NOT sum or transform PL line items to\n"
            "produce a different EBITDA. If a methodology variant you want\n"
            "to reference is missing from the input, say so explicitly\n"
            "(\"cash-view EBITDA was not computed for this period\") rather\n"
            "than approximating one.\n\n"
            "If `briefing_facts.operating_ebitda > 0` you MUST NOT describe\n"
            "the company as posting an operating loss. The operational-view\n"
            "EBITDA (excluding 722) can be negative even when the\n"
            "operating-view EBITDA is positive — explain the 628↔722 wash\n"
            "if relevant, but lead with the operating-view headline.\n\n"
            "CRITICAL: Apply industry-appropriate thresholds.\n"
            " - Real estate: 4-8× Debt/EBITDA is normal; do NOT recommend deleveraging below 8×.\n"
            " - SaaS: focus on rule-of-40, ARR growth, gross margin >70%.\n"
            " - FMCG: working-capital efficiency, inventory turn, thin margins are normal.\n"
            " - Manufacturing: capex intensity, fixed-cost leverage are normal.\n\n"
            "═══════════════════════════════════════════════════════════════\n"
            "INDUSTRY-APPROPRIATE LANGUAGE — STRICTLY ENFORCED\n"
            "═══════════════════════════════════════════════════════════════\n"
            "Recommendation titles, rationales, and actions MUST use the\n"
            "vocabulary of the company's actual industry. Generic templates\n"
            "(\"exit unprofitable SKUs/customers\", \"renegotiate top suppliers\",\n"
            "\"inventory turnover\", \"product line review\") map ONLY to\n"
            "distribution / retail / manufacturing. Misapplying them to a\n"
            "CRE vehicle (one tenant, one property, one mortgage) makes the\n"
            "output unusable and damages user trust.\n\n"
            "For `industry_key = real_estate_*`:\n"
            "  • There are NO \"SKUs\", NO \"customers\" in the retail sense,\n"
            "    NO \"inventory\", NO \"product lines\".\n"
            "  • Use: tenants, leases, rental income, the property itself,\n"
            "    property operating costs (utilities, insurance, maintenance,\n"
            "    property taxes, property management).\n"
            "  • Cost-reduction language: \"renegotiate property management\n"
            "    contract\", \"review property insurance renewals\", \"challenge\n"
            "    property tax assessment\" — NEVER \"renegotiate top suppliers\".\n"
            "  • Tenant management: \"renew lease early\", \"add CPI indexation\",\n"
            "    \"identify backup tenants\" — NEVER \"exit unprofitable customers\".\n\n"
            "DISTRESS LANGUAGE — only allowed when the genuine signal is present:\n"
            "  • DO NOT recommend \"engage a restructuring advisor\" unless\n"
            "    Altman Z\" (industry-appropriate variant) < 1.10 AND DSCR < 1.0.\n"
            "  • DO NOT recommend \"covenant waiver\" unless DSCR < 1.0.\n"
            "  • DO NOT recommend \"13-week cash forecast\" unless cash <\n"
            "    3 months of debt service.\n"
            "  • The platform reads `briefing_facts.operating_ebitda` and\n"
            "    `briefing_facts.net_income_statutory` — both POSITIVE for a\n"
            "    healthy company. Don't claim \"negative EBITDA\" or \"operating\n"
            "    loss\" when those values are positive. Always cite the\n"
            "    statutory headline.\n\n"
            "VALUATION FRAMING — THIS IS NEW AND MANDATORY:\n"
            "When a `valuation` block is present in the user payload, treat the\n"
            "EBITDA-multiple equity (`valuation.equity_p50`, based on the peer P50\n"
            "multiple) as the company's headline equity value. Reference this\n"
            "single number in the briefing — do NOT invent another. The DCF and\n"
            "EV/Revenue numbers are cross-checks; mention divergence between them\n"
            "and the EBITDA multiple only when material (>30% spread).\n"
            " - You MUST NEVER produce a different valuation number than the one\n"
            "   in `valuation.equity_p50`. The engine is the source of truth.\n"
            " - If `valuation.confidence` is 'low' (negative EBITDA, thin margin,\n"
            "   or generic industry fallback), flag the confidence concern in the\n"
            "   briefing and add a recommendation to refine the inputs (capex,\n"
            "   working capital, industry classification) before the number is\n"
            "   used in a transaction context.\n"
            " - Always cite the source line — `valuation.multiples_source` and\n"
            "   `valuation.multiples_as_of_date` — so readers know which peer set.\n"
            " - Recommendations that talk about valuation must quantify in equity\n"
            "   value terms (e.g. 'a 10% EBITDA lift expands equity by X RON at\n"
            "   the current peer multiple').\n\n"
            f"LANGUAGE: {lang_instruction} {currency_hint}\n"
            "Translate industry terminology appropriately (e.g. Working capital → Betriebskapital / Fonds de roulement / Capital de lucru / Capital de trabajo).\n"
            "Briefing text, recommendation titles, rationales, actions — all in the output language.\n"
            "Be specific. Quantify recommendations in monetary terms when possible.\n"
            "Output STRICT JSON. No prose outside JSON. The first character of your reply must be '{'."
        )
    else:
        system = (
            "You are a senior CFO advisor reviewing a business document. The user uploaded\n"
            "a non-statement document (invoice register, sales analysis, product catalog,\n"
            "bank statement, etc.) — there is NO income statement or balance sheet to read.\n\n"
            "Your job:\n"
            "  1. Briefing (3 sentences): describe in concrete terms what the document IS,\n"
            "     what it covers, and what the most material rows / aggregates are. Use the\n"
            "     `summary` block in the input — row_count, headline_total, top_records.\n"
            "  2. Recommendations: 1–3 actionable next steps grounded in this specific\n"
            "     document. E.g. 'export accounts receivable aging from this register and\n"
            "     contact the top 3 overdue customers' or 'cross-check this sales analysis\n"
            "     against your trial balance to confirm revenue recognition'.\n"
            "  Do NOT generate alerts — the platform's deterministic rule registry handles\n"
            "  exception detection. Briefing + recommendations only.\n\n"
            "Do NOT fabricate revenue / EBITDA / ratios. Do NOT lecture about leverage.\n"
            f"LANGUAGE: {lang_instruction} {currency_hint}\n"
            "Output STRICT JSON. The first character of your reply must be '{'."
        )

    # ── Canonical briefing facts — single source of truth ─────────────────
    # The frontend's `usePeriodFacts()` hook, P&L tab, BS tab, KPI tiles,
    # and recommendation rules all read these exact fields from the
    # assembled canonical views. Surfacing them as `briefing_facts` (with
    # the same field names) lets the prompt above pin the briefing to the
    # same numbers and removes the prior drift between dashboard and
    # narrative.
    pl_canonical = assembled["statements"].get("assembled_pl", {}) or {}
    bs_canonical = assembled["statements"].get("assembled_bs", {}) or {}

    operating_ebitda = pl_canonical.get("operating_ebitda", 0.0)
    total_operating_revenue = pl_canonical.get("total_operating_revenue", 0.0)
    total_debt = bs_canonical.get("total_debt", 0.0)
    total_equity = bs_canonical.get("total_equity", 0.0)
    cash_val = bs_canonical.get("cash", 0.0)
    ebitda_for_ratios = operating_ebitda if operating_ebitda else 1e-9

    briefing_facts_raw = {
        # P&L — operating view (matches the frontend P&L tab + KPI tiles).
        "total_operating_revenue": total_operating_revenue,
        "operating_ebitda": operating_ebitda,
        "operating_ebit": pl_canonical.get("operating_ebit", 0.0),
        "depreciation": pl_canonical.get("depreciation", 0.0),
        "interest_expense": pl_canonical.get("interest_expense", 0.0),
        "tax": pl_canonical.get("tax", 0.0),
        "net_income_statutory": pl_canonical.get("net_income_statutory", 0.0),
        "net_income_operational": pl_canonical.get("net_income_operational", 0.0),
        "capitalized_own_work_memo": pl_canonical.get("capitalized_own_work_memo", 0.0),
        # BS — closing balances (Solduri finale year-end convention).
        "total_assets": bs_canonical.get("total_assets", 0.0),
        "total_equity": total_equity,
        "total_liabilities": bs_canonical.get("total_liabilities", 0.0),
        "total_debt": total_debt,
        "lt_debt": bs_canonical.get("lt_debt", 0.0),
        "st_debt": bs_canonical.get("st_debt", 0.0),
        "cash": cash_val,
        "ar_net": bs_canonical.get("ar_net", 0.0),
        "ap_trade": bs_canonical.get("ap_trade", 0.0),
        "ap_dividends": bs_canonical.get("ap_dividends", 0.0),
        "intercompany_loans": bs_canonical.get("intercompany_loans", 0.0),
        "ppe_net": bs_canonical.get("ppe_net", 0.0),
        "ppe_under_construction": bs_canonical.get("ppe_under_construction", 0.0),
        "current_year_pnl": bs_canonical.get("current_year_pnl", 0.0),
        "bs_balance_delta": bs_canonical.get("bs_balance_delta", 0.0),
        # Key derived ratios — operating-view based, so the briefing's
        # leverage / coverage commentary stays consistent with the tab.
        "ratios": {
            "ebitda_margin_pct": (
                round(100 * operating_ebitda / total_operating_revenue, 2)
                if total_operating_revenue else 0.0
            ),
            "net_margin_pct": (
                round(100 * pl_canonical.get("net_income_statutory", 0.0) / total_operating_revenue, 2)
                if total_operating_revenue else 0.0
            ),
            "debt_to_ebitda": round(total_debt / ebitda_for_ratios, 2),
            "debt_to_equity": round(total_debt / total_equity, 2) if total_equity else None,
            "net_debt": round(total_debt - cash_val, 2),
        },
    }
    # ── FX conversion ─────────────────────────────────────────────────
    # Convert every monetary value in briefing_facts from the source
    # currency (the trial balance's native currency, almost always RON)
    # to the user's display currency. Ratios + percentages stay as-is —
    # they're dimensionless. The LLM then narrates in the converted
    # currency, so toggling RON → EUR in the top bar produces a
    # regenerated briefing that says "€1,634,000 revenue" instead of
    # "8,121,590 RON revenue".
    briefing_facts = _convert_briefing_facts(
        briefing_facts_raw, source_currency, effective_display, fx_rates
    )

    user_payload = {
        "company": {
            "name": assembled["statements"].get("companyName"),
            "industry_key": industry_key,
            "industry_display_name": industry_display,
            # `currency` is the DISPLAY currency the briefing should cite
            # — already applied to every number in `briefing_facts`. The
            # original source currency is kept on `source_currency` for
            # auditability (e.g. "company filed in RON; report displays in EUR").
            "currency": effective_display,
            "source_currency": source_currency,
            "period_label": assembled["statements"].get("periodLabel"),
        },
        "document": {
            "filename": doc.get("original_filename"),
            "detected_type": detected_type,
            "accounts_extracted": accounts_count,
            "summary": (parsed or {}).get("summary") or {},
        },
        # CANONICAL — cite from here. Read the system prompt first.
        "briefing_facts": briefing_facts,
        "balance_sheet": assembled["statements"]["balanceSheet"],
        "income_statement": assembled["statements"]["incomeStatement"],
        "metrics": [
            {"name": m["name"], "value": m["value"], "unit": m["unit"], "direction": m["direction"]}
            for m in metrics
        ],
        # Server-computed valuation. Briefing must reference equity_p50 and never
        # invent a different headline number. See VALUATION FRAMING in system.
        "valuation": (
            {
                "primary_method": valuation["primary_method"],
                "confidence": valuation["confidence"],
                "industry_key_used": valuation.get("industry_key_used"),
                "multiples_source": valuation["multiples_source"],
                "multiples_as_of_date": valuation["multiples_as_of_date"],
                "ebitda_used": valuation["ebitda_used"],
                "total_debt_used": valuation["total_debt_used"],
                "cash_used": valuation["cash_used"],
                "multiple_p25": valuation["multiple_ebitda_p25"],
                "multiple_p50": valuation["multiple_ebitda_p50"],
                "multiple_p75": valuation["multiple_ebitda_p75"],
                "equity_p25": valuation["equity_ebitda_p25"],
                "equity_p50": valuation["equity_ebitda_p50"],
                "equity_p75": valuation["equity_ebitda_p75"],
                "ev_revenue_equity_p50": valuation["ev_revenue_equity_p50"],
                "dcf_equity_value": valuation["dcf_equity_value"],
            }
            if valuation
            else None
        ),
        "schema": {
            "briefing": "3 sentences. Industry-aware. RON-denominated. Reference specific metrics from the input.",
            "recommendations": [
                {
                    "severity": "critical|high|medium|low",
                    "category": "financial|operational|data_quality",
                    "title": "string (8 words max)",
                    "rationale": "string (1-2 sentences explaining why)",
                    "actions": ["string action 1", "string action 2"],
                    "estimated_ron_impact": "number or null",
                    "metric_referenced": "name of the metric this is grounded in",
                }
            ],
            # NOTE: alerts are NOT generated by the LLM. The platform's
            # deterministic rule registry (stage_validate) is the single
            # source for alerts. LLMs produce briefing + recommendation
            # narrative only.
        },
    }

    try:
        resp = client.messages.create(
            model="claude-opus-4-7",
            max_tokens=4096,
            system=system,
            messages=[{"role": "user", "content": json.dumps(user_payload)}],
            output_config={"effort": "high"},
        )
    except Exception as e:  # noqa: BLE001
        logger.warning("Opus narrate failed: %s", e)
        return {"briefing": f"Narrative unavailable: {e}", "recommendations": [], "alerts": []}

    text = "".join(getattr(b, "text", "") for b in resp.content if getattr(b, "type", None) == "text").strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.startswith("json"):
            text = text[4:]
        text = text.strip()
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return {"briefing": text[:500] or "Narrative unavailable.", "recommendations": [], "alerts": []}

    return {
        "briefing": data.get("briefing", "Narrative unavailable."),
        "recommendations": data.get("recommendations", []) or [],
        "alerts": data.get("alerts", []) or [],
    }


def stage_persist_narrative(
    doc: Dict[str, Any],
    period_id: str,
    narrate: Dict[str, Any],
    validation_alerts: List[Dict[str, Any]],
) -> None:
    org_id = doc["org_id"]
    document_id = doc["id"]
    with _supabase.admin() as admin_client:
        # Defensive: if a concurrent DELETE /api/period/{id} removed the
        # period between stage_persist and now, every FK-bound write below
        # will 409. Detect that here and skip narrative persistence — the
        # pipeline reports success because the analysis itself completed,
        # the user just lost the briefing/recommendations layer.
        period_still_exists = admin_client.select(
            "financial_periods",
            filters={"id": f"eq.{period_id}"},
            single=True,
            columns="id",
        )
        if not period_still_exists:
            logger.warning(
                "[stage_persist_narrative] period %s vanished mid-pipeline (race with DELETE); "
                "skipping briefing/recommendations write",
                period_id,
            )
            return

        # Briefing — upsert one row per period
        admin_client.upsert(
            "briefings",
            {
                "period_id": period_id,
                "org_id": org_id,
                "body": narrate["briefing"],
                "language": "en",
                "model": "claude-opus-4-7",
            },
            on_conflict="period_id",
            returning=False,
        )

        # Recommendations — wipe per-PERIOD, re-insert with period_id.
        #
        # Phase D-backend: the prior `delete WHERE org_id = ?` wiped EVERY
        # period's recommendations on every re-run, then re-inserted only
        # the current period's. Result: only the most-recently-analysed
        # period had recs; every other period's were silently wiped on
        # the next upload. The new scope is `WHERE period_id = ?` so a
        # re-run only replaces THIS period's recs; sibling periods are
        # untouched. The schema migration at
        # supabase/schema_phase_notes_period_scope.sql adds the period_id
        # column + index this query depends on.
        admin_client.delete("recommendations", filters={"period_id": f"eq.{period_id}"})
        recs = []
        for r in narrate["recommendations"]:
            actions = r.get("actions") or []
            explanation = r.get("rationale", "") + ("\n\nActions:\n• " + "\n• ".join(actions) if actions else "")
            urgency_map = {"critical": "critical", "high": "high", "medium": "medium", "low": "low"}
            severity = (r.get("severity") or "medium").lower()
            recs.append({
                "org_id": org_id,
                "period_id": period_id,
                "target_type": "dataset",
                "target_id": str(period_id),
                "title": r.get("title", "Untitled recommendation"),
                "explanation": explanation,
                "expected_cash_impact_kron": r.get("estimated_ron_impact"),
                "urgency": urgency_map.get(severity, "medium"),
                "status": "new",
            })
        if recs:
            admin_client.insert("recommendations", recs, returning=False)

        # ── Alerts — DETERMINISTIC rules only (read from canonical views) ──
        # LLM-generated alerts are NO LONGER persisted. The "15 duplicate
        # critical alerts" problem came from the LLM emitting many slight
        # variations of the same concern, each landing as a fresh row.
        # `stage_validate` is now the single source — each rule has a
        # unique rule_key, structurally deduped before this step.
        #
        # Phase D-backend: dedup scope is now (period_id, alert_key), not
        # (org_id, alert_key). The pre-D-backend setup had alerts as
        # org-scoped rows — `/api/period/:id` then returned the union of
        # every period's alerts on every period view, producing the
        # "80 on file" duplicate pile the user reported. Now:
        #
        #   1. DELETE every alert for THIS period_id before inserting
        #      (covers re-runs and old document_ids that previously
        #      attached to this period via documents.period_id).
        #   2. INSERT the deduped set with period_id stamped on each
        #      row, with a (period_id, alert_key) unique constraint
        #      enforcing idempotency at the DB level.
        #
        # The schema migration that adds `period_id` + the new unique
        # is at supabase/schema_phase_notes_period_scope.sql — both
        # must ship together.
        admin_client.delete("alerts", filters={"period_id": f"eq.{period_id}"})

        rows: List[Dict[str, Any]] = []
        seen_keys: set[str] = set()
        # validation_alerts already came from stage_validate deduped — but
        # double-check at the persist boundary in case a callsite added more.
        for a in validation_alerts:
            key = a.get("alert_key")
            if not key or key in seen_keys:
                continue
            seen_keys.add(key)
            severity = (a.get("severity") or "medium").lower()
            if severity not in ("critical", "high", "medium", "low", "info"):
                severity = "medium"
            category = (a.get("category") or "data_quality").lower()
            # `risk_inventory` is the Section-7 category — 5-8 named structural
            # risks the deterministic engine identifies (receivables quality,
            # liquidity tightness, raw-material exposure, etc.). Added to the
            # allowlist alongside the existing categories so the FE can filter
            # to it for the Comprehensive Report's risk inventory section.
            if category not in ("liquidity", "leverage", "margin", "inventory", "compliance",
                                 "data_quality", "working_capital", "customer", "supplier",
                                 "opportunity", "risk_inventory"):
                category = "data_quality"
            # Carry facts_cited + industry on the payload column for the
            # FE's "Facts backing this alert" expander.
            rows.append({
                "org_id": org_id,
                "period_id": period_id,
                "alert_key": key,
                "severity": severity,
                "category": category,
                "title": a.get("title", "Untitled alert"),
                "body": a.get("body", ""),
                "document_id": document_id,
                "payload": {
                    "rule_key": a.get("rule_key"),
                    "facts_cited": a.get("facts_cited"),
                    "industry": a.get("industry"),
                },
            })
        if rows:
            admin_client.upsert("alerts", rows, on_conflict="period_id,alert_key", returning=False)


# ─── Orchestrator ───────────────────────────────────────────────────────────


def _run_sales_dataset_pipeline(doc: Dict[str, Any]) -> Optional[str]:
    """Branch for trading-analysis / sales XLSX uploads. Reads the workbook
    via openpyxl, persists native-granularity rows into sku_lines, rolls up
    into sku_aggregates, classifies via the engine, and creates the
    sales_datasets row that ties it all together.

    Returns the new dataset_id on success, None when the file isn't sales-
    shaped (caller falls back to the generic LLM-summary path).
    """
    from . import _supabase
    from ._sales_extract import (
        aggregate_sku_lines,
        extract_category_dio,
        extract_category_dio_from_analysis_sheet,
        is_sales_dataset,
        stream_sales_rows,
    )
    from ._sku_classify import classify_portfolio

    # Mint a signed URL + download the bytes so openpyxl can read them.
    with _supabase.admin() as ac:
        signed = ac.signed_url("documents", doc["storage_path"], expires_in=300)
    r = httpx.get(signed, timeout=60.0)
    r.raise_for_status()
    xlsx_bytes = r.content

    detected, info = is_sales_dataset(xlsx_bytes)
    if not detected or not info:
        return None

    # Stream rows + materialize so we can both insert AND aggregate.
    lines = list(stream_sales_rows(xlsx_bytes, info))
    if not lines:
        return None

    period_label = info.get("period_label") or "Imported dataset"

    with _supabase.admin() as ac:
        # 1. sales_datasets row (upsert on document_id — re-running for the
        # same doc replaces the dataset).
        existing = ac.select("sales_datasets", filters={"document_id": f"eq.{doc['id']}"}, single=True)
        if existing:
            dataset_id = existing[0]["id"]
            ac.update("sales_datasets", {
                "label": period_label,
                "source_filename": doc["original_filename"],
                "row_count": len(lines),
            }, filters={"id": f"eq.{dataset_id}"})
            ac.delete("sku_lines", filters={"dataset_id": f"eq.{dataset_id}"})
            ac.delete("sku_aggregates", filters={"dataset_id": f"eq.{dataset_id}"})
        else:
            inserted = ac.insert("sales_datasets", {
                "org_id": doc["org_id"],
                "document_id": doc["id"],
                "label": period_label,
                "source_filename": doc["original_filename"],
                "row_count": len(lines),
                "is_active": True,
            }, returning=True)
            dataset_id = inserted[0]["id"]

        # 2. sku_lines — batched insert.
        #
        # CRITICAL: PostgREST's bulk-insert endpoint (`POST /sku_lines`
        # with an array body) requires every object in the array to have
        # the SAME key set — error PGRST102 "All object keys must match"
        # otherwise. Previously this code did `{k: v for k, v in line.items()
        # if v is not None}` which stripped Nones row-by-row, producing
        # heterogeneous shapes and triggering a PGRST102 on every batch.
        # The pipeline then fell back to per-row inserts (~400 sequential
        # HTTP round-trips) and stalled the upload for minutes.
        #
        # Fix: build a stable UNION key set across all rows, then fill
        # missing keys with explicit None on every row. PostgREST accepts
        # explicit nulls as long as the keys are uniform.
        # `inventory_value` and `cogs` are parser-only fields driving the
        # DIO calculation; they are aggregated server-side BEFORE the
        # per-line insert. Excluded here so the sku_lines table schema
        # (which doesn't carry these columns) keeps accepting bulk
        # inserts. The aggregator below still sees them on the in-memory
        # `lines` list.
        _LINE_INSERT_EXCLUDE = {"inventory_value", "cogs"}
        raw_rows = [
            {
                "dataset_id": dataset_id,
                "org_id": doc["org_id"],
                **{k: v for k, v in line.items() if k not in _LINE_INSERT_EXCLUDE},
            }
            for line in lines
        ]
        all_keys: set[str] = set()
        for row in raw_rows:
            all_keys.update(row.keys())
        line_rows = [
            {k: row.get(k) for k in all_keys}
            for row in raw_rows
        ]
        BATCH = 200
        inserted = 0
        for i in range(0, len(line_rows), BATCH):
            batch = line_rows[i:i+BATCH]
            try:
                ac.insert("sku_lines", batch, returning=False)
                inserted += len(batch)
            except Exception as e:  # noqa: BLE001
                logger.warning("[sales] batch %d failed, falling back per-row: %s", i // BATCH, e)
                for row in batch:
                    try:
                        ac.insert("sku_lines", row, returning=False)
                        inserted += 1
                    except Exception as e2:  # noqa: BLE001
                        logger.warning("[sales] dropped row product=%r reason=%s", row.get("product_name"), e2)
        logger.info("[sales] inserted %d/%d sku_lines into dataset %s", inserted, len(line_rows), dataset_id)

        # 3. Aggregate + classify.
        # Per-category DIO is read from TWO possible sources, preferring
        # the operator-set targets in the Analysis sheet over the
        # computed actuals in the standalone DIO sheet's Region 2:
        #
        #   1. Analysis sheet (cols 60-63 in Trading_analysis_*.xlsx) —
        #      operator-chosen realistic targets like LEGUME=60d,
        #      JELEURI=180d. Built by the analyst when reasoning about
        #      working capital, so they reflect business reality.
        #   2. DIO sheet Region 2 — naive (inventory ÷ COGS × 365)
        #      computations that can produce extreme outliers
        #      (JELEURI=1001.7d ≈ 3 years), useful only when the
        #      Analysis sheet doesn't carry a value for that category.
        #
        # The dicts are merged with Analysis winning per-key; DIO sheet
        # entries fill in any categories the Analysis sheet missed.
        # Both downstream paths (aggregate_sku_lines fallback chain,
        # FE category-card display) just see one merged dict.
        analysis_dio_map = extract_category_dio_from_analysis_sheet(xlsx_bytes)
        dio_sheet_map = extract_category_dio(xlsx_bytes)
        category_dio_map = {**dio_sheet_map, **analysis_dio_map}
        logger.info(
            "[sales] category DIO sourced: analysis_sheet=%d cats, dio_sheet=%d cats, merged=%d cats",
            len(analysis_dio_map), len(dio_sheet_map), len(category_dio_map),
        )
        agg_rows = aggregate_sku_lines(lines, category_dio=category_dio_map)
        # classify_portfolio expects 'sku' + 'revenue' keys; bridge to its API.
        for a in agg_rows:
            a["sku"] = a["product_name"]
            a["revenue"] = a["niv_krn"]
            a["cogs"] = max(0.0, (a["niv_krn"] or 0) - (a["gm_krn"] or 0))
        classified = classify_portfolio(agg_rows)

        agg_inserts = [{
            "dataset_id": dataset_id,
            "org_id": doc["org_id"],
            "product_name": c["product_name"],
            "brand": c.get("brand"),
            "category": c.get("category"),
            "volume_tons": c.get("volume_tons"),
            "niv_krn": c.get("niv_krn"),
            "gm_krn": c.get("gm_krn"),
            "gm_pct": c.get("gm_pct"),
            "real_margin_krn": c.get("real_margin"),
            "real_margin_pct": c.get("real_margin_pct"),
            # DIO inputs + output (optional). Files without inventory/
            # COGS columns yield None across all three → DIO null → FE
            # renders "not available" honestly. The `sku_aggregates`
            # table's `inventory_value_krn` / `cogs_krn` /
            # `days_inventory_on_hand` columns are nullable; this
            # insertion is therefore backward-compatible with the
            # previous Products upload format. `cogs_krn` is included
            # so the WC roll-up panel can compute company DIO as
            # sum(inv) / sum(cogs) * 365 (the spec's preferred formula)
            # rather than a weighted-average fallback.
            "inventory_value_krn": c.get("inventory_value_krn"),
            "cogs_krn": c.get("cogs_krn"),
            "days_inventory_on_hand": c.get("days_inventory_on_hand"),
            "classification": c.get("classification") or "keep",
            "classification_reason": c.get("classification_reason"),
            "line_row_count": c.get("line_row_count"),
            "channels_present": c.get("channels_present") or [],
            "clients_present": c.get("clients_present") or [],
        } for c in classified]
        any_dio_drops_this_upload = False
        for i in range(0, len(agg_inserts), 400):
            try:
                ac.insert("sku_aggregates", agg_inserts[i:i+400], returning=False)
            except Exception as e:  # noqa: BLE001
                # Defensive: if the deployed schema is older and lacks
                # the inventory_value_krn / days_inventory_on_hand
                # columns, retry the chunk WITHOUT those fields rather
                # than failing the whole upload. This guarantees
                # backward-compatibility on environments that haven't
                # yet had the (nullable) DIO columns added.
                msg = str(e).lower()
                if (
                    "inventory_value_krn" in msg
                    or "cogs_krn" in msg
                    or "days_inventory_on_hand" in msg
                ):
                    fallback = [
                        {k: v for k, v in row.items()
                         if k not in (
                             "inventory_value_krn",
                             "cogs_krn",
                             "days_inventory_on_hand",
                         )}
                        for row in agg_inserts[i:i+400]
                    ]
                    ac.insert("sku_aggregates", fallback, returning=False)
                    # 2026-05-26 — loud-failure plumbing. The defensive
                    # retry preserved the upload but silently dropped
                    # every DIO value, and the warning scrolled past in
                    # the log noise for weeks. Now we ALSO mark the
                    # service as degraded so /api/health surfaces it
                    # and the FE shows a banner. Operator runs the SQL
                    # migration → flag clears on the next clean upload.
                    any_dio_drops_this_upload = True
                    _record_dio_drop(
                        dataset_id=dataset_id,
                        rows_dropped=len(fallback),
                        first_error=str(e)[:300],
                    )
                    logger.error(
                        "[sales] DIO PERSISTENCE DEGRADED — sku_aggregates "
                        "is missing days_inventory_on_hand / inventory_value_krn / "
                        "cogs_krn columns. Inserted %d rows WITHOUT DIO. "
                        "Run supabase/schema_phase_sku_dio_columns.sql to fix. "
                        "Dataset=%s. PostgREST error: %s",
                        len(fallback), dataset_id, str(e)[:300],
                    )
                else:
                    raise

        ac.update("sales_datasets", {"sku_count": len(agg_inserts)},
                  filters={"id": f"eq.{dataset_id}"})

        # Clean upload — clears the degraded flag so /api/health flips
        # back to ok and the FE banner disappears on next poll. Operator
        # has run the schema migration; the silent-drop path is gone.
        if not any_dio_drops_this_upload:
            reset_dio_persistence_state()

    return dataset_id


def _persist_sku_analysis(doc: Dict[str, Any], parsed: Dict[str, Any], narrative: Dict[str, Any]) -> None:
    """Persist a SKU analysis + classified per-SKU rollups.

    SKU/inventory uploads (XLSX trading analysis, sales-by-product exports,
    invoice registers) are intentionally separate from the financial-statement
    data model — they should never appear on Dashboard / Cash / Profit.

    Two writes per upload:
      1. sku_analyses (one row per document) — briefing + recommendations
      2. sku_aggregates (N rows) — engine-classified per-SKU rollups
    """
    from ._sku_classify import classify_portfolio

    raw_skus = parsed.get("skus") or []
    classified = classify_portfolio(raw_skus)
    period_label = parsed.get("period_label") or "Imported period"

    with _supabase.admin() as client:
        client.upsert(
            "sku_analyses",
            {
                "org_id": doc["org_id"],
                "document_id": doc["id"],
                "briefing": narrative.get("briefing", ""),
                "summary": parsed.get("summary") or {},
                "recommendations": narrative.get("recommendations") or [],
                "language": "en",
                "model": "claude-opus-4-7",
            },
            on_conflict="document_id",
            returning=False,
        )

        # Legacy: the OLD sku_aggregates table was keyed on document_id +
        # produced from the LLM's synthetic category roll-ups. After the
        # sales-dataset refactor the table is keyed on dataset_id and is
        # populated by _run_sales_dataset_pipeline directly. _persist_sku_
        # analysis no longer manages sku_aggregates — it only writes the
        # briefing/summary into sku_analyses. The deletes below are a no-op
        # in the new schema (we skip them).
        if False and classified:  # legacy path retained as dead code for clarity
            rows = [
                {
                    "org_id": doc["org_id"],
                    "document_id": doc["id"],
                    "period_label": period_label,
                    "sku": s.get("sku"),
                    "brand": s.get("brand"),
                    "category": s.get("category"),
                    "channel": s.get("channel"),
                    "volume": s.get("volume"),
                    "volume_unit": s.get("volume_unit") or "tons",
                    "units_sold": s.get("units_sold"),
                    "revenue": s.get("revenue") or 0,
                    "cogs": s.get("cogs") or 0,
                    "gross_margin": s.get("gross_margin"),
                    "gross_margin_pct": s.get("gross_margin_pct"),
                    "real_margin": s.get("real_margin"),
                    "real_margin_pct": s.get("real_margin_pct"),
                    "inventory_value": s.get("inventory_value") or 0,
                    "days_inventory_on_hand": s.get("days_inventory_on_hand"),
                    "capital_tied_up": s.get("capital_tied_up"),
                    "classification": s.get("classification") or "keep",
                    "classification_reason": s.get("classification_reason"),
                    "classification_confidence": s.get("classification_confidence"),
                }
                for s in classified
            ]
            for i in range(0, len(rows), 500):
                client.insert("sku_aggregates", rows[i:i+500], returning=False)


def _commit_pipeline_quota(document_id: str, *, success: bool) -> None:
    """Pricing V3 (refined-spec gap D) — convert the upload-time
    reservation into either a consumed slot (success) or a release
    (failure).

    Runs on the orchestrator daemon thread with no HTTP context, so
    we recover the user_id + `metered_extra` flag from the documents
    row directly. The reservation was made at /api/pipeline/run via
    `_usage_gate.reserve_document` (or `confirm_extra_document` for
    extras); committing/releasing here is the terminal half of that
    transaction.

    Best-effort: any failure here is LOGGED but never re-raised.
    Quota correctness is downstream of user-visible analysis state,
    not the other way around.
    """
    from . import _usage_gate as _ug
    if not _ug.enforcement_enabled():
        return
    try:
        with _supabase.admin() as ac:
            rows = ac.select(
                "documents",
                filters={"id": f"eq.{document_id}"},
                columns="id,uploaded_by,metered_extra",
                single=True,
            )
        if not rows:
            return
        row = rows[0]
        user_id = row.get("uploaded_by")
        if not user_id:
            return
        was_extra = bool(row.get("metered_extra"))
        if success:
            _ug.commit_document(user_id, was_extra=was_extra)
            # WS2 — when the doc succeeded AND was flagged as a paid extra,
            # record one usage unit against the user's Stripe metered item.
            # Idempotency key = document_id so any retry of this commit
            # path (e.g. orchestrator restart between commit + return) is
            # a no-op on the Stripe side. Best-effort: errors are logged
            # but don't propagate — the local quota was already committed
            # and the operator can reconcile from billing logs.
            if was_extra:
                try:
                    from . import _billing
                    result = _billing.record_metered_extra_doc(
                        user_id=user_id,
                        reservation_id=document_id,
                    )
                    if not result.get("ok"):
                        logger.error(
                            "[pipeline] record_metered_extra_doc returned non-ok for "
                            "doc=%s user=%s — manual reconciliation may be needed. "
                            "Result: %s",
                            document_id, user_id, result,
                        )
                    elif not result.get("billed"):
                        logger.info(
                            "[pipeline] metered extra-doc not billed (doc=%s reason=%s)",
                            document_id, result.get("reason"),
                        )
                except Exception:  # noqa: BLE001
                    logger.exception(
                        "[pipeline] record_metered_extra_doc raised — local "
                        "extras_billed_period already bumped, Stripe side missed "
                        "this charge. doc=%s user=%s",
                        document_id, user_id,
                    )
        else:
            _ug.release_document(user_id, was_extra=was_extra)
    except Exception:
        logger.exception(
            "[pipeline] _commit_pipeline_quota(%s, success=%s) failed",
            document_id, success,
        )


def _run_pipeline_sync(document_id: str) -> None:
    t0 = time.time()
    try:
        with _supabase.admin() as admin_client:
            doc_rows = admin_client.select("documents", filters={"id": f"eq.{document_id}"}, single=True)
            if not doc_rows:
                logger.warning("[pipeline] document %s vanished mid-run", document_id)
                return
            doc = doc_rows[0]

            org_rows = admin_client.select("organizations", filters={"id": f"eq.{doc['org_id']}"}, single=True)
            org = org_rows[0] if org_rows else {"id": doc["org_id"], "name": "Unknown", "industry_key": None, "industry_display_name": None}

        scope = (doc.get("scope") or "financial").lower()

        _admin_set_status(document_id, "extracting", pipeline_started_at=_now_iso())
        parsed = stage_extract(doc)

        # ── Public-records short-circuit ────────────────────────────────
        # `stage_extract` returns `detected_type='public_records_summary'`
        # when it recognized the PDF as a listafirme.ro / termene.ro
        # multi-year aggregate table. These are NOT trial balances — they
        # have 6 numbers per year, not 800+ accounts. Pushing them through
        # the TB pipeline silently produces nonsense (the PRO TV regression
        # case: revenue = EBITDA = net income = 1.14B). Persist the parsed
        # years to the sku_analyses table (re-using the existing JSONB
        # column to avoid a migration) and mark the doc analyzed cleanly.
        # The `documents.detected_type` column has a CHECK constraint so we
        # tag the doc via `briefing.kind` instead — the FE filters by that.
        if (parsed or {}).get("detected_type") == "public_records_summary":
            persisted = False
            try:
                with _supabase.admin() as admin_client:
                    admin_client.upsert(
                        "sku_analyses",
                        {
                            "org_id": doc["org_id"],
                            "document_id": document_id,
                            "briefing": {
                                "kind": "public_records_summary",
                                "company_name": parsed.get("company_name"),
                                "cui": parsed.get("cui"),
                                "reg_com": parsed.get("reg_com"),
                                "caen_code": parsed.get("caen_code"),
                                "caen_description": parsed.get("caen_description"),
                                "source_site": parsed.get("source_site"),
                                "confidence": parsed.get("confidence"),
                                "years": parsed.get("years") or [],
                            },
                            "summary": {
                                "kind": "public_records_summary",
                                "company_name": parsed.get("company_name"),
                                "year_count": len(parsed.get("years") or []),
                            },
                            "recommendations": [],
                            "language": "en",
                            "model": "public_records_parser_v1",
                        },
                        on_conflict="document_id",
                        returning=False,
                    )
                persisted = True
                logger.info(
                    "[pipeline] %s public_records_summary persisted: %d years",
                    document_id, len(parsed.get("years") or []),
                )
            except Exception:  # noqa: BLE001
                logger.exception("[pipeline] public_records persistence failed")
            if persisted:
                # Mark the doc analyzed with NO period (this isn't a TB).
                _admin_set_status(
                    document_id, "analyzed",
                    duration_ms=int((time.time() - t0) * 1000),
                    period_id=None,
                )
                # Best-effort tag the detected_type. The CHECK constraint
                # may reject `public_records_summary` — that's fine, the
                # briefing.kind discriminator is the canonical signal.
                try:
                    with _supabase.admin() as admin_client:
                        admin_client.update(
                            "documents",
                            {"detected_type": "public_records_summary"},
                            filters={"id": f"eq.{document_id}"},
                        )
                except Exception:  # noqa: BLE001
                    pass  # CHECK constraint rejection is non-fatal
                return  # Short-circuit — no TB stages for this doc.

        # Persist the deterministic detected_type back to the documents
        # row. Upload-time detection uses filename heuristics only — once
        # `stage_extract` has inspected the content (TB anchors, F30/F10
        # row layout, Claude classification), the result here is far more
        # reliable and drives the FE banner. We only overwrite when
        # stage_extract surfaced a strong type so e.g. an SKU sales-export
        # routed via Claude doesn't clobber an existing "xlsx_workbook"
        # tag with "unknown".
        try:
            extracted_type = (parsed or {}).get("detected_type")
            if extracted_type in (
                "trial_balance", "statutory_f30_f10",
                "bilant", "pl", "annual_report",
            ):
                with _supabase.admin() as _ac:
                    _ac.update(
                        "documents",
                        {"detected_type": extracted_type},
                        filters={"id": f"eq.{document_id}"},
                    )
        except Exception:
            logger.exception("[pipeline] detected_type persist failed (non-fatal)")

        # Run the multi-country detector against the extracted text. Result is
        # stored on the document row so the mapper can pick the right COA and
        # the UI can decide whether to surface a "confirm detection?" card.
        # Best-effort: failures are non-fatal — extraction proceeds either way.
        try:
            ocr_text = parsed.get("raw_text") or "\n".join(
                f"{a.get('code','')} {a.get('name','')}"
                for a in (parsed.get("accounts") or [])
            )
            if ocr_text:
                det = _detect.detect_format(
                    ocr_text,
                    filename=doc.get("original_filename"),
                )
                with _supabase.admin() as ac:
                    ac.update(
                        "documents",
                        {
                            "detected_coa": det.get("coa_key"),
                            "detected_country": det.get("country_code"),
                            "detected_language": det.get("language"),
                            "detection_confidence": det.get("confidence"),
                        },
                        filters={"id": f"eq.{document_id}"},
                    )
                # Inject detection hints into parsed payload so downstream
                # stages (mapping, narrate) can use them.
                parsed.setdefault("detection", det)
                logger.info(
                    "[pipeline] %s detected: coa=%s country=%s lang=%s conf=%.2f decided_by=%s",
                    document_id, det.get("coa_key"), det.get("country_code"),
                    det.get("language"), det.get("confidence") or 0.0,
                    det.get("decided_by"),
                )
        except Exception:  # noqa: BLE001
            logger.exception("[pipeline] detection stage failed (non-fatal)")

        # SKU branch — completely independent of financial_periods. Two
        # sub-paths:
        #   1. Trading-analysis XLSX with native per-SKU rows → openpyxl
        #      extraction into sku_lines + sku_aggregates (full 406-row
        #      portfolio for the user's actual file).
        #   2. Anything else (PDF, CSV without recognizable sales shape) →
        #      the LLM-summary briefing path (sku_analyses).
        if scope == "sku":
            _admin_set_status(document_id, "mapping")
            dataset_id: Optional[str] = None
            try:
                dataset_id = _run_sales_dataset_pipeline(doc)
            except Exception as e:  # noqa: BLE001
                logger.warning("[pipeline] sales dataset path failed, falling back to summary: %s", e)

            # Always also produce a briefing — useful even with sku_lines,
            # gives the user a 3-sentence executive summary alongside the
            # raw portfolio. Skips if extraction returned nothing.
            _admin_set_status(document_id, "narrating")
            assembled = stage_map(doc, parsed, org.get("industry_display_name") or org.get("industry_key"))
            narrative = stage_narrate(doc, assembled, [], org, period_id="-", parsed=parsed)
            _persist_sku_analysis(doc, parsed, narrative)

            _admin_set_status(
                document_id,
                "analyzed",
                duration_ms=int((time.time() - t0) * 1000),
                period_id=None,
            )
            logger.info(
                "[pipeline] %s (sku scope) complete in %dms, dataset_id=%s",
                document_id, int((time.time() - t0) * 1000), dataset_id,
            )
            return

        # Financial branch — existing path.
        _admin_set_status(document_id, "mapping")
        assembled = stage_map(doc, parsed, org.get("industry_display_name") or org.get("industry_key"))
        period_id = stage_persist(doc, parsed, assembled)

        _admin_set_status(document_id, "computing")
        # For non-financial docs (no accounts extracted), skip ratio
        # computation + the empty-PL alert — they'd just produce noise. The
        # narrative stage covers what the document actually contains.
        accounts_count = len(parsed.get("accounts") or [])
        valuation_payload: Optional[Dict[str, Any]] = None
        if accounts_count > 0:
            metrics = stage_compute(doc, assembled, period_id)
            # Statutory anchor override — the TB parser captures account
            # 121's closing balance directly (the legally filed net profit
            # on Romanian books). stage_compute writes a stand-in
            # `net_income_statutory = operational + 722`, which is correct
            # for asset-heavy companies (EEI's case) where the 722 carry
            # IS the statutory gap. For manufacturers without 722 the gap
            # comes from class-65 provision movements and 781 reversals
            # — small enough that the reconstruction is within 1-2% of 121
            # for the oracle but can drift up to 6% on the platform.
            # The 121 closing balance is the authoritative number — patch
            # `net_income_statutory` to that when available so the FE +
            # briefing cite the same figure the user sees on their filings.
            anchor = (parsed or {}).get("statutory_net_profit_anchor")
            if anchor and abs(anchor) > 0.01:
                try:
                    with _supabase.admin() as ac:
                        ac.delete(
                            "calculated_metrics",
                            filters={"period_id": f"eq.{period_id}", "name": "eq.net_income_statutory"},
                        )
                        ac.insert("calculated_metrics", [{
                            "period_id": period_id,
                            "org_id": doc["org_id"],
                            "name": "net_income_statutory",
                            "value": round(float(anchor), 2),
                            "unit": "RON",
                            "direction": "higher",
                        }], returning=False)
                    logger.info(
                        "[pipeline] net_income_statutory overridden with ct 121 anchor: %s",
                        f"{float(anchor):,.0f}",
                    )
                except Exception:  # noqa: BLE001
                    logger.exception("[pipeline] statutory anchor override failed (non-fatal)")
            validation_alerts = stage_validate(doc, assembled, period_id)
            # AI Council — advisory extraction-integrity review (2026-07-20).
            # A panel of independent Claude personas scans the extraction and a
            # deterministic chair returns a consensus verdict. Non-blocking:
            # findings are appended to validation_alerts as 'data_quality'
            # advisories and surface in the existing alerts UI. stage_council
            # never raises; a missing API key degrades to a deterministic
            # rule-based verdict.
            council_result = stage_council(doc, parsed, assembled)
            if council_result:
                from . import _ai_council
                logger.info(
                    "[pipeline] ai_council verdict=%s confidence=%.2f findings=%d (doc=%s)",
                    council_result.get("verdict"),
                    council_result.get("confidence", 0.0),
                    len(council_result.get("findings", []) or []),
                    document_id,
                )
                validation_alerts = list(validation_alerts) + \
                    _ai_council.council_findings_as_alerts(council_result)
            # Industry classification fallback. When the org's industry_key is
            # unset or "generic", run the auto-classifier on the assembled
            # statements — for EEI this detects real_estate_commercial from
            # account 215 (investment property) and account 706 (rental income)
            # dominance, which gates the valuation method choice below.
            stored_industry_key = (org.get("industry_key") or "").lower().strip() or None
            classification = _ro_pack().classify_industry({"assembled": assembled})
            detected_industry_key = classification.get("industry_key") if classification.get("confidence", 0) >= 0.5 else None
            effective_industry_key = stored_industry_key if stored_industry_key and stored_industry_key != "generic" else (detected_industry_key or stored_industry_key)
            if classification.get("confidence", 0) >= 0.5:
                logger.info(
                    "[pipeline] industry classified as %s (confidence=%s, stored=%s, effective=%s)",
                    classification.get("industry_key"),
                    classification.get("confidence"),
                    stored_industry_key,
                    effective_industry_key,
                )

            # EBITDA-multiple valuation (primary) + DCF + EV/Revenue cross-checks.
            # For CRE / negative-EBITDA cases, _valuation.compute_valuation
            # demotes EV/EBITDA and uses asset-based as primary (Step 5 guard).
            # Pure math — never blocks the rest of the pipeline if it errors.
            try:
                valuation_payload = _valuation.compute_valuation(
                    industry_key=effective_industry_key,
                    statements=assembled["statements"],
                )
                # Surface the detection result on the valuation payload so the
                # frontend can display "Industry: Commercial Real Estate ·
                # auto-classified · confidence 0.85 · [Change]" badge.
                if valuation_payload is not None:
                    valuation_payload["industry_classification"] = classification
                    valuation_payload["industry_key_effective"] = effective_industry_key
                    valuation_payload["industry_key_stored"] = stored_industry_key
                _valuation.persist_valuation(period_id, doc["org_id"], valuation_payload)
            except Exception:  # noqa: BLE001
                logger.exception("[pipeline] valuation compute failed (non-fatal)")
        else:
            metrics = []
            validation_alerts = []

        _admin_set_status(document_id, "narrating")
        narrative = stage_narrate(
            doc, assembled, metrics, org, period_id,
            parsed=parsed, valuation=valuation_payload,
        )
        stage_persist_narrative(doc, period_id, narrative, validation_alerts)

        # Persist the assembled statements blob on financial_periods so the
        # period read endpoint can return it without re-deriving.
        # F4.3 — also persist the detection envelope (country/standard/
        # doc_type/industry detection + methodology pin) per
        # CANONICAL_SCHEMA_V1.md §7. Best-effort: if the JSONB column
        # doesn't exist yet (pre-F4.3 migration), the UPDATE fails and
        # we log + continue.
        envelope_payload: Dict[str, Any] = {}
        try:
            from engine.detection import build_detection_envelope  # type: ignore
            envelope = build_detection_envelope(
                classification=None,  # full upload classification only available in read path
                assembled=assembled,
                methodology_id="ro_ras_2025_v1",
                industry_key=effective_industry_key,
                period_start=parsed.get("period_end"),
                period_end=parsed.get("period_end"),
                currency=parsed.get("currency") or "RON",
            )
            envelope_payload = {
                "detection_envelope": envelope,
                "methodology_version": envelope.get("methodology_version") or "",
            }
        except Exception:  # noqa: BLE001
            logger.exception(
                "[pipeline] detection envelope build failed (non-fatal); "
                "read path will recompute"
            )
        with _supabase.admin() as admin_client:
            try:
                admin_client.update(
                    "financial_periods",
                    {"updated_at": _now_iso(), **envelope_payload},  # touch to bust cache + persist envelope
                    filters={"id": f"eq.{period_id}"},
                )
            except Exception:  # noqa: BLE001
                # F4.3 column may not exist yet on this DB — retry without it.
                logger.exception(
                    "[pipeline] period UPDATE with envelope failed; "
                    "retrying without envelope columns"
                )
                admin_client.update(
                    "financial_periods",
                    {"updated_at": _now_iso()},
                    filters={"id": f"eq.{period_id}"},
                )

        _admin_set_status(
            document_id,
            "analyzed",
            duration_ms=int((time.time() - t0) * 1000),
            period_id=period_id,
        )
        # Pricing V3 (gap D) — analysis SUCCEEDED. Convert the
        # reservation made at /api/pipeline/run into a consumed slot.
        # If the doc was flagged `metered_extra`, ALSO bump the
        # extra-docs-billed tally — that's the only path that triggers
        # an actual charge. No-op when USAGE_LIMITS_ENABLED is off.
        try:
            _commit_pipeline_quota(document_id, success=True)
        except Exception:
            logger.exception("[pipeline] commit_document_usage failed (non-fatal)")
        logger.info("[pipeline] %s complete in %dms", document_id, int((time.time() - t0) * 1000))
    except Exception as exc:  # noqa: BLE001
        logger.exception("[pipeline] %s failed", document_id)
        msg = f"{type(exc).__name__}: {exc}"
        try:
            _admin_set_status(document_id, "failed", error=msg, duration_ms=int((time.time() - t0) * 1000))
        except Exception:
            logger.exception("[pipeline] also failed to mark failed")
        # Pricing V3 (gap D) — analysis FAILED. Release the
        # reservation so the doc doesn't count against quota and the
        # user is not billed for an extra. No-op when disabled.
        try:
            _commit_pipeline_quota(document_id, success=False)
        except Exception:
            logger.exception("[pipeline] release_document_reservation failed (non-fatal)")


def _rebuild_assembled(line_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Reconstruct the `assembled.statements`-shaped dict the valuation
    engine expects, given the persisted statement_line_items for a period.
    Mirrors the bucket map used in /api/period/:id."""
    bs_buckets = {
        "cash": "cash", "ar": "accountsReceivable", "inventory": "inventory",
        "otherCurrentAssets": "otherCurrentAssets",
        "ppe": "propertyPlantEquipment", "intangibles": "intangibles",
        "otherNonCurrentAssets": "otherNonCurrentAssets",
        "ap": "accountsPayable", "stDebt": "shortTermDebt", "otherCurrentLiab": "otherCurrentLiabilities",
        "ltDebt": "longTermDebt", "otherNonCurrentLiab": "otherNonCurrentLiabilities",
        "shareCapital": "shareCapital", "retainedEarnings": "retainedEarnings", "otherEquity": "otherEquity",
    }
    pl_buckets = {
        "revenue": "revenue", "cogs": "costOfGoodsSold", "operatingExpenses": "operatingExpenses",
        "depreciation": "depreciationAmortization", "interestExpense": "interestExpense",
        "otherIncome": "otherIncome", "financialIncome": "financialIncome",
        "financialExpense": "financialExpense", "taxExpense": "taxExpense",
    }
    bs: Dict[str, float] = {v: 0.0 for v in bs_buckets.values()}
    pl: Dict[str, float] = {v: 0.0 for v in pl_buckets.values()}
    for item in line_items:
        bucket = item["bucket"]
        amount = float(item["amount"] or 0)
        if bucket in bs_buckets:
            bs[bs_buckets[bucket]] += amount
        elif bucket in pl_buckets:
            pl[pl_buckets[bucket]] += amount
    return {"balanceSheet": bs, "incomeStatement": pl}


def _rebuild_assembled_for_briefing(
    line_items: List[Dict[str, Any]],
    period: Dict[str, Any],
    org: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    """Rebuild the FULL canonical-shaped `assembled` envelope the way
    /api/period/{period_id} does — i.e. with `statements.assembled_pl`,
    `assembled_bs`, `assembled_cf`, `companyName`, `currency`,
    `periodLabel` populated — so `stage_narrate` can read its
    `briefing_facts` from operating-view canonical values rather than
    the coarse bucket sums returned by `_rebuild_assembled`.

    Mirrors the canonical-reassembly block in `/api/period/{period_id}`
    (search "canonical re-assembly" in this file). Used by the F2.8
    `POST /api/period/{period_id}/briefing/regenerate` endpoint so the
    regenerated briefing cites the same numbers the dashboard renders.
    """
    # F3.1c: dispatch via the Romania country pack rather than direct
    # `_ro_coa` import. `_coa_mod` alias preserved so the call sites
    # below read identically — the alias now points at the pack.
    _coa_mod = _ro_pack()

    # ── Legacy bucket aggregates (same map as _rebuild_assembled and
    # the /api/period reconstruction). Needed so the response carries
    # balanceSheet + incomeStatement alongside the canonical views.
    bs_buckets = {
        "cash": "cash", "ar": "accountsReceivable", "inventory": "inventory",
        "otherCurrentAssets": "otherCurrentAssets",
        "ppe": "propertyPlantEquipment", "intangibles": "intangibles",
        "otherNonCurrentAssets": "otherNonCurrentAssets",
        "ap": "accountsPayable", "stDebt": "shortTermDebt",
        "otherCurrentLiab": "otherCurrentLiabilities",
        "ltDebt": "longTermDebt",
        "otherNonCurrentLiab": "otherNonCurrentLiabilities",
        "shareCapital": "shareCapital",
        "retainedEarnings": "retainedEarnings",
        "otherEquity": "otherEquity",
    }
    pl_buckets = {
        "revenue": "revenue", "cogs": "costOfGoodsSold",
        "operatingExpenses": "operatingExpenses",
        "depreciation": "depreciationAmortization",
        "interestExpense": "interestExpense",
        "otherIncome": "otherIncome",
        "financialIncome": "financialIncome",
        "financialExpense": "financialExpense",
        "taxExpense": "taxExpense",
    }
    bs: Dict[str, float] = {v: 0.0 for v in bs_buckets.values()}
    pl: Dict[str, float] = {v: 0.0 for v in pl_buckets.values()}
    inv_var_memo = 0.0
    for item in line_items:
        bucket = item["bucket"]
        amount = float(item["amount"] or 0)
        code = (item.get("ro_account_code") or "").strip()
        if bucket in bs_buckets:
            bs[bs_buckets[bucket]] += amount
        elif bucket in pl_buckets:
            if bucket == "otherIncome" and code.startswith("711"):
                inv_var_memo += amount
            else:
                pl[pl_buckets[bucket]] += amount

    statements: Dict[str, Any] = {
        "companyName": (org or {}).get("name") if org else None,
        "industry": (org or {}).get("industry_display_name") if org else None,
        "currency": period.get("currency", "RON"),
        "periodLabel": period.get("period_end"),
        "balanceSheet": {k: round(v, 2) for k, v in bs.items()},
        "incomeStatement": {
            **{k: round(v, 2) for k, v in pl.items()},
            "inventoryVariationMemo": round(inv_var_memo, 2),
        },
        "supplementary": {"periodDays": 365},
    }

    # ── Canonical reassembly via assemble_statements ─────────────────
    # Same flow as /api/period: rebuild `recovered_accounts` from
    # line_items (preserving bucket_override for side-flipped rows),
    # re-apply sign for rule.sign==-1, then call assemble_statements.
    try:
        recovered_accounts: List[Dict[str, Any]] = []
        for li in (line_items or []):
            code = li.get("ro_account_code") or ""
            if not code:
                continue
            stored_bucket = li.get("bucket") or li.get("canonical_bucket") or ""
            rule = _coa_mod.bucket_for(code)
            bucket_override = None
            if rule and stored_bucket and stored_bucket != rule.bucket:
                legacy = _coa_mod.persistence_bucket(rule.bucket)
                if stored_bucket != legacy:
                    bucket_override = stored_bucket
            row = {
                "code": code,
                "name": li.get("ro_account_name") or "",
                "amount": float(li.get("amount") or 0),
            }
            if bucket_override:
                row["bucket_override"] = bucket_override
            recovered_accounts.append(row)
        # Re-apply sign so assemble_statements sees the raw input.
        for acct in recovered_accounts:
            rule = _coa_mod.bucket_for(acct["code"])
            if rule and rule.sign == -1:
                acct["amount"] = -acct["amount"]
        assembled_full = _coa_mod.assemble_statements(
            recovered_accounts,
            company_name=statements["companyName"] or "Entity",
            currency=statements["currency"],
            period_label=str(statements["periodLabel"]) if statements["periodLabel"] else "Period",
            industry=(org or {}).get("industry_key") if org else None,
        )
        statements["assembled_bs"] = assembled_full["statements"].get("assembled_bs")
        statements["assembled_pl"] = assembled_full["statements"].get("assembled_pl")
        statements["assembled_cf"] = assembled_full["statements"].get("assembled_cf")
        statements["assembled_bands"] = assembled_full["statements"].get("assembled_bands")
        statements["assembled_piotroski"] = assembled_full["statements"].get("assembled_piotroski")
        statements["subAggregates"] = assembled_full["statements"].get("subAggregates")
        # F4.1e — surface the country-agnostic canonical envelope on the
        # response statements. Lives at top level of `assembled_full`
        # (sibling of `statements`), NOT under statements.assembled_*.
        # Source: chart_of_accounts.assemble_statements line ~1817.
        statements["assembled_canonical_v1"] = assembled_full.get("assembled_canonical_v1")
        # F4.3 — surface the detection envelope (country/standard/doc_type/
        # industry + methodology pin per CANONICAL_SCHEMA_V1.md §7).
        try:
            from engine.detection import build_detection_envelope as _bde  # type: ignore
            statements["detection_envelope"] = _bde(
                classification=None, assembled=assembled_full,
                methodology_id="ro_ras_2025_v1",
                industry_key=(org or {}).get("industry_key") if org else None,
                currency=statements.get("currency"),
                period_end=str(statements.get("periodLabel") or ""),
            )
        except Exception:  # noqa: BLE001
            logger.exception("[briefing/regenerate] detection envelope build failed (non-fatal)")
    except Exception:  # noqa: BLE001
        logger.exception(
            "[briefing/regenerate] canonical re-assembly failed (non-fatal); "
            "falling back to bucket aggregates only"
        )

    # F4.6 — surface deprecated-field warnings on the response so
    # consumers can migrate ahead of the 2Q sunset (~Nov 2026).
    payload = {"statements": statements}
    try:
        from .deprecated_fields import attach_deprecated_fields
        attach_deprecated_fields(payload)
    except Exception:  # noqa: BLE001
        logger.exception("[briefing/regenerate] deprecated_fields attach failed (non-fatal)")
    return payload


def _serialize_valuation(valuation: Optional[Dict[str, Any]],
                          user_assumptions: Optional[Dict[str, Any]],
                          statements: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
    """Shape the raw valuations row for the dashboard. Returns None when the
    pipeline never produced a valuation (non-financial doc, or failure).

    When `statements` is passed, we re-run `_valuation.compute_valuation`
    against the canonical assembled views so the FE gets the same
    operating-view / asset-based payload the pipeline computed at write
    time — including FCF breakdown with REAL CapEx, asset-based primary
    for CRE, and the explicit ebitda_statutory / _operational / _operating_view
    triple. The persisted DB row alone doesn't carry those fields, but
    they're deterministic from the canonical statements.
    """
    if not valuation:
        return None

    # ── Recompute the full valuation against canonical statements ────────
    # When statements (with assembled_*) are available, prefer the fresh
    # recomputation over the row — the row loses the asset-based card,
    # FCF breakdown, and method warnings on round-trip. Reapply user
    # assumptions on top so manual overrides still take effect.
    fresh: Optional[Dict[str, Any]] = None
    if statements and statements.get("assembled_pl"):
        try:
            industry_key = None
            if isinstance(statements.get("industry"), str):
                industry_key = statements["industry"]
            ua_dict = None
            if user_assumptions:
                ua_dict = {
                    k: user_assumptions.get(k)
                    for k in ("ebitda_used", "multiple_used", "debt_used", "cash_used")
                    if user_assumptions.get(k) is not None
                }
            fresh = _valuation.compute_valuation(
                industry_key=industry_key,
                statements=statements,
                user_assumptions=ua_dict,
            )
        except Exception:  # noqa: BLE001
            logger.exception("[/api/period] valuation fresh recompute failed (non-fatal)")
            fresh = None

    # Choose the source of truth: fresh recomputation if available, else the row.
    src = fresh if fresh is not None else valuation

    def f(key: str) -> Optional[float]:
        v = src.get(key)
        return None if v is None else float(v)

    multiple_p50 = f("multiple_ebitda_p50")
    ebitda = f("ebitda_used")
    debt = f("total_debt_used")
    cash = f("cash_used")
    primary_method = src.get("primary_method") or "ev_ebitda"
    is_asset_based_primary = primary_method == "asset_based"

    def _fmt(n: Optional[float]) -> str:
        if n is None:
            return "—"
        sign = "-" if n < 0 else ""
        a = abs(n)
        if a >= 1_000_000:
            return f"{sign}{a/1_000_000:.2f}M"
        if a >= 1_000:
            return f"{sign}{a/1_000:.0f}K"
        return f"{sign}{a:.0f}"

    if is_asset_based_primary:
        formula_text = (
            src.get("formula_text")
            or f"Equity = Book equity ({_fmt(src.get('total_equity_used'))}) "
               f"+ RE markup (1.2-1.5× book) − Debt + Cash"
        )
    else:
        formula_text = (
            f"Equity = EBITDA ({_fmt(ebitda)}) × {multiple_p50}× − Debt ({_fmt(debt)}) + Cash ({_fmt(cash)})"
            if multiple_p50 is not None
            else "Insufficient benchmark data for an EBITDA-multiple valuation."
        )

    football_field: List[Dict[str, Any]] = []
    # Prefer the engine's pre-built football field when fresh is available
    # — it already places asset-based as primary for CRE and demotes
    # EV/EBITDA. Fall back to row-derived rows for legacy paths.
    if fresh and isinstance(fresh.get("football_field"), list):
        football_field = list(fresh["football_field"])
    elif f("equity_ebitda_p25") is not None and not is_asset_based_primary:
        football_field.append({
            "method": "EV / EBITDA (peers)",
            "primary": True,
            "low": f("equity_ebitda_p25"),
            "mid": f("equity_ebitda_p50"),
            "high": f("equity_ebitda_p75"),
            "subtitle": (
                f"{f('multiple_ebitda_p25')}× — {f('multiple_ebitda_p50')}× — {f('multiple_ebitda_p75')}×"
                if f("multiple_ebitda_p25") is not None else None
            ),
        })
    # Only append from the row when fresh isn't available (else we'd
    # duplicate the rows fresh already produced).
    if not fresh and f("ev_revenue_equity_p25") is not None:
        football_field.append({
            "method": "EV / Revenue (peers)",
            "primary": False,
            "low": f("ev_revenue_equity_p25"),
            "mid": f("ev_revenue_equity_p50"),
            "high": f("ev_revenue_equity_p75"),
            "subtitle": (
                f"{f('multiple_revenue_p25')}× — {f('multiple_revenue_p50')}× — {f('multiple_revenue_p75')}×"
                if f("multiple_revenue_p25") is not None else None
            ),
        })
    if not fresh and f("dcf_equity_value") is not None:
        football_field.append({
            "method": "DCF (WACC + Gordon)",
            "primary": False,
            "low": f("dcf_sensitivity_low"),
            "mid": f("dcf_equity_value"),
            "high": f("dcf_sensitivity_high"),
            "subtitle": (
                f"WACC {float(src['dcf_wacc'])*100:.1f}%, g {float(src['dcf_terminal_growth'])*100:.1f}%"
                if src.get("dcf_wacc") is not None else None
            ),
        })

    return {
        "primary_method": primary_method,
        "primary_label": src.get("primary_label"),
        "primary_equity_value": f("primary_equity_value"),
        "primary_equity_low": f("primary_equity_low"),
        "primary_equity_high": f("primary_equity_high"),
        "method_warnings": src.get("method_warnings") or [],
        # FCF breakdown — Valuation tab tiles read these verbatim. Real
        # CapEx, statutory net income. Only available when fresh recompute
        # ran (canonical statements present).
        "fcf_breakdown": (fresh or {}).get("fcf_breakdown"),
        # Three EBITDA views so the FE can show which one EV/EBITDA used.
        "ebitda_statutory": f("ebitda_statutory"),
        "ebitda_operational": f("ebitda_operational"),
        "ebitda_operating_view": f("ebitda_operating_view"),
        "confidence": src.get("confidence"),
        "multiples_source": src.get("multiples_source"),
        "multiples_as_of_date": str(src.get("multiples_as_of_date")) if src.get("multiples_as_of_date") else None,
        "formula_text": formula_text,
        "inputs": {
            "ebitda_used": ebitda,
            "revenue_used": f("revenue_used"),
            "total_debt_used": debt,
            "cash_used": cash,
        },
        "primary": {
            "method": primary_method,
            "multiple_p25": f("multiple_ebitda_p25"),
            "multiple_p50": multiple_p50,
            "multiple_p75": f("multiple_ebitda_p75"),
            "ev_p25": f("ev_ebitda_p25"),
            "ev_p50": f("ev_ebitda_p50"),
            "ev_p75": f("ev_ebitda_p75"),
            "equity_p25": f("equity_ebitda_p25"),
            "equity_p50": f("equity_ebitda_p50"),
            "equity_p75": f("equity_ebitda_p75"),
        },
        "cross_checks": {
            "revenue_multiple": {
                "multiple_p25": f("multiple_revenue_p25"),
                "multiple_p50": f("multiple_revenue_p50"),
                "multiple_p75": f("multiple_revenue_p75"),
                "equity_p25": f("ev_revenue_equity_p25"),
                "equity_p50": f("ev_revenue_equity_p50"),
                "equity_p75": f("ev_revenue_equity_p75"),
            },
            "dcf": {
                "wacc": f("dcf_wacc"),
                "terminal_growth": f("dcf_terminal_growth"),
                "enterprise_value": f("dcf_enterprise_value"),
                "equity_value": f("dcf_equity_value"),
                "sensitivity_low": f("dcf_sensitivity_low"),
                "sensitivity_high": f("dcf_sensitivity_high"),
            },
        },
        "football_field": football_field,
        "user_assumptions": user_assumptions and {
            "ebitda_used": user_assumptions.get("ebitda_used"),
            "multiple_used": user_assumptions.get("multiple_used"),
            "debt_used": user_assumptions.get("debt_used"),
            "cash_used": user_assumptions.get("cash_used"),
        },
    }


def _maybe_drop_empty_period(period_id: str) -> None:
    """Hard-delete the period row when NO documents (live OR soft-deleted)
    reference it. Skipped when the period was created < 5 minutes ago
    (safety window — a freshly-uploaded doc may not yet have stage_persist
    pinned its period_id, and we don't want to race-delete its parent).

    Bug-A fix (May 2026): previously this NULLed sibling documents'
    period_id BEFORE dropping the period, then deleted the period. With
    the pre-Bug-A 2-col period collision, two companies could share one
    period_id; soft-deleting one company's doc would null the OTHER
    company's period_id reference, orphaning its analyses. The fix here:
      · Don't manually NULL anything. documents.period_id has ON DELETE
        SET NULL (phase3.sql:530), so the FK handles cleanup atomically.
      · Only drop the period when EVERY doc (live OR soft-deleted) that
        could reference it is gone. Conservative — periods linger on the
        soft-delete shelf instead of vanishing the moment all live docs
        are gone, but never orphans data.

    Cascade: the financial_periods row has ON DELETE CASCADE on
    statement_line_items, calculated_metrics, briefings, benchmark_reports,
    valuations, user_valuation_assumptions — all drop atomically with it.
    """
    from datetime import datetime, timedelta, timezone

    with _supabase.admin() as client:
        period_rows = client.select(
            "financial_periods",
            filters={"id": f"eq.{period_id}"},
            single=True,
        )
        if not period_rows:
            return
        created_at = period_rows[0].get("created_at")
        if created_at:
            try:
                created = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                if datetime.now(timezone.utc) - created < timedelta(minutes=5):
                    return  # too young — pipeline may still be attaching docs
            except (TypeError, ValueError):
                pass

        # Bug-A fix: only drop the period when EVERY doc (live OR soft-deleted)
        # referencing it is gone. Don't manually NULL siblings — documents.
        # period_id is ON DELETE SET NULL (phase3.sql:530), the FK handles
        # cleanup atomically when the parent period row is finally deleted.
        any_docs = client.select(
            "documents",
            filters={"period_id": f"eq.{period_id}"},
            limit=1,
        )
        if any_docs:
            return  # period still has at least one doc (live or sd) — keep it

        client.delete("financial_periods", filters={"id": f"eq.{period_id}"})
        logger.info("[docs] dropped orphan period %s", period_id)


def _enqueue(document_id: str) -> None:
    """Run the pipeline on a daemon thread. Production should swap this for a
    real queue (Inngest, Supabase Edge functions, BullMQ-equivalent), but for
    a single-server MVP a thread per upload is fine — uploads are infrequent
    and each pipeline takes <60s.
    """
    t = threading.Thread(target=_run_pipeline_sync, args=(document_id,), daemon=True)
    t.start()


# ─── HTTP shape ─────────────────────────────────────────────────────────────


class RunRequest(BaseModel):
    document_id: str
    # Optional ISO 639-1 language code (en, ro, de, fr, es, …). When provided,
    # stage_narrate replies in that language; defaults to English.
    output_language: Optional[str] = None


class RunResponse(BaseModel):
    document_id: str
    status: str


class ReviewReanalyzeRequest(BaseModel):
    """F3.4 — Body for POST /api/period/{id}/review/reanalyze.

    Carries the user's per-session bucket overrides (`account_buckets`),
    their confirmation of country / accounting standard, and the
    `propose_as_calibration_rules` flag that promotes the overrides
    to `org_coa_mappings_overrides` (org-wide).
    """
    account_buckets: Dict[str, str] = Field(default_factory=dict)
    confirmed_country_code: Optional[str] = None
    confirmed_standard: Optional[str] = None
    notes: str = ""
    propose_as_calibration_rules: bool = False


def build_router() -> APIRouter:
    router = APIRouter(tags=["pipeline"])

    # ── FX rates endpoint (currency toggle, BNR proxy) ────────────────
    # Returns {base, rates, source, as_of, fetched_at, stale}. Backend
    # proxy avoids browser CORS on bnr.ro + caches 24h across all FE clients.
    @router.get("/api/fx-rates")
    def get_fx_rates_endpoint(refresh: bool = False) -> Dict[str, Any]:
        from .fx_rates import get_fx_rates
        return get_fx_rates(force_refresh=refresh)

    # F3-UX-2 follow-up: /api/paste-trial-balance endpoint removed after the
    # FE Paste-Trial-Balance UI was deleted (no remaining FE caller; no
    # external integrations found via grep). The request/response Pydantic
    # models and the handler block were deleted together. If a third-party
    # integration needs deterministic paste-parsing later, restore via:
    #   1. Pydantic models PasteTrialBalanceRequest / PasteTrialBalanceResponse
    #   2. Handler body — calls pack.parse_pasted_trial_balance + accounts_to_canonical_tsv
    # The country-pack methods themselves (parse_pasted_trial_balance,
    # accounts_to_canonical_tsv) remain available in the Romania pack.

    @router.post("/api/pipeline/run", response_model=RunResponse, status_code=202)
    def run_pipeline(req: RunRequest, authorization: Optional[str] = Header(None)) -> RunResponse:
        jwt = _require_jwt(authorization)
        doc = _verify_user_owns_document(jwt, req.document_id)
        # Pricing V3 (refined spec gaps C + D) — atomic reserve, success-only consume.
        #
        # Legacy `_usage_limits.check_quota` remains as a safety rail.
        # The V3 path uses `_usage_gate.reserve_document` which performs
        # an atomic Postgres conditional UPDATE: two concurrent uploads
        # at the boundary cannot both pass (gap C).
        #
        # The reservation is PROVISIONAL — `commit_document` runs only
        # when the orchestrator's daemon thread reports analysis
        # success (`_admin_set_status("analyzed", ...)`); on failure
        # (`_admin_set_status("failed", ...)`) the orchestrator calls
        # `release_document` and no quota is consumed / billed
        # (gap D — "consumed" = success only).
        #
        # The `was_extra` flag is stamped onto the documents row so
        # the daemon thread can recover it without an HTTP context.
        user_id = _user_id_from_jwt(jwt)
        _usage_limits.check_quota(user_id, "upload")
        from . import _usage_gate as _ug
        decision = _ug.reserve_document(user_id)
        if decision.kind == "blocked":
            raise HTTPException(
                status_code=429,
                detail={
                    "code": "doc_quota_blocked",
                    "plan_key": decision.plan_key,
                    "docs_used": decision.used,
                    "docs_included": decision.cap,
                    "message": decision.message,
                    "upgrade_url": "/pricing",
                },
            )
        if decision.kind == "extra_required":
            # FE must surface the confirm dialog and then call
            # POST /api/plan/confirm-extra-doc which routes through
            # `confirm_extra_document(user_id)` and reserves the slot
            # as billable. The repeat /api/pipeline/run call then
            # sees an `allowed` reservation (or the FE bypasses by
            # going straight to /api/pipeline/run after the confirm
            # endpoint succeeds — both flows valid).
            raise HTTPException(
                status_code=402,
                detail={
                    "code": "extra_doc_confirmation_required",
                    "plan_key": decision.plan_key,
                    "docs_used": decision.used,
                    "docs_included": decision.cap,
                    "extra_doc_eur": decision.extra_doc_eur,
                    "message": decision.message,
                    "confirm_url": "/api/plan/confirm-extra-doc",
                },
            )
        # `allowed` or `disabled` — proceed with enqueue.

        # Stash the was_extra flag onto the document row so the
        # daemon thread's commit/release call passes the right value.
        # Reusing an existing column would be cleaner; for now we
        # serialize a tiny meta blob into `documents.notes` (an
        # existing free-text column). The orchestrator parses it back.
        is_extra_reservation = decision.was_extra
        if req.output_language or is_extra_reservation:
            patch: Dict[str, Any] = {}
            if req.output_language:
                patch["detected_language"] = req.output_language
            if is_extra_reservation:
                # Use a column dedicated to this — `metered_extra` is
                # added by the V3 migration to documents below.
                patch["metered_extra"] = True
            with _supabase.admin() as ac:
                ac.update("documents", patch, filters={"id": f"eq.{req.document_id}"})

        _admin_set_status(req.document_id, "queued", pipeline_started_at=_now_iso())
        _enqueue(req.document_id)
        # Legacy monthly counter — preserve existing behavior.
        # Gap-D commit / release lives in the orchestrator (success +
        # failure terminal states). This `record_usage` is a soft
        # historical counter for the legacy /api/billing/usage view.
        _usage_limits.record_usage(user_id, "upload")
        return RunResponse(document_id=req.document_id, status="queued")

    @router.get("/api/sku-analysis/latest")
    def get_latest_sku_analysis(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Return the latest SKU analysis for the caller's organization.

        Strictly scoped to scope='sku' documents — never returns financial
        statement data. Used by the Products page to render the briefing +
        recommendations from the most recent inventory/sales upload.
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            rows = client.select(
                "sku_analyses",
                columns="*,documents!inner(id,original_filename,status,scope,created_at,error)",
                order="created_at.desc",
                limit=1,
            )
            if not rows:
                return {"analysis": None}
            r = rows[0]
            doc = r.get("documents") or {}
            return {
                "analysis": {
                    "id": r["id"],
                    "document": {
                        "id": doc.get("id"),
                        "filename": doc.get("original_filename"),
                        "status": doc.get("status"),
                        "scope": doc.get("scope"),
                        "created_at": doc.get("created_at"),
                        "error": doc.get("error"),
                    },
                    "briefing": r.get("briefing"),
                    "summary": r.get("summary"),
                    "recommendations": r.get("recommendations") or [],
                    "model": r.get("model"),
                    "created_at": r.get("created_at"),
                },
            }

    @router.get("/api/public-records/latest")
    def get_latest_public_records(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Return the most recent public-records-summary extract for the
        caller's org. The data lives in `sku_analyses` (reused — same JSONB
        shape) and is keyed by `briefing.kind == 'public_records_summary'`.
        The Multi-year History FE page reads from here.
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            # We over-fetch by 10 since `sku_analyses` mixes SKU briefings
            # with public-records extracts; the JSON `kind` discriminator
            # is the filter. Postgres JSONB path filter isn't trivially
            # exposed via PostgREST, so client-side filter is the cleanest.
            rows = client.select(
                "sku_analyses",
                columns="*,documents!inner(id,original_filename,status,scope,detected_type,created_at,error)",
                order="created_at.desc",
                limit=10,
            )
            for r in rows:
                briefing = r.get("briefing")
                # PostgREST returns JSONB as either dict or string depending
                # on Accept header negotiation — handle both shapes.
                if isinstance(briefing, str):
                    try:
                        briefing = json.loads(briefing)
                    except (json.JSONDecodeError, TypeError):
                        briefing = {}
                briefing = briefing or {}
                if isinstance(briefing, dict) and briefing.get("kind") == "public_records_summary":
                    doc = r.get("documents") or {}
                    return {
                        "extract": {
                            "id": r["id"],
                            "document": {
                                "id": doc.get("id"),
                                "filename": doc.get("original_filename"),
                                "status": doc.get("status"),
                                "detected_type": doc.get("detected_type"),
                                "created_at": doc.get("created_at"),
                            },
                            "company_name": briefing.get("company_name"),
                            "cui": briefing.get("cui"),
                            "reg_com": briefing.get("reg_com"),
                            "caen_code": briefing.get("caen_code"),
                            "caen_description": briefing.get("caen_description"),
                            "source_site": briefing.get("source_site"),
                            "confidence": briefing.get("confidence"),
                            "years": briefing.get("years") or [],
                            "created_at": r.get("created_at"),
                        },
                    }
            return {"extract": None}

    @router.get("/api/public-records/by-document/{document_id}")
    def get_public_records_by_document(
        document_id: str,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """Fetch the public-records extract for a specific document. Lets
        the FE deep-link `/multi-year-history?doc=<id>` to a specific
        upload rather than always rendering the latest."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            rows = client.select(
                "sku_analyses",
                filters={"document_id": f"eq.{document_id}"},
                columns="*,documents!inner(id,original_filename,status,detected_type,created_at)",
                single=True,
            )
            if not rows:
                raise HTTPException(404, "No analysis for that document.")
            r = rows[0]
            briefing = r.get("briefing")
            if isinstance(briefing, str):
                try:
                    briefing = json.loads(briefing)
                except (json.JSONDecodeError, TypeError):
                    briefing = {}
            briefing = briefing or {}
            if not isinstance(briefing, dict) or briefing.get("kind") != "public_records_summary":
                raise HTTPException(404, "Document is not a public-records summary.")
            doc = r.get("documents") or {}
            return {
                "extract": {
                    "id": r["id"],
                    "document": {
                        "id": doc.get("id"),
                        "filename": doc.get("original_filename"),
                        "status": doc.get("status"),
                        "detected_type": doc.get("detected_type"),
                        "created_at": doc.get("created_at"),
                    },
                    "company_name": briefing.get("company_name"),
                    "cui": briefing.get("cui"),
                    "reg_com": briefing.get("reg_com"),
                    "caen_code": briefing.get("caen_code"),
                    "caen_description": briefing.get("caen_description"),
                    "source_site": briefing.get("source_site"),
                    "confidence": briefing.get("confidence"),
                    "years": briefing.get("years") or [],
                    "created_at": r.get("created_at"),
                },
            }

    @router.get("/api/org/periods-with-documents")
    def list_periods_with_documents(
        authorization: Optional[str] = Header(None),
        x_org_id: Optional[str] = Header(None, alias="X-Org-Id"),
    ) -> Dict[str, Any]:
        """Right-anchored Docs panel feed: every financial_period the user's
        org has, with the source documents that produced it, ordered newest
        first. Plus a `recently_deleted` shelf for soft-deleted docs that
        haven't hit the 30-day hard-delete cutoff.

        Single endpoint per panel-open (no re-fetch on toggle).
        """
        from datetime import datetime, timezone, timedelta

        jwt = _require_jwt(authorization)
        # Caller's ACTIVE workspace — validated against membership, not the
        # first row. See _org.resolve_org.
        try:
            _user_id, org_id = _org.resolve_org(jwt, x_org_id)
        except HTTPException as exc:
            if exc.status_code == 404:
                return {"active_period_id": None, "periods": [], "recently_deleted": []}
            raise
        with _supabase.per_user(jwt) as client:
            periods = client.select(
                "financial_periods",
                filters={"org_id": f"eq.{org_id}"},
                order="period_end.desc,created_at.desc",
            )
            docs = client.select(
                "documents",
                filters={
                    "org_id": f"eq.{org_id}",
                    "scope": "eq.financial",
                    "deleted_at": "is.null",
                },
                order="created_at.desc",
            )
            deleted = client.select(
                "documents",
                filters={
                    "org_id": f"eq.{org_id}",
                    "deleted_at": "not.is.null",
                },
                order="deleted_at.desc",
            )

        # Index docs by period_id (one doc → one period). Documents not yet
        # attached to a period (mid-pipeline or sku-scope) are omitted; the
        # panel surfaces only analyzed financial periods.
        docs_by_period: Dict[str, List[Dict[str, Any]]] = {}
        for d in docs:
            pid = d.get("period_id")
            if not pid:
                continue
            docs_by_period.setdefault(pid, []).append({
                "id": d["id"],
                "display_name": d.get("display_name") or d["original_filename"],
                "original_filename": d["original_filename"],
                "storage_path": d["storage_path"],
                "mime_type": d.get("mime_type"),
                "detected_type": d.get("detected_type"),
                "size_bytes": d["size_bytes"],
                "uploaded_at": d["created_at"],
                "status": d["status"],
                "is_active": d.get("is_active", True),
                "confidence": d.get("extraction_confidence"),
                "error": d.get("error"),
            })

        # Most recent analyzed period = the dashboard default.
        #
        # F3.15 Chunk 2 — prefer periods with meaningful data so a fresh
        # dashboard load lands on the operator's most recent USEFUL
        # analysis, not an empty SKU-misuploaded period that happens to
        # be the newest. The data-status check needs the per-period
        # has_meaningful_data computed below, so we do a quick pre-pass
        # here using the same logic (total_assets metric + sku_analyses
        # existence) — duplicates a few lines but avoids restructuring
        # the whole function. Falls back to the legacy newest-with-docs
        # rule if every period is empty (so the dashboard still has
        # SOMETHING to show on edge accounts).
        pre_period_ids = [pp["id"] for pp in periods if docs_by_period.get(pp["id"])]
        pre_doc_ids = [dd["id"] for plist in docs_by_period.values() for dd in plist]
        with _supabase.admin() as ac_pre:
            pre_metrics = ac_pre.select(
                "calculated_metrics",
                filters={
                    "period_id": "in.(" + ",".join(pre_period_ids) + ")",
                    "name": "eq.total_assets",
                },
            ) if pre_period_ids else []
            pre_sku = ac_pre.select(
                "sku_analyses",
                filters={"document_id": "in.(" + ",".join(pre_doc_ids) + ")"} if pre_doc_ids else {},
            ) if pre_doc_ids else []
        pre_ta_by_period: Dict[str, float] = {}
        for mm in (pre_metrics or []):
            try:
                pre_ta_by_period[mm["period_id"]] = float(mm.get("value") or 0)
            except (TypeError, ValueError):
                continue
        pre_sku_doc_ids = {ss.get("document_id") for ss in (pre_sku or []) if ss.get("document_id")}

        def _period_has_data(pp: Dict[str, Any]) -> bool:
            dlist = docs_by_period.get(pp["id"], [])
            if not dlist:
                return False
            status = (dlist[0].get("status") or "").lower()
            if status != "analyzed":
                # In-flight uploads count as "has data" (don't skip them).
                return True
            return pre_ta_by_period.get(pp["id"], 0.0) > 0 or any(d["id"] in pre_sku_doc_ids for d in dlist)

        active_period_id = next(
            (p["id"] for p in periods if _period_has_data(p)),
            None,
        )
        # Final fallback — if every period is empty, still pick the
        # newest-with-docs so the dashboard isn't completely blank.
        if active_period_id is None:
            active_period_id = next(
                (p["id"] for p in periods if docs_by_period.get(p["id"])),
                None,
            )

        # Skip periods with zero live documents. The pipeline's `stage_persist`
        # always attaches a document on insert, so a doc-less period can only
        # appear when (a) every attached doc was soft-deleted or (b) a row
        # leaked through pre-fix code paths. Either way the UI never wants to
        # render them. (Step 5 of the docs-panel fix.)
        #
        # F3.15 Chunk 2 — `has_meaningful_data` per period. Bulk-load the
        # total_assets metric for every period in one query (no N+1 risk
        # even with hundreds of periods) plus an existence check on
        # sku_analyses for the period's documents. The FE filters by this
        # flag by default and shows "N empty periods hidden · show" at
        # the picker's bottom for one-click reveal. The `status` guard
        # keeps in-flight uploads visible: only `analyzed` periods with
        # zero TB output AND no SKU output are flagged as empty.
        period_ids = [p["id"] for p in periods if docs_by_period.get(p["id"])]
        all_doc_ids = [d["id"] for plist in docs_by_period.values() for d in plist]
        with _supabase.admin() as ac_bulk:
            metric_rows = ac_bulk.select(
                "calculated_metrics",
                filters={
                    "period_id": "in.(" + ",".join(period_ids) + ")",
                    "name": "eq.total_assets",
                },
            ) if period_ids else []
            sku_rows = ac_bulk.select(
                "sku_analyses",
                filters={"document_id": "in.(" + ",".join(all_doc_ids) + ")"} if all_doc_ids else {},
            ) if all_doc_ids else []
        total_assets_by_period: Dict[str, float] = {}
        for m in (metric_rows or []):
            try:
                total_assets_by_period[m["period_id"]] = float(m.get("value") or 0)
            except (TypeError, ValueError):
                continue
        sku_doc_ids = {s.get("document_id") for s in (sku_rows or []) if s.get("document_id")}

        period_rows = []
        for p in periods:
            doc_list = docs_by_period.get(p["id"], [])
            if not doc_list:
                continue
            # has_meaningful_data: True unless ALL these hold:
            #   1. The primary doc is in 'analyzed' status (in-flight stays visible)
            #   2. total_assets is zero/missing on this period
            #   3. None of the period's docs have an sku_analyses row
            # Conservative default: show. Hide only on proven emptiness.
            primary_doc_status = (doc_list[0].get("status") or "").lower()
            ta = total_assets_by_period.get(p["id"], 0.0)
            has_sku = any(d["id"] in sku_doc_ids for d in doc_list)
            if primary_doc_status == "analyzed" and ta == 0 and not has_sku:
                has_data = False
            else:
                has_data = True
            period_rows.append({
                "period_id": p["id"],
                "period_label": p.get("period_end") or "Imported period",
                "period_start": p.get("period_start"),
                "period_end": p.get("period_end"),
                "is_active": p["id"] == active_period_id,
                "currency": p.get("currency"),
                "documents": doc_list,
                "extraction_confidence": p.get("extraction_confidence"),
                "has_meaningful_data": has_data,
            })

        # Soft-delete window is 30 days — surface deleted docs that fall
        # within that range so the user can restore them.
        cutoff = datetime.now(timezone.utc) - timedelta(days=30)
        deleted_rows = []
        for d in deleted:
            try:
                deleted_at = datetime.fromisoformat(d["deleted_at"].replace("Z", "+00:00"))
            except (KeyError, ValueError, AttributeError):
                continue
            if deleted_at < cutoff:
                continue
            deleted_rows.append({
                "id": d["id"],
                "display_name": d.get("display_name") or d["original_filename"],
                "deleted_at": d["deleted_at"],
                "restorable_until": (deleted_at + timedelta(days=30)).isoformat(),
                # `scope` lets the FE filter this shared array per panel:
                #   · DocsPanel (dashboard, financial domain) shows only
                #     scope=='financial' rows
                #   · DatasetsPanel (/products, SKU domain) shows only
                #     scope=='sku' rows
                # Without this field, BOTH panels would render the union
                # — leaking trial-balance / public-records deletes into
                # the Products UI (the user-reported "93 deleted finance
                # files appearing in Products" bug). Domain isolation is
                # client-side here because the endpoint is shared; the
                # backend keeps returning the union for backwards-compat.
                "scope": d.get("scope"),
            })

        # Public-records uploads (listafirme.ro / termene.ro / firme.info)
        # don't create a financial_period — they live in sku_analyses with
        # briefing.kind = 'public_records_summary'. The DocsPanel switcher
        # needs them too so the user can flip between trial-balance and
        # public-records analyses for any company they've uploaded. Add a
        # parallel `public_records` array keyed only on `documents.id`.
        public_records: List[Dict[str, Any]] = []
        try:
            with _supabase.admin() as ac:
                # Live (non-soft-deleted) sku_analyses rows joined with their
                # parent document, filtered to this org.
                sku_rows = ac.select(
                    "sku_analyses",
                    columns="document_id,briefing,created_at,documents!inner(id,original_filename,display_name,status,detected_type,created_at,deleted_at,org_id)",
                    order="created_at.desc",
                    limit=50,
                )
                for r in sku_rows:
                    doc = r.get("documents") or {}
                    if doc.get("deleted_at"):
                        continue
                    if doc.get("org_id") != org_id:
                        continue
                    briefing = r.get("briefing")
                    if isinstance(briefing, str):
                        try:
                            briefing = json.loads(briefing)
                        except (json.JSONDecodeError, TypeError):
                            briefing = {}
                    if not isinstance(briefing, dict):
                        continue
                    if briefing.get("kind") != "public_records_summary":
                        continue
                    years = briefing.get("years") or []
                    year_min = min((y.get("year") for y in years if y.get("year")), default=None)
                    year_max = max((y.get("year") for y in years if y.get("year")), default=None)
                    public_records.append({
                        "document_id": doc.get("id"),
                        "display_name": doc.get("display_name") or doc.get("original_filename"),
                        "original_filename": doc.get("original_filename"),
                        "status": doc.get("status"),
                        "detected_type": doc.get("detected_type"),
                        "created_at": doc.get("created_at"),
                        "company_name": briefing.get("company_name"),
                        "cui": briefing.get("cui"),
                        "caen_code": briefing.get("caen_code"),
                        "caen_description": briefing.get("caen_description"),
                        "years_count": len(years),
                        "year_min": year_min,
                        "year_max": year_max,
                        "confidence": briefing.get("confidence"),
                    })
        except Exception:
            logger.exception("[periods-with-documents] public-records enumeration failed (non-fatal)")

        return {
            "active_period_id": active_period_id,
            "periods": period_rows,
            "public_records": public_records,
            "recently_deleted": deleted_rows,
        }

    @router.patch("/api/documents/{document_id}")
    def patch_document(document_id: str, payload: Dict[str, Any], authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Rename / mark-inactive / soft-delete a document. RLS scoped — the
        caller must be a member of the document's org for the update to
        succeed (per-user client respects member-scoped policies)."""
        jwt = _require_jwt(authorization)
        allowed = {"display_name", "is_active"}
        patch = {k: v for k, v in payload.items() if k in allowed}
        if not patch:
            raise HTTPException(400, "No allowed fields provided. Allowed: display_name, is_active.")
        with _supabase.per_user(jwt) as client:
            client.update("documents", patch, filters={"id": f"eq.{document_id}"})
            rows = client.select("documents", filters={"id": f"eq.{document_id}"}, single=True)
            if not rows:
                raise HTTPException(404, "Document not found or not visible.")
            return rows[0]

    # Static-path routes MUST be registered before the
    # `/api/documents/{document_id}` parameter route below — otherwise
    # FastAPI matches the static segment as a `document_id` value.

    @router.post("/api/documents/clear-mine")
    def clear_my_uploads(
        authorization: Optional[str] = Header(None),
        x_org_id: Optional[str] = Header(None, alias="X-Org-Id"),
    ) -> Dict[str, Any]:
        """Settings → Data → "Clear all my uploaded documents."
        Soft-deletes every live document in the caller's org with a
        single UPDATE: sets `deleted_at = NOW()` where
        `org_id = caller's org AND deleted_at IS NULL`.

        Safety contract — what this endpoint INTENTIONALLY DOES NOT DO:
          · No hard-delete. No `DELETE FROM`. Rows stay in the documents
            table with `deleted_at` set so they remain restorable.
          · No `financial_periods` mutation. No row removed, no column
            updated. Orphaned empty periods left behind are Bug A's
            domain and are intentionally NOT cleaned up here.
          · No call to `_maybe_drop_empty_period`. That helper triggers
            the sibling-NULL cascade that Bug A is about; reproducing
            that cascade from a Settings button would create the exact
            data-bleed pattern Bug A is meant to fix.
          · No `statement_line_items`/`calculated_metrics` mutation.
            FK ON DELETE SET NULL handles `documents.period_id` if a
            period is later dropped (separate path), but that's not
            this endpoint's job.

        This is the only SAFE batch soft-delete in the codebase. The
        adjacent `DELETE /api/documents/{id}` endpoint DOES call
        `_maybe_drop_empty_period` per-doc — Settings deliberately
        does NOT loop that endpoint for this reason.

        Returns: {"deleted_count": N, "org_id": "..."}.
        """
        jwt = _require_jwt(authorization)
        # Caller's ACTIVE workspace. This used to be inlined (to dodge a
        # circular import from _benchmarks.py) and took the first membership;
        # _org has no router deps, so it can be imported safely.
        try:
            _user_id, org_id = _org.resolve_org(jwt, x_org_id)
        except HTTPException as exc:
            if exc.status_code == 404:
                return {"deleted_count": 0, "org_id": None, "deleted_at": None}
            raise

        now = _now_iso()
        # Single UPDATE via PostgREST — atomic, org-scoped via the
        # filter (defense-in-depth on top of RLS), no cascade.
        with _supabase.admin() as ac:
            # Fetch first (for the count to return); update second.
            live = ac.select(
                "documents",
                filters={
                    "org_id":     f"eq.{org_id}",
                    "deleted_at": "is.null",
                },
                columns="id",
            )
            if live:
                ac.update(
                    "documents",
                    {"deleted_at": now},
                    filters={
                        "org_id":     f"eq.{org_id}",
                        "deleted_at": "is.null",
                    },
                )
        return {
            "deleted_count": len(live),
            "org_id": org_id,
            "deleted_at": now,
        }

    @router.delete("/api/documents/clear-deleted")
    def clear_recently_deleted(
        period_id: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """Hard-delete every soft-deleted document visible to the caller.

        When `period_id` is supplied, scope is limited to that period.
        Otherwise, all soft-deleted documents in the caller's org are
        wiped. Uses per_user select to enforce RLS scoping, then admin
        cleanup for storage + cascade — same pattern as the per-doc
        endpoint below.
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            filters: Dict[str, str] = {"deleted_at": "not.is.null"}
            if period_id:
                filters["period_id"] = f"eq.{period_id}"
            visible = client.select("documents", filters=filters)

        deleted_ids: List[str] = []
        with _supabase.admin() as admin:
            for doc in visible:
                doc_id = doc["id"]
                storage_path = doc.get("storage_path")
                if storage_path:
                    try:
                        admin.delete_object("documents", storage_path)
                    except Exception:  # noqa: BLE001
                        logger.exception(
                            "[docs] clear-deleted: failed to remove blob %s", storage_path
                        )
                for table in ("alerts", "statement_line_items", "calculated_metrics",
                              "sales_datasets", "sku_lines", "sku_aggregates"):
                    try:
                        admin.delete(table, filters={"document_id": f"eq.{doc_id}"})
                    except Exception:  # noqa: BLE001
                        logger.debug("[docs] clear-deleted: no cascade on %s", table)
                admin.delete("documents", filters={"id": f"eq.{doc_id}"})
                deleted_ids.append(doc_id)

        return {"deleted_count": len(deleted_ids), "deleted_ids": deleted_ids}

    @router.delete("/api/documents/{document_id}")
    def soft_delete_document(document_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Soft-delete a document. Sets deleted_at = now(). Restorable via
        POST /api/documents/:id/restore within 30 days; a cron sweep
        hard-deletes after.

        Side effect: if soft-deleting this document leaves the parent
        financial_period with zero live documents, drop the period row
        and its derived data so the panel never shows a doc-less period.
        Skips periods younger than 5 minutes (safety window: a doc may
        be mid-pipeline and not yet pinned to period_id). (Step 5 of the
        docs-panel fix.)
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            doc_rows = client.select("documents", filters={"id": f"eq.{document_id}"}, single=True)
            client.update("documents", {"deleted_at": _now_iso()}, filters={"id": f"eq.{document_id}"})

        # Cleanup orphan period using admin so RLS doesn't block the cascade.
        doc = doc_rows[0] if doc_rows else None
        period_id = doc and doc.get("period_id")
        if period_id:
            try:
                _maybe_drop_empty_period(period_id)
            except Exception:  # noqa: BLE001
                logger.exception("[docs] orphan-period cleanup failed for %s", period_id)
        return {"document_id": document_id, "deleted_at": _now_iso()}

    @router.post("/api/documents/{document_id}/restore")
    def restore_document(document_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Restore a soft-deleted document."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            client.update("documents", {"deleted_at": None}, filters={"id": f"eq.{document_id}"})
            return {"document_id": document_id, "restored": True}

    @router.delete("/api/documents/{document_id}/permanent")
    def permanent_delete_document(
        document_id: str, authorization: Optional[str] = Header(None)
    ) -> Dict[str, Any]:
        """Hard-delete a previously soft-deleted document.

        Two-step gate: the document MUST already have `deleted_at` set —
        accidental deletes on active documents are not possible through
        this endpoint. The flow is always
            DELETE /api/documents/:id            (soft, sets deleted_at)
            DELETE /api/documents/:id/permanent  (hard, removes blob + row)

        Side effects:
          1. Removes the underlying blob from the `documents` storage
             bucket (path = `documents.storage_path`).
          2. Cascade-deletes derived rows (statement_line_items,
             calculated_metrics, briefings, valuations, alerts) that
             are FK-tied to the document or its period.
          3. Hard-deletes the document row itself.
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            rows = client.select("documents", filters={"id": f"eq.{document_id}"}, single=True)
            if not rows:
                raise HTTPException(404, "Document not found.")
            doc = rows[0]
            if not doc.get("deleted_at"):
                raise HTTPException(
                    400,
                    "Cannot permanently delete a document that has not been soft-deleted first.",
                )
            storage_path = doc.get("storage_path")

        # Storage + cascade cleanup use the admin client (RLS doesn't gate
        # us once we've passed the membership check above via per_user).
        with _supabase.admin() as admin:
            # 1) Remove the underlying blob. Log + continue on failure — a
            # missing blob shouldn't block the DB cleanup.
            if storage_path:
                try:
                    admin.delete_object("documents", storage_path)
                except Exception:  # noqa: BLE001
                    logger.exception(
                        "[docs] failed to remove storage blob %s (continuing with DB cleanup)",
                        storage_path,
                    )

            # 2) Cascade-delete derived rows that reference this doc. Most
            # FKs are ON DELETE CASCADE at the schema level, but some
            # (alerts.document_id, sales_datasets.document_id) are
            # ON DELETE SET NULL / restrict — wipe those explicitly so the
            # next user doesn't see ghost rows pointing at a deleted doc.
            for table in ("alerts", "statement_line_items", "calculated_metrics",
                          "sales_datasets", "sku_lines", "sku_aggregates"):
                try:
                    admin.delete(table, filters={"document_id": f"eq.{document_id}"})
                except Exception:  # noqa: BLE001
                    # Some tables won't have a document_id column; that's
                    # fine — the FK simply doesn't exist there.
                    logger.debug("[docs] no document_id cascade on %s (skipping)", table)

            # 3) Hard-delete the document row.
            admin.delete("documents", filters={"id": f"eq.{document_id}"})

        return {"document_id": document_id, "permanently_deleted": True}

    @router.patch("/api/sales-datasets/{dataset_id}")
    def patch_sales_dataset(dataset_id: str, payload: Dict[str, Any], authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Rename a sales dataset's user-visible label."""
        jwt = _require_jwt(authorization)
        allowed = {"label", "is_active"}
        patch = {k: v for k, v in payload.items() if k in allowed}
        if not patch:
            raise HTTPException(400, "No allowed fields. Allowed: label, is_active.")
        with _supabase.per_user(jwt) as client:
            client.update("sales_datasets", patch, filters={"id": f"eq.{dataset_id}"})
            rows = client.select("sales_datasets", filters={"id": f"eq.{dataset_id}"}, single=True)
            if not rows:
                raise HTTPException(404, "Dataset not found.")
            return rows[0]

    @router.delete("/api/sales-datasets/{dataset_id}")
    def delete_sales_dataset(dataset_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Soft-delete: marks the parent document as deleted_at=now() so the
        dataset list filter (which joins documents where deleted_at IS NULL)
        hides it. The sku_lines + sku_aggregates rows cascade-delete only
        when the dataset row itself is hard-deleted by the 30-day cron."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            ds = client.select("sales_datasets", filters={"id": f"eq.{dataset_id}"}, single=True)
            if not ds:
                raise HTTPException(404, "Dataset not found.")
            doc_id = ds[0]["document_id"]
            client.update("documents", {"deleted_at": _now_iso()}, filters={"id": f"eq.{doc_id}"})
            return {"dataset_id": dataset_id, "document_id": doc_id, "deleted_at": _now_iso()}

    @router.patch("/api/sku-aggregates/{sku_id}/decision")
    def patch_sku_decision(
        sku_id: str,
        payload: Dict[str, Any],
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """Persist the operator decision on a single SKU. Used by the Products
        detail drawer's two action buttons:
          - 'eliminate_approved' — operator confirms the engine's cut signal
          - 'strategic_override' — operator overrides eliminate/wind_down,
                                   marking the SKU as strategic (keep)
        Pass `null` to clear an existing override and fall back to the engine
        classification. The engine `classification` column is NOT mutated —
        we only write `user_override` so the override is reversible and the
        original engine reasoning stays auditable.
        """
        allowed = {"eliminate_approved", "strategic_override"}
        raw = payload.get("user_override")
        if raw is not None and raw not in allowed:
            raise HTTPException(
                400,
                f"Invalid user_override. Allowed: {sorted(allowed)} or null to clear.",
            )

        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            existing = client.select("sku_aggregates", filters={"id": f"eq.{sku_id}"}, single=True)
            if not existing:
                raise HTTPException(404, "SKU not found.")
            client.update(
                "sku_aggregates",
                {"user_override": raw},
                filters={"id": f"eq.{sku_id}"},
            )
            rows = client.select("sku_aggregates", filters={"id": f"eq.{sku_id}"}, single=True)
            return {"sku": rows[0]}

    @router.post("/api/sales-datasets/{dataset_id}/rerun")
    def rerun_sales_dataset(dataset_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Re-classify a dataset AND re-extract DIO from the source XLSX.
        Useful after engine rule changes OR parser upgrades.

        2026-05-26 — extended from re-classify-only. Originally this
        endpoint just re-ran `classify_portfolio` on existing aggregates,
        so any improvement to the upload-time parsers (e.g. the new
        Analysis-sheet DIO extraction) was invisible to existing
        datasets — users had to re-upload from scratch to see new
        category DIO values. That's unintuitive: "rerun" should mean
        "run the pipeline again", not "re-color the existing buckets".
        Now: fetch the original XLSX → re-extract the DIO map (Analysis
        sheet preferred, DIO sheet fallback) → update each SKU row's
        `days_inventory_on_hand` from the category map → reclassify.
        """
        from ._sales_extract import (
            extract_category_dio,
            extract_category_dio_from_analysis_sheet,
            _normalize_category,
        )
        from ._sku_classify import classify_portfolio
        import re as _re

        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            ds = client.select("sales_datasets", filters={"id": f"eq.{dataset_id}"}, single=True)
            if not ds:
                raise HTTPException(404, "Dataset not found.")
            ds_row = ds[0]

        with _supabase.admin() as ac:
            aggs = ac.select("sku_aggregates", filters={"dataset_id": f"eq.{dataset_id}"})
            if not aggs:
                return {"reclassified": 0}

            # ── DIO re-extraction pass ────────────────────────────
            # The dataset row carries `document_id` linking to the
            # original upload. If we can fetch the storage_path we
            # can re-run the modern DIO parsers on the source bytes
            # and refresh the per-SKU `days_inventory_on_hand` column.
            dio_updated = 0
            try:
                doc_id = ds_row.get("document_id")
                if doc_id:
                    docs = ac.select("documents", filters={"id": f"eq.{doc_id}"}, single=True)
                    if docs:
                        storage_path = docs[0].get("storage_path")
                        if storage_path:
                            signed = ac.signed_url("documents", storage_path, expires_in=300)
                            r = httpx.get(signed, timeout=60.0)
                            r.raise_for_status()
                            xlsx_bytes = r.content
                            analysis_map = extract_category_dio_from_analysis_sheet(xlsx_bytes)
                            dio_sheet_map = extract_category_dio(xlsx_bytes)
                            cat_dio_map = {**dio_sheet_map, **analysis_map}
                            logger.info(
                                "[rerun] dataset=%s re-extracted category DIO: "
                                "analysis=%d, dio_sheet=%d, merged=%d",
                                dataset_id, len(analysis_map), len(dio_sheet_map),
                                len(cat_dio_map),
                            )
                            # Apply per-aggregate, mirroring aggregate_sku_lines
                            # fallback order: explicit DIO (kept as-is) >
                            # computed inv/cogs (kept as-is) > category map.
                            # We only OVERWRITE when the existing value is
                            # null — preserves any per-SKU DIO that was
                            # already correct from a richer-shaped workbook.
                            for a in aggs:
                                if a.get("days_inventory_on_hand") is not None:
                                    continue
                                cat = _normalize_category(a.get("category") or "")
                                dio = None
                                if cat and cat in cat_dio_map:
                                    dio = cat_dio_map[cat]
                                elif cat:
                                    base = _re.sub(r"\s+FILE$", "", cat).strip()
                                    if base and base in cat_dio_map:
                                        dio = cat_dio_map[base]
                                if dio is not None:
                                    ac.update(
                                        "sku_aggregates",
                                        {"days_inventory_on_hand": round(float(dio), 2)},
                                        filters={"id": f"eq.{a['id']}"},
                                    )
                                    a["days_inventory_on_hand"] = round(float(dio), 2)
                                    dio_updated += 1
                            logger.info(
                                "[rerun] dataset=%s applied category DIO to %d SKU rows",
                                dataset_id, dio_updated,
                            )
            except Exception:  # noqa: BLE001
                # DIO refresh is best-effort — don't fail the reclassify
                # if the source workbook is unreadable. Log and continue.
                logger.exception("[rerun] DIO re-extraction failed; reclassify only")

            # ── Reclassify pass (existing behaviour) ──────────────
            # Bridge column names: classifier expects 'sku' + 'revenue'.
            for a in aggs:
                a["sku"] = a["product_name"]
                a["revenue"] = a["niv_krn"] or 0
                a["cogs"] = max(0.0, (a["niv_krn"] or 0) - (a["gm_krn"] or 0))
            classified = classify_portfolio(aggs)
            for c in classified:
                ac.update(
                    "sku_aggregates",
                    {
                        "classification": c["classification"],
                        "classification_reason": c.get("classification_reason"),
                        "real_margin_krn": c.get("real_margin"),
                        "real_margin_pct": c.get("real_margin_pct"),
                    },
                    filters={"id": f"eq.{c['id']}"},
                )
            return {
                "reclassified": len(classified),
                "dio_refreshed": dio_updated,
            }

    @router.get("/api/sales-datasets/compare")
    def compare_sales_datasets(
        a: str,
        b: str,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """Side-by-side comparison of two datasets. Matches SKUs by product_name;
        returns per-SKU deltas, top movers (winners + losers), and SKUs that
        exist in one dataset but not the other."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            datasets = client.select(
                "sales_datasets",
                filters={"id": f"in.({a},{b})"},
            )
            if len(datasets) != 2:
                raise HTTPException(404, "One or both datasets not found.")
            ds_a = next(d for d in datasets if d["id"] == a)
            ds_b = next(d for d in datasets if d["id"] == b)

            skus_a = client.select("sku_aggregates", filters={"dataset_id": f"eq.{a}"})
            skus_b = client.select("sku_aggregates", filters={"dataset_id": f"eq.{b}"})

        by_name_a = {s["product_name"]: s for s in skus_a}
        by_name_b = {s["product_name"]: s for s in skus_b}
        all_names = sorted(set(by_name_a) | set(by_name_b))

        rows: List[Dict[str, Any]] = []
        for name in all_names:
            sa = by_name_a.get(name)
            sb = by_name_b.get(name)
            niv_a = float(sa["niv_krn"]) if sa and sa.get("niv_krn") else 0.0
            niv_b = float(sb["niv_krn"]) if sb and sb.get("niv_krn") else 0.0
            gm_a = float(sa["gm_krn"]) if sa and sa.get("gm_krn") else 0.0
            gm_b = float(sb["gm_krn"]) if sb and sb.get("gm_krn") else 0.0
            vol_a = float(sa["volume_tons"]) if sa and sa.get("volume_tons") else 0.0
            vol_b = float(sb["volume_tons"]) if sb and sb.get("volume_tons") else 0.0
            rows.append({
                "product_name": name,
                "brand": (sa or sb or {}).get("brand"),
                "category": (sa or sb or {}).get("category"),
                "niv_a": round(niv_a, 2),
                "niv_b": round(niv_b, 2),
                "niv_delta": round(niv_a - niv_b, 2),
                "gm_a": round(gm_a, 2),
                "gm_b": round(gm_b, 2),
                "gm_delta": round(gm_a - gm_b, 2),
                "volume_a": round(vol_a, 2),
                "volume_b": round(vol_b, 2),
                "volume_delta": round(vol_a - vol_b, 2),
                "classification_a": sa.get("classification") if sa else None,
                "classification_b": sb.get("classification") if sb else None,
                "new_in_a": not sb,
                "new_in_b": not sa,
            })

        # Movers: top 10 winners (gm_delta > 0) + top 10 losers (gm_delta < 0)
        ranked = sorted(rows, key=lambda r: r["gm_delta"], reverse=True)
        winners = [r for r in ranked if r["gm_delta"] > 0][:10]
        losers = [r for r in reversed(ranked) if r["gm_delta"] < 0][:10]
        new_in_active = [r for r in rows if r["new_in_a"]]

        return {
            "active": {
                "id": ds_a["id"],
                "label": ds_a["label"],
                "source_filename": ds_a.get("source_filename"),
            },
            "compared": {
                "id": ds_b["id"],
                "label": ds_b["label"],
                "source_filename": ds_b.get("source_filename"),
            },
            "totals": {
                "niv_a": round(sum(r["niv_a"] for r in rows), 2),
                "niv_b": round(sum(r["niv_b"] for r in rows), 2),
                "gm_a": round(sum(r["gm_a"] for r in rows), 2),
                "gm_b": round(sum(r["gm_b"] for r in rows), 2),
                "sku_count_a": len(skus_a),
                "sku_count_b": len(skus_b),
                "new_in_active": len(new_in_active),
            },
            "winners": winners,
            "losers": losers,
            "new_in_active": new_in_active[:20],
            "rows": rows,
        }

    @router.get("/api/sales-datasets")
    def list_sales_datasets(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Right-anchored Datasets panel feed for /products."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            rows = client.select(
                "sales_datasets",
                order="uploaded_at.desc",
                columns="*,documents!inner(original_filename,status,deleted_at)",
            )
        active = [r for r in rows if not (r.get("documents") or {}).get("deleted_at")]
        return {
            "active_dataset_id": active[0]["id"] if active else None,
            "datasets": [{
                "id": r["id"],
                "label": r["label"],
                "source_filename": r.get("source_filename"),
                "row_count": r.get("row_count"),
                "sku_count": r.get("sku_count"),
                "uploaded_at": r["uploaded_at"],
                "is_active": r.get("is_active", True),
                "document_status": (r.get("documents") or {}).get("status"),
                "deleted_at": (r.get("documents") or {}).get("deleted_at"),
            } for r in active],
        }

    @router.get("/api/sales-datasets/{dataset_id}/skus")
    def list_skus_for_dataset(dataset_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """All SKU aggregates for one dataset — the full 406-row portfolio.
        Frontend virtualizes the table, so this endpoint can return the
        whole list (typical files are 400-2000 SKUs, ~150 KB JSON)."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            ds = client.select("sales_datasets", filters={"id": f"eq.{dataset_id}"}, single=True)
            if not ds:
                raise HTTPException(404, "Dataset not found.")
            skus = client.select(
                "sku_aggregates",
                filters={"dataset_id": f"eq.{dataset_id}"},
                order="gm_krn.desc.nullslast",
            )
        # Totals for the KPI strip — computed server-side once instead of
        # the frontend mapping over the full list per render.
        cls_counts: Dict[str, int] = {}
        total_volume = total_niv = total_gm = total_losses = 0.0
        categories: set = set()
        brands: set = set()
        for s in skus:
            cls = s.get("classification") or "keep"
            cls_counts[cls] = cls_counts.get(cls, 0) + 1
            total_volume += s.get("volume_tons") or 0
            total_niv += s.get("niv_krn") or 0
            total_gm += s.get("gm_krn") or 0
            if (s.get("gm_krn") or 0) < 0:
                total_losses += s.get("gm_krn") or 0
            if s.get("category"):
                categories.add(s["category"])
            if s.get("brand"):
                brands.add(s["brand"])
        return {
            "dataset": {
                "id": ds[0]["id"],
                "label": ds[0]["label"],
                "source_filename": ds[0].get("source_filename"),
                "row_count": ds[0].get("row_count"),
                "sku_count": ds[0].get("sku_count"),
                "uploaded_at": ds[0]["uploaded_at"],
            },
            "totals": {
                "sku_count": len(skus),
                "classification_counts": cls_counts,
                "volume_tons": round(total_volume, 2),
                "niv_krn": round(total_niv, 2),
                "gm_krn": round(total_gm, 2),
                "losses_krn": round(total_losses, 2),
                "category_count": len(categories),
                "brand_count": len(brands),
                "categories": sorted(categories),
                "brands": sorted(brands),
            },
            "skus": skus,
        }

    @router.get("/api/sku-analysis/portfolio")
    def get_sku_portfolio(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Consolidated payload for the Products page: totals, classification
        counts, full SKU list (frontend virtualizes), categories, briefing,
        recommendations. Returns the latest active SKU document's analysis.
        """
        from ._sku_classify import portfolio_totals
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            # Latest active sku document
            docs = client.select(
                "documents",
                filters={"scope": "eq.sku", "is_active": "eq.true", "deleted_at": "is.null"},
                order="created_at.desc",
                limit=1,
            )
            if not docs:
                return {"document": None, "totals": None, "skus": [], "analysis": None}
            doc = docs[0]
            skus = client.select(
                "sku_aggregates",
                filters={"document_id": f"eq.{doc['id']}"},
                order="real_margin.desc.nullslast",
            )
            analyses = client.select(
                "sku_analyses",
                filters={"document_id": f"eq.{doc['id']}"},
            )
            analysis = analyses[0] if analyses else None

        totals = portfolio_totals(skus)
        return {
            "document": {
                "id": doc["id"],
                "filename": doc.get("display_name") or doc["original_filename"],
                "status": doc["status"],
                "created_at": doc["created_at"],
                "is_active": doc.get("is_active", True),
            },
            "totals": totals,
            "skus": skus,
            "analysis": analysis and {
                "briefing": analysis.get("briefing"),
                "recommendations": analysis.get("recommendations") or [],
                "summary": analysis.get("summary") or {},
                "model": analysis.get("model"),
            },
        }

    @router.get("/api/sku-analysis/documents")
    def list_sku_documents(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Document history for /products. Excludes soft-deleted docs by default."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            docs = client.select(
                "documents",
                filters={"scope": "eq.sku", "deleted_at": "is.null"},
                order="created_at.desc",
            )
            counts: Dict[str, int] = {}
            for d in docs:
                rows = client.select(
                    "sku_aggregates",
                    filters={"document_id": f"eq.{d['id']}"},
                    columns="id",
                )
                counts[d["id"]] = len(rows)
        return {
            "documents": [
                {
                    "id": d["id"],
                    "filename": d.get("display_name") or d["original_filename"],
                    "original_filename": d["original_filename"],
                    "status": d["status"],
                    "is_active": d.get("is_active", True),
                    "created_at": d["created_at"],
                    "size_bytes": d["size_bytes"],
                    "sku_count": counts.get(d["id"], 0),
                    "error": d.get("error"),
                }
                for d in docs
            ],
        }

    @router.get("/api/sku-analysis/inflight")
    def get_inflight_sku_doc(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Return the most recent in-flight (not analyzed) SKU document, if any.
        Lets the Products page resume showing a progress card after a refresh.

        Watchdog: if the doc has been sitting at status='queued' with
        pipeline_started_at=None — i.e. the upload completed but
        /api/pipeline/run was never delivered (network blip, backend was
        down at upload time) — auto-enqueue it on the next poll. This
        prevents the "stuck at Step 0 of 6 forever" failure the user
        otherwise has no way to recover from without a manual retry.
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            rows = client.select(
                "documents",
                filters={"scope": "eq.sku"},
                columns="id,original_filename,status,error,created_at,pipeline_started_at",
                order="created_at.desc",
                limit=1,
            )
            if not rows:
                return {"document": None}
            d = rows[0]
            if d.get("status") in ("analyzed", "failed"):
                return {"document": None}

            # ── Auto-recover stuck uploads ──────────────────────────────
            # If the doc is queued AND the pipeline never started (no
            # pipeline_started_at), AND the row is older than 5 seconds
            # (race-condition guard against the normal upload→enqueue
            # gap), kick it now.
            try:
                from datetime import datetime, timezone, timedelta
                if d.get("status") == "queued" and not d.get("pipeline_started_at"):
                    created = d.get("created_at")
                    if created:
                        created_dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                        if datetime.now(timezone.utc) - created_dt > timedelta(seconds=5):
                            logger.warning(
                                "[pipeline] watchdog: doc %s stuck at queued with no pipeline_started_at — auto-enqueuing",
                                d["id"],
                            )
                            _admin_set_status(d["id"], "queued", pipeline_started_at=_now_iso())
                            _enqueue(d["id"])
            except Exception:  # noqa: BLE001
                logger.exception("[pipeline] watchdog auto-enqueue failed (non-fatal)")
            return {"document": d}

    @router.post("/api/pipeline/recover-stuck", status_code=200)
    def recover_stuck_pipelines(authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Generic watchdog — find every doc the caller's org has at
        status='queued' with NO pipeline_started_at AND older than 5s, then
        re-enqueue all of them. Same recovery the SKU-only watchdog in
        /api/sku-analysis/inflight does, but covers *every* scope so the
        FinancialStatements page can also escape the "Step 0 of 6 forever"
        trap when /api/pipeline/run silently fails at upload moment (env
        crash, network blip, backend restart between FE upload and FE
        enqueue, etc.).

        Idempotent — calling it twice in a row is safe; the second pass sees
        pipeline_started_at != None and skips. RLS-scoped so a user can only
        recover their own org's docs.

        Returns the list of recovered doc IDs + a count so the FE can show
        a toast like "Recovered 2 stuck uploads".
        """
        jwt = _require_jwt(authorization)
        from datetime import datetime, timezone, timedelta

        recovered: List[Dict[str, Any]] = []
        with _supabase.per_user(jwt) as client:
            rows = client.select(
                "documents",
                filters={"status": "eq.queued"},
                columns="id,original_filename,scope,created_at,pipeline_started_at",
                order="created_at.desc",
                limit=20,
            )
            if rows:
                logger.info("[pipeline] recover-stuck: scanning %d queued doc(s) for caller", len(rows))
            now = datetime.now(timezone.utc)
            cutoff_min = now - timedelta(seconds=5)
            # Hard age cap: don't re-enqueue zombie docs older than 24h —
            # the user has clearly moved on (likely uploaded a fresh copy
            # in the meantime), and silently spawning a new period on next
            # page-load would pollute the dashboard with duplicates. Mark
            # those as failed instead so they leave the inflight tracker.
            cutoff_max = now - timedelta(hours=24)
            stale_failed: List[Dict[str, Any]] = []
            for d in rows:
                if d.get("pipeline_started_at"):
                    continue  # worker thread already kicked
                created = d.get("created_at")
                if not created:
                    continue
                try:
                    created_dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                except (ValueError, AttributeError):
                    continue
                if created_dt > cutoff_min:
                    continue  # too fresh — normal upload→enqueue race window
                if created_dt < cutoff_max:
                    # Zombie — mark failed so it stops showing as inflight.
                    logger.warning(
                        "[pipeline] recover-stuck: doc %s (%s, scope=%s) >24h stale — marking failed",
                        d["id"], d.get("original_filename"), d.get("scope"),
                    )
                    try:
                        _admin_set_status(
                            d["id"], "failed",
                            error="Pipeline never started (upload-time backend hiccup). Please re-upload.",
                        )
                        stale_failed.append({
                            "id": d["id"],
                            "filename": d.get("original_filename"),
                            "scope": d.get("scope"),
                        })
                    except Exception:  # noqa: BLE001
                        logger.exception("[pipeline] failed to mark stale doc as failed")
                    continue
                logger.warning(
                    "[pipeline] recover-stuck: doc %s (%s, scope=%s) stuck — re-enqueuing",
                    d["id"], d.get("original_filename"), d.get("scope"),
                )
                _admin_set_status(d["id"], "queued", pipeline_started_at=_now_iso())
                _enqueue(d["id"])
                recovered.append({
                    "id": d["id"],
                    "filename": d.get("original_filename"),
                    "scope": d.get("scope"),
                })
        return {
            "recovered_count": len(recovered),
            "recovered": recovered,
            "stale_failed_count": len(stale_failed),
            "stale_failed": stale_failed,
        }

    @router.post("/api/pipeline/retry", response_model=RunResponse, status_code=202)
    def retry_pipeline(req: RunRequest, authorization: Optional[str] = Header(None)) -> RunResponse:
        jwt = _require_jwt(authorization)
        doc = _verify_user_owns_document(jwt, req.document_id)
        # Wipe prior derivatives via cascade — deleting the financial_periods
        # row removes statement_line_items, calculated_metrics, briefings,
        # AND alerts (alerts.document_id has on delete set null, we explicitly
        # wipe by document below for that one).
        if doc.get("period_id"):
            with _supabase.admin() as admin_client:
                admin_client.delete("financial_periods", filters={"id": f"eq.{doc['period_id']}"})
        with _supabase.admin() as admin_client:
            admin_client.delete("alerts", filters={"document_id": f"eq.{req.document_id}"})
            admin_client.update(
                "documents",
                {"period_id": None, "error": None, "duration_ms": None},
                filters={"id": f"eq.{req.document_id}"},
            )
        _admin_set_status(req.document_id, "queued", pipeline_started_at=_now_iso())
        _enqueue(req.document_id)
        return RunResponse(document_id=req.document_id, status="queued")

    @router.get("/api/period/{period_id}")
    def get_period(period_id: str, authorization: Optional[str] = Header(None)) -> Dict[str, Any]:
        """Return the consolidated payload for one period: statements (assembled
        from line items), metrics, briefing, recommendations, alerts.

        RLS-scoped: uses the caller's JWT so they only see their own data.
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            periods = client.select("financial_periods", filters={"id": f"eq.{period_id}"}, single=True)
            if not periods:
                raise HTTPException(404, "Period not found.")
            period = periods[0]

            line_items = client.select(
                "statement_line_items",
                filters={"period_id": f"eq.{period_id}"},
            )
            metrics = client.select(
                "calculated_metrics",
                filters={"period_id": f"eq.{period_id}"},
            )
            briefings = client.select("briefings", filters={"period_id": f"eq.{period_id}"})
            briefing = briefings[0] if briefings else None

            doc_rows = client.select(
                "documents",
                filters={"id": f"eq.{period['source_document_id']}"},
                single=True,
            ) if period.get("source_document_id") else []
            doc = doc_rows[0] if doc_rows else None
            org_id = period["org_id"]

            # Phase D-backend: scope alerts + recommendations to this
            # period_id rather than the whole org. Pre-fix this query
            # returned every alert/rec in the org on every period view,
            # which is why a single period showed "(80 on file)" — that
            # was 5 periods' worth of rules + stale rows from old
            # document_ids accumulated in one list.
            #
            # Backward compat: legacy rows have period_id IS NULL (the
            # migration backfills via documents.period_id when possible
            # but can't attribute every historic row). Those legacy
            # NULL rows are intentionally excluded here — the FE never
            # surfaces them. They remain in the table for audit until
            # someone wants to garbage-collect.
            recs = client.select(
                "recommendations",
                filters={"period_id": f"eq.{period_id}"},
                order="urgency.desc,created_at.desc",
            )
            alerts = client.select(
                "alerts",
                filters={
                    "period_id": f"eq.{period_id}",
                    "resolved_at": "is.null",
                },
            )

            org_rows = client.select("organizations", filters={"id": f"eq.{org_id}"}, single=True)
            org = org_rows[0] if org_rows else None

            # Valuation row (one per period). Returned at the top level so the
            # dashboard can render the EBITDA-multiple primary card.
            valuation_rows = client.select(
                "valuations",
                filters={"period_id": f"eq.{period_id}"},
                single=True,
            )
            valuation = valuation_rows[0] if valuation_rows else None

            # Per-user assumption overrides — only returned if the user has
            # saved any. Dashboard merges these on top of engine defaults.
            try:
                ua_rows = client.select(
                    "user_valuation_assumptions",
                    filters={"period_id": f"eq.{period_id}"},
                    single=True,
                )
                user_assumptions = ua_rows[0] if ua_rows else None
            except Exception:  # noqa: BLE001
                user_assumptions = None

        # Re-assemble Statements from line_items so the frontend doesn't need
        # to know how to rebuild them.
        bs_buckets = {
            "cash": "cash", "ar": "accountsReceivable", "inventory": "inventory",
            "otherCurrentAssets": "otherCurrentAssets",
            "ppe": "propertyPlantEquipment", "intangibles": "intangibles",
            "otherNonCurrentAssets": "otherNonCurrentAssets",
            "ap": "accountsPayable", "stDebt": "shortTermDebt", "otherCurrentLiab": "otherCurrentLiabilities",
            "ltDebt": "longTermDebt", "otherNonCurrentLiab": "otherNonCurrentLiabilities",
            "shareCapital": "shareCapital", "retainedEarnings": "retainedEarnings", "otherEquity": "otherEquity",
        }
        pl_buckets = {
            "revenue": "revenue", "cogs": "costOfGoodsSold", "operatingExpenses": "operatingExpenses",
            "depreciation": "depreciationAmortization", "interestExpense": "interestExpense",
            "otherIncome": "otherIncome", "financialIncome": "financialIncome",
            "financialExpense": "financialExpense", "taxExpense": "taxExpense",
        }
        bs: Dict[str, float] = {v: 0.0 for v in bs_buckets.values()}
        pl: Dict[str, float] = {v: 0.0 for v in pl_buckets.values()}
        # SPLIT-OUT bucket for RAS account 711 (Variația stocurilor —
        # inventory variation). `_ro_coa._persistence_bucket()` writes
        # 711 to `statement_line_items.bucket = 'otherIncome'` so the
        # existing DB CHECK constraint accepts it, but cash-view EBITDA
        # MUST exclude 711 (it's a non-cash production-variation
        # accrual). Without this split, the FE's `deriveTotals`
        # computes `ebitda = grossProfit - opex + otherIncome` and
        # picks up the 711 inflation — for Scandia FY2025 that's
        # +630M, turning a 13% EBITDA margin into a nonsensical 165%.
        # Identify 711 lines by their ro_account_code prefix and route
        # them into a separate `inventoryVariationMemo` field that the
        # FE knows to exclude from cash EBITDA.
        inv_var_memo = 0.0
        for item in line_items:
            bucket = item["bucket"]
            amount = float(item["amount"] or 0)
            code = (item.get("ro_account_code") or "").strip()
            if bucket in bs_buckets:
                bs[bs_buckets[bucket]] += amount
            elif bucket in pl_buckets:
                # Carve out 711 from otherIncome before the bucket sum.
                if bucket == "otherIncome" and code.startswith("711"):
                    inv_var_memo += amount
                else:
                    pl[pl_buckets[bucket]] += amount

        statements = {
            "companyName": (org or {}).get("name") if org else None,
            "industry": (org or {}).get("industry_display_name") if org else None,
            "currency": period.get("currency", "RON"),
            "periodLabel": period.get("period_end"),
            "balanceSheet": {k: round(v, 2) for k, v in bs.items()},
            "incomeStatement": {
                **{k: round(v, 2) for k, v in pl.items()},
                # Surfaced separately so the FE can render the
                # production-variation footnote AND so cash EBITDA
                # never re-inflates from it.
                "inventoryVariationMemo": round(inv_var_memo, 2),
            },
            # Required by the TS Statements interface so computeRatios() can
            # read supplementary.periodDays. Real enrichment values arrive
            # later via Settings.
            "supplementary": {"periodDays": 365},
        }

        # ── Re-assemble the canonical views from line items ──────────────
        # The DB row only carries the legacy bucket-level aggregates, but
        # the Valuation tab + briefing now consume `assembled_pl`,
        # `assembled_bs`, `assembled_cf` (operating-view EBITDA, real
        # CapEx, etc.). Reconstruct them deterministically from the
        # per-account `line_items` so the page-load response carries the
        # same canonical facts the pipeline computed at write time.
        try:
            # F3.1c: dispatch via the country pack rather than direct
            # `_ro_coa` import. `_coa_mod` alias preserved so the call
            # sites below read identically.
            _coa_mod = _ro_pack()
            recovered_accounts = []
            for li in (line_items or []):
                code = li.get("ro_account_code") or ""
                if not code:
                    continue
                stored_bucket = li.get("bucket") or li.get("canonical_bucket") or ""
                rule = _coa_mod.bucket_for(code)
                # Detect side-flip / inventory-contra: if the persisted bucket
                # diverges from what bucket_for(code) would now return, the
                # write-path must have routed this row via `bucket_override`
                # (typical for 418 customer-accrual C-side, 451 affiliated
                # C-side, etc.). Preserve that override so the re-assembled
                # canonical view round-trips correctly. Without this, every
                # /api/period reassembly silently moves the side-flipped
                # amounts BACK to the natural-side bucket — a BS asset bloom
                # plus a matching liability hole of identical magnitude.
                bucket_override = None
                if rule and stored_bucket and stored_bucket != rule.bucket:
                    # Also tolerate the legacy-bucket bridge (e.g. canonical
                    # `ar_intercompany` persists as `otherCurrentAssets`).
                    legacy = _coa_mod.persistence_bucket(rule.bucket)
                    if stored_bucket != legacy:
                        bucket_override = stored_bucket
                row = {
                    "code": code,
                    "name": li.get("ro_account_name") or "",
                    # `line_items.amount` is already signed (mapping rule
                    # applied at write time). assemble_statements expects
                    # raw amounts and re-applies the sign via the mapping
                    # rule. To avoid double-signing, divide back out.
                    "amount": float(li.get("amount") or 0),
                }
                if bucket_override:
                    row["bucket_override"] = bucket_override
                recovered_accounts.append(row)
            # Re-apply sign so assemble_statements sees the raw input.
            for acct in recovered_accounts:
                rule = _coa_mod.bucket_for(acct["code"])
                if rule and rule.sign == -1:
                    acct["amount"] = -acct["amount"]
            assembled_full = _coa_mod.assemble_statements(
                recovered_accounts,
                company_name=statements["companyName"] or "Entity",
                currency=statements["currency"],
                period_label=str(statements["periodLabel"]) if statements["periodLabel"] else "Period",
                industry=(org or {}).get("industry_key") if org else None,
            )
            # Surface canonical views on the response statements.
            statements["assembled_bs"] = assembled_full["statements"].get("assembled_bs")
            statements["assembled_pl"] = assembled_full["statements"].get("assembled_pl")
            statements["assembled_cf"] = assembled_full["statements"].get("assembled_cf")
            # F1.f / F1.g — bands + Piotroski live on the re-assembled
            # canonical view; carry them through to the response.
            statements["assembled_bands"] = assembled_full["statements"].get("assembled_bands")
            statements["assembled_piotroski"] = assembled_full["statements"].get("assembled_piotroski")
            statements["subAggregates"] = assembled_full["statements"].get("subAggregates")
            # F4.1e — surface the country-agnostic canonical envelope on the
            # response statements. Lives at top level of `assembled_full`
            # (sibling of `statements`), NOT under statements.assembled_*.
            # Source: chart_of_accounts.assemble_statements line ~1817.
            statements["assembled_canonical_v1"] = assembled_full.get("assembled_canonical_v1")

            # F3.27-DRIFT-TRANSFORMATION-GLUE — Fix A1.
            # The re-assembled `statements["assembled_bs"]["bs_balance_delta"]`
            # above comes from `_coa_mod.assemble_statements(recovered_accounts)`.
            # Because `recovered_accounts` is reconstructed from persisted
            # `statement_line_items` — which omits _IGNORE_BUCKETS accounts
            # (121, 581) and loses semantic-fallback routing context — the
            # round-trip produces a different bs_balance_delta than the engine
            # emitted at write time. `scripts/measure_bs_drift_roundtrip.py`
            # proves this round-trip discrepancy ranges from 0.44% (Frozen) to
            # 35.47% (RealEstate) across the 8 prod fixtures, and confirms
            # that subtracting `methodology.totals` from the canonical
            # envelope reproduces the engine's authoritative bs_balance_delta
            # to the cent on every fixture. Override the round-tripped value
            # with the envelope-true value so the FE consumes the engine's
            # emission rather than a re-assembly artifact. All other
            # re-assembled assembled_bs.* fields (totals, sub-aggregates,
            # breakouts) are preserved unchanged.
            # CRITICAL: read from the PERSISTED envelope (`period[...]`,
            # the DB row written at upload time), NOT `assembled_full[...]`
            # (the re-assembled output of the same lossy round-trip we are
            # trying to override). `assembled_full["assembled_canonical_v1"]`
            # carries RE-COMPUTED totals from `recovered_accounts`, which by
            # definition reproduce the same round-trip delta we want to
            # replace. The PERSISTED envelope at `period["assembled_canonical_v1"]`
            # carries the engine's WRITE-TIME totals — those are what the
            # F-A3.1 canary measures, and the only correct source for the
            # override. (Browser-verified: reading from assembled_full
            # gave 3.49% drift identical to the bug; reading from period
            # gives the engine truth of 0.0125%.)
            try:
                _env = period.get("assembled_canonical_v1") or {}
                _methodology = _env.get("methodology") or {}
                _totals = _methodology.get("totals") or {}
                if all(k in _totals for k in ("total_assets", "total_liabilities", "total_equity")):
                    _ta = float(_totals["total_assets"])
                    _tl = float(_totals["total_liabilities"])
                    _te = float(_totals["total_equity"])
                    _a_bs = statements.get("assembled_bs") or {}
                    _a_bs["bs_balance_delta"] = round(_ta - (_tl + _te), 2)
                    statements["assembled_bs"] = _a_bs
            except Exception:  # noqa: BLE001
                logger.exception(
                    "[/api/period] Fix A1 bs_balance_delta override failed (non-fatal)"
                )

            # F4.3 — surface the detection envelope (country/standard/doc_type/
            # industry + methodology pin per CANONICAL_SCHEMA_V1.md §7).
            try:
                from engine.detection import build_detection_envelope as _bde  # type: ignore
                statements["detection_envelope"] = _bde(
                    classification=None, assembled=assembled_full,
                    methodology_id="ro_ras_2025_v1",
                    industry_key=(org or {}).get("industry_key") if org else None,
                    currency=statements.get("currency"),
                    period_end=str(statements.get("periodLabel") or ""),
                )
            except Exception:  # noqa: BLE001
                logger.exception("[/api/period] detection envelope build failed (non-fatal)")
        except Exception:  # noqa: BLE001
            logger.exception("[/api/period] canonical re-assembly failed (non-fatal)")

        # ── F3.3 — Per-upload confidence report ─────────────────────
        # Surface the country-detection / layout / reconciliation /
        # Review-Mode-trigger envelope on every analysis page. Uses
        # the assembled output (already produced above) plus a
        # synthesised content blob (line-items code+name concat) for
        # detection — the original document bytes aren't always
        # available at read time, but the line-items carry enough RAS
        # signal for the pack's detect_from_content() to score
        # confidently.
        confidence_dict: Optional[Dict[str, Any]] = None
        try:
            from engine.core.upload_classifier import classify_upload
            from engine.core.confidence_engine import (
                build_confidence_report,
                confidence_report_to_dict,
            )
            # Synthesise a content blob from line items + period
            # metadata so the pack's existing text-pattern detector
            # has something to scan. This is sufficient for high-
            # confidence Romanian detection (account codes alone
            # carry 0.30 weight, language labels 0.25). Real
            # uploads at write time will use the original document
            # bytes (future enhancement once we plumb them through).
            synth_lines = []
            for li in (line_items or [])[:200]:
                code = li.get("ro_account_code") or ""
                name = li.get("ro_account_name") or ""
                synth_lines.append(f"{code} {name}")
            synth_lines.append(f"Moneda: {period.get('currency', 'RON')}")
            synth_lines.append("Sume totale credit  Sold final debitor")  # SAGA header marker
            synth = "\n".join(synth_lines).encode("utf-8", errors="ignore")
            fn_hint = (doc or {}).get("original_filename") or "balanta.xlsx"
            classification = classify_upload(synth, fn_hint)
            report = build_confidence_report(classification, assembled_full)
            confidence_dict = confidence_report_to_dict(report)
            # F4.4-int — run fan-out routing on top of the classification
            # so the detection envelope's routing_decision field gets
            # populated. With one pack registered, fan-out collapses to
            # fast_path; with multiple, the highest-coverage pack is
            # auto-picked. Either way we rebuild the detection envelope
            # with full classification + routing_decision attached
            # (the earlier build at line ~5078 used classification=None).
            try:
                from engine.routing import route_with_fan_out as _route, routing_decision_dict as _rdd
                from engine.detection import build_detection_envelope as _bde2
                _routing_result = _route(
                    classification, synth, fn_hint,
                    company_name=statements.get("companyName") or "Entity",
                    currency=statements.get("currency") or "RON",
                    period_label=str(statements.get("periodLabel") or ""),
                    industry=(org or {}).get("industry_key") if org else None,
                )
                statements["detection_envelope"] = _bde2(
                    classification=classification,
                    assembled=assembled_full,
                    methodology_id="ro_ras_2025_v1",
                    industry_key=(org or {}).get("industry_key") if org else None,
                    currency=statements.get("currency"),
                    period_end=str(statements.get("periodLabel") or ""),
                    routing_decision=_rdd(_routing_result),
                )
            except Exception:  # noqa: BLE001
                logger.exception("[/api/period] fan-out routing integration failed (non-fatal)")
        except Exception:  # noqa: BLE001
            logger.exception("[/api/period] confidence report build failed (non-fatal)")

        # ── F1.i — typed `assembled_metrics` envelope. Single bundled
        # object the FE imports + reads; eliminates string lookups
        # against `metrics[]` and removes the case-by-case fallback
        # chains the canonical-conformance audit catalogued. Composed
        # here at read time from data already on the response (no new
        # math, no new persistence).
        _m_by_name = {m["name"]: m for m in (metrics or [])}
        def _m(name: str) -> Optional[float]:
            row = _m_by_name.get(name)
            return None if row is None else row.get("value")
        assembled_metrics_envelope = {
            "pl": statements.get("assembled_pl") or {},
            "bs": statements.get("assembled_bs") or {},
            "cf": statements.get("assembled_cf") or {},
            "ratios": {
                "liquidity": {
                    "current_ratio":         _m("current_ratio"),
                    "quick_ratio":           _m("quick_ratio"),
                    "cash_ratio":            _m("cash_ratio"),
                    "working_capital":       _m("working_capital"),
                    "net_debt":              _m("net_debt"),
                    "net_debt_to_ebitda":    _m("net_debt_to_ebitda"),
                },
                "leverage": {
                    "equity_ratio":          _m("equity_ratio"),
                    "debt_to_assets":        _m("debt_to_assets"),
                    "debt_to_equity":        _m("debt_to_equity"),
                    "debt_to_ebitda":        _m("debt_to_ebitda"),
                    "lt_debt_to_equity":     _m("lt_debt_to_equity"),
                },
                "coverage": {
                    "interest_coverage":     _m("interest_coverage"),
                    "ebitda_to_interest":    _m("ebitda_to_interest"),
                    "dscr":                  _m("dscr"),
                    "dscr_with_lt_principal":_m("dscr_with_lt_principal"),
                },
                "profitability": {
                    "gross_margin":          _m("gross_margin"),
                    "operating_margin":      _m("operating_margin"),
                    "ebitda_margin":         _m("ebitda_margin"),
                    "core_ebitda_margin":    _m("core_ebitda_margin"),
                    "net_margin":            _m("net_margin"),
                    "roa":                   _m("roa"),
                    "roe":                   _m("roe"),
                    "roic":                  _m("roic"),
                },
                "efficiency": {
                    "asset_turnover":        _m("asset_turnover"),
                    "dso":                   _m("dso"),
                    "dio":                   _m("dio"),
                    "dpo":                   _m("dpo"),
                    "ccc":                   _m("ccc"),
                    "inventory_turnover":    _m("inventory_turnover"),
                },
            },
            "bands": statements.get("assembled_bands"),
            "credit": {
                "altman_z_score":            _m("altman_z_score"),
                "altman_variant":            "Z\"",
                "altman_components": {
                    "x1": _m("altman_x1"),
                    "x2": _m("altman_x2"),
                    "x3": _m("altman_x3"),
                    "x4": _m("altman_x4"),
                },
                "composite_score":           _m("credit_composite"),
                # F1.h — letter grade emitted as a canonical field on the
                # envelope so the FE can be a pure reader (no FE-side
                # mapping). Uses the same _composite_to_letter_grade helper
                # as stage_compute, so engine logging and envelope agree.
                "letter_grade": (
                    None if _m("credit_composite") is None
                    else _composite_to_letter_grade(float(_m("credit_composite")))
                ),
                "letter_grade_bands": [
                    {"min": 90, "grade": "AAA"},
                    {"min": 80, "grade": "AA"},
                    {"min": 70, "grade": "A"},
                    {"min": 60, "grade": "BBB"},
                    {"min": 50, "grade": "BB"},
                    {"min": 40, "grade": "B"},
                    {"min": 25, "grade": "CCC"},
                    {"min": 0,  "grade": "CC"},
                ],
                "composite_weights": {
                    "altman":        0.30,
                    "profitability": 0.20,
                    "leverage":      0.15,
                    "coverage":      0.10,
                    "dscr":          0.10,
                    "liquidity":     0.10,
                    "equity":        0.05,
                },
                "subscores": {
                    "altman":        _m("credit_subscore_altman"),
                    "profitability": _m("credit_subscore_profitability"),
                    "leverage":      _m("credit_subscore_leverage"),
                    "coverage":      _m("credit_subscore_coverage"),
                    "dscr":          _m("credit_subscore_dscr"),
                    "liquidity":     _m("credit_subscore_liquidity"),
                    "equity":        _m("credit_subscore_equity"),
                },
            },
            "piotroski": statements.get("assembled_piotroski"),
            # `valuation` is deferred to F1.j (new override endpoint) —
            # the existing `valuation` key on the response carries the
            # current data the FE can read meanwhile.
        }

        return {
            # F1.k — canonical_version stamp. v2.0 = the F1 contract
            # extensions (assembled_metrics envelope, ratio expansion,
            # bands, piotroski, F1.a/b/c canonical extras). v2.1 = F1.e
            # (FE source switch on ebitda_margin / net_margin). FE readers
            # can branch on this if v1 / v2 differ; the cache integrity
            # check in _benchmarks.py uses an analogous gate.
            "canonical_version": "v2.1",
            "assembled_metrics": assembled_metrics_envelope,
            # F3.3 — per-upload country-detection + confidence engine.
            # Drives the Confidence Indicator badge on every FE
            # analysis page and the Review Mode trigger (F3.4).
            "confidence": confidence_dict,
            "period": {
                "id": period["id"],
                "period_end": period["period_end"],
                "currency": period["currency"],
                "extraction_confidence": period.get("extraction_confidence"),
                "source_document": doc and {
                    "id": doc["id"],
                    "filename": doc["original_filename"],
                    "status": doc["status"],
                    # NEW: surfaces "statutory_f30_f10" vs "trial_balance"
                    # so the UI can render the appropriate accuracy /
                    # limitation banner. Existing TB-routed periods still
                    # carry "trial_balance" here — no banner needed.
                    "detected_type": doc.get("detected_type"),
                },
            },
            "organization": org and {
                "id": org["id"],
                "name": org["name"],
                "industry_key": org.get("industry_key"),
                "industry_display_name": org.get("industry_display_name"),
            },
            "statements": statements,
            # Per-account line items — drives the reference-format P&L
            # renderer (account codes + per-line drill-down). Each entry
            # carries the RO account code, name, bucket, statement (BS/PL),
            # and amount (positive after sign-correction by the mapping
            # rule). The frontend's buildPLStatement() consumes this list.
            "line_items": [
                {
                    "statement": li.get("statement"),
                    "bucket": li.get("bucket"),
                    "ro_account_code": li.get("ro_account_code"),
                    "ro_account_name": li.get("ro_account_name"),
                    "amount": float(li.get("amount") or 0),
                    "is_derived": bool(li.get("is_derived")),
                }
                for li in (line_items or [])
            ],
            "metrics": [
                {
                    "name": m["name"],
                    "value": m["value"],
                    "unit": m["unit"],
                    "direction": m.get("direction"),
                }
                for m in metrics
            ],
            "briefing": briefing and {
                "body": briefing["body"],
                "language": briefing.get("language", "en"),
                "model": briefing.get("model"),
            },
            "recommendations": [
                {
                    "id": r["id"],
                    "title": r["title"],
                    "explanation": r.get("explanation"),
                    "urgency": r.get("urgency"),
                    "expected_cash_impact_kron": r.get("expected_cash_impact_kron"),
                    "status": r.get("status"),
                }
                for r in recs
            ],
            "alerts": [
                {
                    "id": a["id"],
                    "alert_key": a["alert_key"],
                    # `rule_key` / `facts_cited` / `industry` are carried on
                    # the `payload` JSONB column at persist time. Surface
                    # them at the top level for the FE's facts-backing
                    # expander.
                    "rule_key": (a.get("payload") or {}).get("rule_key")
                                 if isinstance(a.get("payload"), dict) else None,
                    "facts_cited": (a.get("payload") or {}).get("facts_cited")
                                 if isinstance(a.get("payload"), dict) else None,
                    "industry": (a.get("payload") or {}).get("industry")
                                 if isinstance(a.get("payload"), dict) else None,
                    "severity": a["severity"],
                    "category": a["category"],
                    "title": a["title"],
                    "body": a.get("body"),
                }
                for a in alerts
            ],
            "valuation": _serialize_valuation(valuation, user_assumptions, statements),
            # F4.6 — list of legacy fields slated for removal at the 2Q
            # deprecation horizon (~Nov 2026 per F3.15 §3e). Consumers
            # should switch to the canonical replacements before sunset.
            "deprecated_fields": _deprecated_fields_for_response(),
        }

    @router.put("/api/period/{period_id}/valuation-assumptions")
    def save_valuation_assumptions(
        period_id: str,
        body: Dict[str, Any],
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """Save per-user overrides on the valuation inputs (EBITDA / multiple /
        debt / cash). RLS keys on auth.uid() == user_id so users can only
        manage their own row. POST/PUT both upsert by (user_id, period_id).
        """
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            # Resolve the caller's auth.uid() via the auth endpoint so we can
            # populate user_id without trusting the body.
            user = client.get_user(jwt)
            user_id = user["id"]
            # Make sure the user can read the period (RLS check).
            periods = client.select(
                "financial_periods",
                filters={"id": f"eq.{period_id}"},
                single=True,
            )
            if not periods:
                raise HTTPException(404, "Period not found.")

            payload = {
                "user_id": user_id,
                "period_id": period_id,
                "ebitda_used": body.get("ebitda_used"),
                "multiple_used": body.get("multiple_used"),
                "debt_used": body.get("debt_used"),
                "cash_used": body.get("cash_used"),
                "notes": body.get("notes"),
                "updated_at": _now_iso(),
            }
            client.upsert(
                "user_valuation_assumptions",
                payload,
                on_conflict="user_id,period_id",
            )
        # Recompute and re-persist the valuation row reflecting the user's
        # overrides, so the dashboard can re-read /api/period/:id and see
        # the new numbers without any client-side math.
        try:
            with _supabase.admin() as admin_client:
                periods_admin = admin_client.select(
                    "financial_periods",
                    filters={"id": f"eq.{period_id}"},
                    single=True,
                )
                period = periods_admin[0] if periods_admin else None
                if period:
                    org_rows = admin_client.select(
                        "organizations",
                        filters={"id": f"eq.{period['org_id']}"},
                        single=True,
                    )
                    org = org_rows[0] if org_rows else {}
                    line_items = admin_client.select(
                        "statement_line_items",
                        filters={"period_id": f"eq.{period_id}"},
                    )
                    assembled = _rebuild_assembled(line_items)
                    result = _valuation.compute_valuation(
                        industry_key=org.get("industry_key"),
                        statements=assembled,
                        user_assumptions={
                            "ebitda_used": payload["ebitda_used"],
                            "multiple_used": payload["multiple_used"],
                            "debt_used": payload["debt_used"],
                            "cash_used": payload["cash_used"],
                        },
                    )
                    _valuation.persist_valuation(period_id, period["org_id"], result)
        except Exception:  # noqa: BLE001
            logger.exception("[pipeline] valuation recompute failed (non-fatal)")
        return {"ok": True}

    @router.delete("/api/period/{period_id}/valuation-assumptions")
    def reset_valuation_assumptions(
        period_id: str,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """Reset to engine defaults: drop the user_valuation_assumptions row
        and re-compute the valuations row from raw statements."""
        jwt = _require_jwt(authorization)
        with _supabase.per_user(jwt) as client:
            user = client.get_user(jwt)
            user_id = user["id"]
            client.delete(
                "user_valuation_assumptions",
                filters={
                    "user_id": f"eq.{user_id}",
                    "period_id": f"eq.{period_id}",
                },
            )
        try:
            with _supabase.admin() as admin_client:
                periods_admin = admin_client.select(
                    "financial_periods",
                    filters={"id": f"eq.{period_id}"},
                    single=True,
                )
                period = periods_admin[0] if periods_admin else None
                if period:
                    org_rows = admin_client.select(
                        "organizations",
                        filters={"id": f"eq.{period['org_id']}"},
                        single=True,
                    )
                    org = org_rows[0] if org_rows else {}
                    line_items = admin_client.select(
                        "statement_line_items",
                        filters={"period_id": f"eq.{period_id}"},
                    )
                    assembled = _rebuild_assembled(line_items)
                    result = _valuation.compute_valuation(
                        industry_key=org.get("industry_key"),
                        statements=assembled,
                    )
                    _valuation.persist_valuation(period_id, period["org_id"], result)
        except Exception:  # noqa: BLE001
            logger.exception("[pipeline] valuation reset recompute failed (non-fatal)")
        return {"ok": True}

    @router.post("/api/period/{period_id}/valuation/recompute")
    def recompute_valuation(
        period_id: str,
        body: Dict[str, Any],
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F1.j — Stateless interactive DCF recompute (SPEC §11).

        Accepts the spec'd override body:
          {
            forecast_years?:        number,   // default 5
            forecast_growth?:       number,   // default 0.035
            terminal_growth?:       number,   // default 0.030
            beta?:                  number,   // default 1.0
            equity_risk_premium?:   number,   // default 0.075
            rf?:                    number,   // default 0.0675 (extension)
            property_market_value?: number,   // schema-only, round-tripped
            annual_lease_expense?:  number,   // schema-only, round-tripped
            shares_outstanding?:    number,   // schema-only, round-tripped
          }

        Auth: same JWT + ownership check as GET /api/period/:id. RLS on
        financial_periods scopes the row to the caller's org.

        Response shape: the SAME `assembled_metrics.valuation` envelope
        that GET /api/period/:id emits — so the FE can swap the rendered
        result for the recomputed one with no shape translation.

        Stateless: this endpoint does NOT persist. The
        `user_valuation_assumptions` table is only written by
        PUT /api/period/{id}/valuation-assumptions (the dedicated
        save-overrides endpoint). The recompute endpoint is the
        what-if calculator.
        """
        jwt = _require_jwt(authorization)
        # Whitelist + type-coerce the body so the FE can't sneak unexpected
        # keys into compute_valuation. Anything else is ignored silently.
        ALLOWED_OVERRIDES = {
            "forecast_years",
            "forecast_growth",
            "terminal_growth",
            "beta",
            "equity_risk_premium",
            "rf",
            "property_market_value",
            "annual_lease_expense",
            "shares_outstanding",
        }
        overrides: Dict[str, Any] = {}
        for k in ALLOWED_OVERRIDES:
            v = (body or {}).get(k)
            if v is None:
                continue
            try:
                # forecast_years is an int; everything else is a float.
                overrides[k] = int(v) if k == "forecast_years" else float(v)
            except (TypeError, ValueError):
                raise HTTPException(
                    400, f"valuation/recompute: '{k}' must be numeric (got {v!r})."
                )

        with _supabase.per_user(jwt) as client:
            periods = client.select(
                "financial_periods",
                filters={"id": f"eq.{period_id}"},
                single=True,
            )
            if not periods:
                raise HTTPException(404, "Period not found.")
            period = periods[0]

        with _supabase.admin() as admin_client:
            org_rows = admin_client.select(
                "organizations",
                filters={"id": f"eq.{period['org_id']}"},
                single=True,
            )
            org = org_rows[0] if org_rows else {}
            line_items = admin_client.select(
                "statement_line_items",
                filters={"period_id": f"eq.{period_id}"},
            )
            assembled = _rebuild_assembled(line_items)

            # Layer the user's saved EBITDA/multiple/debt/cash overrides
            # underneath the recompute's DCF overrides — same precedence
            # the GET endpoint uses, so toggling DCF inputs doesn't
            # silently revert the user's persisted EBITDA choice.
            ua_rows = admin_client.select(
                "user_valuation_assumptions",
                filters={"period_id": f"eq.{period_id}"},
            )
            user_assumptions = None
            if ua_rows:
                ua = ua_rows[0]
                user_assumptions = {
                    "ebitda_used":   ua.get("ebitda_used"),
                    "multiple_used": ua.get("multiple_used"),
                    "debt_used":     ua.get("debt_used"),
                    "cash_used":     ua.get("cash_used"),
                }

            try:
                result = _valuation.compute_valuation(
                    industry_key=org.get("industry_key"),
                    statements=assembled,
                    user_assumptions=user_assumptions,
                    dcf_overrides=overrides,
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception(
                    "[/api/period/{id}/valuation/recompute] compute failed"
                )
                raise HTTPException(
                    500, f"Valuation recompute failed: {type(exc).__name__}"
                ) from exc

        return {
            "valuation": result,
            "overrides_applied": result.get("overrides_applied"),
        }

    @router.post("/api/period/{period_id}/briefing/regenerate")
    def regenerate_briefing(
        period_id: str,
        currency: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F2.8 — Regenerate the LLM CFO Briefing for an existing period
        against the current canonical statements + metrics. Used after the
        F1.e basis decisions (cash vs statutory EBITDA, statutory ct.121
        net profit, Z″ Altman) shifted the canonical numbers — cached
        briefings still cite the pre-F1.e values until regenerated.

        `currency` (optional query param): RON/EUR/USD. When provided,
        every monetary number in `briefing_facts` is FX-converted to that
        currency before the LLM sees them, and the prompt asks the LLM
        to cite the converted currency. Defaults to the period's source
        currency (typically RON). Wired from the FE TopHeader currency
        toggle so the briefing prose follows RON ↔ EUR ↔ USD switches.

        Calls `stage_narrate` against the rebuilt assembled statements +
        the period's `calculated_metrics` rows, then upserts the resulting
        briefing body on the `briefings` table. Recommendations + alerts
        are NOT touched (those have their own deterministic generation
        path via `stage_validate`).
        """
        jwt = _require_jwt(authorization)
        # Ownership check via RLS.
        with _supabase.per_user(jwt) as client:
            periods = client.select(
                "financial_periods",
                filters={"id": f"eq.{period_id}"},
                single=True,
            )
            if not periods:
                raise HTTPException(404, "Period not found.")
            period = periods[0]

        with _supabase.admin() as admin_client:
            org_rows = admin_client.select(
                "organizations",
                filters={"id": f"eq.{period['org_id']}"},
                single=True,
            )
            org = org_rows[0] if org_rows else {}
            line_items = admin_client.select(
                "statement_line_items",
                filters={"period_id": f"eq.{period_id}"},
            )
            metric_rows = admin_client.select(
                "calculated_metrics",
                filters={"period_id": f"eq.{period_id}"},
            )
            # calculated_metrics columns are name/value/unit/direction
            # (NOT metric_name/metric_value — the earlier shape was an F2.8
            # transcription error against the live schema, caught by the
            # post-deploy KeyError on the first regenerate call).
            metrics = [
                {
                    "name": m["name"],
                    "value": m["value"],
                    "unit": m.get("unit"),
                    "direction": m.get("direction"),
                }
                for m in metric_rows
            ]
            valuation_rows = admin_client.select(
                "valuations",
                filters={"period_id": f"eq.{period_id}"},
            )
            valuation = valuation_rows[0] if valuation_rows else None
            doc_rows = admin_client.select(
                "documents",
                filters={"period_id": f"eq.{period_id}"},
            )
            doc = doc_rows[0] if doc_rows else {
                "id": "regenerate",
                "org_id": period["org_id"],
                "language": "en",
            }
            # F2.8 fix: stage_narrate reads from `assembled["statements"]
            # .assembled_pl` etc., not the bucket-only shape that
            # `_rebuild_assembled` returns. Use the canonical-shaped
            # rebuilder so the regenerated briefing cites
            # operating-view EBITDA, statutory net income, and the
            # rest of the briefing_facts envelope.
            assembled = _rebuild_assembled_for_briefing(line_items, period, org)

            # FX rates for currency conversion. Skip the fetch when the
            # caller wants the period's native currency (the no-op case).
            display_currency = (currency or "").upper() or None
            fx_payload: Optional[Dict[str, Any]] = None
            if display_currency and display_currency not in ("", "RON"):
                try:
                    from .fx_rates import get_fx_rates as _get_fx_rates
                    fx_payload = _get_fx_rates()
                except Exception:  # noqa: BLE001
                    logger.warning(
                        "[/api/period/{id}/briefing/regenerate] fx_rates fetch failed; "
                        "falling back to source currency"
                    )
            fx_rates_map = (fx_payload or {}).get("rates") if fx_payload else None

            try:
                narrative = stage_narrate(
                    doc, assembled, metrics, org, period_id,
                    parsed=None, valuation=valuation,
                    display_currency=display_currency,
                    fx_rates=fx_rates_map,
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception(
                    "[/api/period/{id}/briefing/regenerate] stage_narrate failed"
                )
                raise HTTPException(
                    500, f"Briefing regeneration failed: {type(exc).__name__}"
                ) from exc

            # Upsert briefing only; do NOT touch recommendations/alerts
            # (deterministic rules own those — regen would create
            # duplicates and disagree with the validation pipeline).
            #
            # Currency-conversion regenerations (display_currency != source)
            # are NOT persisted — the DB row stays the canonical
            # source-currency briefing. If a user toggles RON → EUR, the
            # EUR briefing is returned in the response only; on next
            # session load they'd see RON again until they toggle. This
            # avoids the alternative-currency briefing becoming "sticky"
            # and confusing users who later toggle back.
            should_persist = not display_currency or display_currency.upper() == "RON"
            if should_persist:
                admin_client.upsert(
                    "briefings",
                    {
                        "period_id": period_id,
                        "org_id": period["org_id"],
                        "body": narrative.get("briefing", ""),
                        "language": "en",
                        "model": "claude-opus-4-7",
                    },
                    on_conflict="period_id",
                    returning=False,
                )
                admin_client.update(
                    "financial_periods",
                    {"updated_at": _now_iso()},
                    filters={"id": f"eq.{period_id}"},
                )

        return {
            "ok": True,
            "period_id": period_id,
            "briefing_length": len(narrative.get("briefing", "")),
            "briefing": narrative.get("briefing", ""),
            "currency": display_currency or "RON",
        }

    # ── F3.4 — Review Mode ─────────────────────────────────────────
    # `ReviewReanalyzeRequest` lives at module level (see above);
    # nested-class Pydantic bodies don't resolve correctly through
    # FastAPI's type inspector.

    @router.get("/api/period/{period_id}/review")
    def get_review_state(
        period_id: str,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F3.4 — return the Review Mode state for a period.

        Surfaces:
          - `unmapped_accounts` — accounts the pack's chart-of-
            accounts couldn't bucket. These are what the user
            assigns manually in Review Mode.
          - `org_overrides` — any persisted org-wide overrides
            (`org_coa_mappings_overrides`) that already apply to
            this org's uploads.

        Romanian fixtures should return `unmapped_accounts: []`.
        """
        jwt = _require_jwt(authorization)
        from engine.core import review_mode as _rm
        from engine.core.country_pack_registry import get_pack

        with _supabase.per_user(jwt) as client:
            periods = client.select(
                "financial_periods",
                filters={"id": f"eq.{period_id}"},
                single=True,
            )
            if not periods:
                raise HTTPException(404, "Period not found.")
            period = periods[0]

        with _supabase.admin() as ac:
            line_items = ac.select(
                "statement_line_items",
                filters={"period_id": f"eq.{period_id}"},
            )
            # Load any org-level overrides previously saved.
            try:
                org_overrides_rows = ac.select(
                    "org_coa_mappings_overrides",
                    filters={"org_id": f"eq.{period['org_id']}"},
                )
            except Exception:  # noqa: BLE001
                org_overrides_rows = []

        pack = get_pack("RO")  # F3.7+ will dispatch the active pack
        accounts_for_unmapped = []
        for li in (line_items or []):
            code = (li.get("ro_account_code") or "").strip()
            if not code:
                continue
            accounts_for_unmapped.append({
                "code": code,
                "name": li.get("ro_account_name") or "",
                "amount": float(li.get("amount") or 0),
            })
        unmapped = _rm.collect_unmapped_accounts(pack, accounts_for_unmapped) if pack else []

        return {
            "period_id": period_id,
            "unmapped_accounts": unmapped,
            "org_overrides": [
                {
                    "account_code": r["account_code"],
                    "standardized_bucket": r["standardized_bucket"],
                    "sign": r.get("sign", 1),
                    "coa_key": r.get("coa_key"),
                }
                for r in (org_overrides_rows or [])
            ],
            "review_state_version": "f3.4",
        }

    @router.post("/api/period/{period_id}/review/reanalyze")
    def reanalyze_with_overrides(
        period_id: str,
        payload: ReviewReanalyzeRequest,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F3.4 — Rerun assemble_statements with user-supplied
        overrides applied. Overrides come in the request body
        (transient per-session); when `propose_as_calibration_rules`
        is True they are ALSO upserted to org_coa_mappings_overrides
        for future uploads.

        Returns the post-reanalysis confidence report so the FE can
        show "Review Mode trigger went from True → False" once the
        residuals are within tolerance.
        """
        jwt = _require_jwt(authorization)
        from engine.core import review_mode as _rm
        from engine.core.country_pack_registry import get_pack
        from engine.core.upload_classifier import classify_upload
        from engine.core.confidence_engine import (
            build_confidence_report, confidence_report_to_dict,
        )

        with _supabase.per_user(jwt) as client:
            periods = client.select(
                "financial_periods",
                filters={"id": f"eq.{period_id}"},
                single=True,
            )
            if not periods:
                raise HTTPException(404, "Period not found.")
            period = periods[0]

        overrides = _rm.ReviewOverrides(
            period_id=period_id,
            account_buckets=payload.account_buckets,
            confirmed_country_code=payload.confirmed_country_code,
            confirmed_standard=payload.confirmed_standard,
            notes=payload.notes,
            propose_as_calibration_rules=payload.propose_as_calibration_rules,
        )

        with _supabase.admin() as ac:
            line_items = ac.select(
                "statement_line_items",
                filters={"period_id": f"eq.{period_id}"},
            )
            if not line_items:
                raise HTTPException(400, "Period has no line items to reanalyse.")
            org_rows = ac.select(
                "organizations",
                filters={"id": f"eq.{period['org_id']}"},
                single=True,
            )
            org = org_rows[0] if org_rows else {}

            pack_country = overrides.confirmed_country_code or "RO"
            pack = get_pack(pack_country)
            if pack is None:
                raise HTTPException(
                    400,
                    f"No country pack registered for {pack_country!r}.",
                )

            # ── Merge org-level overrides on top of user input ─────
            # This lets users layer a Review session on top of
            # previously-saved org defaults.
            merged_overrides = dict(overrides.account_buckets)
            try:
                org_overrides_rows = ac.select(
                    "org_coa_mappings_overrides",
                    filters={"org_id": f"eq.{period['org_id']}"},
                )
                for r in (org_overrides_rows or []):
                    code = r["account_code"]
                    if code not in merged_overrides:
                        merged_overrides[code] = r["standardized_bucket"]
            except Exception:  # noqa: BLE001
                pass
            merged = _rm.ReviewOverrides(
                period_id=period_id,
                account_buckets=merged_overrides,
                confirmed_country_code=overrides.confirmed_country_code,
                confirmed_standard=overrides.confirmed_standard,
                notes=overrides.notes,
            )
            override_bucket_for = _rm.apply_overrides(pack, merged)

            recovered_accounts: List[Dict[str, Any]] = []
            for li in line_items:
                code = (li.get("ro_account_code") or "").strip()
                if not code:
                    continue
                stored_bucket = li.get("bucket") or ""
                rule = override_bucket_for(code)
                bucket_override = None
                if rule and stored_bucket and stored_bucket != rule.bucket:
                    legacy = pack.persistence_bucket(rule.bucket)
                    if stored_bucket != legacy:
                        bucket_override = stored_bucket
                row = {
                    "code": code,
                    "name": li.get("ro_account_name") or "",
                    "amount": float(li.get("amount") or 0),
                }
                if bucket_override:
                    row["bucket_override"] = bucket_override
                recovered_accounts.append(row)
            for acct in recovered_accounts:
                rule = override_bucket_for(acct["code"])
                if rule and rule.sign == -1:
                    acct["amount"] = -acct["amount"]

            original_bucket_for = pack.bucket_for
            pack.bucket_for = override_bucket_for  # type: ignore[assignment]
            try:
                assembled_full = pack.assemble_statements(
                    recovered_accounts,
                    company_name=org.get("name") or "Entity",
                    currency=period.get("currency", "RON"),
                    period_label=str(period.get("period_end")) or "Period",
                    industry=(org or {}).get("industry_key"),
                )
            finally:
                pack.bucket_for = original_bucket_for  # type: ignore[assignment]

            synth = "\n".join(
                f"{a['code']} {a.get('name', '')}" for a in recovered_accounts[:200]
            )
            synth += f"\nMoneda: {period.get('currency', 'RON')}\n"
            synth_bytes = synth.encode("utf-8", errors="ignore")
            fn_hint = period.get("period_end") or ""
            classification = classify_upload(synth_bytes, fn_hint)
            report = build_confidence_report(classification, assembled_full)
            confidence_dict = confidence_report_to_dict(report)
            # F4.4-int — run fan-out routing + rebuild detection envelope
            # with full classification + routing_decision (mirrors site A).
            try:
                from engine.routing import route_with_fan_out as _route, routing_decision_dict as _rdd
                from engine.detection import build_detection_envelope as _bde2
                _routing_result = _route(
                    classification, synth_bytes, fn_hint,
                    company_name=org.get("name") or "Entity",
                    currency=period.get("currency", "RON"),
                    period_label=str(period.get("period_end")) or "Period",
                    industry=(org or {}).get("industry_key"),
                )
                # Reanalyze path doesn't populate `statements` the same way;
                # attach routing_decision to the response payload below.
                confidence_dict = confidence_dict or {}
                confidence_dict["routing_decision"] = _rdd(_routing_result)
            except Exception:  # noqa: BLE001
                logger.exception("[/api/period reanalyze] fan-out routing integration failed (non-fatal)")

            # ── F3.5 hand-off: propose as calibration rules ────────
            # Insert into `calibration_rules` with status='pending'.
            # Admin approves via /api/admin/calibration/rules/{id}/
            # approve, which copies to org_coa_mappings_overrides
            # (the engine's actual read source).
            calibration_rules_proposed = 0
            if payload.propose_as_calibration_rules and overrides.account_buckets:
                rows_to_insert: List[Dict[str, Any]] = []
                for code, bucket in overrides.account_buckets.items():
                    rows_to_insert.append({
                        "org_id": period["org_id"],
                        "coa_key": "omfp_1802",  # F3.7+ will reflect pack
                        "account_code": code,
                        "standardized_bucket": bucket,
                        "sign": 1,
                        "source": "review_mode",
                        "status": "pending",
                        "period_id": period_id,
                        "notes": overrides.notes,
                    })
                try:
                    ac.insert(
                        "calibration_rules",
                        rows_to_insert,
                        returning=False,
                    )
                    calibration_rules_proposed = len(rows_to_insert)
                except Exception:  # noqa: BLE001
                    logger.warning(
                        "[review/reanalyze] proposing to calibration_rules "
                        "failed (table may not be migrated yet): %s",
                        rows_to_insert[0] if rows_to_insert else None,
                    )

        return {
            "ok": True,
            "period_id": period_id,
            "overrides_applied": len(overrides.account_buckets),
            "merged_overrides_applied": len(merged_overrides),
            "calibration_rules_proposed": calibration_rules_proposed,
            "post_reanalysis_confidence": confidence_dict,
        }

    # ── F3.5 / F3.6 — Calibration store + admin dashboard ──────────

    def _require_engine_admin(authorization: Optional[str]) -> None:
        """Gate admin endpoints behind the ENGINE_API_TOKEN.
        Mirrors the pattern used by billing's cron endpoints — admin
        endpoints are scheduler/operator-only, never user-callable.
        """
        token = os.environ.get("ENGINE_API_TOKEN")
        if not token:
            # No token configured → no admin endpoints accessible.
            raise HTTPException(503, "Admin endpoints disabled (ENGINE_API_TOKEN unset).")
        provided = _require_jwt(authorization)
        if provided != token:
            raise HTTPException(401, "Invalid admin token.")

    @router.get("/api/admin/calibration/coverage")
    def admin_calibration_coverage(
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F3.6 — Coverage matrix for the admin dashboard.

        Returns, per country pack:
          - calibration_tier (from the pack's metadata)
          - fixture count (from `calibration_fixtures`)
          - pending-rule count (from `calibration_rules` where
            status='pending')
          - last F-A3.1 verdict (from `calibration_results`)
        """
        _require_engine_admin(authorization)
        from engine.core.country_pack_registry import all_packs

        rows: List[Dict[str, Any]] = []
        with _supabase.admin() as ac:
            # Per-pack counts — tolerate missing tables.
            for pack in all_packs():
                fixture_count = 0
                pending_count = 0
                approved_count = 0
                latest_result: Optional[Dict[str, Any]] = None
                try:
                    fixtures = ac.select(
                        "calibration_fixtures",
                        filters={"country_code": f"eq.{pack.country_code}"},
                    )
                    fixture_count = len(fixtures or [])
                except Exception:  # noqa: BLE001
                    pass
                try:
                    pending = ac.select(
                        "calibration_rules",
                        filters={"coa_key": f"eq.omfp_1802", "status": "eq.pending"},
                    )
                    pending_count = len(pending or [])
                    approved = ac.select(
                        "calibration_rules",
                        filters={"coa_key": f"eq.omfp_1802", "status": "eq.approved"},
                    )
                    approved_count = len(approved or [])
                except Exception:  # noqa: BLE001
                    pass
                rows.append({
                    "country_code": pack.country_code,
                    "country_name": pack.country_name,
                    "accounting_standard": pack.accounting_standard,
                    "pack_version": pack.pack_version,
                    "calibration_tier": pack.calibration_tier.value,
                    "fixture_count": fixture_count,
                    "pending_rule_count": pending_count,
                    "approved_rule_count": approved_count,
                    "regression_fixtures_declared": list(pack.regression_fixtures),
                })
        return {
            "packs": rows,
            "summary": {
                "total_packs": len(rows),
                "deeply_calibrated": sum(1 for r in rows if r["calibration_tier"] == "deeply_calibrated"),
                "partially_calibrated": sum(1 for r in rows if r["calibration_tier"] == "partially_calibrated"),
                "experimental": sum(1 for r in rows if r["calibration_tier"] == "experimental"),
            },
        }

    @router.get("/api/admin/calibration/rules")
    def admin_list_calibration_rules(
        status: Optional[str] = None,
        coa_key: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F3.6 — List calibration rules. Supports `status` filter
        (pending / approved / rejected) and `coa_key` filter.
        """
        _require_engine_admin(authorization)
        filters: Dict[str, str] = {}
        if status:
            filters["status"] = f"eq.{status}"
        if coa_key:
            filters["coa_key"] = f"eq.{coa_key}"
        try:
            with _supabase.admin() as ac:
                rows = ac.select(
                    "calibration_rules",
                    filters=filters,
                    order="created_at.desc",
                )
        except Exception as e:  # noqa: BLE001
            logger.warning("[admin/calibration/rules] table miss: %s", e)
            return {"rules": [], "error": "calibration_rules table not migrated yet"}
        return {"rules": rows or []}

    @router.post("/api/admin/calibration/rules/{rule_id}/approve")
    def admin_approve_calibration_rule(
        rule_id: str,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F3.6 — Approve a pending calibration rule. Side effects:
          1. Mark `calibration_rules.status='approved'`.
          2. Upsert the rule into `org_coa_mappings_overrides` so the
             engine picks it up on future uploads.
        """
        _require_engine_admin(authorization)
        try:
            with _supabase.admin() as ac:
                rows = ac.select(
                    "calibration_rules",
                    filters={"id": f"eq.{rule_id}"},
                    single=True,
                )
                if not rows:
                    raise HTTPException(404, f"Calibration rule {rule_id} not found.")
                rule = rows[0]
                if rule["status"] != "pending":
                    raise HTTPException(
                        400,
                        f"Rule status is {rule['status']!r}; only pending rules can be approved.",
                    )
                # Mark approved.
                ac.update(
                    "calibration_rules",
                    {
                        "status": "approved",
                        "approved_at": _now_iso(),
                        "updated_at": _now_iso(),
                    },
                    filters={"id": f"eq.{rule_id}"},
                )
                # Apply to org_coa_mappings_overrides — this is what
                # the engine actually reads.
                if rule.get("org_id"):
                    ac.upsert(
                        "org_coa_mappings_overrides",
                        {
                            "org_id": rule["org_id"],
                            "coa_key": rule["coa_key"],
                            "account_code": rule["account_code"],
                            "standardized_bucket": rule["standardized_bucket"],
                            "sign": rule.get("sign", 1),
                        },
                        on_conflict="org_id,coa_key,account_code",
                        returning=False,
                    )
                return {
                    "ok": True,
                    "rule_id": rule_id,
                    "applied_to_org": rule.get("org_id"),
                    "account_code": rule["account_code"],
                    "standardized_bucket": rule["standardized_bucket"],
                }
        except HTTPException:
            raise
        except Exception as e:  # noqa: BLE001
            logger.exception("[admin/calibration/rules/approve] failed")
            raise HTTPException(500, f"Approval failed: {type(e).__name__}")

    @router.post("/api/admin/calibration/rules/{rule_id}/reject")
    def admin_reject_calibration_rule(
        rule_id: str,
        reason: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F3.6 — Reject a pending calibration rule."""
        _require_engine_admin(authorization)
        try:
            with _supabase.admin() as ac:
                ac.update(
                    "calibration_rules",
                    {
                        "status": "rejected",
                        "rejection_reason": reason or "",
                        "updated_at": _now_iso(),
                    },
                    filters={"id": f"eq.{rule_id}", "status": "eq.pending"},
                )
        except Exception as e:  # noqa: BLE001
            logger.exception("[admin/calibration/rules/reject] failed")
            raise HTTPException(500, f"Rejection failed: {type(e).__name__}")
        return {"ok": True, "rule_id": rule_id, "status": "rejected"}

    @router.get("/api/admin/calibration/fixtures")
    def admin_list_fixtures(
        country_code: Optional[str] = None,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """F3.6 — List calibration fixtures. Optional country_code
        filter (e.g. 'RO' → only Romanian fixtures)."""
        _require_engine_admin(authorization)
        filters: Dict[str, str] = {}
        if country_code:
            filters["country_code"] = f"eq.{country_code}"
        try:
            with _supabase.admin() as ac:
                rows = ac.select(
                    "calibration_fixtures",
                    filters=filters,
                    order="created_at.desc",
                )
        except Exception as e:  # noqa: BLE001
            logger.warning("[admin/calibration/fixtures] table miss: %s", e)
            return {"fixtures": [], "error": "calibration_fixtures table not migrated yet"}
        return {"fixtures": rows or []}

    @router.delete("/api/period/{period_id}")
    def delete_period(
        period_id: str,
        authorization: Optional[str] = Header(None),
    ) -> Dict[str, Any]:
        """Hard-delete a period and ALL of its derivatives + soft-delete the
        attached documents. Surfaced behind the "Reset (clear period)" item
        in the Replace dropdown so the user can recover the dashboard back
        to an empty state without leaving stale rows behind.

        The actual deletion happens through the admin client (service role)
        so RLS doesn't get in the way; ownership is verified via the
        per-user client first.
        """
        jwt = _require_jwt(authorization)
        # 1. Ownership check via the user's own RLS scope.
        with _supabase.per_user(jwt) as user_client:
            visible = user_client.select(
                "financial_periods",
                filters={"id": f"eq.{period_id}"},
                single=True,
            )
            if not visible:
                raise HTTPException(404, "Period not found or not visible to you.")
            org_id = visible[0]["org_id"]

        # 2. Soft-delete every document attached to the period — keeps the
        # underlying Storage blob recoverable from "Recently deleted" for 30
        # days. (The cleanup-cron handles the hard-delete after that.)
        with _supabase.admin() as ac:
            attached = ac.select(
                "documents",
                filters={"period_id": f"eq.{period_id}"},
                columns="id",
            )
            now = _now_iso()
            for d in attached:
                ac.update(
                    "documents",
                    {"deleted_at": now, "period_id": None},
                    filters={"id": f"eq.{d['id']}"},
                )

            # 3. Hard-delete period derivatives. Order matters where foreign
            # keys exist; explicit per-table is safer than relying on cascade.
            # `alerts` is scoped by document_id (not period_id) — we handle
            # it separately via the document soft-delete loop above which
            # leaves alert rows intact under the soft-deleted document.
            for table in (
                "statement_line_items",
                "calculated_metrics",
                "briefings",
                "valuations",
                "user_valuation_assumptions",
            ):
                try:
                    ac.delete(table, filters={"period_id": f"eq.{period_id}"})
                except Exception:  # noqa: BLE001
                    logger.exception("[delete_period] cascade delete failed on %s", table)
            # Recommendations are org-scoped (not period-scoped) but the period
            # they were generated for is gone; safest to leave them — re-running
            # a future upload will overwrite. (Don't accidentally wipe other
            # periods' recommendations.)
            ac.delete("financial_periods", filters={"id": f"eq.{period_id}"})

        return {
            "ok": True,
            "period_id": period_id,
            "documents_soft_deleted": len(attached),
        }

    return router
