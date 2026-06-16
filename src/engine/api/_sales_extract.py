"""Sales-dataset extraction — XLSX → sku_lines → sku_aggregates.

The earlier path for sku-scope uploads asked Opus to summarize the workbook
into a briefing + a handful of synthetic category roll-up rows. That's the
wrong abstraction for trading-analysis files which already have one row per
(sku × channel × client) at native granularity. This module reads the
spreadsheet rows directly via openpyxl, maps columns to a normalized schema,
streams them into sku_lines, then rolls up to sku_aggregates.

Pipeline branch: when a document's scope is 'sku' AND the workbook has
a recognizable sales shape (column synonyms match), `extract_sales_dataset`
takes over from the generic LLM-briefing path.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any, Dict, Iterator, List, Optional, Tuple

from openpyxl import load_workbook  # type: ignore


# ─── Column synonyms ────────────────────────────────────────────────────────
# Each canonical field has a list of regexes; first match wins. Headers are
# normalized (lowercased + stripped) before matching. Order in the list
# expresses preference when multiple variants might match.

COLUMN_SYNONYMS: Dict[str, List[str]] = {
    "channel": [r"^canal$", r"^channel$"],
    "client_type": [r"^tip[\s_]client$", r"^client[\s_]type$"],
    "client_parent": [r"^client[_\s]?parinte$", r"^client[\s_]parent$", r"^cliente?$", r"^customer$"],
    "business_unit": [r"^bu$", r"^business[_\s]unit$", r"^div(?:erse)?$"],
    "category": [r"^categ[_\s]?pr$", r"^categori[ea]?$", r"^categ$", r"^category$", r"^specie$"],
    "brand": [r"^brand$", r"^marca$", r"^marque$"],
    "product_name": [r"^denumire[_\s]?produs$", r"^denumire$", r"^product[\s_]name$", r"^sku$", r"^article$", r"^product$"],
    "pack_size": [r"^pack[\s_]?size$", r"^pack$", r"^cantitate$"],
    "volume_kg": [r"^volume[\s_]?kg$", r"^volum[\s_]?kg$", r"^sold[\s_]?in[\s_]?kg$"],
    "volume_tons": [r"^volume\s*\(to\)$", r"^volume[\s_]?to$", r"^volum[\s_]?to$", r"^volume$"],
    "gross_revenue": [r"^gross[\s_]revenue$", r"^vanzari[\s_]brute$"],
    "niv_turnover": [
        r"^niv[\s_]turnover.*",
        r"^niv[\s_]\(?\s*krn?\s*\)?$",
        r"^niv$",
        r"^net[\s_]invoiced.*",
    ],
    "nip_revenue": [r"^nip$", r"^nip[\s_]revenue$", r"^net[\s_]in[\s_]pocket$"],
    "raw_materials": [r"^raw[\s_]materials?$", r"^materii[\s_]prime$"],
    "pack_materials": [r"^pack[\s_]materials?$", r"^ambalaje$"],
    "secondary_pack_materials": [r"^secondary[\s_]pack.*", r"^ambalaje[\s_]secundare$"],
    "direct_material_cost": [r"^direct[\s_]material[\s_]cost$", r"^cost[\s_]material[\s_]direct$"],
    "transportation_cost": [r"^transport.*$", r"^transportation[\s_]cost$"],
    "conversion_cost": [r"^conversion[\s_]cost$", r"^cost[\s_]conversie$"],
    "warehousing_cost": [r"^warehousing.*$", r"^depozitare$"],
    "environment_tax": [r"^environment[\s_]tax$", r"^env[\s_]tax$", r"^taxa[\s_]de[\s_]mediu$"],
    "depreciation": [r"^depreciation$", r"^amortizare$"],
    "direct_margin": [r"^direct[\s_]margin$", r"^marja[\s_]directa$"],
    "gm_krn": [
        r"^gm\s*\(?\s*krn?\s*\)?$",
        r"^gm2[\s_]?with[\s_]?dep$",
        r"^gross[\s_]margin$",
        r"^gm$",
        r"^profit$",
    ],
    "gm_pct": [r"^gm2[\s_]?pct$", r"^gm[\s_]?pct$", r"^gm[\s_]?%$", r"^%\s*gm$"],
    # Inventory value + COGS — OPTIONAL columns introduced for per-SKU/
    # per-category DIO and the company-level CCC roll-up. Backward
    # compatible: existing files without these columns parse identically
    # and yield `null` DIO (rendered honestly as "not available" in the
    # UI, never as 0). Romanian variants accepted to match the rest of
    # the parser's tolerance.
    "inventory_value": [
        r"^inventory[\s_]?value$",
        r"^inventory$",
        r"^stock[\s_]?value$",
        r"^valoare[\s_]?stoc$",
        r"^stoc[\s_]?valoare$",
        r"^stocuri[\s_]?valoare$",
    ],
    "cogs": [
        r"^cogs$",
        r"^cost[\s_]?of[\s_]?goods[\s_]?sold$",
        r"^cost[\s_]?marfuri[\s_]?vandute$",
        r"^cmv$",
        r"^cost[\s_]?marfa$",
    ],
    # Pre-computed DIO column — some operator workbooks already include
    # the metric ready-to-go (e.g. "DIO days" with values like 69, 60,
    # 90, 180). Accept these directly. When present, `aggregate_sku_lines`
    # uses the explicit value instead of computing DIO from
    # inventory_value / COGS — that fallback path remains for files
    # that only provide the raw inputs. Pattern set covers the common
    # spellings: "DIO", "DIO days", "DIO (days)", "Days inventory
    # outstanding", "days inventory on hand", plus Romanian variants.
    "dio_days": [
        r"^dio$",
        r"^dio[\s_]?days$",
        r"^dio\s*\(?\s*days?\s*\)?$",
        r"^days?[\s_]?inventory[\s_]?outstanding$",
        r"^days?[\s_]?inventory[\s_]?on[\s_]?hand$",
        r"^days?[\s_]?inventory$",
        r"^days?[\s_]?in[\s_]?stock$",
        r"^zile[\s_]?stoc$",
        r"^durata[\s_]?stoc$",
    ],
}


def _normalize_header(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _normalize_category(s: Any) -> str:
    """Normalise a category string for comparison: uppercase, strip
    diacritics, collapse internal whitespace. Used to match SKU.category
    against the DIO sheet's category column ("LEGUME CONSERVATE",
    "MURATURI" etc.) — the analyst's free-text labels in either sheet
    can have stray accents or doubled spaces."""
    import unicodedata
    if s is None:
        return ""
    nfkd = unicodedata.normalize("NFD", str(s))
    stripped = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", stripped.upper().strip())


def extract_category_dio_from_analysis_sheet(xlsx_bytes: bytes) -> Dict[str, float]:
    """Read the "Analysis" sheet and return {category: dio_days} from
    the operator-set DIO/DSO/DPO/CCC block (typically cols 60-63 in the
    Trading_analysis_* workbooks). These are the realistic operational
    targets the analyst chose (e.g. LEGUME=60, JELEURI=180) — vastly
    different from, and preferred over, the extreme computed values in
    the standalone DIO sheet's Region 2 (LEGUME=73.5, JELEURI=1001.7).

    Detection: the header row near the top has 'DIO' / 'DSO' / 'DPO' /
    'CCC' (case-insensitive) in adjacent cells. We walk the next ~40
    rows pulling (category-name, dio-days) pairs whenever the row has
    a category label in col B/C and a number under the DIO column.

    Returns {} when no Analysis sheet exists or the structure doesn't
    match — caller falls back to `extract_category_dio` (DIO sheet R2).
    """
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception:
        return {}
    # Case-insensitive search for an "Analysis" sheet.
    target = next((s for s in wb.sheetnames if s.strip().upper() == "ANALYSIS"), None)
    if not target:
        return {}
    ws = wb[target]

    # ── Locate the DIO column index by scanning the first 6 rows for
    # the literal header "DIO". The header is typically on row 2 in the
    # known fixture (Trading_analysis_YTDMar_26), at column index 60,
    # alongside 'DSO', 'DPO', 'CCC'. We scan defensively so a future
    # column-layout shift doesn't silently break the extraction.
    dio_col_idx: Optional[int] = None
    label_col_idx: Optional[int] = None
    header_row_idx: Optional[int] = None
    sniff_rows: List[List[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 8:
            break
        sniff_rows.append(list(row))
    for row_idx, row in enumerate(sniff_rows):
        for col_idx, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().upper() == "DIO":
                dio_col_idx = col_idx
                header_row_idx = row_idx
                break
        if dio_col_idx is not None:
            break
    if dio_col_idx is None or header_row_idx is None:
        return {}
    # Category labels in the Analysis sheet sit in column B (index 1).
    # Defend against a layout where they live in column A or C by
    # scanning for the first non-empty string column in the next 5
    # data rows after the header.
    for col_candidate in (1, 0, 2):
        sample_rows = list(ws.iter_rows(
            min_row=header_row_idx + 2 + 1,  # 1-indexed; +2 for header rows; +1 for the data row
            max_row=header_row_idx + 8,
            values_only=True,
        ))
        if any(
            isinstance(r[col_candidate], str) and r[col_candidate].strip()
            for r in sample_rows
            if len(r) > col_candidate
        ):
            label_col_idx = col_candidate
            break
    if label_col_idx is None:
        label_col_idx = 1  # fall back to canonical layout

    # Pull (category, dio_days) pairs from the rows immediately after
    # the header. Walk up to ~40 rows; stop when we hit a long run of
    # empty category cells (end of the block).
    result: Dict[str, float] = {}
    blank_streak = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx + 1:
            continue
        if i > header_row_idx + 45:
            break
        label = row[label_col_idx] if len(row) > label_col_idx else None
        dio = row[dio_col_idx] if len(row) > dio_col_idx else None
        if label is None or str(label).strip() == "":
            blank_streak += 1
            if blank_streak > 5:
                break
            continue
        blank_streak = 0
        if not isinstance(dio, (int, float)) or dio <= 0:
            continue
        # Skip subtotal/total rollups — we want LEAF categories so the
        # downstream SKU join finds a match. "TOTAL", "DIVERSE",
        # "PESTE" are the rollup labels; "CATEGORY" + "PESTE / CATEGORY"
        # are phantom header artifacts that surface when openpyxl reads
        # merged cells in the Analysis sheet (the visible label is the
        # rollup name like "DIVERSE", but the underlying cell value is
        # the column header "Category").
        norm = _normalize_category(str(label))
        if (
            norm in {"TOTAL", "DIVERSE", "PESTE", "CATEGORY"}
            or "CATEGORY" in norm
        ):
            continue
        result[norm] = float(dio)
    return result


def extract_category_dio(xlsx_bytes: bytes) -> Dict[str, float]:
    """Read the standalone "DIO" sheet and return {category_normalised:
    days} from REGION 2 (the per-category DIO breakdown).

    Source-file shape: REGION 2 sits below REGION 1 (category inventory
    snapshot) and follows the pattern: column A empty, column B holds
    the category name (string), column D holds DIO days (number).
    Empty / header rows simply don't match the pattern and get skipped.

    Returns {} when no "DIO" sheet exists OR no rows match — that's
    the graceful-missing path; downstream `aggregate_sku_lines` falls
    back to (a) per-SKU explicit DIO column, (b) inventory_value ÷
    cogs × 365 computation, (c) None.

    PREFERENCE ORDER (set in the pipeline orchestrator, not here):
      1. Analysis sheet operator-set DIO (cols 60-63 typically) —
         realistic targets the analyst chose (e.g. LEGUME=60).
      2. DIO sheet REGION 2 — computed actuals (e.g. LEGUME=73.5,
         JELEURI=1001.7) that can be extreme outliers. Used as a
         fallback when the Analysis sheet doesn't carry a value.
    """
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception:
        return {}
    if "DIO" not in wb.sheetnames:
        # Case-insensitive search for a sheet named DIO with any whitespace.
        target = next((s for s in wb.sheetnames if s.strip().upper() == "DIO"), None)
        if not target:
            return {}
        ws = wb[target]
    else:
        ws = wb["DIO"]

    result: Dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        # Need at least 4 columns to inspect (A, B, C, D).
        if not row or len(row) < 4:
            continue
        col_a, col_b, _col_c, col_d = row[0], row[1], row[2], row[3]
        # REGION 2 row signature: col A empty, col B is a non-empty
        # string category, col D is a positive number.
        if col_a is None and isinstance(col_b, str) and isinstance(col_d, (int, float)):
            if col_d <= 0:
                continue
            cat = _normalize_category(col_b)
            if not cat:
                continue
            # When the same category appears more than once in the sheet,
            # the LAST occurrence wins (matches the analyst's "most
            # recent" calculation; REGION 2 is below REGION 1, so this
            # naturally picks the computed-DIO not the inventory row).
            result[cat] = float(col_d)
    return result


def _match_synonym(header: str, patterns: List[str]) -> bool:
    for p in patterns:
        if re.search(p, header):
            return True
    return False


def _build_column_map(headers: List[Any]) -> Dict[str, int]:
    """Map canonical field name → column index. Required: product_name +
    niv_turnover. Others optional (file may omit cost breakdown columns).
    """
    normalized = [_normalize_header(h) for h in headers]
    col_map: Dict[str, int] = {}
    for canonical, patterns in COLUMN_SYNONYMS.items():
        for idx, h in enumerate(normalized):
            if not h:
                continue
            if _match_synonym(h, patterns):
                col_map[canonical] = idx
                break
    return col_map


# ─── Sheet selection ────────────────────────────────────────────────────────

YTD_PATTERN = re.compile(r"\b(ytd|q[1-4]|h[12]|fy)\b", re.IGNORECASE)


def _pick_data_sheet(sheet_names: List[str]) -> str:
    """Prefer YTD/Q* sheets; fall back to the first non-summary sheet, then
    the first sheet."""
    for name in sheet_names:
        if YTD_PATTERN.search(name):
            return name
    for name in sheet_names:
        if not re.search(r"^(summary|index|toc|notes?)$", name, re.IGNORECASE):
            return name
    return sheet_names[0]


def _find_header_row(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    """Some XLSX files have title rows above the actual headers. Scan the
    first 12 rows; return the index of the row that yields the richest
    column map (i.e. matches the most canonical fields)."""
    best_idx = 0
    best_map: Dict[str, int] = {}
    for i in range(min(12, len(rows))):
        candidate = _build_column_map(rows[i])
        if len(candidate) > len(best_map):
            best_idx = i
            best_map = candidate
    return best_idx, best_map


# ─── Detection ──────────────────────────────────────────────────────────────


def is_sales_dataset(xlsx_bytes: bytes) -> Tuple[bool, Optional[Dict[str, Any]]]:
    """Returns (is_sales, info). Info includes sheet name + column map +
    inferred period label when sales-shaped. The pipeline calls this before
    deciding which branch to take; a False here keeps the existing generic
    LLM-summary path."""
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception:
        return False, None
    if not wb.sheetnames:
        return False, None
    sheet_name = _pick_data_sheet(wb.sheetnames)
    ws = wb[sheet_name]
    # Read up to first 14 rows to find headers.
    sniff: List[List[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        sniff.append(list(row))
        if i >= 13:
            break
    if not sniff:
        return False, None
    header_idx, col_map = _find_header_row(sniff)
    required = {"product_name", "niv_turnover"}
    if not required.issubset(col_map.keys()):
        return False, None
    return True, {
        "sheet_name": sheet_name,
        "header_row_index": header_idx,
        "column_map": col_map,
        "period_label": _infer_period_label(sheet_name),
    }


def _infer_period_label(sheet_name: str) -> str:
    # "YTD October '25" / "YTD Oct '25" / "Q4 2025" — normalize.
    m = re.search(r"(YTD|Q[1-4]|H[12]|FY)[^\d\w]*([A-Za-z]{3,9})?[^\d]*'?(\d{2,4})?", sheet_name, re.IGNORECASE)
    if m:
        return " ".join(part for part in [m.group(1).upper(), m.group(2), m.group(3) and f"'{m.group(3)[-2:]}"] if part)
    return sheet_name


# ─── Row streaming ──────────────────────────────────────────────────────────


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        # Romanian decimal: "1.234,56" → 1234.56
        s = str(v).strip().replace("\xa0", "")
        if "," in s and s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
        return float(s)
    except (TypeError, ValueError):
        return None


def stream_sales_rows(xlsx_bytes: bytes, info: Dict[str, Any]) -> Iterator[Dict[str, Any]]:
    """Yield one dict per data row from the sheet identified by `info`.
    Skips rows where product_name is empty. Coerces numbers, handles
    Romanian decimal separators."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb[info["sheet_name"]]
    col_map: Dict[str, int] = info["column_map"]
    header_idx: int = info["header_row_index"]

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        row_list = list(row)
        prod_idx = col_map.get("product_name")
        if prod_idx is None or prod_idx >= len(row_list):
            continue
        product = row_list[prod_idx]
        if not product or not str(product).strip():
            continue
        rec: Dict[str, Any] = {}
        for field, idx in col_map.items():
            if idx >= len(row_list):
                continue
            val = row_list[idx]
            if field in ("product_name", "brand", "category", "channel", "client_type",
                          "client_parent", "business_unit", "pack_size"):
                rec[field] = str(val).strip() if val is not None else None
            else:
                rec[field] = _num(val)
        # Derived: volume_tons + niv_krn + gm_krn. Files vary in which
        # version they provide; we normalize so downstream code has a
        # single source of truth.
        if rec.get("volume_tons") is None and rec.get("volume_kg") is not None:
            rec["volume_tons"] = (rec["volume_kg"] or 0) / 1000.0
        if rec.get("niv_krn") is None and rec.get("niv_turnover") is not None:
            rec["niv_krn"] = rec["niv_turnover"]
        if rec.get("gm_krn") is None and rec.get("direct_margin") is not None:
            rec["gm_krn"] = rec["direct_margin"]
        # gm_pct fallback: gm_krn / niv_krn
        if rec.get("gm_pct") is None and rec.get("gm_krn") and rec.get("niv_krn"):
            try:
                rec["gm_pct"] = float(rec["gm_krn"]) / float(rec["niv_krn"])
            except (TypeError, ValueError, ZeroDivisionError):
                pass
        yield rec


# ─── Roll-up: lines → aggregates ────────────────────────────────────────────


@dataclass
class _Agg:
    product_name: str
    brand: Optional[str] = None
    category: Optional[str] = None
    volume_tons: float = 0.0
    niv_krn: float = 0.0
    gm_krn: float = 0.0
    # Optional new fields — DIO inputs. Tracked as Optional sums so that
    # a product with NO inventory/COGS data anywhere across its lines
    # stays `None` (→ DIO not-available) rather than collapsing to 0
    # (which would be a fabricated "perfect inventory turnover" reading).
    # Any line with a value bumps these to a concrete float; lines
    # without a value are left out of the sum.
    inventory_krn: Optional[float] = None
    cogs_krn: Optional[float] = None
    inventory_lines_with_data: int = 0
    cogs_lines_with_data: int = 0
    # Pre-computed DIO from the source workbook (when a "DIO days"
    # column exists). When the column is present we average it across
    # the lines for this SKU; when absent we leave it None and fall
    # back to computing DIO from inventory_krn / cogs_krn below.
    dio_explicit_sum: float = 0.0
    dio_explicit_count: int = 0
    line_row_count: int = 0
    channels: set = None  # type: ignore[assignment]
    clients: set = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.channels is None:
            self.channels = set()
        if self.clients is None:
            self.clients = set()


def aggregate_sku_lines(
    lines: List[Dict[str, Any]],
    category_dio: Optional[Dict[str, float]] = None,
) -> List[Dict[str, Any]]:
    """Roll lines up to one row per product_name. Returns list of dicts
    ready for insertion into sku_aggregates (without dataset_id / org_id —
    the caller stamps those).

    DIO resolution order (first non-null wins):
      1. Pre-computed per-SKU "DIO days" column in the source line rows
      2. inventory_value ÷ cogs × 365 when both are present
      3. Category-level DIO from the standalone "DIO" sheet
         (`category_dio` dict). Matches sku.category (normalised) with
         a " FILE" suffix-strip fallback so "MACROU FILE" picks up
         "MACROU" when no exact match exists.
      4. None — rendered as "not available" in the UI, NEVER 0.

    The category fallback (3) lets the engine populate DIO for every SKU
    in a workbook whose analyst only provided the per-category DIO sheet
    (no per-SKU inventory_value/COGS columns). All SKUs in the same
    category inherit the category DIO — that's by design and matches
    how the source analyst computed it.
    """
    cat_dio_map = category_dio or {}
    bucket: Dict[str, _Agg] = {}
    for line in lines:
        name = (line.get("product_name") or "").strip()
        if not name:
            continue
        agg = bucket.get(name)
        if agg is None:
            agg = _Agg(product_name=name, brand=line.get("brand"), category=line.get("category"))
            bucket[name] = agg
        agg.volume_tons += line.get("volume_tons") or 0.0
        agg.niv_krn += line.get("niv_krn") or 0.0
        agg.gm_krn += line.get("gm_krn") or 0.0
        # Inventory & COGS — only accumulate when the line actually has
        # a value. A None on either side keeps that side's sum at None
        # (preserves "not provided" semantics) unless and until some
        # later line for this product provides a real number.
        inv = line.get("inventory_value")
        if inv is not None:
            try:
                agg.inventory_krn = (agg.inventory_krn or 0.0) + float(inv)
                agg.inventory_lines_with_data += 1
            except (TypeError, ValueError):
                pass
        cog = line.get("cogs")
        if cog is not None:
            try:
                agg.cogs_krn = (agg.cogs_krn or 0.0) + float(cog)
                agg.cogs_lines_with_data += 1
            except (TypeError, ValueError):
                pass
        # Explicit DIO (when the source workbook ships pre-computed days).
        # Accumulate the sum + count so we can take the mean per SKU at
        # emit time. Negatives and zeros are kept (a SKU that genuinely
        # sells out instantly should read 0, not be silently dropped).
        dio_explicit = line.get("dio_days")
        if dio_explicit is not None:
            try:
                agg.dio_explicit_sum += float(dio_explicit)
                agg.dio_explicit_count += 1
            except (TypeError, ValueError):
                pass
        agg.line_row_count += 1
        ch = line.get("channel")
        if ch:
            agg.channels.add(ch)
        cl = line.get("client_parent")
        if cl:
            agg.clients.add(cl)
        if not agg.brand and line.get("brand"):
            agg.brand = line.get("brand")
        if not agg.category and line.get("category"):
            agg.category = line.get("category")

    rows: List[Dict[str, Any]] = []
    for a in bucket.values():
        gm_pct = (a.gm_krn / a.niv_krn) if a.niv_krn else 0.0
        # DIO resolution order:
        #   1. Pre-computed DIO column in the source workbook ("DIO" /
        #      "DIO days" / "Days inventory outstanding" etc.) — when
        #      present, we average it across the lines for this SKU
        #      (some workbooks repeat the same DIO on every line, some
        #      vary it per period; the mean is the honest summary).
        #   2. Computed DIO = (Inventory ÷ COGS) × 365 — fallback for
        #      workbooks that ship only the raw inputs.
        #   3. None — both paths absent, rendered as "not available".
        # Order matters: an explicit DIO column reflects operator
        # judgement (e.g. seasonal adjustment, period basis) that the
        # naive (inv ÷ cogs × 365) computation can't reproduce.
        dio: Optional[float] = None
        if a.dio_explicit_count > 0:
            try:
                dio = round(a.dio_explicit_sum / a.dio_explicit_count, 2)
            except (TypeError, ValueError, ZeroDivisionError):
                dio = None
        elif (
            a.inventory_krn is not None
            and a.cogs_krn is not None
            and a.cogs_krn > 0
        ):
            try:
                dio = round((a.inventory_krn / a.cogs_krn) * 365.0, 2)
            except (TypeError, ValueError, ZeroDivisionError):
                dio = None
        # Step 3 fallback: category-level DIO from the standalone "DIO"
        # sheet. Try the SKU's category as-is; if no match, strip a
        # trailing " FILE" (Romanian for fillet — same product family,
        # e.g. "MACROU FILE" inherits "MACROU"'s DIO). Last resort the
        # SKU stays at None (not 0) per the no-fabrication rule.
        if dio is None and cat_dio_map:
            cat = _normalize_category(a.category or "")
            if cat and cat in cat_dio_map:
                dio = round(cat_dio_map[cat], 2)
            elif cat:
                base = re.sub(r"\s+FILE$", "", cat).strip()
                if base and base in cat_dio_map:
                    dio = round(cat_dio_map[base], 2)
        rows.append({
            "product_name": a.product_name,
            "brand": a.brand,
            "category": a.category,
            "volume_tons": round(a.volume_tons, 4),
            "niv_krn": round(a.niv_krn, 2),
            "gm_krn": round(a.gm_krn, 2),
            "gm_pct": round(gm_pct, 6),
            # DIO inputs + output. Persisted alongside the existing
            # aggregates so the FE can render per-SKU DIO without a
            # second roundtrip and roll up category/company aggregates
            # honestly (covered rows only).
            "inventory_value_krn": (
                round(a.inventory_krn, 2) if a.inventory_krn is not None else None
            ),
            "cogs_krn": (
                round(a.cogs_krn, 2) if a.cogs_krn is not None else None
            ),
            "days_inventory_on_hand": dio,
            "line_row_count": a.line_row_count,
            "channels_present": sorted(a.channels),
            "clients_present": sorted(a.clients),
        })
    return rows
